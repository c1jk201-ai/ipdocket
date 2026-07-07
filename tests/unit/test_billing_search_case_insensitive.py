from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def _seed_basic_invoice(
    app,
    *,
    client_name: str = "ALPHA CLIENT",
    invoice_number: str = "INV-ALPHA-001",
    notes: str = "UPPERCASE NOTE",
) -> None:
    from app.blueprints.billing_invoices.db import get_db, init_db

    with app.app_context():
        init_db()
        conn = get_db()
        try:
            conn.execute(
                """
                INSERT INTO business_profile (id, name, currency, vat_rate, next_invoice_no)
                VALUES (?, ?, ?, ?, ?)
                """,
                (1, "BP-ALPHA", "USD", 10.0, 1),
            )
            conn.execute(
                "INSERT INTO clients (id, name, email) VALUES (?, ?, ?)",
                (1, client_name, "alpha@example.com"),
            )
            conn.execute(
                """
                INSERT INTO invoices (
                    id,
                    client_id,
                    business_profile_id,
                    number,
                    issue_date,
                    due_date,
                    status,
                    billing_status,
                    payment_status,
                    currency,
                    subtotal,
                    tax,
                    total,
                    subtotal_minor,
                    tax_minor,
                    total_minor,
                    notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    1,
                    1,
                    1,
                    invoice_number,
                    "2026-01-15",
                    "2026-02-15",
                    "sent",
                    "sent",
                    "unpaid",
                    "USD",
                    1000,
                    100,
                    1100,
                    1000,
                    100,
                    1100,
                    notes,
                ),
            )
            conn.commit()
        finally:
            conn.close()


def _seed_bank_activity_transaction(app, *, remark1: str = "ALPHA PAYMENT") -> None:
    from app.blueprints.billing_invoices.db import _actual_table_name, get_db, init_db

    with app.app_context():
        init_db()
        conn = get_db()
        try:
            tx_table = _actual_table_name("bank_transactions")
            conn.execute(
                f"""
                INSERT INTO {tx_table} (
                    tid, bank_code, account_number, trdate, trdt, acc_in, acc_out, balance,
                    remark1, remark2, remark3, memo
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "tid-alpha-1",
                    "004",
                    "1234567890",
                    "20260309",
                    "20260309120000",
                    150000,
                    0,
                    700000,
                    remark1,
                    "MATCH",
                    "",
                    "UPPER MEMO",
                ),
            )
            conn.commit()
        finally:
            conn.close()


def test_billing_clients_list_search_is_case_insensitive(
    admin_client, app, clean_legacy_invoice_db
):
    _seed_basic_invoice(app)

    response = admin_client.get("/accounting/invoice-system/clientsNewq=alpha")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "ALPHA CLIENT" in html


def test_billing_invoices_list_search_is_case_insensitive(
    admin_client, app, clean_legacy_invoice_db
):
    _seed_basic_invoice(app, invoice_number="INV-MIXED-CASE-001")

    response = admin_client.get("/accounting/invoice-system/invoicesNewq=inv-mixed-case-001")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "INV-MIXED-CASE-001" in html


def test_billing_aging_search_matches_invoice_number_case_insensitively(
    admin_client, app, clean_legacy_invoice_db
):
    _seed_basic_invoice(app, invoice_number="AGING-UPPER-001")

    response = admin_client.get(
        "/accounting/invoice-system/aging/agingNewq=aging-upper-001&as_of=2026-02-20"
    )

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "ALPHA CLIENT" in html


def test_billing_taxinvoice_drafts_redirect_to_tax_issue_queue(
    admin_client, app, clean_legacy_invoice_db
):
    response = admin_client.get(
        "/accounting/invoice-system/taxinvoice/draftsNewq=draft-upper-001",
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    assert "/accounting/invoice-system/invoices/tax_issue" in (
        response.headers.get("Location") or ""
    )


def test_bank_activity_export_search_is_case_insensitive(admin_client, app, clean_legacy_invoice_db):
    _seed_bank_activity_transaction(app, remark1="ALPHA PAYMENT")

    response = admin_client.get(
        "/accounting/invoice-system/bank_activity/export"
        "?sdate=20260301"
        "&edate=20260331"
        "&accounts=004%7C1234567890"
        "&tradeType=I"
        "&searchString=alpha"
        "&order=D"
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook.active
    assert sheet.max_row == 2
    assert sheet["I2"].value == "ALPHA PAYMENT"
