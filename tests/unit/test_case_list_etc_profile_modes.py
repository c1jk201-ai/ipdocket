from __future__ import annotations

import io
import uuid

from bs4 import BeautifulSoup
from openpyxl import load_workbook


def _listed_refs(html: str) -> set[str]:
    soup = BeautifulSoup(html, "html.parser")
    return {
        link.get_text(" ", strip=True)
        for link in soup.select("#caseListTableBody tr td:nth-of-type(3) a")
    }


def _add_overview_row(
    db_session,
    *,
    matter_id: str,
    our_ref: str,
    right_group: str,
    matter_type: str,
    right_name: str,
) -> None:
    from app.models.ip_records import VMatterOverview

    db_session.add(
        VMatterOverview(
            matter_id=matter_id,
            our_ref=our_ref,
            right_group=right_group,
            matter_type=matter_type,
            right_name=right_name,
            entered_at="2026-03-18",
        )
    )
    db_session.commit()


def test_madrid_list_uses_class_column_mode(admin_client, db_session):
    from app.models.ip_records import Matter, MatterCustomField, VMatterOverview
    from app.services.case.case_parameter_service import CaseParameterService

    mid = uuid.uuid4().hex
    ref = f"TEST-MADRID-LIST-{uuid.uuid4().hex[:8].upper()}"

    db_session.add(
        Matter(
            matter_id=mid,
            our_ref=ref,
            right_group="ETC",
            matter_type="MADRID",
            is_deleted=False,
        )
    )
    db_session.add(
        VMatterOverview(
            matter_id=mid,
            our_ref=ref,
            right_group="ETC",
            matter_type="MADRID",
            entered_at="2026-03-18",
        )
    )
    db_session.add(
        MatterCustomField(
            matter_id=mid,
            namespace=CaseParameterService.get_namespace("ETC", "MADRID"),
            data={"application_classes": "35"},
        )
    )
    db_session.commit()

    resp = admin_client.get(f"/case/etc/madridNewq={ref}", follow_redirects=True)

    assert resp.status_code == 200
    html = resp.data.decode("utf-8")
    assert ref in _listed_refs(html)
    assert "35" in html


def test_madrid_list_excludes_plain_out_trademark_without_marker(admin_client, db_session):
    ref = f"TEST-PLAIN-TM-{uuid.uuid4().hex[:8].upper()}"
    _add_overview_row(
        db_session,
        matter_id=uuid.uuid4().hex,
        our_ref=ref,
        right_group="OUT",
        matter_type="TRADEMARK",
        right_name="Plain foreign trademark",
    )

    resp = admin_client.get(f"/case/etc/madrid?q={ref}", follow_redirects=True)

    assert resp.status_code == 200
    assert ref not in _listed_refs(resp.data.decode("utf-8"))


def test_hague_list_excludes_plain_out_design_without_marker(admin_client, db_session):
    ref = f"TEST-PLAIN-DESIGN-{uuid.uuid4().hex[:8].upper()}"
    _add_overview_row(
        db_session,
        matter_id=uuid.uuid4().hex,
        our_ref=ref,
        right_group="OUT",
        matter_type="DESIGN",
        right_name="Plain foreign design",
    )

    resp = admin_client.get(f"/case/etc/hague?q={ref}", follow_redirects=True)

    assert resp.status_code == 200
    assert ref not in _listed_refs(resp.data.decode("utf-8"))


def test_copyright_list_excludes_plain_misc_without_marker(admin_client, db_session):
    ref = f"TEST-PLAIN-MISC-{uuid.uuid4().hex[:8].upper()}"
    _add_overview_row(
        db_session,
        matter_id=uuid.uuid4().hex,
        our_ref=ref,
        right_group="ETC",
        matter_type="MISC",
        right_name="Plain miscellaneous matter",
    )

    resp = admin_client.get(f"/case/etc/copyright?q={ref}", follow_redirects=True)

    assert resp.status_code == 200
    assert ref not in _listed_refs(resp.data.decode("utf-8"))


def test_case_list_xlsx_uses_trademark_classes_for_madrid_in_mixed_mode(app):
    from app.blueprints.case.routes.list import _case_list_xlsx_response
    from app.models.ip_records import VMatterOverview

    matter_id = uuid.uuid4().hex
    case = VMatterOverview(
        matter_id=matter_id,
        our_ref="TEST-MADRID-XLSX",
        right_group="ETC",
        matter_type="MADRID",
    )
    extras = {
        matter_id: {
            "display_red": "",
            "display_blue": "",
            "trademark_classes": "09, 35",
            "inventor_name": "Inventor Name",
            "client_name": "",
            "proposal_title": "",
            "applicant_name": "",
            "application_no": "",
            "application_date": "",
        }
    }

    with app.app_context():
        resp = _case_list_xlsx_response(
            cases=[case],
            case_extras=extras,
            inventor_column_mode="mixed",
            inventor_column_label="INVENTER/CLASS",
            filter_title="test",
            export_scope="page",
            total=1,
        )

    wb = load_workbook(io.BytesIO(resp.get_data()))
    ws = wb.active
    assert ws["G1"].value == "INVENTER/CLASS"
    assert ws["G2"].value == "09, 35"
