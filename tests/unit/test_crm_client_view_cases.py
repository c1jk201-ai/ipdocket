import json
import uuid


def test_crm_case_status_export_zips_one_workbook_per_selected_client(admin_client, db_session):
    import io
    import zipfile

    from openpyxl import load_workbook

    from app.models.client import Client
    from app.models.ip_records import Matter, MatterCustomField

    client_a = Client(name="TextA", extra={})
    client_b = Client(name="TextB", extra={})
    client_without_cases = Client(name="Text", extra={})
    db_session.add_all([client_a, client_b, client_without_cases])
    db_session.commit()

    matter_a_id = uuid.uuid4().hex
    db_session.add(
        Matter(
            matter_id=matter_a_id,
            our_ref="CRM-EXPORT-A-1",
            your_ref="YA-1",
            right_name="TextA Text",
            right_group="DOM",
            matter_type="PATENT",
            status_red="Text",
            status_blue="Text Text Text",
            is_deleted=False,
        )
    )
    db_session.add(
        MatterCustomField(
            matter_id=matter_a_id,
            namespace="domestic_patent",
            data={
                "client_id": str(client_a.id),
                "client_name": client_a.name,
                "applicant_name": "TextA",
                "application_no": "10-2026-0000001",
                "application_date": "2026-05-01",
            },
        )
    )

    matter_b_id = uuid.uuid4().hex
    db_session.add(
        Matter(
            matter_id=matter_b_id,
            our_ref="CRM-EXPORT-B-1",
            right_name="TextB Text",
            right_group="INC",
            matter_type="TRADEMARK",
            status_blue="Text Text Text",
            is_deleted=False,
        )
    )
    db_session.add(
        MatterCustomField(
            matter_id=matter_b_id,
            namespace="incoming_trademark",
            data={"client_id": str(client_b.id), "client_name": client_b.name},
        )
    )
    db_session.commit()

    resp = admin_client.post(
        "/crm/clients/case-status-export",
        data={
            "export_scope": "selected",
            "client_ids": [
                str(client_a.id),
                str(client_b.id),
                str(client_without_cases.id),
            ],
        },
    )

    assert resp.status_code == 200
    assert resp.mimetype == "application/zip"

    with zipfile.ZipFile(io.BytesIO(resp.data)) as zf:
        names = sorted(zf.namelist())
        assert len(names) == 2
        assert all(str(client_without_cases.id) not in name for name in names)
        first = zf.read(names[0])
        second = zf.read(names[1])

    wb_a = load_workbook(io.BytesIO(first))
    wb_b = load_workbook(io.BytesIO(second))
    rows_a = list(wb_a["Case Status"].iter_rows(values_only=True))
    rows_b = list(wb_b["Case Status"].iter_rows(values_only=True))
    flattened_a = " ".join(str(value or "") for row in rows_a for value in row)
    flattened_b = " ".join(str(value or "") for row in rows_b for value in row)

    assert "CRM-EXPORT-A-1" in flattened_a
    assert "TextA Text" in flattened_a
    assert "10-2026-0000001" in flattened_a
    assert "CRM-EXPORT-B-1" in flattened_b
    assert "Matter ID" not in flattened_a
    assert matter_a_id not in flattened_a
    assert matter_b_id not in flattened_b
    assert wb_a["Meta"]["B2"].value == "TextA"
    assert wb_b["Meta"]["B2"].value == "TextB"
    assert resp.headers["X-CRM-Exported-Clients"] == "2"
    assert resp.headers["X-CRM-Skipped-No-Case-Clients"] == "1"


def test_crm_case_status_export_all_uses_current_client_filter(admin_client, db_session):
    import io
    import zipfile

    from app.models.client import Client
    from app.models.ip_records import Matter, MatterCustomField

    included = Client(name="Text", email="target@example.com", extra={})
    excluded = Client(name="Text-Text", email="target-nocase@example.com", extra={})
    db_session.add_all([included, excluded])
    db_session.commit()

    matter_id = uuid.uuid4().hex
    db_session.add(
        Matter(
            matter_id=matter_id,
            our_ref="CRM-EXPORT-ALL-FILTERED-1",
            right_name="Text Text",
            right_group="DOM",
            matter_type="PATENT",
            is_deleted=False,
        )
    )
    db_session.add(
        MatterCustomField(
            matter_id=matter_id,
            namespace="domestic_patent",
            data={"client_id": str(included.id), "client_name": included.name},
        )
    )
    db_session.commit()

    resp = admin_client.post(
        "/crm/clients/case-status-export",
        data={"export_scope": "all", "q": "Text", "sort": "id", "direction": "asc"},
    )

    assert resp.status_code == 200
    with zipfile.ZipFile(io.BytesIO(resp.data)) as zf:
        names = zf.namelist()

    assert len(names) == 1
    assert str(included.id) in names[0]
    assert str(excluded.id) not in names[0]
    assert resp.headers["X-CRM-Exported-Clients"] == "1"
    assert resp.headers["X-CRM-Skipped-No-Case-Clients"] == "1"


def test_crm_client_view_shows_linked_matters(authenticated_client, sample_user, db_session):
    from app.models.client import Client
    from app.models.ip_records import Matter, MatterCustomField, MatterStaffAssignment

    # Ensure sample_user is attached to the active session (teardown hooks may detach instances).
    sample_user = db_session.merge(sample_user)

    crm_client = Client(name="Text", extra={})
    db_session.add(crm_client)
    db_session.commit()

    matter_id = uuid.uuid4().hex
    matter = Matter(
        matter_id=matter_id,
        our_ref="TEST-CRM-CLIENT-CASE-1",
        right_name="Text Text",
        right_group="DOM",
        matter_type="PATENT",
        is_deleted=False,
    )
    db_session.add(matter)
    db_session.add(
        MatterCustomField(
            matter_id=matter_id,
            namespace="domestic_patent",
            data={"client_id": str(crm_client.id), "client_name": crm_client.name},
        )
    )

    # Permission model: non-admin users must be assigned (directly or via team) to view linked matters.
    if not (sample_user.staff_party_id or "").strip():
        sample_user.staff_party_id = "test_staff_01"
        db_session.add(sample_user)
    if not MatterStaffAssignment.query.filter_by(
        matter_id=matter_id, staff_party_id=sample_user.staff_party_id
    ).first():
        db_session.add(
            MatterStaffAssignment(
                matter_id=matter_id,
                staff_party_id=sample_user.staff_party_id,
                staff_role_code="attorney",
            )
        )
    db_session.commit()

    resp = authenticated_client.get(f"/crm/clients/{crm_client.id}")
    assert resp.status_code == 200
    html = resp.data.decode("utf-8")

    assert "Text Text" in html
    assert "TEST-CRM-CLIENT-CASE-1" in html


def test_crm_client_view_shows_linked_matters_via_party_role(
    authenticated_client, sample_user, db_session
):
    from app.models.client import Client
    from app.models.ip_records import Matter, MatterPartyRole, MatterStaffAssignment

    # Ensure sample_user is attached to the active session (teardown hooks may detach instances).
    sample_user = db_session.merge(sample_user)

    crm_client = Client(name="Text-Text", party_id="test_party_01", extra={})
    db_session.add(crm_client)
    db_session.commit()

    matter_id = uuid.uuid4().hex
    matter = Matter(
        matter_id=matter_id,
        our_ref="TEST-CRM-CLIENT-CASE-PARTY-1",
        right_name="Text Text (party-role)",
        right_group="DOM",
        matter_type="PATENT",
        is_deleted=False,
    )
    db_session.add(matter)
    db_session.add(
        MatterPartyRole(
            matter_id=matter_id,
            party_id=crm_client.party_id,
            role_code="client",
            seq=1,
            raw_text=crm_client.name,
        )
    )

    # Permission model: non-admin users must be assigned (directly or via team) to view linked matters.
    if not (sample_user.staff_party_id or "").strip():
        sample_user.staff_party_id = "test_staff_01"
        db_session.add(sample_user)
    if not MatterStaffAssignment.query.filter_by(
        matter_id=matter_id, staff_party_id=sample_user.staff_party_id
    ).first():
        db_session.add(
            MatterStaffAssignment(
                matter_id=matter_id,
                staff_party_id=sample_user.staff_party_id,
                staff_role_code="attorney",
            )
        )
    db_session.commit()

    resp = authenticated_client.get(f"/crm/clients/{crm_client.id}")
    assert resp.status_code == 200
    html = resp.data.decode("utf-8")

    assert "Text Text" in html
    assert "TEST-CRM-CLIENT-CASE-PARTY-1" in html


def test_crm_client_view_linked_matters_sorted_by_latest_our_ref_desc(
    authenticated_client, sample_user, db_session
):
    from app.models.client import Client
    from app.models.ip_records import Matter, MatterCustomField, MatterStaffAssignment

    sample_user = db_session.merge(sample_user)

    crm_client = Client(name="Text", extra={})
    db_session.add(crm_client)
    db_session.commit()

    refs = [
        ("25TD0127US", "2026-02-01"),
        ("23TO0105US", "2026-01-15"),
        ("24PD0111US", "2025-12-30"),
    ]

    for ref, entered_at in refs:
        matter_id = uuid.uuid4().hex
        db_session.add(
            Matter(
                matter_id=matter_id,
                our_ref=ref,
                right_name=f"Text Text {ref}",
                right_group="DOM",
                matter_type="TRADEMARK",
                entered_at=entered_at,
                is_deleted=False,
            )
        )
        db_session.add(
            MatterCustomField(
                matter_id=matter_id,
                namespace="domestic_trademark",
                data={"client_id": str(crm_client.id), "client_name": crm_client.name},
            )
        )

        if not (sample_user.staff_party_id or "").strip():
            sample_user.staff_party_id = "test_staff_01"
            db_session.add(sample_user)
        if not MatterStaffAssignment.query.filter_by(
            matter_id=matter_id, staff_party_id=sample_user.staff_party_id
        ).first():
            db_session.add(
                MatterStaffAssignment(
                    matter_id=matter_id,
                    staff_party_id=sample_user.staff_party_id,
                    staff_role_code="attorney",
                )
            )

    db_session.commit()

    resp = authenticated_client.get(f"/crm/clients/{crm_client.id}")
    assert resp.status_code == 200
    html = resp.data.decode("utf-8")

    idx_25 = html.find("25TD0127US")
    idx_24 = html.find("24PD0111US")
    idx_23 = html.find("23TO0105US")
    assert idx_25 != -1 and idx_24 != -1 and idx_23 != -1
    assert idx_25 < idx_24 < idx_23


def test_crm_client_view_includes_bizreg_upload_modal(authenticated_client, db_session):
    from app.models.client import Client

    crm_client = Client(name="Text-Text-Text", extra={})
    db_session.add(crm_client)
    db_session.commit()
    client_id = int(crm_client.id)

    resp = authenticated_client.get(f"/crm/clients/{client_id}")
    assert resp.status_code == 200
    html = resp.data.decode("utf-8")

    assert 'id="bizRegUploadModal"' in html
    assert f'data-upload-url="/crm/clients/{client_id}/biz-reg/upload"' in html
    assert f'data-download-url="/crm/clients/{client_id}/biz-reg/shared-download"' in html
    assert 'data-bs-target="#bizRegUploadModal"' in html
    assert "Business document Registration" in html
    assert "Client File Upload LLM" in html
    assert "Automatically extracts Tax ID / EIN" in html


def test_applicant_code_debug_payload_sanitizes_customer_data():
    from app.blueprints.crm.routes import _sanitize_applicant_code_debug_payload

    payload = {
        "client_id": 123,
        "client_name": "Acme US",
        "existing_codes": ["CUST-001"],
        "slots": 2,
        "matter_count": 3,
        "matter_ids_preview": ["MATTER-SECRET-1", "MATTER-SECRET-2"],
        "name_variants": ["Acme", "Acme US"],
        "candidates_primary_count": 0,
        "candidates_name_match_count": 1,
        "reason": "no_candidates",
        "suggestion": {"codes": ["CUST-002"], "slots": 2, "extra_count": 0},
    }

    sanitized = _sanitize_applicant_code_debug_payload(payload)
    dumped = json.dumps(sanitized, ensure_ascii=False)

    assert sanitized["client_id_present"] is True
    assert sanitized["client_name_present"] is True
    assert sanitized["existing_codes_count"] == 1
    assert sanitized["matter_ids_preview_count"] == 2
    assert sanitized["name_variants_count"] == 2
    assert sanitized["suggestion"]["codes_count"] == 1
    assert "Acme" not in dumped
    assert "MATTER-SECRET" not in dumped
    assert "CUST-00" not in dumped
