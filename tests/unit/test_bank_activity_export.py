from io import BytesIO

from openpyxl import load_workbook

from app.blueprints.billing_invoices.db import _actual_table_name, get_db, init_db


def test_bank_activity_export_supports_local_search_filters(app, admin_client):
    with app.app_context():
        init_db()
        conn = get_db()
        tx_table = _actual_table_name("bank_transactions")
        conn.execute(
            f"""
            INSERT INTO {tx_table} (
                tid, bank_code, account_number, trdate, trdt, acc_in, acc_out, balance,
                remark1, remark2, remark3, memo
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "tid-local-export-1",
                "004",
                "1234567890",
                "20260309",
                "20260309120000",
                150000,
                0,
                700000,
                "Text Text",
                "Text",
                "",
                "Text-A",
            ),
        )
        conn.execute(
            f"""
            INSERT INTO {tx_table} (
                tid, bank_code, account_number, trdate, trdt, acc_in, acc_out, balance,
                remark1, remark2, remark3, memo
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "tid-local-export-2",
                "004",
                "1234567890",
                "20260308",
                "20260308110000",
                0,
                25000,
                550000,
                "Text Text",
                "Text",
                "",
                "Text-B",
            ),
        )
        conn.commit()

    response = admin_client.get(
        "/accounting/invoice-system/bank_activity/export"
        "?sdate=20260301"
        "&edate=20260331"
        "&accounts=004%7C1234567890"
        "&currency=USD"
        "&tradeType=I"
        "&searchString=Text-A"
        "&order=D"
    )

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook.active

    assert sheet.max_row == 2
    assert sheet["B2"].value == "004 1234567890"
    assert sheet["E2"].value == 1500
    assert sheet["F2"].value == 0
    assert sheet["I2"].value == "Text Text"
    assert sheet["L2"].value == "Text-A"
    assert sheet["P2"].value in (None, "")
