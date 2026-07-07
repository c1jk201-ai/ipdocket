"""WorkLog routes for task review, completion, notes, and permissions."""

from __future__ import annotations

import io
import json
import re
from datetime import date, datetime, timedelta

from flask import current_app, jsonify, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, case, func, not_, or_
from sqlalchemy.orm import load_only

from app.blueprints.worklog import bp
from app.extensions import db
from app.models.ip_records import Communication, DocketItem, Matter, MatterCustomField
from app.models.user import User
from app.models.workflow import Workflow
from app.models.workflow_assignment_request import WorkflowAssignmentRequest
from app.models.worklog import WorkLog
from app.services.core.staff_options import build_staff_assignment_lists
from app.services.docket_manual_state import (
    clear_docket_manual_abandoned,
    mark_docket_manual_abandoned,
)
from app.services.matter.matter_status_cache import apply_auto_status_cache_to_matter
from app.services.workflow.assignment_requests import (
    AssignmentRequestForbidden,
    pending_assignment_request_badges_by_workflow,
    respond_assignment_request,
    serialize_assignment_request,
    sync_assignment_requests_for_changed_roles,
    workflow_assignment_state,
)
from app.services.workflow.sync_requests import (
    enqueue_docket_sync_for_item,
    enqueue_workflow_sync,
    enqueue_workflow_task_sync,
)
from app.services.workflow.task_sync import persist_manual_workflow_assignment_override
from app.services.worklog.due_filters import effective_docket_due_expr as _effective_docket_due_expr
from app.services.worklog.due_filters import (
    normalize_worklog_due_axis as _normalize_worklog_due_axis,
)
from app.services.worklog.due_filters import today_in_app_timezone as _today_in_app_timezone
from app.services.worklog.due_filters import (
    worklog_calendar_due_range_condition as _worklog_calendar_due_range_condition,
)
from app.services.worklog.due_filters import (
    worklog_calendar_query_options as _worklog_calendar_query_options,
)
from app.services.worklog.due_filters import worklog_due_expr_for_axis as _worklog_due_expr_for_axis
from app.services.worklog.due_filters import worklog_final_due_expr as _worklog_final_due_expr
from app.services.worklog.due_filters import worklog_internal_due_expr as _worklog_internal_due_expr
from app.services.worklog.matter_display import (
    load_worklog_matter_display_context as _load_worklog_matter_display_context,
)
from app.services.worklog.task_merge import (
    _dedupe_staff_rows,
    _filter_hidden_workflow_rows,
    _filter_hidden_workflows,
    _finalize_merged_task_bucket,
    _merge_csv_names,
    _merge_docket_autogen_tasks,
    _merge_staff_rows,
    _merge_task_row_into_groups,
    _merge_task_statuses,
    _staff_row_names,
    _task_has_completion_recommendation,
    _workflow_category_badge_values,
    _workflow_docket_id,
    _workflow_is_auto_docket_generated,
    _worklog_group_key_for_workflow,
)
from app.services.worklog.task_search import (
    _append_search_terms,
    _flatten_field_search_values,
    _load_intake_case_access_recommendation_by_task_id,
    _matches_worklog_search_query,
    _parse_worklog_search_expression,
    _search_text_matches_query,
    _task_matches_search_query,
    _task_search_field_values,
    _task_search_text,
)
from app.services.worklog.visibility import worklog_role_scope_flags
from app.services.worklog.workflow_task_sync import sync_workflow_task_immediately
from app.utils.docket_dates import adjusted_legal_due_for_docket, internal_due_for_docket
from app.utils.error_logging import report_swallowed_exception
from app.utils.permissions import managed_matter_ids_select, resolve_role_scope
from app.utils.policy_sql import policy_text as text

# Import category constants from central location
from app.utils.task_classification import MGMT_CATEGORIES, WORK_CATEGORIES
from app.utils.task_distribution_rules import DistributionDecision, resolve_distribution_decision
from app.utils.workflow_deadline_labels import strip_workflow_deadline_title_suffix
from app.utils.workflow_roles import workflow_user_filter
from app.utils.workflow_semantics import workflow_primary_owner_user_id

_OA_ID_FROM_NAME_REF_RE = re.compile(r"^(?:MGMT:)?NOTICE:OA:([^:]+)", re.IGNORECASE)
URGENT_WINDOW_DAYS = 7
_MGMT_CATEGORIES_UPPER = frozenset({str(c or "").upper() for c in MGMT_CATEGORIES})
_WORK_CATEGORIES_UPPER = frozenset({str(c or "").upper() for c in WORK_CATEGORIES})
_VALID_EVIDENCE_TYPES = frozenset({"memo", "file", "mail", "number"})


def _date_to_iso(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value or "").strip()
    return raw or None


def _workflow_display_task_name(wf: Workflow | None) -> str:
    raw_name = str(getattr(wf, "name", None) or "").strip()
    if not raw_name:
        return ""
    if getattr(wf, "legal_due_date", None) or _workflow_docket_id(wf):
        return strip_workflow_deadline_title_suffix(raw_name) or ""
    return raw_name


def _workflow_list_status(
    *,
    wf: Workflow,
    today: date,
    urgent_date: date,
    due_date: date | None = None,
) -> str:
    raw = (getattr(wf, "status", None) or "").strip()
    if raw == "Completed":
        return "completed"
    if raw == "Abandoned":
        return "abandoned"
    target_due = due_date if due_date is not None else getattr(wf, "due_date", None)
    if target_due and target_due < today:
        return "overdue"
    if target_due and target_due <= urgent_date:
        return "urgent"
    return "pending"


def _normalized_workflow_name_token(value: str | None) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def _workflow_is_intake_confirmation_task(wf: Workflow) -> bool:
    business_code = str(getattr(wf, "business_code", None) or "").strip().upper()
    if business_code.startswith("INTAKE:"):
        return True
    return "Confirm" in _normalized_workflow_name_token(getattr(wf, "name", None))


def _user_to_staff_row(user) -> dict | None:
    if not user:
        return None
    row_id = (
        str(getattr(user, "staff_party_id", None) or "").strip()
        or str(getattr(user, "id", None) or "").strip()
    )
    if not row_id:
        return None
    row_name = (
        str(getattr(user, "display_name", None) or "").strip()
        or str(getattr(user, "username", None) or "").strip()
        or row_id
    )
    return {"id": row_id, "name": row_name}


def _workflow_assignee_filter_ids(
    *,
    assignee_id: int | None = None,
    assignee_user=None,
    attorney_assignee_id: int | None = None,
    attorney_assignee_user=None,
    inspector_id: int | None = None,
    inspector_user=None,
) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()

    def _add(raw) -> None:
        value = str(raw or "").strip()
        if not value or value in seen:
            return
        seen.add(value)
        values.append(value)

    for raw_id, user in (
        (assignee_id, assignee_user),
        (attorney_assignee_id, attorney_assignee_user),
        (inspector_id, inspector_user),
    ):
        _add(raw_id)
        if user is not None:
            _add(getattr(user, "id", None))
            _add(getattr(user, "staff_party_id", None))

    return values


def _extract_task_source_from_docket_item(docket_item: DocketItem | None) -> str | None:
    if not docket_item:
        return None
    raw_source = (getattr(docket_item, "source", None) or "").strip()
    if raw_source:
        return raw_source.lower()

    memo = (getattr(docket_item, "memo", None) or "").strip()
    payload: dict[str, object] = {}
    if memo:
        try:
            parsed = json.loads(memo)
        except Exception:
            parsed = {}
        if isinstance(parsed, dict):
            payload = parsed
            src = str(payload.get("source") or "").strip()
            if src:
                return src.lower()
            if isinstance(payload.get("events"), list):
                return "upload_automation"
            trigger = str(payload.get("trigger") or "").strip()
            if trigger in {"REGISTRATION_CERTIFICATE", "Notice of allowance", "Final rejection"}:
                return "uspto_notice"

    category = (getattr(docket_item, "category", None) or "").strip().upper()
    name_ref = (getattr(docket_item, "name_ref", None) or "").strip().upper()
    if name_ref.startswith("USPTO:"):
        return "uspto_notice"
    if name_ref.startswith("USPTO_OA:"):
        return "upload_automation"
    if category == "USPTO_OA":
        return "upload_automation"
    return None


def _normalize_distribution_role(role_code: str | None) -> str | None:
    role = (role_code or "").strip().lower()
    if role in ("manager", "mgmt"):
        return "manager"
    if role in ("attorney", "retainer"):
        return "attorney"
    if role in ("handler", "staff", "draftsman"):
        return "handler"
    if role in ("owner", "fallback"):
        return "owner"
    return None


def _user_is_manager_like(user) -> bool:
    if user is None:
        return False
    try:
        checker = getattr(user, "is_mgmt_role", None)
        if callable(checker):
            return bool(checker())
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="worklog.routes._user_is_manager_like",
            log_key="worklog.routes._user_is_manager_like",
            log_window_seconds=300,
        )
    role = str(getattr(user, "role", "") or "").strip().lower()
    return role in {"manager", "mgmt_staff", "mgmt_director"}


def _workflow_has_explicit_flow_assignments(wf: Workflow) -> bool:
    for raw in (
        getattr(wf, "assignee_id", None),
        getattr(wf, "attorney_assignee_id", None),
        getattr(wf, "inspector_id", None),
    ):
        try:
            if int(raw or 0) > 0:
                return True
        except Exception:
            if raw not in (None, ""):
                return True
    return False


def _resolve_effective_staff_rows_for_task(
    *,
    wf: Workflow,
    attorneys_list: list[dict],
    handlers_list: list[dict],
    managers_list: list[dict],
    assignee_user=None,
    attorney_assignee_user=None,
    inspector_user=None,
) -> dict[str, list[dict]]:
    handler_row = _user_to_staff_row(assignee_user)
    attorney_row = _user_to_staff_row(attorney_assignee_user)
    manager_row = _user_to_staff_row(inspector_user)

    if _workflow_has_explicit_flow_assignments(wf):
        return {
            "handler": _merge_staff_rows([handler_row] if handler_row else []),
            "attorney": _merge_staff_rows([attorney_row] if attorney_row else []),
            "manager": _merge_staff_rows([manager_row] if manager_row else []),
        }

    return {
        "handler": _merge_staff_rows(handlers_list, [handler_row] if handler_row else []),
        "attorney": _merge_staff_rows(attorneys_list, [attorney_row] if attorney_row else []),
        "manager": _merge_staff_rows(managers_list, [manager_row] if manager_row else []),
    }


def _decision_is_manager_only(decision: DistributionDecision | None) -> bool:
    if not decision or decision.distribute_to != "role_set":
        return False
    roles = {
        normalized
        for normalized in (_normalize_distribution_role(r) for r in (decision.role_codes or ()))
        if normalized
    }
    if not roles:
        return True
    return roles.issubset({"manager"})


def _resolve_owner_rows_for_task(
    *,
    wf: Workflow,
    linked_docket_item: DocketItem | None,
    owner_user,
    attorneys_list: list[dict],
    handlers_list: list[dict],
    managers_list: list[dict],
    assignee_user=None,
    attorney_assignee_user=None,
    inspector_user=None,
) -> tuple[list[dict], DistributionDecision]:
    category = (
        getattr(linked_docket_item, "category", None)
        if linked_docket_item is not None
        else getattr(wf, "category", None)
    )
    name_ref = (
        getattr(linked_docket_item, "name_ref", None) if linked_docket_item is not None else None
    )
    name_free = (
        getattr(linked_docket_item, "name_free", None)
        if linked_docket_item is not None
        else getattr(wf, "name", None)
    )
    source = _extract_task_source_from_docket_item(linked_docket_item)
    decision = resolve_distribution_decision(
        category=category,
        name_ref=name_ref,
        name_free=name_free,
        source=source,
    )

    if assignee_user is None:
        assignee_user = getattr(wf, "assignee", None)
    if attorney_assignee_user is None:
        attorney_assignee_user = getattr(wf, "attorney_assignee", None)
    if inspector_user is None:
        inspector_user = getattr(wf, "inspector", None)

    owner_row = _user_to_staff_row(owner_user)
    role_rows = _resolve_effective_staff_rows_for_task(
        wf=wf,
        attorneys_list=attorneys_list,
        handlers_list=handlers_list,
        managers_list=managers_list,
        assignee_user=assignee_user,
        attorney_assignee_user=attorney_assignee_user,
        inspector_user=inspector_user,
    )

    owners_by_role: dict[str, list[dict]] = {
        "owner": _merge_staff_rows([owner_row] if owner_row else []),
        "handler": list(role_rows.get("handler") or []),
        "attorney": list(role_rows.get("attorney") or []),
        "manager": list(role_rows.get("manager") or []),
    }

    if decision.distribute_to == "all_staff":
        rows = _merge_staff_rows(
            owners_by_role["owner"],
            owners_by_role["handler"],
            owners_by_role["attorney"],
            owners_by_role["manager"],
        )
        return rows, decision

    if decision.distribute_to == "role_set":
        target_roles = {
            normalized
            for normalized in (_normalize_distribution_role(r) for r in decision.role_codes)
            if normalized
        }
        if not target_roles:
            target_roles = {"manager"}
        rows: list[dict] = []
        for role_key in ("owner", "handler", "attorney", "manager"):
            if role_key in target_roles:
                role_rows = owners_by_role.get(role_key, [])
                if not role_rows and role_key in {"handler", "attorney"}:
                    # Legacy rows can encode the target role only in assignee(owner) fields.
                    role_rows = owners_by_role.get("owner", [])
                if not role_rows and role_key == "manager" and _user_is_manager_like(owner_user):
                    role_rows = owners_by_role.get("owner", [])
                rows = _merge_staff_rows(rows, role_rows)
        return rows, decision

    if decision.distribute_to == "none":
        return [], decision

    # owner(default): keep owner semantics as a single task owner if possible.
    rows = owners_by_role["owner"]
    if rows:
        return rows, decision
    rows = _merge_staff_rows(
        owners_by_role["handler"],
        owners_by_role["attorney"],
        owners_by_role["manager"],
    )
    return rows, decision


def _resolve_owner_rows_for_workflows(workflows: list[Workflow]) -> dict[int, list[dict]]:
    if not workflows:
        return {}

    workflow_users_by_id: dict[int, User] = {}
    user_ids: set[int] = set()
    for wf in workflows:
        for raw in (
            getattr(wf, "assignee_id", None),
            getattr(wf, "attorney_assignee_id", None),
            getattr(wf, "inspector_id", None),
        ):
            try:
                uid = int(raw or 0)
            except Exception:
                uid = 0
            if uid > 0:
                user_ids.add(uid)
    if user_ids:
        workflow_users_by_id = {
            int(user.id): user
            for user in User.query.filter(User.id.in_(sorted(user_ids))).all()
            if int(getattr(user, "id", 0) or 0) > 0
        }

    matter_ids = {
        str(getattr(wf, "case_id", "") or "").strip()
        for wf in workflows
        if str(getattr(wf, "case_id", "") or "").strip()
    }
    matter_staff_map: dict[str, dict[str, list[dict]]] = {}

    if matter_ids:
        try:
            from app.models.party import Party, PartyStaff
            from app.models.ip_records import MatterStaffAssignment

            role_expr = func.lower(func.trim(MatterStaffAssignment.staff_role_code))
            msa_rows = (
                db.session.query(
                    MatterStaffAssignment.matter_id,
                    MatterStaffAssignment.staff_role_code,
                    MatterStaffAssignment.staff_party_id,
                    Party.name_display,
                )
                .join(PartyStaff, PartyStaff.party_id == MatterStaffAssignment.staff_party_id)
                .join(Party, Party.party_id == PartyStaff.party_id)
                .filter(MatterStaffAssignment.matter_id.in_(sorted(matter_ids)))
                .filter(
                    role_expr.in_(
                        (
                            "attorney",
                            "retainer",
                            "handler",
                            "staff",
                            "draftsman",
                            "manager",
                            "mgmt",
                        )
                    )
                )
                .all()
            )
            for mid, role, spid, name in msa_rows:
                mid_s = str(mid or "").strip()
                if not mid_s:
                    continue
                if mid_s not in matter_staff_map:
                    matter_staff_map[mid_s] = {"attorney": [], "handler": [], "manager": []}
                sid = str(spid or "").strip()
                n = str(name or "").strip()
                if not sid or not n:
                    continue
                entry = {"id": sid, "name": n}
                r = str(role or "").strip().lower()
                if r in ("attorney", "retainer"):
                    if sid not in {p.get("id") for p in matter_staff_map[mid_s]["attorney"]}:
                        matter_staff_map[mid_s]["attorney"].append(entry)
                elif r in ("handler", "staff", "draftsman"):
                    if sid not in {p.get("id") for p in matter_staff_map[mid_s]["handler"]}:
                        matter_staff_map[mid_s]["handler"].append(entry)
                elif r in ("manager", "mgmt"):
                    if sid not in {p.get("id") for p in matter_staff_map[mid_s]["manager"]}:
                        matter_staff_map[mid_s]["manager"].append(entry)
        except Exception as exc:
            report_swallowed_exception(
                exc,
                context="worklog.owner_rows_for_workflows.staff_assignments",
                log_key="worklog.owner_rows_for_workflows.staff_assignments",
                log_window_seconds=300,
            )

    docket_ids = {did for did in (_workflow_docket_id(wf) for wf in workflows) if did}
    docket_item_by_docket_id: dict[str, DocketItem] = {}
    if docket_ids:
        try:
            docket_q = DocketItem.query.filter(DocketItem.docket_id.in_(sorted(docket_ids)))
            if hasattr(DocketItem, "is_deleted"):
                docket_q = docket_q.filter(
                    or_(DocketItem.is_deleted.is_(False), DocketItem.is_deleted.is_(None))
                )
            for di in docket_q.all():
                did = str(getattr(di, "docket_id", "") or "").strip()
                if did:
                    docket_item_by_docket_id[did] = di
        except Exception as exc:
            report_swallowed_exception(
                exc,
                context="worklog.owner_rows_for_workflows.docket_lookup",
                log_key="worklog.owner_rows_for_workflows.docket_lookup",
                log_window_seconds=300,
            )

    out: dict[int, list[dict]] = {}
    for wf in workflows:
        wf_id = getattr(wf, "id", None)
        if wf_id is None:
            continue
        case_id = str(getattr(wf, "case_id", "") or "").strip()
        staff_bucket = matter_staff_map.get(case_id, {})
        linked_docket_id = _workflow_docket_id(wf)
        linked_di = docket_item_by_docket_id.get(linked_docket_id) if linked_docket_id else None
        if linked_di and str(getattr(linked_di, "matter_id", "") or "") != case_id:
            linked_di = None
        assignee_user = workflow_users_by_id.get(int(getattr(wf, "assignee_id", 0) or 0))
        attorney_assignee_user = workflow_users_by_id.get(
            int(getattr(wf, "attorney_assignee_id", 0) or 0)
        )
        inspector_user = workflow_users_by_id.get(int(getattr(wf, "inspector_id", 0) or 0))
        owner_user_id = workflow_primary_owner_user_id(
            category=getattr(wf, "category", None),
            handler_id=getattr(wf, "assignee_id", None),
            attorney_id=getattr(wf, "attorney_assignee_id", None),
            manager_id=getattr(wf, "inspector_id", None),
        )
        owner_user = (
            (
                workflow_users_by_id.get(int(owner_user_id or 0))
                if owner_user_id is not None
                else None
            )
            or assignee_user
            or attorney_assignee_user
            or inspector_user
        )
        owner_rows, _decision = _resolve_owner_rows_for_task(
            wf=wf,
            linked_docket_item=linked_di,
            owner_user=owner_user,
            attorneys_list=list(staff_bucket.get("attorney") or []),
            handlers_list=list(staff_bucket.get("handler") or []),
            managers_list=list(staff_bucket.get("manager") or []),
            assignee_user=assignee_user,
            attorney_assignee_user=attorney_assignee_user,
            inspector_user=inspector_user,
        )
        out[int(wf_id)] = owner_rows
    return out


def _summary_counts_from_workflow_rows(
    rows: list[tuple[Workflow, Matter]],
    *,
    today: date,
    urgent_date: date,
    end_date: date,
    search_query: str = "",
    owner_filter_active: bool = False,
    owner_filter_ids: set[str] | None = None,
    display_context: dict[str, object] | None = None,
) -> dict[str, int]:
    rows = _filter_hidden_workflow_rows(rows, today=today)
    workflows = [wf for wf, _m in rows]
    owner_rows_by_wf: dict[int, list[dict]] = {}
    owner_rows_needed = owner_filter_active or bool(str(search_query or "").strip())
    if owner_rows_needed:
        owner_rows_by_wf = _resolve_owner_rows_for_workflows(workflows)

    display_context = display_context or {}
    matter_applicant_map = dict(display_context.get("matter_applicant_map") or {})
    matter_attorney_map = dict(display_context.get("matter_attorney_map") or {})
    matter_handler_map = dict(display_context.get("matter_handler_map") or {})
    matter_manager_map = dict(display_context.get("matter_manager_map") or {})
    matter_staff_map = dict(display_context.get("matter_staff_map") or {})

    grouped: dict[str, dict[str, object]] = {}
    ordered_keys: list[str] = []

    for wf, matter in rows:
        key = _worklog_group_key_for_workflow(wf)
        bucket = grouped.get(key)
        if bucket is None:
            bucket = {
                "statuses": set(),
                "categories": set(),
                "due_dates": [],
                "done_dates": [],
                "owners": [],
                "task_names": [],
                "our_ref": str(getattr(matter, "our_ref", "") or "").strip(),
                "applicant_terms": [],
                "owner_terms": [],
                "attorney_terms": [],
                "handler_terms": [],
                "manager_terms": [],
                "note_terms": [],
                "search_terms": [],
            }
            grouped[key] = bucket
            ordered_keys.append(key)

        status = _workflow_list_status(wf=wf, today=today, urgent_date=urgent_date)
        cast_statuses = bucket["statuses"]
        if isinstance(cast_statuses, set):
            cast_statuses.add(status)
        cast_categories = bucket["categories"]
        if isinstance(cast_categories, set):
            cat_type, cat_display = _workflow_category_badge_values(getattr(wf, "category", None))
            for value in (getattr(wf, "category", None), cat_type, cat_display):
                token = str(value or "").strip()
                if token:
                    cast_categories.add(token)

        due_date = getattr(wf, "due_date", None)
        if due_date is not None and isinstance(bucket["due_dates"], list):
            bucket["due_dates"].append(due_date)

        done_date = getattr(wf, "completed_date", None)
        if done_date is not None and isinstance(bucket["done_dates"], list):
            bucket["done_dates"].append(done_date)

        task_name = _workflow_display_task_name(wf)
        if (
            task_name
            and isinstance(bucket["task_names"], list)
            and task_name not in bucket["task_names"]
        ):
            bucket["task_names"].append(task_name)

        matter_id = str(getattr(wf, "case_id", "") or "").strip()
        search_terms = (
            bucket["search_terms"] if isinstance(bucket.get("search_terms"), list) else []
        )
        applicant_terms = (
            bucket["applicant_terms"] if isinstance(bucket.get("applicant_terms"), list) else []
        )
        attorney_terms = (
            bucket["attorney_terms"] if isinstance(bucket.get("attorney_terms"), list) else []
        )
        handler_terms = (
            bucket["handler_terms"] if isinstance(bucket.get("handler_terms"), list) else []
        )
        manager_terms = (
            bucket["manager_terms"] if isinstance(bucket.get("manager_terms"), list) else []
        )
        note_terms = bucket["note_terms"] if isinstance(bucket.get("note_terms"), list) else []
        _append_search_terms(
            applicant_terms,
            matter_applicant_map.get(matter_id),
        )
        _append_search_terms(
            search_terms,
            matter_applicant_map.get(matter_id),
            matter_attorney_map.get(matter_id),
            matter_handler_map.get(matter_id),
            matter_manager_map.get(matter_id),
            getattr(wf, "note", None),
        )
        _append_search_terms(
            attorney_terms,
            matter_attorney_map.get(matter_id),
        )
        _append_search_terms(
            handler_terms,
            matter_handler_map.get(matter_id),
        )
        _append_search_terms(
            manager_terms,
            matter_manager_map.get(matter_id),
        )
        _append_search_terms(note_terms, getattr(wf, "note", None))
        for role_key in ("attorney", "handler", "manager"):
            for row in (matter_staff_map.get(matter_id, {}) or {}).get(role_key, []) or []:
                if not isinstance(row, dict):
                    continue
                _append_search_terms(search_terms, row.get("name"), row.get("id"))
                if role_key == "attorney":
                    _append_search_terms(attorney_terms, row.get("name"), row.get("id"))
                elif role_key == "handler":
                    _append_search_terms(handler_terms, row.get("name"), row.get("id"))
                elif role_key == "manager":
                    _append_search_terms(manager_terms, row.get("name"), row.get("id"))

        if owner_rows_needed:
            owner_terms = (
                bucket["owner_terms"] if isinstance(bucket.get("owner_terms"), list) else []
            )
            owner_rows = owner_rows_by_wf.get(int(getattr(wf, "id", 0) or 0), [])
            if owner_rows and isinstance(bucket["owners"], list):
                bucket["owners"].extend(owner_rows)
            elif isinstance(bucket["owners"], list):
                fallback_owner = _user_to_staff_row(
                    wf.assignee
                    or getattr(wf, "attorney_assignee", None)
                    or getattr(wf, "inspector", None)
                )
                if fallback_owner:
                    bucket["owners"].append(fallback_owner)
                    _append_search_terms(
                        owner_terms, fallback_owner.get("name"), fallback_owner.get("id")
                    )
            for row in owner_rows:
                if not isinstance(row, dict):
                    continue
                _append_search_terms(search_terms, row.get("name"), row.get("id"))
                _append_search_terms(owner_terms, row.get("name"), row.get("id"))

    week_ago = today - timedelta(days=7)

    counts = {
        "pending": 0,
        "urgent": 0,
        "overdue": 0,
        "completed_week": 0,
    }

    search_expression = _parse_worklog_search_expression(search_query) if search_query else None

    for key in ordered_keys:
        bucket = grouped[key]
        owners = _dedupe_staff_rows(list(bucket.get("owners") or []))
        owner_id = None
        if len(owners) == 1 and str(owners[0].get("id") or "").strip():
            owner_id = str(owners[0]["id"]).strip()

        if owner_filter_active:
            if not _task_matches_owner_filter(
                {"owners": owners, "owner_id": owner_id},
                owner_filter_ids or set(),
            ):
                continue

        task_names = [n for n in list(bucket.get("task_names") or []) if str(n or "").strip()]
        merged_task_name = task_names[0] if len(task_names) == 1 else " / ".join(task_names)
        if search_query:
            merged_status = _merge_task_statuses(set(bucket.get("statuses") or set()))
            field_values = {
                "our_ref": [str(bucket.get("our_ref") or "").strip()],
                "task_name": [merged_task_name],
                "applicant_name": list(bucket.get("applicant_terms") or []),
                "owner_name": [
                    str(row.get("name") or "").strip()
                    for row in owners
                    if str(row.get("name") or "").strip()
                ],
                "attorney_names": list(bucket.get("attorney_terms") or []),
                "handler_names": list(bucket.get("handler_terms") or []),
                "manager_names": list(bucket.get("manager_terms") or []),
                "note": list(bucket.get("note_terms") or []),
                "status": [merged_status, *list(bucket.get("statuses") or set())],
                "category": list(bucket.get("categories") or set()),
            }
            field_values["staff"] = list(
                {
                    *field_values["owner_name"],
                    *field_values["attorney_names"],
                    *field_values["handler_names"],
                    *field_values["manager_names"],
                }
            )
            search_text = _flatten_field_search_values(field_values)
            if not _matches_worklog_search_query(
                search_text=search_text,
                search_query=search_query,
                field_values=field_values,
                search_expression=search_expression,
            ):
                continue

        merged_status = _merge_task_statuses(set(bucket.get("statuses") or set()))
        due_dates = [d for d in list(bucket.get("due_dates") or []) if d is not None]
        done_dates = [d for d in list(bucket.get("done_dates") or []) if d is not None]
        due = min(due_dates) if due_dates else None
        done = max(done_dates) if done_dates else None
        is_done = merged_status in ("completed", "abandoned")

        if (not is_done) and (due is not None) and due <= end_date:
            counts["pending"] += 1
        if (not is_done) and (due is not None) and due <= end_date and today <= due <= urgent_date:
            counts["urgent"] += 1
        if (not is_done) and (due is not None) and due < today:
            counts["overdue"] += 1
        if is_done and (done is not None) and done >= week_ago:
            counts["completed_week"] += 1

    return counts


def _linked_docket_item_for_workflow(
    wf: Workflow,
    *,
    docket_item_by_docket_id: dict[str, DocketItem] | None = None,
) -> DocketItem | None:
    """If this workflow is docket-backed (business_code=DOCKET:...), return the linked DocketItem."""
    docket_id = _workflow_docket_id(wf)
    if not docket_id:
        return None
    di = None
    if docket_item_by_docket_id is not None:
        di = docket_item_by_docket_id.get(docket_id)
    if di is None:
        di = DocketItem.query.filter_by(docket_id=docket_id).first()
    if not di:
        return None
    if hasattr(di, "is_deleted") and bool(getattr(di, "is_deleted", False)):
        return None
    if str(getattr(di, "matter_id", "")) != str(getattr(wf, "case_id", "")):
        return None
    return di


def _recalc_matter_status(*, matter: Matter, memo: str | None = None) -> None:
    if not matter:
        return
    apply_auto_status_cache_to_matter(
        matter=matter,
        memo=memo or getattr(matter, "memo", None),
        empty_as_none=True,
    )


def _sync_docket_task_immediately(di: DocketItem, *, actor_id: int | None = None) -> bool:
    """Best-effort synchronous docket->workflow/worklog sync for immediate UI consistency."""
    if not di:
        return False
    try:
        from app.services.workflow.task_sync import sync_from_docket_item

        sync_from_docket_item(docket_item=di, actor_id=actor_id)
        return True
    except Exception as exc:
        current_app.logger.warning(
            "Immediate docket sync failed for %s: %s",
            getattr(di, "docket_id", None),
            exc,
        )
        return False


def _extract_office_action_id(di: DocketItem) -> str | None:
    memo_raw = (getattr(di, "memo", None) or "").strip()
    if memo_raw:
        try:
            memo = json.loads(memo_raw)
        except Exception:
            memo = {}
        if isinstance(memo, dict) and (memo.get("trigger") or "").strip() == "office_action_due":
            oa_id = (memo.get("oa_id") or "").strip()
            if oa_id:
                return oa_id

    name_ref = (getattr(di, "name_ref", None) or "").strip()
    if name_ref:
        m = _OA_ID_FROM_NAME_REF_RE.match(name_ref)
        if m:
            oa_id = (m.group(1) or "").strip()
            if oa_id:
                return oa_id
    return None


def _classify_category(
    *,
    category: str | None,
    name_ref: str | None,
    name_free: str | None,
    staff_role: str | None = None,
    owner_role: str | None = None,
) -> str:
    """
    Classify docket item as 'mgmt' or 'work'.

    Delegates to unified classify_task_type() for consistent classification.
    """
    from app.utils.task_classification import classify_task_type

    return classify_task_type(
        staff_role=staff_role,
        owner_role=owner_role,
        category=category,
        name_ref=name_ref,
        name_free=name_free,
    )


def _can_view_all_mgmt() -> bool:
    """Check if current user can view all mgmt tasks."""
    if not current_user.is_authenticated:
        return False
    flags = worklog_role_scope_flags(current_user)
    return bool(flags.get("show_all_mgmt"))


def _can_view_all_work() -> bool:
    """Check if current user can view all work tasks."""
    if not current_user.is_authenticated:
        return False
    flags = worklog_role_scope_flags(current_user)
    return bool(flags.get("show_all_work"))


def _get_staff_party_id() -> str | None:
    """Get current user's staff_party_id."""
    if not current_user.is_authenticated:
        return None
    return current_user.staff_party_id


def _current_worklog_mine_owner_value() -> str:
    if not current_user.is_authenticated:
        return ""
    return (
        str(getattr(current_user, "staff_party_id", None) or "").strip()
        or str(getattr(current_user, "id", None) or "").strip()
    )


def _resolve_owner_user_id(owner_filter: str | None) -> int | None:
    if not owner_filter:
        return None
    try:
        from app.models.user import User

        target_user = User.query.filter_by(staff_party_id=str(owner_filter)).first()
        if not target_user and str(owner_filter).isdigit():
            target_user = db.session.get(User, int(owner_filter))
        return target_user.id if target_user else None
    except Exception:
        return None


def _normalize_owner_role(raw: str | None) -> str:
    role = (raw or "").strip().lower()
    if role in ("owner", "assignee", "workflow", "wf"):
        return "owner"
    if role in ("attorney", "patent_attorney"):
        return "attorney"
    if role in ("handler", "staff", "draftsman"):
        return "handler"
    if role in ("manager", "mgmt"):
        return "manager"
    if role in ("any", "case", "case_staff"):
        return "any"
    return "owner"


def _resolve_staff_party_id(owner_filter: str | None) -> str | None:
    """
    Resolve a staff_party_id from a UI filter value.

    - Prefer exact match on User.staff_party_id
    - If numeric and maps to User.id, return that user's staff_party_id
    - Otherwise treat it as a staff_party_id-like value
    """
    if not owner_filter:
        return None
    s = str(owner_filter).strip()
    if not s:
        return None

    try:
        from app.models.user import User

        target_user = User.query.filter_by(staff_party_id=s).first()
        if target_user and (target_user.staff_party_id or "").strip():
            return (target_user.staff_party_id or "").strip()

        if s.isdigit():
            u = db.session.get(User, int(s))
            spid = (getattr(u, "staff_party_id", None) or "").strip() if u else ""
            if spid:
                return spid
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="worklog._resolve_staff_party_id",
            log_key="worklog._resolve_staff_party_id",
            log_window_seconds=300,
        )
        return s

    return s


def _resolve_staff_display_name(staff_party_id: str | None) -> str | None:
    if not staff_party_id:
        return None
    spid = str(staff_party_id).strip()
    if not spid:
        return None
    try:
        from app.models.party import Party

        p = Party.query.filter_by(party_id=spid).first()
        name = (getattr(p, "name_display", None) or "").strip() if p else ""
        return name or None
    except Exception:
        return None


def _owner_filter_candidate_ids(owner_filter: str | None) -> set[str]:
    out: set[str] = set()
    raw = str(owner_filter or "").strip()
    if not raw:
        return out
    out.add(raw)
    spid = _resolve_staff_party_id(raw)
    if spid:
        out.add(spid)
    uid = _resolve_owner_user_id(raw)
    if uid is not None:
        out.add(str(uid))
    return {v for v in out if str(v).strip()}


def _task_matches_owner_filter(task: dict, owner_filter_ids: set[str]) -> bool:
    if not owner_filter_ids:
        return False
    owner_rows = task.get("owners") or []
    row_ids = {
        str(row.get("id") or "").strip()
        for row in owner_rows
        if isinstance(row, dict) and str(row.get("id") or "").strip()
    }
    if not row_ids:
        fallback = str(task.get("owner_id") or "").strip()
        if fallback:
            row_ids.add(fallback)
    return bool(row_ids.intersection(owner_filter_ids))


def _task_matches_mine_filter(task: dict, owner_filter_ids: set[str]) -> bool:
    if _task_matches_owner_filter(task, owner_filter_ids):
        return True
    direct_ids = {
        str(raw or "").strip()
        for raw in task.get("_workflow_assignee_filter_ids") or []
        if str(raw or "").strip()
    }
    return bool(direct_ids.intersection(owner_filter_ids))


def _task_matches_staff_role_filter(
    task: dict,
    owner_filter_ids: set[str],
    *,
    owner_role: str,
) -> bool:
    if not owner_filter_ids:
        return False
    normalized_role = _normalize_owner_role(owner_role)
    role_keys_by_owner_role: dict[str, tuple[str, ...]] = {
        "attorney": ("attorneys",),
        "handler": ("handlers",),
        "manager": ("managers",),
        "any": ("attorneys", "handlers", "managers"),
    }
    role_keys = role_keys_by_owner_role.get(normalized_role, ())
    if not role_keys:
        return False

    row_ids: set[str] = set()
    for key in role_keys:
        for row in task.get(key) or []:
            row_id = str((row or {}).get("id") or "").strip()
            if row_id:
                row_ids.add(row_id)
    return bool(row_ids.intersection(owner_filter_ids))


def _strip_internal_task_filter_fields(task: dict) -> dict:
    task.pop("_workflow_assignee_filter_ids", None)
    return task


def _workflow_id_from_task_payload(task: dict) -> int | None:
    for key in ("workflow_link_id", "id"):
        token = str(task.get(key) or "").strip()
        if token.startswith("wf_"):
            token = token[3:]
        try:
            parsed = int(token)
        except Exception:
            continue
        if parsed > 0:
            return parsed
    return None


def _attach_assignment_request_badges(tasks: list[dict]) -> list[dict]:
    workflow_ids = {
        int(wf_id)
        for wf_id in (_workflow_id_from_task_payload(task) for task in tasks)
        if wf_id is not None
    }
    if not workflow_ids:
        return tasks
    badges_by_wf = pending_assignment_request_badges_by_workflow(
        workflow_ids,
        current_user_id=getattr(current_user, "id", None),
    )
    for task in tasks:
        wf_id = _workflow_id_from_task_payload(task)
        task["assignment_pending_roles"] = list(badges_by_wf.get(int(wf_id or 0), []))
    return tasks


def _case_staff_match_condition(*, owner_role: str, staff_party_id: str, staff_name: str | None):
    """
    Build a SQL condition for "case staff role" filters.

    owner_role:
      - attorney | handler | manager | any
    """
    owner_role = _normalize_owner_role(owner_role)
    if owner_role == "owner":
        return text("1=1")

    # Map UI role -> MSA staff_role_code buckets
    msa_roles_by_owner_role: dict[str, tuple[str, ...]] = {
        "attorney": ("attorney", "retainer"),
        "handler": ("handler", "staff", "draftsman"),
        "manager": ("manager", "mgmt"),
        "any": ("attorney", "retainer", "handler", "staff", "draftsman", "manager", "mgmt"),
    }
    cf_keys_by_owner_role: dict[str, tuple[str, ...]] = {
        "attorney": ("attorney",),
        "handler": ("handler",),
        "manager": ("manager",),
        "any": ("attorney", "handler", "manager"),
    }

    role_codes = msa_roles_by_owner_role.get(owner_role, ())
    cf_keys = cf_keys_by_owner_role.get(owner_role, ())

    if not role_codes:
        return text("1=0")

    from app.models.ip_records import MatterStaffAssignment

    role_expr = func.lower(func.trim(MatterStaffAssignment.staff_role_code))
    msa_exists = (
        db.session.query(MatterStaffAssignment.msa_id)
        .filter(MatterStaffAssignment.matter_id == Matter.matter_id)
        .filter(MatterStaffAssignment.staff_party_id == staff_party_id)
        .filter(role_expr.in_(role_codes))
        .exists()
    )

    conditions = [msa_exists]

    # Custom field fallback (best-effort): match name string inside JSON values.
    # This helps when MSA data is missing, but basic namespace strings are present.
    staff_name = (staff_name or "").strip()
    if staff_name and cf_keys:
        term = f"%{staff_name}%"
        for key in cf_keys:
            cf_value = MatterCustomField.data[key].as_string()
            cf_exists = (
                db.session.query(MatterCustomField.id)
                .filter(MatterCustomField.matter_id == Matter.matter_id)
                .filter(MatterCustomField.namespace == "basic")
                .filter(cf_value.isnot(None))
                .filter(cf_value.ilike(term))
                .exists()
            )
            conditions.append(cf_exists)

    return or_(*conditions) if conditions else text("1=0")


def _parse_days_param(raw: str | None, *, default: int = 30, max_days: int = 9999) -> int:
    """
    Parse ?days= query param safely.
    - 'all' or '*' => max_days
    - invalid / empty => default
    - negative => 0
    - too large => max_days
    """
    s = (raw or "").strip().lower()
    if not s:
        return default
    if s in ("all", "*"):
        return max_days
    try:
        n = int(s)
    except Exception:
        return default
    if n < 0:
        return 0
    if n > max_days:
        return max_days
    return n


def _parse_iso_date_param(raw: str | None) -> date | None:
    s = (raw or "").strip()
    if not s:
        return None
    token = s.split("T", 1)[0].strip()
    if not token:
        return None
    try:
        return date.fromisoformat(token)
    except Exception:
        return None


def _normalize_due_range(
    due_from: date | None, due_to: date | None
) -> tuple[date | None, date | None]:
    if due_from and due_to and due_from > due_to:
        return due_to, due_from
    return due_from, due_to


def _workflow_task_due_dates(
    *,
    wf: Workflow,
    linked_docket_item: DocketItem | None = None,
) -> tuple[str | None, str | None]:
    final_due = _date_to_iso(getattr(wf, "legal_due_date", None)) or _date_to_iso(
        getattr(wf, "due_date", None)
    )
    internal_due = None
    workflow_internal_due = _date_to_iso(getattr(wf, "due_date", None))
    legal_due = _date_to_iso(getattr(wf, "legal_due_date", None))
    if workflow_internal_due and legal_due and workflow_internal_due != legal_due:
        internal_due = workflow_internal_due

    # Legacy/imported linked workflows can still fall back to docket due dates when
    # the workflow row has never been initialized.
    if linked_docket_item is not None and final_due is None and internal_due is None:
        final_due = _date_to_iso(
            adjusted_legal_due_for_docket(
                getattr(linked_docket_item, "due_date", None),
                getattr(linked_docket_item, "extended_due_date", None),
            )
        )
        internal_due = _date_to_iso(
            internal_due_for_docket(
                getattr(linked_docket_item, "due_date", None),
                getattr(linked_docket_item, "extended_due_date", None),
            )
        )

    if final_due and internal_due and final_due == internal_due:
        internal_due = None
    return final_due, internal_due


def _workflow_task_primary_due_date(
    *,
    final_due_date: str | None,
    internal_due_date: str | None,
    due_axis: str,
) -> str | None:
    normalized_axis = _normalize_worklog_due_axis(due_axis, default="all")
    if normalized_axis == "final":
        return final_due_date
    if normalized_axis == "internal":
        return internal_due_date
    return internal_due_date or final_due_date


def _split_csv_tokens(raw: str | None) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for token in str(raw or "").split(","):
        item = token.strip()
        if not item or item in seen:
            continue
        seen.add(item)
        out.append(item)
    return out


def _staff_names_for_export(rows: list[dict] | None, fallback: str | None = None) -> str:
    names: list[str] = []
    seen: set[str] = set()
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        name = str(row.get("name") or "").strip()
        if not name or name in seen:
            continue
        seen.add(name)
        names.append(name)
    if names:
        return ", ".join(names)
    if fallback:
        return ", ".join(_split_csv_tokens(fallback))
    return ""


def _worklog_tasks_xlsx_response(
    *,
    tasks: list[dict],
    total_count: int,
    scope: str,
    filter_type: str,
    category_filter: str,
    due_axis: str,
    owner_role: str,
    owner_filter: str,
) -> object:
    status_label = {
        "pending": "Waiting",
        "urgent": "",
        "overdue": "Deadline Overdue",
        "completed": "Done",
        "abandoned": "Task ",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Task"

    headers = [
        "Matter reference",
        "Task",
        "category",
        "Status",
        "Display Due date",
        "Final Due date",
        "Internal Due date",
        "Task Contact",
        "Responsible attorney",
        "Handler",
        "Manager",
        "Matter ID",
        "Task ID",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for task in tasks or []:
        owners = _staff_names_for_export(
            task.get("owners"), str(task.get("owner_name") or task.get("owner_id") or "")
        )
        attorneys = _staff_names_for_export(
            task.get("attorneys"), str(task.get("attorney_names") or "")
        )
        handlers = _staff_names_for_export(
            task.get("handlers"), str(task.get("handler_names") or "")
        )
        managers = _staff_names_for_export(
            task.get("managers"), str(task.get("manager_names") or "")
        )

        raw_status = str(task.get("status") or "").strip().lower()
        ws.append(
            [
                str(task.get("our_ref") or "").strip(),
                str(task.get("task_name") or "").strip(),
                str(task.get("category_display") or task.get("category_type") or "").strip(),
                status_label.get(raw_status, raw_status),
                str(task.get("due_date") or "").strip(),
                str(task.get("final_due_date") or "").strip(),
                str(task.get("internal_due_date") or "").strip(),
                owners,
                attorneys,
                handlers,
                managers,
                str(task.get("matter_id") or "").strip(),
                str(task.get("workflow_link_id") or task.get("id") or "").strip(),
            ]
        )

    ws.freeze_panes = "A2"

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            value_len = len(str(getattr(cell, "value", "") or ""))
            if value_len > max_length:
                max_length = value_len
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    meta = wb.create_sheet("Meta")
    meta.append(["Scope", "all" if scope == "all" else "page"])
    meta.append(["Matched Total", int(total_count or 0)])
    meta.append(["Exported Rows", len(tasks or [])])
    meta.append(["Filter", filter_type or "todo"])
    meta.append(["Category", category_filter or ""])
    meta.append(["Due Axis", due_axis or "all"])
    meta.append(["Owner Role", owner_role or "owner"])
    meta.append(["Owner", owner_filter or ""])
    meta.append(["Export Date", datetime.utcnow().isoformat(timespec="seconds") + "Z"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"worklog_tasks_{scope}_{ts}.xlsx"
    resp = current_app.response_class(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _worklog_visibility_scope() -> tuple[str | None, dict[str, bool], list[object]]:
    user_role = current_user.role if current_user.is_authenticated else None
    flags = worklog_role_scope_flags(current_user)
    show_all_mgmt = flags["show_all_mgmt"]
    show_all_work = flags["show_all_work"]
    show_own_mgmt = flags["show_own_mgmt"]
    show_own_work = flags["show_own_work"]

    visibility_conditions: list[object] = []
    cat_upper = func.upper(Workflow.category)
    is_mine = workflow_user_filter(getattr(current_user, "id", None))

    if show_all_mgmt:
        visibility_conditions.append(cat_upper.in_(MGMT_CATEGORIES))
    elif show_own_mgmt:
        visibility_conditions.append(and_(cat_upper.in_(MGMT_CATEGORIES), is_mine))

    if show_all_work:
        visibility_conditions.append(cat_upper.in_(WORK_CATEGORIES))
    elif show_own_work:
        visibility_conditions.append(and_(cat_upper.in_(WORK_CATEGORIES), is_mine))

    if not show_all_work:
        # H-1 fix: managed_ids fallback show_own_work  (Administrator Role)
        #  Add . show_own_work True is_mine itemsto ,
        # managed matters OR Add General Contact   Data .
        if not show_own_work:
            try:
                managed_ids = managed_matter_ids_select(current_user)
                visibility_conditions.append(
                    and_(cat_upper.in_(WORK_CATEGORIES), Workflow.case_id.in_(managed_ids))
                )
            except Exception as exc:
                report_swallowed_exception(
                    exc,
                    context="worklog.routes.visibility_scope.managed_matter_ids_select",
                    log_key="worklog.routes.visibility_scope.managed_matter_ids_select",
                    log_window_seconds=300,
                )

    return user_role, flags, visibility_conditions


def _build_worklog_workflow_query(
    *,
    today: date,
    urgent_date: date,
    end_date: date,
    category_filter: str = "",
    owner_filter: str = "",
    owner_role: str = "owner",
    owner_name: str | None = None,
    due_axis: str = "all",
    due_from: date | None = None,
    due_to: date | None = None,
    filter_type: str | None = None,
    bucket: str = "",
    sort_col: str | None = None,
    sort_dir: str | None = None,
    raw_limit: int | None = None,
) -> tuple[object, set[str], str | None]:
    due_expr = _worklog_due_expr_for_axis(due_axis)
    q = db.session.query(Workflow, Matter).join(Matter, Workflow.case_id == Matter.matter_id)

    q = q.filter(
        or_(
            Workflow.business_code.is_(None),
            not_(Workflow.business_code.like("ANNUITY:%")),
        )
    )

    if filter_type == "todo" and not bucket:
        # H-3 fix: bucket   bucket  Status+ items 
        # filter_type=todo end_date Filter Duplicate Apply .
        q = q.filter(Workflow.status.notin_(["Completed", "Abandoned"]))
        q = q.filter(due_expr.isnot(None))
        q = q.filter(due_expr <= end_date)
    elif filter_type == "todo" and bucket:
        # bucket  Status Filter Apply ( bucket from Process)
        q = q.filter(Workflow.status.notin_(["Completed", "Abandoned"]))
    elif filter_type == "completed":
        q = q.filter(Workflow.status.in_(["Completed", "Abandoned"]))

    if due_from is not None or due_to is not None:
        normalized_axis = _normalize_worklog_due_axis(due_axis, default="all")
        if normalized_axis == "all":
            # When due_axis=all, match if EITHER the final or internal due date
            # falls within the requested range.  Using only the coalesced
            # expression would silently hide tasks whose *other* due date is
            # in range (e.g. final=2026-04-20 hidden because internal=2026-07-20).
            final_expr = _worklog_final_due_expr()
            internal_expr = _worklog_internal_due_expr()

            def _axis_in_range(expr):
                conditions = [expr.isnot(None)]
                if due_from is not None:
                    conditions.append(expr >= due_from)
                if due_to is not None:
                    conditions.append(expr <= due_to)
                return and_(*conditions)

            q = q.filter(or_(_axis_in_range(final_expr), _axis_in_range(internal_expr)))
        else:
            # For explicit axis (final / internal), the single expression is correct.
            if due_from is not None:
                q = q.filter(due_expr.isnot(None))
                q = q.filter(due_expr >= due_from)
            if due_to is not None:
                q = q.filter(due_expr.isnot(None))
                q = q.filter(due_expr <= due_to)

    if bucket == "urgent":
        q = q.filter(Workflow.status.notin_(["Completed", "Abandoned"]))
        q = q.filter(due_expr.isnot(None))
        q = q.filter(due_expr >= today, due_expr <= urgent_date)
    elif bucket == "overdue":
        q = q.filter(Workflow.status.notin_(["Completed", "Abandoned"]))
        q = q.filter(due_expr.isnot(None))
        q = q.filter(due_expr < today)
    elif bucket == "completed_week":
        week_ago = today - timedelta(days=7)
        q = q.filter(Workflow.status.in_(["Completed", "Abandoned"]))
        q = q.filter(Workflow.completed_date.isnot(None))
        q = q.filter(Workflow.completed_date >= week_ago)

    cat_upper = func.upper(Workflow.category)
    if category_filter == "mgmt":
        q = q.filter(cat_upper.in_(MGMT_CATEGORIES))
    elif category_filter == "work":
        q = q.filter(cat_upper.in_(WORK_CATEGORIES))

    owner_filter_ids_for_owner_role: set[str] = set()
    if owner_filter:
        if owner_role == "owner":
            owner_filter_ids_for_owner_role = _owner_filter_candidate_ids(owner_filter)
            if not owner_filter_ids_for_owner_role:
                q = q.filter(text("1=0"))
        else:
            staff_party_id = _resolve_staff_party_id(owner_filter)
            staff_name = (
                _resolve_staff_display_name(staff_party_id) or (owner_name or "").strip() or None
            )
            owner_user_id = _resolve_owner_user_id(owner_filter)
            filter_conditions = []
            if staff_party_id:
                filter_conditions.append(
                    _case_staff_match_condition(
                        owner_role=owner_role,
                        staff_party_id=staff_party_id,
                        staff_name=staff_name,
                    )
                )
            if owner_user_id is not None:
                if owner_role == "attorney":
                    filter_conditions.append(Workflow.attorney_assignee_id == owner_user_id)
                elif owner_role == "handler":
                    filter_conditions.append(Workflow.assignee_id == owner_user_id)
                elif owner_role == "manager":
                    filter_conditions.append(Workflow.inspector_id == owner_user_id)
                elif owner_role == "any":
                    filter_conditions.append(
                        or_(
                            Workflow.attorney_assignee_id == owner_user_id,
                            Workflow.assignee_id == owner_user_id,
                            Workflow.inspector_id == owner_user_id,
                        )
                    )
            if filter_conditions:
                q = q.filter(or_(*filter_conditions))
            else:
                q = q.filter(text("1=0"))

    user_role, flags, visibility_conditions = _worklog_visibility_scope()
    if not (flags["show_all_mgmt"] and flags["show_all_work"]):
        if visibility_conditions:
            q = q.filter(or_(*visibility_conditions))
        else:
            q = q.filter(text("1=0"))

    if sort_col == "our_ref":
        q = q.order_by(Matter.our_ref.desc() if sort_dir == "desc" else Matter.our_ref.asc())
    elif sort_col:
        q = q.order_by(due_expr.desc() if sort_dir == "desc" else due_expr.asc())

    if raw_limit is not None and raw_limit > 0:
        q = q.limit(raw_limit)

    return q, owner_filter_ids_for_owner_role, user_role


def _iter_worklog_row_batches(
    q,
    *,
    today: date,
    batch_size: int = 200,
):
    batch: list[tuple[Workflow, Matter]] = []
    for row in q.yield_per(batch_size):
        batch.append(row)
        if len(batch) < batch_size:
            continue
        filtered = _filter_hidden_workflow_rows(batch, today=today)
        if filtered:
            yield filtered
        batch = []

    if batch:
        filtered = _filter_hidden_workflow_rows(batch, today=today)
        if filtered:
            yield filtered


def _load_docket_task_context_for_rows(
    rows: list[tuple[Workflow, Matter]],
    *,
    include_completion_recommendations: bool,
) -> dict[str, object]:
    workflow_docket_id_by_wf_id: dict[int, str] = {}
    workflow_case_ids_by_docket_id: dict[str, set[str]] = {}

    for wf, _m in rows:
        did = _workflow_docket_id(wf)
        if not did:
            continue
        workflow_docket_id_by_wf_id[int(wf.id)] = did
        workflow_case_ids_by_docket_id.setdefault(did, set()).add(
            str(getattr(wf, "case_id", "") or "")
        )

    docket_item_by_docket_id: dict[str, DocketItem] = {}
    notice_send_recommendation_by_docket_id: dict[str, dict] = {}

    if workflow_case_ids_by_docket_id:
        docket_q = DocketItem.query.filter(
            DocketItem.docket_id.in_(sorted(workflow_case_ids_by_docket_id.keys()))
        )
        if hasattr(DocketItem, "is_deleted"):
            docket_q = docket_q.filter(
                or_(DocketItem.is_deleted.is_(False), DocketItem.is_deleted.is_(None))
            )

        docket_rows = docket_q.all()
        for di in docket_rows:
            did = str(getattr(di, "docket_id", "") or "").strip()
            if not did:
                continue
            expected_mids = workflow_case_ids_by_docket_id.get(did) or set()
            if expected_mids and str(getattr(di, "matter_id", "") or "") not in expected_mids:
                continue
            docket_item_by_docket_id[did] = di

        if include_completion_recommendations and docket_item_by_docket_id:
            try:
                from app.services.deadlines.notice_send_semi_close import (
                    get_notice_send_recommendation_state,
                    infer_notice_send_prompt_candidate_from_communications,
                    load_notice_send_communications_for_matters,
                )

                for did, di in docket_item_by_docket_id.items():
                    notice_send_recommendation_by_docket_id[did] = (
                        get_notice_send_recommendation_state(di)
                    )

                fallback_docket_ids = [
                    did
                    for did, di in docket_item_by_docket_id.items()
                    if not bool(
                        (notice_send_recommendation_by_docket_id.get(did) or {}).get("recommended")
                    )
                    and not bool(str(getattr(di, "done_date", "") or "").strip())
                ]
                fallback_matter_ids = {
                    str(getattr(docket_item_by_docket_id[did], "matter_id", "") or "").strip()
                    for did in fallback_docket_ids
                    if did in docket_item_by_docket_id
                }
                fallback_matter_ids.discard("")

                comm_rows_by_matter_id: dict[str, list[dict]] = {}
                if fallback_matter_ids:
                    comm_rows_by_matter_id = load_notice_send_communications_for_matters(
                        matter_ids=sorted(fallback_matter_ids)
                    )

                for did in fallback_docket_ids:
                    di = docket_item_by_docket_id.get(did)
                    if not di:
                        continue
                    mid = str(getattr(di, "matter_id", "") or "").strip()
                    if not mid:
                        continue
                    inferred = infer_notice_send_prompt_candidate_from_communications(
                        docket_items=[di],
                        communications=comm_rows_by_matter_id.get(mid) or [],
                        respect_prompted=False,
                    )
                    if not inferred:
                        continue
                    if str(inferred.get("docket_id") or "").strip() != did:
                        continue

                    base = dict(notice_send_recommendation_by_docket_id.get(did) or {})
                    base["recommended"] = True
                    matched_doc_name = str(inferred.get("matched_doc_name") or "").strip()
                    if matched_doc_name and not str(base.get("trigger_doc_name") or "").strip():
                        base["trigger_doc_name"] = matched_doc_name
                    base["inferred"] = True
                    notice_send_recommendation_by_docket_id[did] = base
            except Exception as exc:
                report_swallowed_exception(
                    exc,
                    context="worklog.routes.notice_send_recommendation_lookup",
                    log_key="worklog.routes.notice_send_recommendation_lookup",
                    log_window_seconds=300,
                )

    matter_ids = {
        str(getattr(wf, "case_id", "") or "").strip()
        for wf, _matter in rows
        if str(getattr(wf, "case_id", "") or "").strip()
    }

    return {
        "workflow_docket_id_by_wf_id": workflow_docket_id_by_wf_id,
        "docket_item_by_docket_id": docket_item_by_docket_id,
        "notice_send_recommendation_by_docket_id": notice_send_recommendation_by_docket_id,
        "display_context": _load_worklog_matter_display_context(matter_ids),
    }


def _workflow_needs_linked_docket_due_fallback(wf: Workflow) -> bool:
    return (
        _date_to_iso(getattr(wf, "legal_due_date", None)) is None
        and _date_to_iso(getattr(wf, "due_date", None)) is None
    )


def _load_linked_docket_items_for_rows(
    rows: list[tuple[Workflow, Matter]],
    *,
    due_fallback_only: bool = False,
) -> tuple[dict[int, str], dict[str, DocketItem], dict[int, DocketItem]]:
    workflow_docket_id_by_wf_id: dict[int, str] = {}
    workflow_case_ids_by_docket_id: dict[str, set[str]] = {}

    for wf, _matter in rows:
        did = _workflow_docket_id(wf)
        if not did:
            continue
        wf_id = int(getattr(wf, "id", 0) or 0)
        if wf_id <= 0:
            continue
        workflow_docket_id_by_wf_id[wf_id] = did
        if due_fallback_only and not _workflow_needs_linked_docket_due_fallback(wf):
            continue
        workflow_case_ids_by_docket_id.setdefault(did, set()).add(
            str(getattr(wf, "case_id", "") or "").strip()
        )

    docket_item_by_docket_id: dict[str, DocketItem] = {}
    workflow_docket_item_by_wf_id: dict[int, DocketItem] = {}
    if not workflow_case_ids_by_docket_id:
        return workflow_docket_id_by_wf_id, docket_item_by_docket_id, workflow_docket_item_by_wf_id

    docket_q = DocketItem.query.filter(
        DocketItem.docket_id.in_(sorted(workflow_case_ids_by_docket_id.keys()))
    )
    if hasattr(DocketItem, "is_deleted"):
        docket_q = docket_q.filter(
            or_(DocketItem.is_deleted.is_(False), DocketItem.is_deleted.is_(None))
        )

    for di in docket_q.all():
        did = str(getattr(di, "docket_id", "") or "").strip()
        if not did:
            continue
        expected_mids = workflow_case_ids_by_docket_id.get(did) or set()
        if expected_mids and str(getattr(di, "matter_id", "") or "").strip() not in expected_mids:
            continue
        docket_item_by_docket_id[did] = di

    for wf_id, did in workflow_docket_id_by_wf_id.items():
        di = docket_item_by_docket_id.get(did)
        if di is not None:
            workflow_docket_item_by_wf_id[wf_id] = di

    return workflow_docket_id_by_wf_id, docket_item_by_docket_id, workflow_docket_item_by_wf_id


def _load_workflow_users_for_rows(rows: list[tuple[Workflow, Matter]]) -> dict[int, User]:
    user_ids: set[int] = set()
    for wf, _matter in rows:
        for raw in (
            getattr(wf, "assignee_id", None),
            getattr(wf, "attorney_assignee_id", None),
            getattr(wf, "inspector_id", None),
        ):
            try:
                uid = int(raw or 0)
            except Exception:
                uid = 0
            if uid > 0:
                user_ids.add(uid)
    if not user_ids:
        return {}
    return {
        int(user.id): user
        for user in User.query.filter(User.id.in_(sorted(user_ids))).all()
        if int(getattr(user, "id", 0) or 0) > 0
    }


def _build_worklog_tasks(
    *,
    q,
    today: date,
    urgent_date: date,
    due_axis: str = "all",
    search_query: str = "",
    owner_filter: str = "",
    owner_role: str = "owner",
    owner_filter_ids_for_owner_role: set[str] | None = None,
    mine_only: bool = False,
    include_completion_recommendations: bool = True,
    page_start: int | None = None,
    page_limit: int | None = None,
    recommended_only: bool = False,
) -> list[dict] | tuple[list[dict], int]:
    grouped: dict[str, dict] = {}
    ordered_keys: list[str] = []

    for rows in _iter_worklog_row_batches(q, today=today):
        batch_context = _load_docket_task_context_for_rows(
            rows,
            include_completion_recommendations=include_completion_recommendations,
        )
        workflow_users_by_id = _load_workflow_users_for_rows(rows)
        workflow_docket_id_by_wf_id = dict(batch_context.get("workflow_docket_id_by_wf_id") or {})
        docket_item_by_docket_id = dict(batch_context.get("docket_item_by_docket_id") or {})
        notice_send_recommendation_by_docket_id = dict(
            batch_context.get("notice_send_recommendation_by_docket_id") or {}
        )
        display_context = dict(batch_context.get("display_context") or {})
        matter_applicant_map = dict(display_context.get("matter_applicant_map") or {})
        matter_applicant_client_id_map = dict(
            display_context.get("matter_applicant_client_id_map") or {}
        )
        matter_attorney_map = dict(display_context.get("matter_attorney_map") or {})
        matter_handler_map = dict(display_context.get("matter_handler_map") or {})
        matter_manager_map = dict(display_context.get("matter_manager_map") or {})
        matter_staff_map = dict(display_context.get("matter_staff_map") or {})
        batch_tasks: list[dict] = []

        for wf, m in rows:
            cat_type, cat_display = _workflow_category_badge_values(getattr(wf, "category", None))
            assignee_user = workflow_users_by_id.get(int(getattr(wf, "assignee_id", 0) or 0))
            attorney_assignee_user = workflow_users_by_id.get(
                int(getattr(wf, "attorney_assignee_id", 0) or 0)
            )
            inspector_user = workflow_users_by_id.get(int(getattr(wf, "inspector_id", 0) or 0))
            owner_user_id = workflow_primary_owner_user_id(
                category=getattr(wf, "category", None),
                handler_id=getattr(wf, "assignee_id", None),
                attorney_id=getattr(wf, "attorney_assignee_id", None),
                manager_id=getattr(wf, "inspector_id", None),
            )
            owner_user = (
                (
                    workflow_users_by_id.get(int(owner_user_id or 0))
                    if owner_user_id is not None
                    else None
                )
                or assignee_user
                or attorney_assignee_user
                or inspector_user
            )

            attorneys_list = list(matter_staff_map.get(wf.case_id, {}).get("attorney", []) or [])
            handlers_list = list(matter_staff_map.get(wf.case_id, {}).get("handler", []) or [])
            managers_list = list(matter_staff_map.get(wf.case_id, {}).get("manager", []) or [])
            effective_role_rows = _resolve_effective_staff_rows_for_task(
                wf=wf,
                attorneys_list=attorneys_list,
                handlers_list=handlers_list,
                managers_list=managers_list,
                assignee_user=assignee_user,
                attorney_assignee_user=attorney_assignee_user,
                inspector_user=inspector_user,
            )
            attorneys_list = list(effective_role_rows.get("attorney") or [])
            handlers_list = list(effective_role_rows.get("handler") or [])
            managers_list = list(effective_role_rows.get("manager") or [])
            attorneys = _staff_row_names(attorneys_list) or matter_attorney_map.get(wf.case_id, "")
            handlers = _staff_row_names(handlers_list) or matter_handler_map.get(wf.case_id, "")
            managers = _staff_row_names(managers_list) or matter_manager_map.get(wf.case_id, "")
            linked_docket_id = workflow_docket_id_by_wf_id.get(int(wf.id))
            linked_docket_item = (
                docket_item_by_docket_id.get(linked_docket_id) if linked_docket_id else None
            )
            final_due_date, internal_due_date = _workflow_task_due_dates(
                wf=wf,
                linked_docket_item=linked_docket_item,
            )
            primary_due_date = _workflow_task_primary_due_date(
                final_due_date=final_due_date,
                internal_due_date=internal_due_date,
                due_axis=due_axis,
            )
            status = _workflow_list_status(
                wf=wf,
                today=today,
                urgent_date=urgent_date,
                due_date=_parse_iso_date_param(primary_due_date),
            )
            owner_rows, owner_decision = _resolve_owner_rows_for_task(
                wf=wf,
                linked_docket_item=linked_docket_item,
                owner_user=owner_user,
                attorneys_list=attorneys_list,
                handlers_list=handlers_list,
                managers_list=managers_list,
                assignee_user=assignee_user,
                attorney_assignee_user=attorney_assignee_user,
                inspector_user=inspector_user,
            )
            if _decision_is_manager_only(owner_decision):
                cat_type, cat_display = ("mgmt", "")
            owner_names = _merge_csv_names(
                [str(p.get("name") or "") for p in owner_rows if str(p.get("name") or "").strip()]
            )
            owner_id = (
                str(owner_rows[0].get("id") or "").strip()
                if len(owner_rows) == 1 and str(owner_rows[0].get("id") or "").strip()
                else None
            )
            notice_send_reco = (
                notice_send_recommendation_by_docket_id.get(linked_docket_id, {})
                if linked_docket_id
                else {}
            )
            completion_recommendation = bool(notice_send_reco.get("recommended"))
            trigger_doc_name = str(notice_send_reco.get("trigger_doc_name") or "").strip()
            completion_recommendation_text = ""
            if completion_recommendation:
                if trigger_doc_name:
                    completion_recommendation_text = f"Suggested completion match: {trigger_doc_name}"
                else:
                    completion_recommendation_text = "Suggested completion based on client notice."

            task = {
                "id": f"wf_{wf.id}",
                "workflow_link_id": f"wf_{wf.id}",
                "matter_id": wf.case_id,
                "our_ref": m.our_ref if m else "",
                "applicant_name": matter_applicant_map.get(str(wf.case_id or "").strip(), ""),
                "applicant_client_id": matter_applicant_client_id_map.get(
                    str(wf.case_id or "").strip(), ""
                ),
                "task_name": _workflow_display_task_name(wf),
                "category": wf.category or "",
                "category_type": cat_type,
                "category_display": cat_display,
                "due_axis": _normalize_worklog_due_axis(due_axis, default="all"),
                "due_date": primary_due_date,
                "original_due": primary_due_date,
                "extended_due": internal_due_date,
                "final_due_date": final_due_date,
                "internal_due_date": internal_due_date,
                "done_date": wf.completed_date.isoformat() if wf.completed_date else None,
                "status": status,
                "owner_id": owner_id,
                "owner_name": owner_names,
                "owners": owner_rows,
                "attorneys": attorneys_list,
                "handlers": handlers_list,
                "managers": managers_list,
                "attorney_names": attorneys,
                "handler_names": handlers,
                "manager_names": managers,
                "_workflow_assignee_filter_ids": _workflow_assignee_filter_ids(
                    assignee_id=getattr(wf, "assignee_id", None),
                    assignee_user=assignee_user,
                    attorney_assignee_id=getattr(wf, "attorney_assignee_id", None),
                    attorney_assignee_user=attorney_assignee_user,
                    inspector_id=getattr(wf, "inspector_id", None),
                    inspector_user=inspector_user,
                ),
                "memo": wf.note or "",
                "worklog_description": wf.note,
                "worklog_completed_at": None,
                "completion_recommendation": completion_recommendation,
                "completion_recommendation_kind": (
                    "notice_send_semi_auto" if completion_recommendation else None
                ),
                "completion_recommendation_text": completion_recommendation_text,
                "_linked_docket_id": linked_docket_id,
                "_auto_docket_generated": _workflow_is_auto_docket_generated(wf),
                "_intake_confirmation_task": _workflow_is_intake_confirmation_task(wf),
            }
            batch_tasks.append(task)

        intake_access_reco_by_task_id = _load_intake_case_access_recommendation_by_task_id(
            batch_tasks
        )
        for task in batch_tasks:
            intake_reco = intake_access_reco_by_task_id.get(str(task.get("id") or "").strip()) or {}
            if intake_reco and not bool(task.get("completion_recommendation")):
                task["completion_recommendation"] = True
                task["completion_recommendation_kind"] = "intake_case_access"
                task["completion_recommendation_text"] = str(intake_reco.get("text") or "").strip()
            _merge_task_row_into_groups(task, grouped=grouped, ordered_keys=ordered_keys)

    normalized_due_axis = _normalize_worklog_due_axis(due_axis, default="all")
    search_expression = _parse_worklog_search_expression(search_query) if search_query else None
    owner_filter_ids: set[str] = set()
    if owner_filter and owner_role in {"attorney", "handler", "manager", "any"}:
        owner_filter_ids = _owner_filter_candidate_ids(owner_filter)
    if search_query:
        search_expression = _parse_worklog_search_expression(search_query)

    def _include_task(task: dict) -> bool:
        if (
            normalized_due_axis in {"final", "internal"}
            and not str(task.get("due_date") or "").strip()
        ):
            return False
        if search_query and not _matches_worklog_search_query(
            search_text=_task_search_text(task),
            search_query=search_query,
            field_values=_task_search_field_values(task),
            search_expression=search_expression,
        ):
            return False
        if mine_only and owner_filter:
            candidate_ids = owner_filter_ids_for_owner_role or set()
            if not candidate_ids or not _task_matches_mine_filter(task, candidate_ids):
                return False
        elif owner_filter and owner_role == "owner":
            candidate_ids = owner_filter_ids_for_owner_role or set()
            if not candidate_ids or not _task_matches_owner_filter(task, candidate_ids):
                return False
        elif owner_filter and owner_role in {"attorney", "handler", "manager", "any"}:
            if not owner_filter_ids or not _task_matches_staff_role_filter(
                task, owner_filter_ids, owner_role=owner_role
            ):
                return False
        if recommended_only and not _task_has_completion_recommendation(task):
            return False
        return True

    if page_start is None:
        tasks: list[dict] = []
        for key in ordered_keys:
            task = _finalize_merged_task_bucket(grouped[key])
            if _include_task(task):
                _strip_internal_task_filter_fields(task)
                tasks.append(task)
        _attach_assignment_request_badges(tasks)
        return tasks

    start = max(0, int(page_start or 0))
    limit = max(0, int(page_limit or 0))
    page_tasks: list[dict] = []
    total_count = 0
    for key in ordered_keys:
        task = _finalize_merged_task_bucket(grouped[key])
        if not _include_task(task):
            continue
        if total_count >= start and len(page_tasks) < limit:
            _strip_internal_task_filter_fields(task)
            page_tasks.append(task)
        total_count += 1
    _attach_assignment_request_badges(page_tasks)
    return page_tasks, total_count


def _build_worklog_tasks_light(
    *,
    rows: list[tuple[Workflow, Matter]],
    today: date,
    urgent_date: date,
    due_axis: str = "all",
    search_query: str = "",
    owner_filter: str = "",
    owner_role: str = "owner",
    owner_filter_ids_for_owner_role: set[str] | None = None,
    mine_only: bool = False,
) -> list[dict]:
    filtered_rows = _filter_hidden_workflow_rows(rows, today=today)
    if not filtered_rows:
        return []

    workflows = [wf for wf, _matter in filtered_rows]
    normalized_due_axis = _normalize_worklog_due_axis(due_axis, default="all")
    need_owner_rows = bool(search_query) or bool(owner_filter and owner_role == "owner")
    owner_rows_by_wf = _resolve_owner_rows_for_workflows(workflows) if need_owner_rows else {}

    display_context: dict[str, object] = {}
    need_staff_context = bool(search_query) or bool(owner_filter)
    if need_staff_context:
        matter_ids = {
            str(getattr(wf, "case_id", "") or "").strip()
            for wf, _matter in filtered_rows
            if str(getattr(wf, "case_id", "") or "").strip()
        }
        display_context = _load_worklog_matter_display_context(matter_ids)
    workflow_users_by_id = (
        _load_workflow_users_for_rows(filtered_rows) if need_staff_context else {}
    )

    matter_applicant_map = dict(display_context.get("matter_applicant_map") or {})
    matter_attorney_map = dict(display_context.get("matter_attorney_map") or {})
    matter_handler_map = dict(display_context.get("matter_handler_map") or {})
    matter_manager_map = dict(display_context.get("matter_manager_map") or {})
    matter_staff_map = dict(display_context.get("matter_staff_map") or {})

    workflow_docket_id_by_wf_id, _docket_item_by_docket_id, docket_item_by_wf_id = (
        _load_linked_docket_items_for_rows(filtered_rows, due_fallback_only=True)
    )

    grouped: dict[str, dict] = {}
    ordered_keys: list[str] = []

    for wf, matter in filtered_rows:
        wf_id = int(getattr(wf, "id", 0) or 0)
        cat_type, cat_display = _workflow_category_badge_values(getattr(wf, "category", None))
        linked_docket_id = workflow_docket_id_by_wf_id.get(wf_id)
        linked_docket_item = docket_item_by_wf_id.get(wf_id)
        final_due_date, internal_due_date = _workflow_task_due_dates(
            wf=wf,
            linked_docket_item=linked_docket_item,
        )
        primary_due_date = _workflow_task_primary_due_date(
            final_due_date=final_due_date,
            internal_due_date=internal_due_date,
            due_axis=normalized_due_axis,
        )
        status = _workflow_list_status(
            wf=wf,
            today=today,
            urgent_date=urgent_date,
            due_date=_parse_iso_date_param(primary_due_date),
        )

        owner_rows = list(owner_rows_by_wf.get(wf_id) or [])
        owner_names = _merge_csv_names(
            [str(p.get("name") or "") for p in owner_rows if str(p.get("name") or "").strip()]
        )
        owner_id = (
            str(owner_rows[0].get("id") or "").strip()
            if len(owner_rows) == 1 and str(owner_rows[0].get("id") or "").strip()
            else None
        )

        matter_id = str(getattr(wf, "case_id", "") or "").strip()
        assignee_user = workflow_users_by_id.get(int(getattr(wf, "assignee_id", 0) or 0))
        attorney_assignee_user = workflow_users_by_id.get(
            int(getattr(wf, "attorney_assignee_id", 0) or 0)
        )
        inspector_user = workflow_users_by_id.get(int(getattr(wf, "inspector_id", 0) or 0))
        attorneys_list = list(matter_staff_map.get(matter_id, {}).get("attorney", []) or [])
        handlers_list = list(matter_staff_map.get(matter_id, {}).get("handler", []) or [])
        managers_list = list(matter_staff_map.get(matter_id, {}).get("manager", []) or [])
        if need_staff_context:
            effective_role_rows = _resolve_effective_staff_rows_for_task(
                wf=wf,
                attorneys_list=attorneys_list,
                handlers_list=handlers_list,
                managers_list=managers_list,
                assignee_user=assignee_user,
                attorney_assignee_user=attorney_assignee_user,
                inspector_user=inspector_user,
            )
            attorneys_list = list(effective_role_rows.get("attorney") or [])
            handlers_list = list(effective_role_rows.get("handler") or [])
            managers_list = list(effective_role_rows.get("manager") or [])
        task = {
            "id": f"wf_{wf_id}",
            "workflow_link_id": f"wf_{wf_id}",
            "matter_id": matter_id,
            "our_ref": getattr(matter, "our_ref", "") if matter else "",
            "applicant_name": matter_applicant_map.get(matter_id, ""),
            "task_name": _workflow_display_task_name(wf),
            "category": getattr(wf, "category", None) or "",
            "category_type": cat_type,
            "category_display": cat_display,
            "due_axis": normalized_due_axis,
            "due_date": primary_due_date,
            "original_due": primary_due_date,
            "extended_due": internal_due_date,
            "final_due_date": final_due_date,
            "internal_due_date": internal_due_date,
            "done_date": (
                getattr(wf, "completed_date", None).isoformat()
                if getattr(wf, "completed_date", None)
                else None
            ),
            "status": status,
            "owner_id": owner_id,
            "owner_name": owner_names,
            "owners": owner_rows,
            "attorneys": attorneys_list,
            "handlers": handlers_list,
            "managers": managers_list,
            "attorney_names": _staff_row_names(attorneys_list)
            or matter_attorney_map.get(matter_id, ""),
            "handler_names": _staff_row_names(handlers_list)
            or matter_handler_map.get(matter_id, ""),
            "manager_names": _staff_row_names(managers_list)
            or matter_manager_map.get(matter_id, ""),
            "_workflow_assignee_filter_ids": _workflow_assignee_filter_ids(
                assignee_id=getattr(wf, "assignee_id", None),
                assignee_user=assignee_user,
                attorney_assignee_id=getattr(wf, "attorney_assignee_id", None),
                attorney_assignee_user=attorney_assignee_user,
                inspector_id=getattr(wf, "inspector_id", None),
                inspector_user=inspector_user,
            ),
            "memo": getattr(wf, "note", None) or "",
            "worklog_description": getattr(wf, "note", None),
            "_linked_docket_id": linked_docket_id,
            "_auto_docket_generated": _workflow_is_auto_docket_generated(wf),
        }
        _merge_task_row_into_groups(task, grouped=grouped, ordered_keys=ordered_keys)

    tasks = [_finalize_merged_task_bucket(grouped[key]) for key in ordered_keys]

    if normalized_due_axis in {"final", "internal"}:
        tasks = [task for task in tasks if str(task.get("due_date") or "").strip()]

    if search_query:
        search_expression = _parse_worklog_search_expression(search_query)
        tasks = [
            task
            for task in tasks
            if _matches_worklog_search_query(
                search_text=_task_search_text(task),
                search_query=search_query,
                field_values=_task_search_field_values(task),
                search_expression=search_expression,
            )
        ]

    if mine_only and owner_filter:
        owner_filter_ids = owner_filter_ids_for_owner_role or set()
        if owner_filter_ids:
            tasks = [task for task in tasks if _task_matches_mine_filter(task, owner_filter_ids)]
        else:
            tasks = []
    elif owner_filter and owner_role == "owner":
        owner_filter_ids = owner_filter_ids_for_owner_role or set()
        if owner_filter_ids:
            tasks = [task for task in tasks if _task_matches_owner_filter(task, owner_filter_ids)]
        else:
            tasks = []
    elif owner_filter and owner_role in {"attorney", "handler", "manager", "any"}:
        owner_filter_ids = _owner_filter_candidate_ids(owner_filter)
        if owner_filter_ids:
            tasks = [
                task
                for task in tasks
                if _task_matches_staff_role_filter(task, owner_filter_ids, owner_role=owner_role)
            ]
        else:
            tasks = []

    for task in tasks:
        _strip_internal_task_filter_fields(task)

    _attach_assignment_request_badges(tasks)
    return tasks


def _summary_counts_from_tasks(
    tasks: list[dict],
    *,
    today: date,
    urgent_date: date,
    end_date: date,
) -> dict[str, int]:
    week_ago = today - timedelta(days=7)
    counts = {
        "pending": 0,
        "urgent": 0,
        "overdue": 0,
        "completed_week": 0,
    }

    for task in tasks:
        status = str(task.get("status") or "").strip().lower()
        due = _parse_iso_date_param(str(task.get("due_date") or ""))
        done = _parse_iso_date_param(str(task.get("done_date") or ""))
        is_done = status in ("completed", "abandoned")

        if (not is_done) and (due is not None) and due <= end_date:
            counts["pending"] += 1
        if (not is_done) and (due is not None) and due <= end_date and today <= due <= urgent_date:
            counts["urgent"] += 1
        if (not is_done) and (due is not None) and due < today:
            counts["overdue"] += 1
        if is_done and (done is not None) and done >= week_ago:
            counts["completed_week"] += 1

    return counts


# ============================================================
# HTML Pages
# ============================================================


@bp.route("/")
@login_required
def index():
    """Task    page."""
    return render_template("worklog/index.html")


# ============================================================
# API Endpoints
# ============================================================


@bp.route("/api/tasks")
@login_required
def api_tasks():
    """
    Task List Search API (Workflow Only).

    Existing DocketItem    Workflow  Search.
    Role  Filter:
    - :  MGMT Task
    - Manager:  MGMT Task
    - table:  WORK Task
    - PatentContact:  WORK Task
    """
    filter_type = request.args.get("filter", "todo")
    category_filter = request.args.get("category", "")
    owner_filter = request.args.get("owner", "")
    owner_role = _normalize_owner_role(request.args.get("owner_role") or request.args.get("role"))
    mine_only = (request.args.get("mine") or "").strip().lower() in ("1", "true", "yes")
    export_format = (request.args.get("format") or "").strip().lower()
    export_scope_raw = request.args.get("export_scope")
    export_scope = (export_scope_raw or "page").strip().lower()
    if export_scope not in ("page", "all"):
        export_scope = "page"
    raw_export_flag = (request.args.get("export") or "").strip().lower()
    export_requested = export_format in ("xlsx", "csv") or raw_export_flag in (
        "1",
        "true",
        "yes",
        "on",
    )
    bucket = (request.args.get("bucket") or "").strip().lower()
    days_param = request.args.get("days", "30")
    days = _parse_days_param(days_param, default=30, max_days=9999)
    due_axis = _normalize_worklog_due_axis(request.args.get("due_axis"), default="all")
    due_from = _parse_iso_date_param(request.args.get("due_from"))
    due_to = _parse_iso_date_param(request.args.get("due_to"))
    due_from, due_to = _normalize_due_range(due_from, due_to)
    if mine_only:
        owner_role = "owner"
        owner_filter = _current_worklog_mine_owner_value()

    today = _today_in_app_timezone()
    urgent_date = today + timedelta(days=URGENT_WINDOW_DAYS)
    end_date = today + timedelta(days=days)
    search_query = request.args.get("search", "").strip()
    sort_col = request.args.get("sort", "due_date")
    sort_dir = request.args.get("order", "asc")

    # Pagination parameters (actual pagination applied after merge).
    page = max(1, request.args.get("page", 1, type=int))
    limit = request.args.get("limit", 50, type=int)
    max_limit = 200
    if export_requested:
        raw_max = current_app.config.get("WORKLOG_XLSX_MAX_ROWS", 10000)
        try:
            max_limit = int(raw_max or 10000)
        except Exception:
            max_limit = 10000
        max_limit = max(100, min(max_limit, 50000))
    limit = max(1, min(limit, max_limit))

    raw_row_limit = max_limit if export_requested else None
    q, owner_filter_ids_for_owner_role, user_role = _build_worklog_workflow_query(
        today=today,
        urgent_date=urgent_date,
        end_date=end_date,
        category_filter=category_filter,
        owner_filter=owner_filter,
        owner_role=owner_role,
        owner_name=request.args.get("owner_name"),
        due_axis=due_axis,
        due_from=due_from,
        due_to=due_to,
        filter_type=filter_type,
        bucket=bucket,
        sort_col=sort_col,
        sort_dir=sort_dir,
        raw_limit=raw_row_limit,
    )
    export_all = export_requested and export_scope == "all"
    page_start = 0 if export_all else (page - 1) * limit
    page_limit = max_limit if export_all else limit
    tasks_result = _build_worklog_tasks(
        q=q,
        today=today,
        urgent_date=urgent_date,
        due_axis=due_axis,
        search_query=search_query,
        owner_filter=owner_filter,
        owner_role=owner_role,
        owner_filter_ids_for_owner_role=owner_filter_ids_for_owner_role,
        mine_only=mine_only,
        include_completion_recommendations=True,
        page_start=page_start,
        page_limit=page_limit,
        recommended_only=bucket == "recommended",
    )
    paged_tasks, total_count = tasks_result
    total_pages = max(1, (total_count + limit - 1) // limit) if limit > 0 else 1
    if not export_all and page > total_pages:
        page = total_pages
        if total_count > 0:
            tasks_result = _build_worklog_tasks(
                q=q,
                today=today,
                urgent_date=urgent_date,
                due_axis=due_axis,
                search_query=search_query,
                owner_filter=owner_filter,
                owner_role=owner_role,
                owner_filter_ids_for_owner_role=owner_filter_ids_for_owner_role,
                mine_only=mine_only,
                include_completion_recommendations=True,
                page_start=(page - 1) * limit,
                page_limit=limit,
                recommended_only=bucket == "recommended",
            )
            paged_tasks, total_count = tasks_result

    if export_requested:
        return _worklog_tasks_xlsx_response(
            tasks=paged_tasks,
            total_count=total_count,
            scope=export_scope,
            filter_type=filter_type,
            category_filter=category_filter,
            due_axis=due_axis,
            owner_role=owner_role,
            owner_filter=owner_filter,
        )

    return jsonify(
        {
            "tasks": paged_tasks,
            "total": total_count,
            "page": page,
            "total_pages": total_pages,
            "filter": filter_type,
            "due_axis": due_axis,
            "owner_role": owner_role,
            "user_role": user_role,
        }
    )


@bp.route("/api/calendar-events")
@login_required
def api_calendar_events():
    due_axis = _normalize_worklog_due_axis(request.args.get("due_axis"), default="all")
    mine_only = (request.args.get("mine") or "").strip().lower() in ("1", "true", "yes")
    include_done = (request.args.get("include_done") or "").strip().lower() in ("1", "true", "yes")
    owner_filter = (request.args.get("owner") or "").strip()
    owner_name = (request.args.get("owner_name") or "").strip() or None
    owner_role = _normalize_owner_role(request.args.get("owner_role") or "owner")
    if mine_only:
        owner_role = "owner"
        owner_filter = _current_worklog_mine_owner_value()

    today = _today_in_app_timezone()
    urgent_date = today + timedelta(days=URGENT_WINDOW_DAYS)
    default_window_days = 60
    start_date = _parse_iso_date_param(request.args.get("start")) or (
        today - timedelta(days=default_window_days)
    )
    end_date = _parse_iso_date_param(request.args.get("end")) or (
        today + timedelta(days=default_window_days)
    )
    try:
        requested_limit = int(request.args.get("limit", 5000))
    except Exception:
        requested_limit = 5000
    raw_limit = max(1, min(requested_limit, 5000))

    # Calendar should reflect effective Task dates, including linked-docket fallback.
    # Avoid DB-side due-date filtering here because some legacy linked workflows have
    # blank workflow due fields and are completed only after task materialization.
    q, owner_filter_ids_for_owner_role, _user_role = _build_worklog_workflow_query(
        today=today,
        urgent_date=urgent_date,
        end_date=end_date,
        owner_filter=owner_filter,
        owner_role=owner_role,
        owner_name=owner_name,
        due_axis=due_axis,
        due_from=None,
        due_to=None,
        filter_type=None,
        raw_limit=None,
    )
    q = q.options(*_worklog_calendar_query_options())
    q = q.filter(
        _worklog_calendar_due_range_condition(
            start_date=start_date,
            end_date=end_date,
            due_axis=due_axis,
        )
    )
    if not include_done:
        status_key = func.lower(func.coalesce(Workflow.status, ""))
        q = q.filter(~status_key.in_(["completed", "abandoned"]))
    q = q.limit(raw_limit)
    tasks = _build_worklog_tasks_light(
        rows=q.all(),
        today=today,
        urgent_date=urgent_date,
        due_axis=due_axis,
        owner_filter=owner_filter,
        owner_role=owner_role,
        owner_filter_ids_for_owner_role=owner_filter_ids_for_owner_role,
        mine_only=mine_only,
    )

    def axis_palette(axis: str) -> dict[str, str]:
        if axis == "internal":
            return {
                "backgroundColor": "#fff7ed",
                "borderColor": "#fed7aa",
                "textColor": "#c2410c",
            }
        return {
            "backgroundColor": "#eff6ff",
            "borderColor": "#bfdbfe",
            "textColor": "#1d4ed8",
        }

    def event_meta(*, status: str, axis: str) -> dict[str, object]:
        normalized_status = (status or "").strip().lower()
        if normalized_status not in {"pending", "urgent", "overdue", "completed", "abandoned"}:
            normalized_status = "pending"
        axis_key = "internal" if axis == "internal" else "final"
        axis_short_label = "Internal" if axis_key == "internal" else "Final"
        axis_label = "Internal Due date" if axis_key == "internal" else "Final Due date"
        status_label_map = {
            "pending": "OpenIn Progress",
            "urgent": "Due Soon",
            "overdue": "Deadline Overdue",
            "completed": "Done",
            "abandoned": "In Progress",
        }
        return {
            "status": normalized_status,
            "status_label": status_label_map.get(normalized_status, "OpenIn Progress"),
            "due_axis": axis_key,
            "due_axis_label": axis_label,
            "axis_short_label": axis_short_label,
            "classNames": [
                "deadline-calendar-event",
                f"deadline-calendar-event--axis-{axis_key}",
                f"deadline-calendar-event--status-{normalized_status}",
            ],
            **axis_palette(axis_key),
        }

    events: list[dict[str, object]] = []
    start_token = start_date.isoformat()
    end_token = end_date.isoformat()

    def _token_in_range(value: str | None) -> bool:
        token = str(value or "").strip()
        if not token:
            return False
        return start_token <= token <= end_token

    for task in tasks:
        workflow_link_id = str(task.get("workflow_link_id") or "").strip()
        workflow_id = (
            workflow_link_id.removeprefix("wf_") if workflow_link_id.startswith("wf_") else ""
        )
        status = str(task.get("status") or "").strip().lower()
        if not include_done and status in {"completed", "abandoned"}:
            continue
        detail_url = (
            url_for("workflow.detail", workflow_id=int(workflow_id))
            if workflow_id.isdigit()
            else url_for("worklog.index")
        )
        our_ref = str(task.get("our_ref") or "").strip()
        task_name = str(task.get("task_name") or "").strip() or str(task.get("id") or "").strip()
        base_title = f"[{our_ref}] {task_name}" if our_ref else task_name
        final_due_date = str(task.get("final_due_date") or "").strip() or None
        internal_due_date = str(task.get("internal_due_date") or "").strip() or None
        event_key = workflow_link_id or str(task.get("id") or "").strip()
        same_day_due = bool(
            final_due_date and internal_due_date and final_due_date == internal_due_date
        )

        if due_axis in {"all", "final"} and _token_in_range(final_due_date):
            events.append(
                {
                    "id": f"{event_key}:final",
                    "workflow_id": workflow_id or None,
                    "title": base_title,
                    "start": final_due_date,
                    "url": detail_url,
                    **event_meta(status=status, axis="final"),
                }
            )

        if (
            due_axis in {"all", "internal"}
            and _token_in_range(internal_due_date)
            and not (due_axis == "all" and same_day_due)
        ):
            events.append(
                {
                    "id": f"{event_key}:internal",
                    "workflow_id": workflow_id or None,
                    "title": base_title,
                    "start": internal_due_date,
                    "url": detail_url,
                    **event_meta(status=status, axis="internal"),
                }
            )

    return jsonify(events)


def _ensure_worklog_row_for_docket(
    *,
    docket_item: DocketItem,
    actor_id: int | None,
) -> WorkLog | None:
    from app.services.workflow.task_sync import ensure_worklog_for_docket

    return ensure_worklog_for_docket(docket_item=docket_item, actor_id=actor_id)


def _ensure_worklog_row_for_workflow(
    *,
    wf: Workflow,
    actor_id: int | None,
) -> WorkLog | None:
    workflow_id = int(getattr(wf, "id", 0) or 0)
    if workflow_id <= 0:
        return None

    wl = (
        WorkLog.query.filter(WorkLog.workflow_id == workflow_id).order_by(WorkLog.id.desc()).first()
    )
    if wl:
        return wl

    matter_id = str(getattr(wf, "case_id", "") or "").strip()
    matter = db.session.get(Matter, matter_id) if matter_id else None
    final_due_date, internal_due_date = _workflow_task_due_dates(wf=wf, linked_docket_item=None)
    wl = WorkLog(
        workflow_id=workflow_id,
        matter_id=matter_id or None,
        our_ref=getattr(matter, "our_ref", None) if matter else None,
        task_name=_workflow_display_task_name(wf),
        task_category=getattr(wf, "category", None),
        due_date=_parse_iso_date_param(internal_due_date or final_due_date),
        action_type="note",
        status="pending",
        description=getattr(wf, "note", None),
        completed_by_id=actor_id if actor_id else None,
    )
    db.session.add(wl)
    return wl


def _upsert_worklog_action_for_workflow(
    *,
    wf: Workflow,
    actor_id: int | None,
    action_type: str,
    status: str,
    description: str | None = None,
) -> WorkLog | None:
    wl = _ensure_worklog_row_for_workflow(wf=wf, actor_id=actor_id)
    if not wl:
        return None

    wl.action_type = action_type
    wl.status = status
    wl.task_name = _workflow_display_task_name(wf)
    wl.task_category = getattr(wf, "category", None)
    final_due_date, internal_due_date = _workflow_task_due_dates(wf=wf, linked_docket_item=None)
    wl.due_date = _parse_iso_date_param(internal_due_date or final_due_date)
    if description is not None:
        wl.description = description

    if status in ("completed", "abandoned"):
        wl.completed_at = datetime.utcnow()
        if actor_id:
            wl.completed_by_id = actor_id
    else:
        wl.completed_at = None
        wl.completed_by_id = None

    db.session.add(wl)
    return wl


def _upsert_worklog_action_for_docket(
    *,
    docket_item: DocketItem,
    actor_id: int | None,
    action_type: str,
    status: str,
    description: str | None = None,
) -> WorkLog | None:
    wl = _ensure_worklog_row_for_docket(docket_item=docket_item, actor_id=actor_id)
    if not wl:
        return None

    wl.action_type = action_type
    wl.status = status
    if description is not None:
        wl.description = description

    if status in ("completed", "abandoned"):
        wl.completed_at = datetime.utcnow()
        if actor_id:
            wl.completed_by_id = actor_id
    else:
        wl.completed_at = None
        wl.completed_by_id = None

    db.session.add(wl)
    return wl


def _append_note_to_worklog(wl: WorkLog, description: str) -> None:
    old_desc = (wl.description or "").strip()
    wl.description = f"{old_desc}\n{description}".strip() if old_desc else description
    action_type = (wl.action_type or "").strip().lower()
    status = (wl.status or "").strip().lower()
    if action_type not in ("completed", "abandoned", "expired") and status not in (
        "completed",
        "abandoned",
    ):
        wl.action_type = "note"
        wl.status = "pending"


def _validate_evidence_type(raw: str | None) -> str | None:
    value = str(raw or "").strip().lower() or "memo"
    if value in _VALID_EVIDENCE_TYPES:
        return value
    return None


def _parse_task_identifier(task_id: str) -> tuple[str, int | str] | None:
    token = str(task_id or "").strip()
    if not token:
        return None
    if token.startswith("wf_"):
        try:
            return ("workflow", int(token[3:]))
        except Exception:
            return None
    return ("docket", token)


def _normalize_task_ids(raw_task_ids: object) -> list[str]:
    values = raw_task_ids
    if isinstance(values, str):
        values = [values]
    if not isinstance(values, list):
        return []

    out: list[str] = []
    seen: set[str] = set()
    for raw in values:
        token = str(raw or "").strip()
        if not token or token in seen:
            continue
        seen.add(token)
        out.append(token)
    return out


def _load_docket_items_by_ids(docket_ids: list[str] | set[str]) -> dict[str, DocketItem]:
    normalized_ids = sorted(
        {str(did or "").strip() for did in (docket_ids or []) if str(did or "").strip()}
    )
    if not normalized_ids:
        return {}

    q = DocketItem.query.filter(DocketItem.docket_id.in_(normalized_ids))
    if hasattr(DocketItem, "is_deleted"):
        q = q.filter(or_(DocketItem.is_deleted.is_(False), DocketItem.is_deleted.is_(None)))
    return {
        str(di.docket_id): di for di in q.all() if str(getattr(di, "docket_id", "") or "").strip()
    }


def _load_linked_docket_items_for_workflows(workflows: list[Workflow]) -> dict[int, DocketItem]:
    workflows = [wf for wf in workflows if wf is not None]
    if not workflows:
        return {}

    docket_ids = {
        str(did or "").strip()
        for did in (_workflow_docket_id(wf) for wf in workflows)
        if str(did or "").strip()
    }
    docket_item_by_docket_id = _load_docket_items_by_ids(docket_ids)
    out: dict[int, DocketItem] = {}
    for wf in workflows:
        did = _workflow_docket_id(wf)
        if not did:
            continue
        di = docket_item_by_docket_id.get(did)
        if not di:
            continue
        if str(getattr(di, "matter_id", "") or "") != str(getattr(wf, "case_id", "") or ""):
            continue
        out[int(wf.id)] = di
    return out


def _commit_session_or_error(*, log_context: str, error_message: str):
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception(log_context)
        return jsonify({"error": error_message}), 500
    return None


def _log_worklog_audit(
    action: str,
    target_type: str,
    *,
    target_id: int | None = None,
    meta: dict[str, object] | None = None,
) -> None:
    try:
        from app.blueprints.billing_invoices.auth import log_audit

        log_audit(
            action,
            target_type,
            target_id,
            json.dumps(meta or {}, ensure_ascii=False),
        )
    except Exception as exc:
        current_app.logger.warning("Failed to log worklog audit %s: %s", action, exc)


def _run_post_commit_actions(post_commit: dict[str, object]) -> None:
    actor_id = int(getattr(current_user, "id", 0) or 0) or None
    for di, docket_actor_id in list(post_commit.get("docket_sync_items") or []):
        try:
            enqueue_docket_sync_for_item(docket_item=di, actor_id=docket_actor_id)
        except Exception as exc:
            current_app.logger.warning(
                "Docket sync enqueue failed for %s: %s",
                getattr(di, "docket_id", None),
                exc,
            )

    for workflow_id in sorted(post_commit.get("workflow_sync_ids") or set()):
        try:
            enqueue_workflow_sync(workflow_id=int(workflow_id))
        except Exception as exc:
            current_app.logger.warning(
                "Workflow sync enqueue failed for wf=%s: %s", workflow_id, exc
            )
        # C-1 fix: sync_workflow_task_immediately commit Done  
        # docket sync race condition .
        try:
            wf_obj = db.session.get(Workflow, int(workflow_id))
            if wf_obj:
                sync_workflow_task_immediately(workflow=wf_obj, actor_id=actor_id)
        except Exception as exc:
            current_app.logger.warning("Workflow task sync failed for wf=%s: %s", workflow_id, exc)

    for entry in list(post_commit.get("audit_entries") or []):
        if not isinstance(entry, dict):
            continue
        _log_worklog_audit(
            str(entry.get("action") or "").strip(),
            str(entry.get("target_type") or "").strip(),
            target_id=entry.get("target_id"),
            meta=dict(entry.get("meta") or {}),
        )


def _queue_audit_entry(
    post_commit: dict[str, object],
    *,
    action: str,
    target_type: str,
    meta: dict[str, object] | None = None,
    target_id: int | None = None,
) -> None:
    entries = post_commit.setdefault("audit_entries", [])
    if isinstance(entries, list):
        entries.append(
            {
                "action": action,
                "target_type": target_type,
                "target_id": target_id,
                "meta": dict(meta or {}),
            }
        )


def _complete_workflow_task(
    *,
    wf: Workflow,
    evidence_type: str,
    description: str,
    today_dt: date,
    today_str: str,
    actor_id: int | None,
    linked_di: DocketItem | None,
    post_commit: dict[str, object],
) -> WorkLog | None:
    evidence_note = f"[Done:{evidence_type}] {description}"
    wf.status = "Completed"
    wf.completed_date = today_dt
    wf.completed_by_id = actor_id
    wf.note = f"{(wf.note or '').strip()}\n{evidence_note}".strip()

    wl = None
    if linked_di:
        linked_di.done_date = today_str
        clear_docket_manual_abandoned(linked_di)
        db.session.add(linked_di)
        wl = _upsert_worklog_action_for_docket(
            docket_item=linked_di,
            actor_id=actor_id,
            action_type="completed",
            status="completed",
            description=evidence_note,
        )
        if not _sync_docket_task_immediately(linked_di, actor_id=actor_id):
            sync_items = post_commit.setdefault("docket_sync_items", [])
            if isinstance(sync_items, list):
                sync_items.append((linked_di, actor_id))
        _sync_office_action_status(linked_di, today_str)
    else:
        wl = _upsert_worklog_action_for_workflow(
            wf=wf,
            actor_id=actor_id,
            action_type="completed",
            status="completed",
            description=evidence_note,
        )

    # C-1 fix: workflow sync post_commit  row  Add.
    # Existing _sync_docket_task_immediately + sync_workflow_task_immediately 
    #  Immediate     race condition  .
    workflow_sync_ids = post_commit.setdefault("workflow_sync_ids", set())
    if isinstance(workflow_sync_ids, set):
        workflow_sync_ids.add(int(wf.id))

    matter = db.session.get(Matter, str(wf.case_id))
    if matter:
        try:
            _recalc_matter_status(matter=matter, memo=getattr(matter, "memo", None))
        except Exception as exc:
            current_app.logger.warning("Failed to recalc matter status after complete: %s", exc)

    _queue_audit_entry(
        post_commit,
        action="worklog.complete",
        target_type="workflow",
        target_id=int(wf.id),
        meta={
            "task_id": f"wf_{wf.id}",
            "workflow_id": int(wf.id),
            "docket_id": str(getattr(linked_di, "docket_id", "") or "").strip() or None,
            "matter_id": str(getattr(wf, "case_id", "") or "").strip() or None,
            "evidence_type": evidence_type,
            "description": description,
        },
    )
    return wl


def _complete_docket_task(
    *,
    di: DocketItem,
    evidence_type: str,
    description: str,
    today_str: str,
    actor_id: int | None,
    post_commit: dict[str, object],
) -> WorkLog | None:
    evidence_note = f"[Done:{evidence_type}] {description}"
    di.done_date = today_str
    clear_docket_manual_abandoned(di)
    wl = _upsert_worklog_action_for_docket(
        docket_item=di,
        actor_id=actor_id,
        action_type="completed",
        status="completed",
        description=evidence_note,
    )
    if not _sync_docket_task_immediately(di, actor_id=actor_id):
        sync_items = post_commit.setdefault("docket_sync_items", [])
        if isinstance(sync_items, list):
            sync_items.append((di, actor_id))
    _sync_office_action_status(di, today_str)

    matter = db.session.get(Matter, str(di.matter_id))
    if matter:
        try:
            _recalc_matter_status(matter=matter, memo=getattr(matter, "memo", None))
        except Exception as exc:
            current_app.logger.warning("Failed to recalc matter status after complete: %s", exc)

    _queue_audit_entry(
        post_commit,
        action="docket.status_change",
        target_type="docket_item",
        meta={
            "docket_id": di.docket_id,
            "matter_id": di.matter_id,
            "old_status": "pending",
            "new_status": "done",
            "name": di.name_free or di.name_ref,
        },
    )
    return wl


def _reopen_workflow_task(
    *,
    wf: Workflow,
    actor_id: int | None,
    linked_di: DocketItem | None,
    post_commit: dict[str, object],
) -> WorkLog | None:
    wf.status = "Pending"
    wf.completed_date = None
    wf.completed_by_id = None

    wl = None
    if linked_di:
        linked_di.done_date = None
        clear_docket_manual_abandoned(linked_di)
        db.session.add(linked_di)
        wl = _upsert_worklog_action_for_docket(
            docket_item=linked_di,
            actor_id=actor_id,
            action_type="reopened",
            status="pending",
        )
        if not _sync_docket_task_immediately(linked_di, actor_id=actor_id):
            sync_items = post_commit.setdefault("docket_sync_items", [])
            if isinstance(sync_items, list):
                sync_items.append((linked_di, actor_id))
    else:
        wl = _upsert_worklog_action_for_workflow(
            wf=wf,
            actor_id=actor_id,
            action_type="reopened",
            status="pending",
        )

    workflow_sync_ids = post_commit.setdefault("workflow_sync_ids", set())
    if isinstance(workflow_sync_ids, set):
        workflow_sync_ids.add(int(wf.id))

    matter = db.session.get(Matter, str(wf.case_id))
    if matter:
        try:
            _recalc_matter_status(matter=matter, memo=getattr(matter, "memo", None))
        except Exception as exc:
            current_app.logger.warning("Failed to recalc matter status after reopen: %s", exc)

    _queue_audit_entry(
        post_commit,
        action="worklog.reopen",
        target_type="workflow",
        target_id=int(wf.id),
        meta={
            "task_id": f"wf_{wf.id}",
            "workflow_id": int(wf.id),
            "docket_id": str(getattr(linked_di, "docket_id", "") or "").strip() or None,
            "matter_id": str(getattr(wf, "case_id", "") or "").strip() or None,
        },
    )
    return wl


def _reopen_docket_task(
    *,
    di: DocketItem,
    actor_id: int | None,
    post_commit: dict[str, object],
) -> tuple[WorkLog | None, str | None]:
    prev_done_date = di.done_date
    di.done_date = None
    wl = _upsert_worklog_action_for_docket(
        docket_item=di,
        actor_id=actor_id,
        action_type="reopened",
        status="pending",
    )
    if not _sync_docket_task_immediately(di, actor_id=actor_id):
        sync_items = post_commit.setdefault("docket_sync_items", [])
        if isinstance(sync_items, list):
            sync_items.append((di, actor_id))

    _queue_audit_entry(
        post_commit,
        action="docket.reopen",
        target_type="docket_item",
        meta={
            "docket_id": di.docket_id,
            "matter_id": di.matter_id,
            "prev_done_date": prev_done_date,
        },
    )
    return wl, prev_done_date


def _abandon_workflow_task(
    *,
    wf: Workflow,
    reason: str,
    today_dt: date,
    done_value: str,
    actor_id: int | None,
    linked_di: DocketItem | None,
    post_commit: dict[str, object],
) -> WorkLog | None:
    wf.status = "Abandoned"
    wf.completed_date = today_dt
    wf.completed_by_id = actor_id
    if reason:
        wf.note = f"{(wf.note or '').strip()}\n[Task ] {reason}".strip()

    wl = None
    if linked_di:
        linked_di.done_date = done_value
        mark_docket_manual_abandoned(linked_di, reason=reason, when=today_dt)
        db.session.add(linked_di)
        wl = _upsert_worklog_action_for_docket(
            docket_item=linked_di,
            actor_id=actor_id,
            action_type="abandoned",
            status="abandoned",
            description=f"Task : {reason}" if reason else "Task ",
        )
        if not _sync_docket_task_immediately(linked_di, actor_id=actor_id):
            sync_items = post_commit.setdefault("docket_sync_items", [])
            if isinstance(sync_items, list):
                sync_items.append((linked_di, actor_id))
        _sync_office_action_status(linked_di, done_value)
    else:
        wl = _upsert_worklog_action_for_workflow(
            wf=wf,
            actor_id=actor_id,
            action_type="abandoned",
            status="abandoned",
            description=f"Task : {reason}" if reason else "Task ",
        )

    workflow_sync_ids = post_commit.setdefault("workflow_sync_ids", set())
    if isinstance(workflow_sync_ids, set):
        workflow_sync_ids.add(int(wf.id))

    matter = db.session.get(Matter, str(wf.case_id))
    if matter:
        try:
            # C-3 fix: reason(Task  Reason) memo   .
            # matter.memo Matter Notes,  reasonto    .
            _recalc_matter_status(matter=matter, memo=getattr(matter, "memo", None))
        except Exception as exc:
            current_app.logger.warning("Failed to recalc matter status after abandon: %s", exc)

    _queue_audit_entry(
        post_commit,
        action="worklog.abandon",
        target_type="workflow",
        target_id=int(wf.id),
        meta={
            "task_id": f"wf_{wf.id}",
            "workflow_id": int(wf.id),
            "docket_id": str(getattr(linked_di, "docket_id", "") or "").strip() or None,
            "matter_id": str(getattr(wf, "case_id", "") or "").strip() or None,
            "reason": reason or None,
            "done_date": done_value,
        },
    )
    return wl


def _abandon_docket_task(
    *,
    di: DocketItem,
    reason: str,
    done_value: str,
    actor_id: int | None,
    post_commit: dict[str, object],
) -> WorkLog | None:
    di.done_date = done_value
    mark_docket_manual_abandoned(di, reason=reason, when=done_value)
    wl = _upsert_worklog_action_for_docket(
        docket_item=di,
        actor_id=actor_id,
        action_type="abandoned",
        status="abandoned",
        description=f"Task : {reason}" if reason else "Task ",
    )
    if not _sync_docket_task_immediately(di, actor_id=actor_id):
        sync_items = post_commit.setdefault("docket_sync_items", [])
        if isinstance(sync_items, list):
            sync_items.append((di, actor_id))
    _sync_office_action_status(di, done_value)

    matter = db.session.get(Matter, str(di.matter_id))
    if matter:
        try:
            _recalc_matter_status(matter=matter, memo=getattr(matter, "memo", None))
        except Exception as exc:
            current_app.logger.warning("Failed to recalc matter status after abandon: %s", exc)

    _queue_audit_entry(
        post_commit,
        action="docket.abandon",
        target_type="docket_item",
        meta={
            "docket_id": di.docket_id,
            "matter_id": di.matter_id,
            "reason": reason,
            "done_date": done_value,
        },
    )
    return wl


def _note_workflow_task(
    *,
    wf: Workflow,
    description: str,
    linked_di: DocketItem | None,
    post_commit: dict[str, object],
) -> WorkLog | None:
    wf.note = f"{(wf.note or '').strip()}\n[Notes] {description}".strip()
    wl = None
    if linked_di:
        wl = _ensure_worklog_row_for_docket(docket_item=linked_di, actor_id=None)
        if wl:
            _append_note_to_worklog(wl, description)
            db.session.add(wl)
    else:
        wl = _ensure_worklog_row_for_workflow(wf=wf, actor_id=None)
        if wl:
            _append_note_to_worklog(wl, description)
            db.session.add(wl)

    _queue_audit_entry(
        post_commit,
        action="worklog.note",
        target_type="workflow",
        target_id=int(wf.id),
        meta={
            "task_id": f"wf_{wf.id}",
            "workflow_id": int(wf.id),
            "docket_id": str(getattr(linked_di, "docket_id", "") or "").strip() or None,
            "matter_id": str(getattr(wf, "case_id", "") or "").strip() or None,
            "note": description,
        },
    )
    return wl


def _note_docket_task(
    *,
    di: DocketItem,
    description: str,
    post_commit: dict[str, object],
) -> WorkLog | None:
    wl = _ensure_worklog_row_for_docket(docket_item=di, actor_id=None)
    if not wl:
        return None
    _append_note_to_worklog(wl, description)
    db.session.add(wl)
    _queue_audit_entry(
        post_commit,
        action="docket.note",
        target_type="docket_item",
        meta={
            "docket_id": di.docket_id,
            "matter_id": di.matter_id,
            "note": description,
        },
    )
    return wl


def _current_user_can_modify_docket_task(di: DocketItem) -> bool:
    role_flags = worklog_role_scope_flags(current_user)
    owner_staff_party_id = str(getattr(di, "owner_staff_party_id", "") or "").strip()
    current_staff_party_id = str(getattr(current_user, "staff_party_id", "") or "").strip()
    current_user_id = str(getattr(current_user, "id", "") or "").strip()
    owner_matches = bool(
        owner_staff_party_id and owner_staff_party_id in {current_staff_party_id, current_user_id}
    )
    category_upper = str(getattr(di, "category", "") or "").strip().upper()
    if category_upper in _MGMT_CATEGORIES_UPPER:
        return bool(
            role_flags.get("show_all_mgmt") or (role_flags.get("show_own_mgmt") and owner_matches)
        )
    return bool(
        role_flags.get("show_all_work") or (role_flags.get("show_own_work") and owner_matches)
    )


def _current_user_can_modify_workflow_task(wf: Workflow) -> bool:
    role_flags = worklog_role_scope_flags(current_user)
    if role_flags.get("show_all_mgmt") or role_flags.get("show_all_work"):
        return True
    current_user_id = int(getattr(current_user, "id", 0) or 0)
    if current_user_id <= 0:
        return False
    owner_ids: set[int] = set()
    for raw in (
        getattr(wf, "assignee_id", None),
        getattr(wf, "attorney_assignee_id", None),
        getattr(wf, "inspector_id", None),
        getattr(wf, "created_by_id", None),
    ):
        try:
            parsed = int(raw or 0)
        except Exception:
            parsed = 0
        if parsed > 0:
            owner_ids.add(parsed)
    return current_user_id in owner_ids


@bp.route("/api/tasks/<docket_id>/complete", methods=["POST"])
@login_required
def api_complete_task(docket_id: str):
    """
    Task Done Process API.

    DocketItem done_date Settings WorkLog Log Create.
    """
    data = request.get_json(silent=True) or {}
    evidence_type = _validate_evidence_type(data.get("evidence_type"))
    description = (data.get("description") or "").strip()

    if not description:
        return jsonify({"error": "evidence_required"}), 400
    if not evidence_type:
        return jsonify({"error": "invalid_evidence_type"}), 400

    today_dt = _today_in_app_timezone()
    today_str = today_dt.isoformat()
    actor_id = int(getattr(current_user, "id", 0) or 0) or None
    post_commit: dict[str, object] = {
        "workflow_sync_ids": set(),
        "docket_sync_items": [],
        "audit_entries": [],
    }

    task_ref = _parse_task_identifier(docket_id)
    if task_ref and task_ref[0] == "workflow":
        try:
            wf = db.session.get(Workflow, int(task_ref[1]))
            if not wf:
                return jsonify({"error": "Task not found"}), 404
            if not _current_user_can_modify_workflow_task(wf):
                return jsonify({"error": "forbidden"}), 403
            linked_di = _linked_docket_item_for_workflow(wf)
            wl = _complete_workflow_task(
                wf=wf,
                evidence_type=evidence_type,
                description=description,
                today_dt=today_dt,
                today_str=today_str,
                actor_id=actor_id,
                linked_di=linked_di,
                post_commit=post_commit,
            )
            error_response = _commit_session_or_error(
                log_context=f"Failed to complete workflow task: wf_{wf.id}",
                error_message="Failed to complete task",
            )
            if error_response:
                return error_response
            _run_post_commit_actions(post_commit)
            return jsonify(
                {
                    "success": True,
                    "message": "Task Done.",
                    "worklog_id": wl.id if wl else None,
                }
            )
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    di = DocketItem.query.filter_by(docket_id=docket_id).first()
    if not di:
        return jsonify({"error": "Task not found"}), 404
    if not _current_user_can_modify_docket_task(di):
        return jsonify({"error": "forbidden"}), 403
    try:
        wl = _complete_docket_task(
            di=di,
            evidence_type=evidence_type,
            description=description,
            today_str=today_str,
            actor_id=actor_id,
            post_commit=post_commit,
        )
        error_response = _commit_session_or_error(
            log_context=f"Failed to complete docket task: {docket_id}",
            error_message="Failed to complete task",
        )
        if error_response:
            return error_response
        _run_post_commit_actions(post_commit)
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

    return jsonify(
        {
            "success": True,
            "message": "Task completed",
            "worklog_id": wl.id if wl else None,
        }
    )


def _sync_office_action_status(di: DocketItem, done_date: str):
    """
    If the task is related to an Office Action (e.g., OA Response),
    mark the corresponding OfficeAction as done.
    """
    if not di or not di.matter_id:
        return

    oa_id = _extract_office_action_id(di)
    try:
        if oa_id:
            db.session.execute(
                text(
                    """
                    UPDATE office_action
                    SET done_date = :done
                    WHERE oa_id = :oid
                      AND matter_id = :mid
                      AND (done_date IS NULL OR TRIM(done_date) = '')
                    """
                ).execution_options(policy_bypass=True),
                {"done": done_date, "oid": oa_id, "mid": di.matter_id},
            )
            return

        # New fallback: match by due date for NOTICE-like tasks only.
        cat = (di.category or "").strip().upper()
        if cat not in {"NOTICE", "USPTO_OA"}:
            return

        triggers = ["", "", "OA", "", "Notice"]
        if not any(k in (di.name_free or "") for k in triggers) and not any(
            k in (di.name_ref or "") for k in triggers
        ):
            return

        due_target = (di.due_date or "")[:10] or (di.extended_due_date or "")[:10]
        if not due_target:
            return

        # H-5 fix: LIMIT 1     OA 
        #  Recent OA  Process.  OA Bulk Process  
        # done_date    .
        rows = db.session.execute(
            text(
                """
                SELECT oa_id
                FROM office_action
                WHERE matter_id = :mid
                  AND (done_date IS NULL OR TRIM(done_date) = '')
                  AND (
                    due_date = :due
                    OR extended_due_date = :due
                  )
                ORDER BY received_date DESC NULLS LAST
                LIMIT 1
                """
            ).execution_options(policy_bypass=True),
            {"mid": di.matter_id, "due": due_target},
        ).all()
        for (matched_oa_id,) in rows:
            if not matched_oa_id:
                continue
            db.session.execute(
                text(
                    """
                    UPDATE office_action
                    SET done_date = :done
                    WHERE oa_id = :oid
                      AND matter_id = :mid
                      AND (done_date IS NULL OR TRIM(done_date) = '')
                    """
                ).execution_options(policy_bypass=True),
                {"done": done_date, "oid": matched_oa_id, "mid": di.matter_id},
            )
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="worklog.routes._sync_office_action_status",
            log_key="worklog.routes._sync_office_action_status",
            log_window_seconds=300,
        )


@bp.route("/api/tasks/<docket_id>/reopen", methods=["POST"])
@login_required
def api_reopen_task(docket_id: str):
    """
    Done Task Reopen Open.

    DocketItem done_date  WorkLog Status .
    """
    actor_id = int(getattr(current_user, "id", 0) or 0) or None
    post_commit: dict[str, object] = {
        "workflow_sync_ids": set(),
        "docket_sync_items": [],
        "audit_entries": [],
    }

    task_ref = _parse_task_identifier(docket_id)
    if task_ref and task_ref[0] == "workflow":
        try:
            wf = db.session.get(Workflow, int(task_ref[1]))
            if not wf:
                return jsonify({"error": "Task not found"}), 404
            if not _current_user_can_modify_workflow_task(wf):
                return jsonify({"error": "forbidden"}), 403
            linked_di = _linked_docket_item_for_workflow(wf)
            wl = _reopen_workflow_task(
                wf=wf,
                actor_id=actor_id,
                linked_di=linked_di,
                post_commit=post_commit,
            )
            error_response = _commit_session_or_error(
                log_context=f"Failed to reopen workflow task: wf_{wf.id}",
                error_message="Failed to reopen task",
            )
            if error_response:
                return error_response
            _run_post_commit_actions(post_commit)
            return jsonify(
                {
                    "success": True,
                    "message": "Task Reopen active.",
                    "worklog_id": wl.id if wl else None,
                }
            )
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    di = DocketItem.query.filter_by(docket_id=docket_id).first()
    if not di:
        return jsonify({"error": "Task not found"}), 404
    if not _current_user_can_modify_docket_task(di):
        return jsonify({"error": "forbidden"}), 403
    try:
        wl, _prev_done_date = _reopen_docket_task(
            di=di,
            actor_id=actor_id,
            post_commit=post_commit,
        )
        error_response = _commit_session_or_error(
            log_context=f"Failed to reopen docket task: {docket_id}",
            error_message="Failed to reopen task",
        )
        if error_response:
            return error_response
        _run_post_commit_actions(post_commit)
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

    return jsonify(
        {
            "success": True,
            "message": "Task reopened",
            "worklog_id": wl.id if wl else None,
        }
    )


@bp.route("/api/tasks/<docket_id>/abandon", methods=["POST"])
@login_required
def api_abandon_task(docket_id: str):
    """
    Task  Process API.

    - Workflow/DocketItem 'Task (Cancel)' Status Change.
    - Matter(Matter)  'Abandoned/Matter closed'  Change, Abandoned/Withdrawn  Createdoes not.
    """
    data = request.get_json(silent=True) or {}
    reason = data.get("reason", "")

    today_dt = _today_in_app_timezone()
    today_str = today_dt.isoformat()
    done_value = f"AUTO_CANCELLED:{today_str}"
    actor_id = int(getattr(current_user, "id", 0) or 0) or None
    post_commit: dict[str, object] = {
        "workflow_sync_ids": set(),
        "docket_sync_items": [],
        "audit_entries": [],
    }

    task_ref = _parse_task_identifier(docket_id)
    if task_ref and task_ref[0] == "workflow":
        try:
            wf = db.session.get(Workflow, int(task_ref[1]))
            if not wf:
                return jsonify({"error": "Task not found"}), 404
            if not _current_user_can_modify_workflow_task(wf):
                return jsonify({"error": "forbidden"}), 403
            linked_di = _linked_docket_item_for_workflow(wf)
            wl = _abandon_workflow_task(
                wf=wf,
                reason=str(reason or "").strip(),
                today_dt=today_dt,
                done_value=done_value,
                actor_id=actor_id,
                linked_di=linked_di,
                post_commit=post_commit,
            )
            error_response = _commit_session_or_error(
                log_context=f"Failed to abandon workflow task: wf_{wf.id}",
                error_message="Failed to abandon task",
            )
            if error_response:
                return error_response
            _run_post_commit_actions(post_commit)
            return jsonify(
                {
                    "success": True,
                    "message": "Task  Process.",
                    "worklog_id": wl.id if wl else None,
                }
            )
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    di = DocketItem.query.filter_by(docket_id=docket_id).first()
    if not di:
        return jsonify({"error": "Task not found"}), 404
    try:
        wl = _abandon_docket_task(
            di=di,
            reason=str(reason or "").strip(),
            done_value=done_value,
            actor_id=actor_id,
            post_commit=post_commit,
        )
        error_response = _commit_session_or_error(
            log_context=f"Failed to abandon docket task: {docket_id}",
            error_message="Failed to abandon task",
        )
        if error_response:
            return error_response
        _run_post_commit_actions(post_commit)
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

    return jsonify(
        {
            "success": True,
            "message": "Task abandoned",
            "worklog_id": wl.id if wl else None,
        }
    )


@bp.route("/api/tasks/<docket_id>/note", methods=["POST"])
@login_required
def api_add_note(docket_id: str):
    """
    Task Notes Add.
    """
    data = request.get_json(silent=True) or {}
    description = data.get("description", "")

    if not description:
        return jsonify({"error": "Description is required"}), 400

    post_commit: dict[str, object] = {
        "workflow_sync_ids": set(),
        "docket_sync_items": [],
        "audit_entries": [],
    }

    task_ref = _parse_task_identifier(docket_id)
    if task_ref and task_ref[0] == "workflow":
        try:
            wf = db.session.get(Workflow, int(task_ref[1]))
            if not wf:
                return jsonify({"error": "Task not found"}), 404
            if not _current_user_can_modify_workflow_task(wf):
                return jsonify({"error": "forbidden"}), 403
            linked_di = _linked_docket_item_for_workflow(wf)
            wl = _note_workflow_task(
                wf=wf,
                description=description,
                linked_di=linked_di,
                post_commit=post_commit,
            )
            error_response = _commit_session_or_error(
                log_context=f"Failed to add note to workflow task: wf_{wf.id}",
                error_message="Failed to add note",
            )
            if error_response:
                return error_response
            _run_post_commit_actions(post_commit)
            return jsonify(
                {
                    "success": True,
                    "message": "Note added",
                    "worklog_id": wl.id if wl else None,
                }
            )
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    di = DocketItem.query.filter_by(docket_id=docket_id).first()
    if not di:
        return jsonify({"error": "Task not found"}), 404
    if not _current_user_can_modify_docket_task(di):
        return jsonify({"error": "forbidden"}), 403

    try:
        wl = _note_docket_task(di=di, description=description, post_commit=post_commit)
        if not wl:
            db.session.rollback()
            return jsonify({"error": "Failed to update worklog"}), 500
        error_response = _commit_session_or_error(
            log_context=f"Failed to add note to docket task: {docket_id}",
            error_message="Failed to add note",
        )
        if error_response:
            return error_response
        _run_post_commit_actions(post_commit)
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

    return jsonify(
        {
            "success": True,
            "message": "Note added",
            "worklog_id": wl.id,
        }
    )


def _apply_bulk_task_action(
    *,
    task_ids: list[str],
    action: str,
    description: str = "",
    evidence_type: str | None = None,
    reason: str = "",
):
    parsed_refs = [(task_id, _parse_task_identifier(task_id)) for task_id in task_ids]
    workflow_ids = sorted(
        {
            int(ref[1])
            for _task_id, ref in parsed_refs
            if ref and ref[0] == "workflow" and int(ref[1]) > 0
        }
    )
    docket_ids = sorted(
        {
            str(ref[1])
            for _task_id, ref in parsed_refs
            if ref and ref[0] == "docket" and str(ref[1]).strip()
        }
    )

    workflow_by_id = (
        {int(wf.id): wf for wf in Workflow.query.filter(Workflow.id.in_(workflow_ids)).all()}
        if workflow_ids
        else {}
    )
    linked_docket_by_wf_id = _load_linked_docket_items_for_workflows(list(workflow_by_id.values()))
    docket_by_id = _load_docket_items_by_ids(docket_ids)

    actor_id = int(getattr(current_user, "id", 0) or 0) or None
    today_dt = _today_in_app_timezone()
    today_str = today_dt.isoformat()
    done_value = f"AUTO_CANCELLED:{today_str}"
    post_commit: dict[str, object] = {
        "workflow_sync_ids": set(),
        "docket_sync_items": [],
        "audit_entries": [],
    }

    processed_task_ids: list[str] = []
    missing_task_ids: list[str] = []

    for task_id, ref in parsed_refs:
        if not ref:
            missing_task_ids.append(task_id)
            continue
        kind, raw_id = ref
        if kind == "workflow":
            wf = workflow_by_id.get(int(raw_id))
            if not wf:
                missing_task_ids.append(task_id)
                continue
            linked_di = linked_docket_by_wf_id.get(int(wf.id))
            if action == "complete":
                _complete_workflow_task(
                    wf=wf,
                    evidence_type=str(evidence_type or "memo"),
                    description=description,
                    today_dt=today_dt,
                    today_str=today_str,
                    actor_id=actor_id,
                    linked_di=linked_di,
                    post_commit=post_commit,
                )
            elif action == "abandon":
                _abandon_workflow_task(
                    wf=wf,
                    reason=reason,
                    today_dt=today_dt,
                    done_value=done_value,
                    actor_id=actor_id,
                    linked_di=linked_di,
                    post_commit=post_commit,
                )
            processed_task_ids.append(task_id)
            continue

        di = docket_by_id.get(str(raw_id))
        if not di:
            missing_task_ids.append(task_id)
            continue
        if action == "complete":
            _complete_docket_task(
                di=di,
                evidence_type=str(evidence_type or "memo"),
                description=description,
                today_str=today_str,
                actor_id=actor_id,
                post_commit=post_commit,
            )
        elif action == "abandon":
            _abandon_docket_task(
                di=di,
                reason=reason,
                done_value=done_value,
                actor_id=actor_id,
                post_commit=post_commit,
            )
        processed_task_ids.append(task_id)

    if not processed_task_ids:
        db.session.rollback()
        return jsonify({"error": "no_valid_tasks", "missing_task_ids": missing_task_ids}), 400

    _queue_audit_entry(
        post_commit,
        action=f"worklog.bulk_{action}",
        target_type="worklog_batch",
        meta={
            "task_ids": processed_task_ids,
            "missing_task_ids": missing_task_ids,
            "count": len(processed_task_ids),
        },
    )

    error_message = "Failed to process bulk task action"
    error_response = _commit_session_or_error(
        log_context=f"Failed bulk worklog action: {action}",
        error_message=error_message,
    )
    if error_response:
        return error_response

    _run_post_commit_actions(post_commit)
    return jsonify(
        {
            "success": True,
            "action": action,
            "processed_count": len(processed_task_ids),
            "missing_count": len(missing_task_ids),
            "processed_task_ids": processed_task_ids,
            "missing_task_ids": missing_task_ids,
        }
    )


@bp.route("/api/tasks/bulk-complete", methods=["POST"])
@login_required
def api_bulk_complete_tasks():
    payload = request.get_json(silent=True) or {}
    task_ids = _normalize_task_ids(payload.get("task_ids"))
    description = str(payload.get("description") or "").strip()
    evidence_type = _validate_evidence_type(payload.get("evidence_type"))

    if not task_ids:
        return jsonify({"error": "no_valid_task_ids"}), 400
    if not description:
        return jsonify({"error": "evidence_required"}), 400
    if not evidence_type:
        return jsonify({"error": "invalid_evidence_type"}), 400

    try:
        return _apply_bulk_task_action(
            task_ids=task_ids,
            action="complete",
            description=description,
            evidence_type=evidence_type,
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 500


@bp.route("/api/tasks/bulk-abandon", methods=["POST"])
@login_required
def api_bulk_abandon_tasks():
    payload = request.get_json(silent=True) or {}
    task_ids = _normalize_task_ids(payload.get("task_ids"))
    reason = str(payload.get("reason") or "").strip()

    if not task_ids:
        return jsonify({"error": "no_valid_task_ids"}), 400

    try:
        return _apply_bulk_task_action(
            task_ids=task_ids,
            action="abandon",
            reason=reason,
        )
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 500


@bp.route("/api/transfer-targets")
@login_required
def api_transfer_targets():
    """Assignable user list for bulk workflow transfer."""
    users: list[User] = []
    try:
        users = list(build_staff_assignment_lists().get("all_users") or [])
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="worklog.routes.api_transfer_targets.staff_options",
            log_key="worklog.routes.api_transfer_targets.staff_options",
            log_window_seconds=300,
        )
        users = (
            User.query.filter(User.is_active.is_(True))
            .order_by(User.department.asc(), User.username.asc())
            .all()
        )

    payload = []
    seen: set[int] = set()
    current_uid = int(getattr(current_user, "id", 0) or 0)
    for u in users:
        try:
            uid = int(getattr(u, "id", 0) or 0)
        except Exception:
            uid = 0
        if uid <= 0 or uid in seen or uid == current_uid:
            continue
        seen.add(uid)
        label = (
            str(getattr(u, "display_name", None) or "").strip()
            or str(getattr(u, "username", None) or "").strip()
            or str(getattr(u, "email", None) or "").strip()
            or f"User #{uid}"
        )
        payload.append({"id": uid, "name": label})

    payload.sort(key=lambda x: (x.get("name") or "").lower())
    return jsonify({"users": payload})


@bp.route("/api/assignment-requests", methods=["GET"])
@login_required
def api_assignment_requests():
    scope = str(request.args.get("scope") or "inbox").strip().lower()
    if scope not in {"inbox", "sent"}:
        return jsonify({"error": "invalid_scope"}), 400
    try:
        limit = int(str(request.args.get("limit") or "50").strip())
    except Exception:
        limit = 50
    limit = max(1, min(limit, 200))
    current_uid = int(getattr(current_user, "id", 0) or 0)
    inbox_pending_count = WorkflowAssignmentRequest.query.filter(
        WorkflowAssignmentRequest.target_user_id == current_uid,
        WorkflowAssignmentRequest.status == WorkflowAssignmentRequest.STATUS_PENDING,
    ).count()
    sent_pending_count = WorkflowAssignmentRequest.query.filter(
        WorkflowAssignmentRequest.requested_by_id == current_uid,
        WorkflowAssignmentRequest.status == WorkflowAssignmentRequest.STATUS_PENDING,
    ).count()
    q = WorkflowAssignmentRequest.query
    if scope == "sent":
        q = q.filter(WorkflowAssignmentRequest.requested_by_id == current_uid)
        pending_count = sent_pending_count
    else:
        q = q.filter(WorkflowAssignmentRequest.target_user_id == current_uid)
        pending_count = inbox_pending_count
    q = q.order_by(
        case(
            (WorkflowAssignmentRequest.status == WorkflowAssignmentRequest.STATUS_PENDING, 0),
            else_=1,
        ),
        WorkflowAssignmentRequest.requested_at.desc(),
        WorkflowAssignmentRequest.id.desc(),
    )
    rows = q.limit(limit).all()
    return jsonify(
        {
            "scope": scope,
            "pending_count": pending_count,
            "counts": {
                "inbox_pending": inbox_pending_count,
                "sent_pending": sent_pending_count,
            },
            "requests": [
                serialize_assignment_request(row, current_user_id=current_uid) for row in rows
            ],
        }
    )


def _enqueue_assignment_response_sync(workflow_id: int | None, actor_id: int | None) -> None:
    if not workflow_id:
        return
    try:
        enqueue_workflow_sync(workflow_id=int(workflow_id))
        enqueue_workflow_task_sync(workflow_id=int(workflow_id), actor_id=actor_id)
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="worklog.routes.assignment_request.enqueue_sync",
            log_key="worklog.routes.assignment_request.enqueue_sync",
            log_window_seconds=300,
        )


@bp.route("/api/assignment-requests/<int:request_id>/accept", methods=["POST"])
@login_required
def api_accept_assignment_request(request_id: int):
    actor_id = int(getattr(current_user, "id", 0) or 0)
    try:
        result = respond_assignment_request(request_id, actor_id, "accept")
        db.session.commit()
    except LookupError:
        db.session.rollback()
        return jsonify({"error": "assignment_request_not_found"}), 404
    except AssignmentRequestForbidden:
        db.session.rollback()
        return jsonify({"error": "forbidden"}), 403
    except Exception:
        db.session.rollback()
        current_app.logger.exception("worklog.assignment_request.accept_failed")
        return jsonify({"error": "assignment_request_accept_failed"}), 500

    if result.workflow_changed:
        _enqueue_assignment_response_sync(result.request.workflow_id, actor_id)
    return jsonify(
        {
            "success": True,
            "request": serialize_assignment_request(result.request, current_user_id=actor_id),
        }
    )


@bp.route("/api/assignment-requests/<int:request_id>/reject", methods=["POST"])
@login_required
def api_reject_assignment_request(request_id: int):
    payload = request.get_json(silent=True) or {}
    reason = str(payload.get("reason") or "").strip() or None
    actor_id = int(getattr(current_user, "id", 0) or 0)
    try:
        result = respond_assignment_request(request_id, actor_id, "reject", reason=reason)
        db.session.commit()
    except LookupError:
        db.session.rollback()
        return jsonify({"error": "assignment_request_not_found"}), 404
    except AssignmentRequestForbidden:
        db.session.rollback()
        return jsonify({"error": "forbidden"}), 403
    except Exception:
        db.session.rollback()
        current_app.logger.exception("worklog.assignment_request.reject_failed")
        return jsonify({"error": "assignment_request_reject_failed"}), 500

    if result.workflow_changed:
        _enqueue_assignment_response_sync(result.request.workflow_id, actor_id)
    return jsonify(
        {
            "success": True,
            "request": serialize_assignment_request(result.request, current_user_id=actor_id),
        }
    )


@bp.route("/api/tasks/bulk-transfer", methods=["POST"])
@login_required
def api_bulk_transfer_workflows():
    """
    Bulk-transfer selected workflow tasks to another user.

    Rule:
    - User can transfer only workflows where they are directly assigned
      (assignee / attorney_assignee / inspector).
    """
    payload = request.get_json(silent=True) or {}
    raw_task_ids = payload.get("task_ids") or []
    raw_target_user_id = payload.get("target_user_id")

    if isinstance(raw_task_ids, str):
        raw_task_ids = [raw_task_ids]
    if not isinstance(raw_task_ids, list):
        return jsonify({"error": "invalid_task_ids"}), 400

    try:
        target_user_id = int(str(raw_target_user_id or "").strip())
    except Exception:
        return jsonify({"error": "invalid_target_user_id"}), 400
    if target_user_id <= 0:
        return jsonify({"error": "invalid_target_user_id"}), 400
    if target_user_id == int(getattr(current_user, "id", 0) or 0):
        return jsonify({"error": "target_must_be_other_user"}), 400

    # M-7 fix: bulk transfer  Task Previous , Administrator Permissions 
    #  Previous  Security . show_all_mgmt  show_all_work Permissions
    #  User(/table) bulk transfer .
    _actor_role = getattr(current_user, "role", None)
    _actor_flags = resolve_role_scope(_actor_role)
    if not (_actor_flags.get("show_all_mgmt") or _actor_flags.get("show_all_work")):
        return jsonify({"error": "insufficient_permission"}), 403

    wf_ids: list[int] = []
    seen_wf_ids: set[int] = set()
    for raw in raw_task_ids:
        token = str(raw or "").strip()
        if token.startswith("wf_"):
            token = token[3:]
        try:
            wf_id = int(token)
        except Exception:
            continue
        if wf_id <= 0 or wf_id in seen_wf_ids:
            continue
        seen_wf_ids.add(wf_id)
        wf_ids.append(wf_id)

    if not wf_ids:
        return jsonify({"error": "no_valid_workflow_ids"}), 400

    target_user = User.query.filter(User.id == target_user_id, User.is_active.is_(True)).first()
    if not target_user:
        return jsonify({"error": "target_user_not_found"}), 404

    # Ensure target is selectable from staff picker set when available.
    try:
        assignable_ids = {
            int(getattr(u, "id", 0) or 0)
            for u in (build_staff_assignment_lists().get("all_users") or [])
            if int(getattr(u, "id", 0) or 0) > 0
        }
    except Exception:
        assignable_ids = set()
    if assignable_ids and target_user_id not in assignable_ids:
        return jsonify({"error": "target_user_not_assignable"}), 400

    workflows = Workflow.query.filter(Workflow.id.in_(wf_ids)).all()
    wf_by_id = {int(wf.id): wf for wf in workflows}

    actor_id = int(getattr(current_user, "id", 0) or 0)

    role_fields = ("assignee_id", "attorney_assignee_id", "inspector_id")
    transferred_ids: list[int] = []
    skipped_ids: list[int] = []
    forbidden_ids: list[int] = []
    missing_ids: list[int] = []

    for wf_id in wf_ids:
        wf = wf_by_id.get(int(wf_id))
        if not wf:
            missing_ids.append(int(wf_id))
            continue

        changed = False
        direct_owner = any(getattr(wf, field, None) == actor_id for field in role_fields)
        if not direct_owner:
            forbidden_ids.append(int(wf_id))
            continue

        # Default: move only roles where current user is assigned.
        assignment_request_before = workflow_assignment_state(wf)
        for field in role_fields:
            if getattr(wf, field, None) == actor_id:
                setattr(wf, field, target_user_id)
                changed = True

        if changed:
            persist_manual_workflow_assignment_override(
                workflow=wf,
                actor_id=actor_id,
            )
            sync_assignment_requests_for_changed_roles(
                wf,
                assignment_request_before,
                requested_by_id=actor_id,
                source="worklog_bulk_transfer",
            )
            db.session.add(wf)
            transferred_ids.append(int(wf_id))
        else:
            skipped_ids.append(int(wf_id))

    if transferred_ids:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            current_app.logger.exception("worklog.bulk_transfer.commit_failed")
            return jsonify({"error": "bulk_transfer_failed"}), 500

        _log_worklog_audit(
            "worklog.bulk_transfer",
            "workflow",
            meta={
                "target_user_id": target_user_id,
                "target_user_name": (
                    str(getattr(target_user, "display_name", None) or "").strip()
                    or str(getattr(target_user, "username", None) or "").strip()
                    or str(getattr(target_user, "email", None) or "").strip()
                    or str(target_user_id)
                ),
                "transferred_workflow_ids": transferred_ids,
                "skipped_workflow_ids": skipped_ids,
                "forbidden_workflow_ids": forbidden_ids,
                "missing_workflow_ids": missing_ids,
            },
        )

        for wf_id in transferred_ids:
            try:
                enqueue_workflow_sync(workflow_id=wf_id)
                enqueue_workflow_task_sync(workflow_id=wf_id, actor_id=actor_id)
            except Exception as exc:
                report_swallowed_exception(
                    exc,
                    context="worklog.routes.api_bulk_transfer_workflows.enqueue_sync",
                    log_key="worklog.routes.api_bulk_transfer_workflows.enqueue_sync",
                    log_window_seconds=300,
                )
    else:
        db.session.rollback()

    return jsonify(
        {
            "success": True,
            "target_user_id": target_user_id,
            "target_user_name": (
                str(getattr(target_user, "display_name", None) or "").strip()
                or str(getattr(target_user, "username", None) or "").strip()
                or str(getattr(target_user, "email", None) or "").strip()
                or str(target_user_id)
            ),
            "transferred_count": len(transferred_ids),
            "skipped_count": len(skipped_ids),
            "forbidden_count": len(forbidden_ids),
            "missing_count": len(missing_ids),
            "transferred_workflow_ids": transferred_ids,
            "skipped_workflow_ids": skipped_ids,
            "forbidden_workflow_ids": forbidden_ids,
            "missing_workflow_ids": missing_ids,
        }
    )


@bp.route("/api/owners")
@login_required
def api_owners():
    """
    Contact List Search API.

    Current User Permissions  Search  Contact List .
    """
    owner_role = _normalize_owner_role(request.args.get("owner_role") or request.args.get("role"))
    # Get filter conditions based on role using shared helpers/logic
    user_role = current_user.role if current_user.is_authenticated else None
    flags = worklog_role_scope_flags(current_user)
    show_all_mgmt = flags["show_all_mgmt"]
    show_all_work = flags["show_all_work"]
    show_own_mgmt = flags["show_own_mgmt"]
    show_own_work = flags["show_own_work"]

    # Construct visibility conditions
    visibility_conditions = []
    cat_upper = func.upper(Workflow.category)
    is_mine = workflow_user_filter(getattr(current_user, "id", None))

    if show_all_mgmt:
        visibility_conditions.append(cat_upper.in_(MGMT_CATEGORIES))
    elif show_own_mgmt:
        visibility_conditions.append(
            and_(
                cat_upper.in_(MGMT_CATEGORIES),
                is_mine,
            )
        )

    if show_all_work:
        visibility_conditions.append(cat_upper.in_(WORK_CATEGORIES))
    elif show_own_work:
        visibility_conditions.append(
            and_(
                cat_upper.in_(WORK_CATEGORIES),
                is_mine,
            )
        )

    # Case managers can view WORK workflows for matters they manage.
    if not show_all_work:
        try:
            managed_ids = managed_matter_ids_select(current_user)
            visibility_conditions.append(
                and_(cat_upper.in_(WORK_CATEGORIES), Workflow.case_id.in_(managed_ids))
            )
        except Exception as exc:
            report_swallowed_exception(
                exc,
                context="worklog.owners_options.managed_matter_ids_select",
            )

    # Base query for selectable staff (owners/attorney/handler/manager) within visible workflows
    if not visibility_conditions:
        return jsonify({"owners": []})

    # Exclude annuity (Renewal) workflows: they belong to the annuity screens, not Worklog owners.
    annuity_excluded = or_(
        Workflow.business_code.is_(None),
        not_(Workflow.business_code.like("ANNUITY:%")),
    )

    today = _today_in_app_timezone()
    try:
        visible_workflows = (
            Workflow.query.options(
                load_only(
                    Workflow.id,
                    Workflow.case_id,
                    Workflow.name,
                    Workflow.category,
                    Workflow.business_code,
                    Workflow.note,
                    Workflow.assignee_id,
                    Workflow.attorney_assignee_id,
                    Workflow.inspector_id,
                )
            )
            .filter(annuity_excluded)
            .filter(or_(*visibility_conditions))
            .all()
        )
        visible_workflows = _filter_hidden_workflows(visible_workflows, today=today)
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="worklog.routes.owner_options_visible_workflows",
            log_key="worklog.routes.owner_options_visible_workflows",
            log_window_seconds=300,
        )
        return jsonify({"owners": []})

    if owner_role == "owner":
        try:
            owner_rows_by_wf = _resolve_owner_rows_for_workflows(visible_workflows)
            owners_by_id: dict[str, str] = {}
            for wf_id, rows in owner_rows_by_wf.items():
                if not wf_id:
                    continue
                for row in rows:
                    sid = str(row.get("id") or "").strip()
                    if not sid:
                        continue
                    label = str(row.get("name") or "").strip() or sid
                    if sid not in owners_by_id:
                        owners_by_id[sid] = label
            owners = [{"id": sid, "name": name} for sid, name in owners_by_id.items()]
            owners.sort(key=lambda x: x["name"])
            return jsonify({"owners": owners})
        except Exception as exc:
            report_swallowed_exception(
                exc,
                context="worklog.routes.owner_options_distribution",
                log_key="worklog.routes.owner_options_distribution",
                log_window_seconds=300,
            )
            return jsonify({"owners": []})

    # Case staff roles (MSA )
    owner_role_to_msa_roles: dict[str, tuple[str, ...]] = {
        "attorney": ("attorney", "retainer"),
        "handler": ("handler", "staff", "draftsman"),
        "manager": ("manager", "mgmt"),
        "any": ("attorney", "retainer", "handler", "staff", "draftsman", "manager", "mgmt"),
    }
    role_codes = owner_role_to_msa_roles.get(owner_role, ())
    if not role_codes:
        return jsonify({"owners": []})

    visible_matter_ids = sorted(
        {
            str(getattr(wf, "case_id", "") or "").strip()
            for wf in visible_workflows
            if str(getattr(wf, "case_id", "") or "").strip()
        }
    )
    if not visible_matter_ids:
        return jsonify({"owners": []})

    try:
        from app.models.party import Party, PartyStaff
        from app.models.ip_records import MatterStaffAssignment

        role_expr = func.lower(func.trim(MatterStaffAssignment.staff_role_code))
        rows = (
            db.session.query(MatterStaffAssignment.staff_party_id, Party.name_display)
            .join(PartyStaff, PartyStaff.party_id == MatterStaffAssignment.staff_party_id)
            .join(Party, Party.party_id == PartyStaff.party_id)
            .filter(MatterStaffAssignment.matter_id.in_(visible_matter_ids))
            .filter(MatterStaffAssignment.staff_party_id.isnot(None))
            .filter(role_expr.in_(role_codes))
            .distinct()
            .all()
        )
        owners = []
        for spid, name in rows:
            sid = (spid or "").strip()
            if not sid:
                continue
            label = (name or "").strip() or sid
            owners.append({"id": sid, "name": label})
        owners.sort(key=lambda x: x["name"])
        return jsonify({"owners": owners})
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="worklog.routes.case_staff_owner_options",
            log_key="worklog.routes.case_staff_owner_options",
            log_window_seconds=300,
        )
        return jsonify({"owners": []})


@bp.route("/api/summary")
@login_required
def api_summary():
    """
    Task   API.

    Role  Task  :
    - Total   
    -  (7 )
    - Overdue (Deadline )
    - Done
    """
    today = _today_in_app_timezone()
    urgent_date = today + timedelta(days=URGENT_WINDOW_DAYS)
    category_filter = request.args.get("category", "")
    owner_filter = request.args.get("owner", "")
    owner_role = _normalize_owner_role(request.args.get("owner_role") or request.args.get("role"))
    mine_only = (request.args.get("mine") or "").strip().lower() in ("1", "true", "yes")
    search_query = request.args.get("search", "").strip()
    days_param = request.args.get("days", "30")
    days = _parse_days_param(days_param, default=30, max_days=9999)
    end_date = today + timedelta(days=days)
    due_axis = _normalize_worklog_due_axis(request.args.get("due_axis"), default="all")
    due_from = _parse_iso_date_param(request.args.get("due_from"))
    due_to = _parse_iso_date_param(request.args.get("due_to"))
    due_from, due_to = _normalize_due_range(due_from, due_to)
    if mine_only:
        owner_role = "owner"
        owner_filter = _current_worklog_mine_owner_value()

    q, owner_filter_ids_for_owner_role, _user_role = _build_worklog_workflow_query(
        today=today,
        urgent_date=urgent_date,
        end_date=end_date,
        category_filter=category_filter,
        owner_filter=owner_filter,
        owner_role=owner_role,
        owner_name=request.args.get("owner_name"),
        due_axis=due_axis,
        due_from=due_from,
        due_to=due_to,
    )
    tasks = _build_worklog_tasks_light(
        rows=q.all(),
        today=today,
        urgent_date=urgent_date,
        due_axis=due_axis,
        search_query=search_query,
        owner_filter=owner_filter,
        owner_role=owner_role,
        owner_filter_ids_for_owner_role=owner_filter_ids_for_owner_role,
        mine_only=mine_only,
    )
    counts = _summary_counts_from_tasks(
        tasks, today=today, urgent_date=urgent_date, end_date=end_date
    )
    return jsonify(counts)
