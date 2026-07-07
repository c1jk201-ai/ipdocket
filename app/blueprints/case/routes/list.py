from __future__ import annotations

import hashlib
import io
import json
import threading
import time
from collections import deque
from datetime import date, timedelta

from flask import abort, current_app, jsonify, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, false, func, or_

from app.blueprints.case import bp
from app.blueprints.case.helpers import (
    _build_case_list_extras,
    _normalize_case_division,
    _normalize_case_type,
)
from app.extensions import db
from app.models.case_flat_index import CaseFlatIndex
from app.models.docket import DocketItem
from app.models.ip_records import (
    Matter,
    MatterCustomField,
    MatterEvent,
    MatterIdentifier,
    MatterPartyRole,
    MatterStaffAssignment,
    RawImportField,
    VMatterOverview,
)
from app.services.case.case_kind import resolve_profile_case_kind
from app.utils.docket_dates import effective_due_text_expr
from app.utils.error_logging import report_swallowed_exception
from app.utils.permissions import get_self_assigned_role_codes, require_permission
from app.utils.policy_sql import policy_text as text
from app.utils.search import (
    build_sqlalchemy_search_filter,
    compact_search_text,
    parse_search_expression,
    sqlalchemy_contains_query,
)


def _uses_compact_index_query(value: object) -> bool:
    return False

_CASE_QUERY_FIELD_ALIASES = {
    "client": "client",
    "Client": "client",
    "customer": "client",
    "applicant": "applicant",
    "Applicant": "applicant",
    "party": "party",
    "": "party",
    "attorney": "attorney",
    "": "attorney",
    "handler": "handler",
    "Contact": "handler",
    "manager": "manager",
    "": "manager",
    "our_ref": "our_ref",
    "ref": "our_ref",
    "Matter reference": "our_ref",
    "your_ref": "your_ref",
    "client_ref": "your_ref",
    "matter_id": "matter_id",
    "id": "matter_id",
    "application_no": "app_no",
    "app_no": "app_no",
    "Application No.": "app_no",
    "status": "status",
    "Status": "status",
    "title": "right_name",
    "name": "right_name",
    "right_name": "right_name",
    "Title": "right_name",
    "inventor": "inventor",
    "Inventor": "inventor",
}


class _TTLCache:
    """Tiny in-process TTL cache (thread-safe).

    Used for hot endpoints like autocomplete to reduce repeated identical DB queries.
    """

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._data: dict[str, tuple[float, object]] = {}

    def get(self, key: str):  # noqa: ANN001
        now = time.monotonic()
        with self._lock:
            row = self._data.get(key)
            if not row:
                return None
            expires_at, value = row
            if expires_at <= now:
                self._data.pop(key, None)
                return None
            return value

    def set(self, key: str, value, ttl_seconds: float) -> None:  # noqa: ANN001
        if ttl_seconds <= 0:
            return
        now = time.monotonic()
        expires_at = now + float(ttl_seconds)
        with self._lock:
            self._data[key] = (expires_at, value)
            # Best-effort pruning: keep cache bounded.
            if len(self._data) > 2048:
                # Remove expired entries first; then drop oldest by expiry.
                expired = [k for k, (exp, _v) in self._data.items() if exp <= now]
                for k in expired:
                    self._data.pop(k, None)
                if len(self._data) > 2048:
                    for k, _row in sorted(self._data.items(), key=lambda item: item[1][0])[:512]:
                        self._data.pop(k, None)


_CASE_SEARCH_CACHE = _TTLCache()
_CASE_LIST_EXTRAS_CACHE = _TTLCache()
_CASE_SEARCH_REDIS_INIT_LOCK = threading.Lock()
_CASE_SEARCH_REDIS = None
_CASE_SEARCH_REDIS_INITED = False
_STATUS_RED_NAME_REF_PREFIX = "MGMT:STATUS_RED:"
_PASSIVE_STATUS_RED_LABELS = (
    "ExaminationWaiting",
    "ExaminationIn Progress",
    "ExaminationIn Progress",
    "FilingExaminationIn Progress",
    "",
)
_STATUS_RED_BLUE_ALIASES = {
    "ForeignFilingDeadline": ("ForeignFiling  In Progress", "ForeignFiling In Progress"),
    "Examination requestDeadline": ("Examination  Billing In Progress", "ExaminationBilling"),
    "RegistrationDeadline": ("RegistrationWaiting In Progress",),
    "RegistrationDeadline": ("RegistrationWaiting In Progress",),
    "Publication": ("Filing Publication In Progress",),
    "ExaminationOpen": ("Examination In progress",),
    "FilingDeadline": ("Filing  In Progress",),
    "Deadline": ("   In Progress",),
}


def _compact_status_query(value: object) -> str:
    return "".join(str(value or "").casefold().split())


def _status_red_labels_matching_query(value: str) -> list[str]:
    compact = _compact_status_query(value)
    if len(compact) < 3:
        return []
    labels: list[str] = []
    for label, aliases in _STATUS_RED_BLUE_ALIASES.items():
        candidates = (label, *aliases)
        if any(
            compact in _compact_status_query(candidate)
            or _compact_status_query(candidate) in compact
            for candidate in candidates
        ):
            labels.append(label)
    return labels


def _status_display_match_exists(matter_id_expr, value: str):
    raw = str(value or "").strip()
    if not raw:
        return false()

    label_expr = func.trim(
        func.substr(
            func.coalesce(DocketItem.name_ref, ""),
            len(_STATUS_RED_NAME_REF_PREFIX) + 1,
        )
    )
    match_clauses = [
        sqlalchemy_contains_query(DocketItem.name_ref, raw),
        sqlalchemy_contains_query(DocketItem.name_free, raw),
        sqlalchemy_contains_query(label_expr, raw),
    ]
    for label in _status_red_labels_matching_query(raw):
        match_clauses.append(label_expr == label)

    q_status = (
        db.session.query(DocketItem.docket_id)
        .filter(DocketItem.matter_id == matter_id_expr)
        .filter(or_(DocketItem.is_deleted.is_(False), DocketItem.is_deleted.is_(None)))
        .filter(or_(DocketItem.done_date.is_(None), func.trim(DocketItem.done_date) == ""))
        .filter(DocketItem.due_date.isnot(None))
        .filter(func.trim(DocketItem.due_date) != "")
        .filter(func.upper(func.trim(func.coalesce(DocketItem.name_ref, ""))).like("MGMT:STATUS_RED:%"))
        .filter(~label_expr.in_(_PASSIVE_STATUS_RED_LABELS))
        .filter(~label_expr.ilike("%In Progress"))
        .filter(~label_expr.ilike("%Waiting"))
        .filter(~label_expr.ilike("%Filing%"))
        .filter(~label_expr.ilike("%RenewalDeadline%"))
        .filter(or_(*match_clauses))
    )
    return q_status.exists()


def _case_search_redis_client():
    """
    Best-effort Redis client for cross-worker caching.
    Falls back to in-process TTL cache if Redis is unavailable/misconfigured.
    """
    global _CASE_SEARCH_REDIS, _CASE_SEARCH_REDIS_INITED
    if _CASE_SEARCH_REDIS_INITED:
        return _CASE_SEARCH_REDIS
    with _CASE_SEARCH_REDIS_INIT_LOCK:
        if _CASE_SEARCH_REDIS_INITED:
            return _CASE_SEARCH_REDIS
        _CASE_SEARCH_REDIS_INITED = True
        try:
            from redis import Redis  # type: ignore[import-not-found]

            url = (current_app.config.get("CASE_SEARCH_CACHE_REDIS_URL") or "").strip()
            if not url:
                url = (current_app.config.get("RATELIMIT_STORAGE_URI") or "").strip()
            if not url.startswith(("redis://", "rediss://")):
                _CASE_SEARCH_REDIS = None
                return None

            _CASE_SEARCH_REDIS = Redis.from_url(
                url,
                socket_connect_timeout=0.15,
                socket_timeout=0.15,
                retry_on_timeout=False,
            )
        except Exception as exc:
            report_swallowed_exception(
                exc,
                context="case.routes.list.search.redis_init",
                log_key="case.routes.list.search.redis_init",
                log_window_seconds=300,
            )
            _CASE_SEARCH_REDIS = None
        return _CASE_SEARCH_REDIS


def _case_list_extras_cache_key(matter_ids: list[str]) -> str:
    raw = ",".join(matter_ids)
    digest = hashlib.sha256(raw.encode("utf-8", errors="ignore")).hexdigest()[:32]
    return f"case.list.extras.v1:{len(matter_ids)}:{digest}"


def _case_list_extras_cache_ttl_seconds() -> int:
    try:
        ttl = int(current_app.config.get("CASE_LIST_EXTRAS_CACHE_TTL_SECONDS", 15) or 15)
    except Exception:
        ttl = 15
    return max(0, min(300, ttl))


def _get_case_list_extras_cached(matter_ids: list[str]) -> dict | None:
    if not matter_ids:
        return {}
    if not bool(current_app.config.get("CASE_LIST_EXTRAS_CACHE_ENABLED", True)):
        return None

    cache_key = _case_list_extras_cache_key(matter_ids)

    cached = _CASE_LIST_EXTRAS_CACHE.get(cache_key)
    if isinstance(cached, dict):
        return cached

    rc = _case_search_redis_client()
    if rc is None:
        return None

    try:
        raw = rc.get(cache_key)
        if not raw:
            return None
        if isinstance(raw, (bytes, bytearray)):
            raw = raw.decode("utf-8", errors="ignore")
        payload = json.loads(raw)
        if isinstance(payload, dict):
            _CASE_LIST_EXTRAS_CACHE.set(cache_key, payload, ttl_seconds=2.0)
            return payload
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="case.routes.list.extras.redis_get",
            log_key="case.routes.list.extras.redis_get",
            log_window_seconds=60,
        )
    return None


def _set_case_list_extras_cached(matter_ids: list[str], payload: dict) -> None:
    if not matter_ids:
        return
    if not bool(current_app.config.get("CASE_LIST_EXTRAS_CACHE_ENABLED", True)):
        return
    if not isinstance(payload, dict):
        return

    ttl = _case_list_extras_cache_ttl_seconds()
    if ttl <= 0:
        return

    cache_key = _case_list_extras_cache_key(matter_ids)
    _CASE_LIST_EXTRAS_CACHE.set(cache_key, payload, ttl_seconds=float(ttl))

    rc = _case_search_redis_client()
    if rc is None:
        return
    try:
        rc.setex(cache_key, ttl, json.dumps(payload, ensure_ascii=False))
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="case.routes.list.extras.redis_set",
            log_key="case.routes.list.extras.redis_set",
            log_window_seconds=60,
        )


def _clamp_page(value, default: int = 1) -> int:
    try:
        page = int(value)
    except (TypeError, ValueError):
        page = default
    return max(1, page)


def _clamp_per_page(value, default: int = 50, max_value: int = 500) -> int:
    try:
        per_page = int(value)
    except (TypeError, ValueError):
        per_page = default
    return max(1, min(per_page, max_value))


def _staff_name_match_exists(matter_id_expr, *, like: str, role_codes: tuple[str, ...]):
    from app.models.party import Party, PartyStaff
    from app.models.user import User

    normalized_roles = tuple(
        role.strip().lower() for role in role_codes if isinstance(role, str) and role.strip()
    )
    if not normalized_roles:
        return text("1=0")

    role_expr = func.lower(func.trim(MatterStaffAssignment.staff_role_code))
    q_staff = (
        db.session.query(MatterStaffAssignment.msa_id)
        .outerjoin(Party, Party.party_id == MatterStaffAssignment.staff_party_id)
        .outerjoin(PartyStaff, PartyStaff.party_id == MatterStaffAssignment.staff_party_id)
        .outerjoin(User, User.staff_party_id == MatterStaffAssignment.staff_party_id)
        .filter(MatterStaffAssignment.matter_id == matter_id_expr)
        .filter(role_expr.in_(normalized_roles))
        .filter(
            or_(
                Party.name_display.ilike(like),
                PartyStaff.staff_code.ilike(like),
                MatterStaffAssignment.raw_text.ilike(like),
                User.display_name.ilike(like),
                User.username.ilike(like),
            )
        )
    )
    return q_staff.exists()


def _default_build_search_text_from_row(row) -> str:
    """
    Safe fallback for compact pagination.
    Some callers pass build_search_text=None expecting DB-side compact filtering.
    If DB fast-path fails, Python fallback must not crash.
    """
    try:
        vals = row.values() if hasattr(row, "values") else []
        parts: list[str] = []
        for v in vals:
            if v is None:
                continue
            s = str(v).strip()
            if s:
                parts.append(s)
        text = " ".join(parts)
        return text[:2000] if len(text) > 2000 else text
    except Exception:
        try:
            s = str(row)
            return s[:2000] if len(s) > 2000 else s
        except Exception:
            return ""


def _paginate_compact_search(
    *,
    base_query: str,
    where_clause: str,
    params: dict,
    order_clause: str,
    search: str,
    page: int,
    per_page: int,
    build_search_text,
    compact_where_sql: str | None = None,
):
    if build_search_text is None:
        build_search_text = _default_build_search_text_from_row
    q_compact = compact_search_text(search)
    # Defensive: avoid accidentally matching everything ('' in any string is True).
    if not q_compact:
        return [], 0, 1, 1

    # used by python fallback loop
    base_params = dict(params) if isinstance(params, dict) else {}

    # [OPTIMIZATION] If DB has `search_compact`, use it directly.
    if compact_where_sql:
        # We construct a count query and a data query for the fast path.
        count_sql = f"SELECT count(1) FROM ({base_query} WHERE {compact_where_sql}) as _cnt"
        # We need to inject the original where clause too if presentNew
        # Actually the caller passes `where_clause` separate from `compact_where_sql`.
        # The `compact_where_sql` is specifically for the compact part.

        # Let's combine correctly:
        full_where = where_clause or ""
        if full_where:
            full_where += f" AND ({compact_where_sql})"
        else:
            full_where = f" WHERE ({compact_where_sql})"

        count_query = f"SELECT COUNT(*) FROM ({base_query} {full_where}) as sub"
        data_query = base_query + full_where + order_clause + " LIMIT :limit OFFSET :offset"

        fast_params = dict(params)
        fast_params["like_compact"] = f"%{q_compact}%"
        try:
            total = int(
                db.session.execute(
                    text(count_query).execution_options(policy_bypass=True),
                    fast_params,
                ).scalar()
                or 0
            )
            total_pages = (total + per_page - 1) // per_page if total else 1
            if total_pages and page > total_pages:
                page = total_pages

            fast_params["limit"] = per_page
            fast_params["offset"] = (page - 1) * per_page

            rows = (
                db.session.execute(
                    text(data_query).execution_options(policy_bypass=True),
                    fast_params,
                )
                .mappings()
                .all()
            )
            return rows, total, total_pages, page
        except Exception as exc:
            # We fall back to python scan, but keep diagnostics.
            try:
                db.session.rollback()
            except Exception as rollback_exc:
                report_swallowed_exception(
                    rollback_exc,
                    context="case.list._paginate_compact_search.fast_path.rollback",
                    log_key="case.list._paginate_compact_search.fast_path.rollback",
                    log_window_seconds=300,
                )
            report_swallowed_exception(
                exc,
                context="case.list._paginate_compact_search.fast_path",
                log_key="case.list._paginate_compact_search.fast_path",
                log_window_seconds=300,
            )

    # 2) Python fallback (slow, but safe)
    if build_search_text is None:

        def build_search_text(row):
            if isinstance(row, dict):
                vals = []
                for v in row.values():
                    if v is None:
                        continue
                    try:
                        s = str(v)
                    except Exception:
                        continue
                    if s:
                        vals.append(s)
                return " ".join(vals)
            try:
                return str(row)
            except Exception:
                return ""

    batch_size = 2000
    offset = 0
    total = 0
    page_rows = []
    last_page_rows = deque(maxlen=per_page)
    start = (page - 1) * per_page
    end = start + per_page

    while True:
        data_query = base_query + where_clause + order_clause + " LIMIT :limit OFFSET :offset"
        batch_params = dict(params)
        batch_params["limit"] = batch_size
        batch_params["offset"] = offset
        rows = (
            db.session.execute(
                text(data_query).execution_options(policy_bypass=True),
                batch_params,
            )
            .mappings()
            .all()
        )
        if not rows:
            break
        for r in rows:
            try:
                search_text = build_search_text(r) or ""
            except Exception:
                search_text = ""
            if q_compact in compact_search_text(search_text):
                idx = total
                total += 1
                last_page_rows.append(r)
                if start <= idx < end:
                    page_rows.append(r)
        offset += batch_size

    total_pages = (total + per_page - 1) // per_page if total else 1
    if total and page > total_pages:
        page = total_pages
        page_rows = list(last_page_rows)
    return page_rows, total, total_pages, page


def _json_text_value(column, key: str):
    dialect = (db.engine.dialect.name or "").lower()
    if dialect.startswith("postgres"):
        return column.op("->>")(key)
    return func.json_extract(column, f"$.{key}")


def _custom_field_search_expr(like: str):
    keys = [
        "client_name",
        "applicant_name",
        "application_applicant_name",
        "applicant_registrant",
    ]
    return (
        db.session.query(MatterCustomField.matter_id)
        .filter(MatterCustomField.matter_id == VMatterOverview.matter_id)
        .filter(or_(*[_json_text_value(MatterCustomField.data, key).ilike(like) for key in keys]))
        .exists()
    )


def _is_nonempty_expr(expr):
    return and_(expr.isnot(None), func.trim(expr) != "")


def _apply_madrid_filter(query):
    # Filter outgoing trademarks to cases that are actually Madrid filings.
    explicit_type_match = and_(
        VMatterOverview.right_group == "ETC",
        VMatterOverview.matter_type == "MADRID",
    )
    try:
        from app.services.case.case_parameter_service import CaseParameterService

        namespace = (CaseParameterService.get_namespace("OUT", "TRADEMARK") or "").strip()
    except Exception:
        namespace = ""

    namespaces = [namespace] if namespace else []
    for fallback in ("outgoing_trademark", "out_trademark"):
        if fallback not in namespaces:
            namespaces.append(fallback)

    custom_match_exprs = []
    for ns in namespaces:
        app_route = _json_text_value(MatterCustomField.data, "app_route")
        madrid_no = _json_text_value(MatterCustomField.data, "madrid_application_no")
        madrid_date = _json_text_value(MatterCustomField.data, "madrid_application_date")

        route_match = or_(
            app_route.ilike("%madrid%"),
            app_route.ilike("%\ub9c8\ub4dc\ub9ac\ub4dc%"),
        )
        custom_match_exprs.append(
            db.session.query(MatterCustomField.matter_id)
            .filter(MatterCustomField.matter_id == VMatterOverview.matter_id)
            .filter(MatterCustomField.namespace == ns)
            .filter(
                or_(
                    route_match,
                    _is_nonempty_expr(madrid_no),
                    _is_nonempty_expr(madrid_date),
                )
            )
            .exists()
        )

    id_match = (
        db.session.query(MatterIdentifier.mid_id)
        .filter(MatterIdentifier.matter_id == VMatterOverview.matter_id)
        .filter(
            or_(
                MatterIdentifier.id_type.ilike("%madrid%"),
                MatterIdentifier.id_type.ilike("%\ub9c8\ub4dc\ub9ac\ub4dc%"),
            )
        )
        .filter(_is_nonempty_expr(MatterIdentifier.id_value))
        .exists()
    )

    event_match = (
        db.session.query(MatterEvent.mevent_id)
        .filter(MatterEvent.matter_id == VMatterOverview.matter_id)
        .filter(
            or_(
                MatterEvent.event_key.ilike("%madrid%"),
                MatterEvent.event_key.ilike("%\ub9c8\ub4dc\ub9ac\ub4dc%"),
            )
        )
        .filter(_is_nonempty_expr(MatterEvent.event_at))
        .exists()
    )

    right_name_match = or_(
        VMatterOverview.right_name.ilike("%madrid%"),
        VMatterOverview.right_name.ilike("%\ub9c8\ub4dc\ub9ac\ub4dc%"),
    )

    exprs = [explicit_type_match, right_name_match, id_match, event_match] + custom_match_exprs
    return query.filter(or_(*exprs))


def _apply_hague_filter(query):
    # Filter outgoing designs to cases that are actually Hague filings.
    explicit_type_match = and_(
        VMatterOverview.right_group == "ETC",
        VMatterOverview.matter_type == "HAGUE",
    )
    try:
        from app.services.case.case_parameter_service import CaseParameterService

        namespace = (CaseParameterService.get_namespace("OUT", "DESIGN") or "").strip()
    except Exception:
        namespace = ""

    namespaces = [namespace] if namespace else []
    for fallback in ("outgoing_design", "out_design"):
        if fallback not in namespaces:
            namespaces.append(fallback)

    custom_match_exprs = []
    for ns in namespaces:
        app_route = _json_text_value(MatterCustomField.data, "app_route")
        hague_no = _json_text_value(MatterCustomField.data, "hague_application_no")
        hague_date = _json_text_value(MatterCustomField.data, "hague_application_date")
        route_match = or_(
            app_route.ilike("%hague%"),
            app_route.ilike("%\ud5e4\uc774\uadf8%"),
        )
        custom_match_exprs.append(
            db.session.query(MatterCustomField.matter_id)
            .filter(MatterCustomField.matter_id == VMatterOverview.matter_id)
            .filter(MatterCustomField.namespace == ns)
            .filter(
                or_(
                    route_match,
                    _is_nonempty_expr(hague_no),
                    _is_nonempty_expr(hague_date),
                )
            )
            .exists()
        )

    id_match = (
        db.session.query(MatterIdentifier.mid_id)
        .filter(MatterIdentifier.matter_id == VMatterOverview.matter_id)
        .filter(
            or_(
                MatterIdentifier.id_type.ilike("%hague%"),
                MatterIdentifier.id_type.ilike("%\ud5e4\uc774\uadf8%"),
            )
        )
        .filter(_is_nonempty_expr(MatterIdentifier.id_value))
        .exists()
    )

    event_match = (
        db.session.query(MatterEvent.mevent_id)
        .filter(MatterEvent.matter_id == VMatterOverview.matter_id)
        .filter(
            or_(
                MatterEvent.event_key.ilike("%hague%"),
                MatterEvent.event_key.ilike("%\ud5e4\uc774\uadf8%"),
            )
        )
        .filter(_is_nonempty_expr(MatterEvent.event_at))
        .exists()
    )

    right_name_match = or_(
        VMatterOverview.right_name.ilike("%hague%"),
        VMatterOverview.right_name.ilike("%\ud5e4\uc774\uadf8%"),
    )

    exprs = [explicit_type_match, right_name_match, id_match, event_match] + custom_match_exprs
    return query.filter(or_(*exprs))


def _apply_copyright_filter(query):
    # Filter misc cases to copyright matters.
    explicit_type_match = and_(
        VMatterOverview.right_group == "ETC",
        VMatterOverview.matter_type == "COPYRIGHT",
    )
    try:
        from app.services.case.case_parameter_service import CaseParameterService

        namespace = (CaseParameterService.get_namespace("", "MISC") or "").strip()
    except Exception:
        namespace = ""

    namespaces = [namespace] if namespace else []
    if "misc" not in namespaces:
        namespaces.append("misc")

    custom_match_exprs = []
    for ns in namespaces:
        right_type = _json_text_value(MatterCustomField.data, "right_type")
        case_kind = _json_text_value(MatterCustomField.data, "case_kind")
        route_match = or_(
            right_type.ilike("%copyright%"),
            right_type.ilike("%\uc800\uc791\uad8c%"),
            case_kind.ilike("%copyright%"),
            case_kind.ilike("%\uc800\uc791\uad8c%"),
        )
        custom_match_exprs.append(
            db.session.query(MatterCustomField.matter_id)
            .filter(MatterCustomField.matter_id == VMatterOverview.matter_id)
            .filter(MatterCustomField.namespace == ns)
            .filter(route_match)
            .exists()
        )

    id_match = (
        db.session.query(MatterIdentifier.mid_id)
        .filter(MatterIdentifier.matter_id == VMatterOverview.matter_id)
        .filter(
            or_(
                MatterIdentifier.id_type.ilike("%copyright%"),
                MatterIdentifier.id_type.ilike("%\uc800\uc791\uad8c%"),
                MatterIdentifier.id_value.ilike("%copyright%"),
                MatterIdentifier.id_value.ilike("%\uc800\uc791\uad8c%"),
            )
        )
        .exists()
    )

    event_match = (
        db.session.query(MatterEvent.mevent_id)
        .filter(MatterEvent.matter_id == VMatterOverview.matter_id)
        .filter(
            or_(
                MatterEvent.event_key.ilike("%copyright%"),
                MatterEvent.event_key.ilike("%\uc800\uc791\uad8c%"),
            )
        )
        .exists()
    )

    right_name_match = or_(
        VMatterOverview.right_name.ilike("%copyright%"),
        VMatterOverview.right_name.ilike("%\uc800\uc791\uad8c%"),
    )

    exprs = [explicit_type_match, right_name_match, id_match, event_match] + custom_match_exprs
    return query.filter(or_(*exprs))


def _case_kind_filter_expr(
    right_group_col, matter_type_col, *, division_code: str | None, type_code: str | None
):
    if not (division_code or "").strip() and not (type_code or "").strip():
        return None

    div = _normalize_case_division(division_code)
    typ = _normalize_case_type(type_code)
    if not typ:
        try:
            from app.services.case.case_menu_config import (
                find_case_menu_item,
                normalize_case_menu_division,
                normalize_case_menu_type,
            )

            menu_item = find_case_menu_item(division_code, type_code)
            if menu_item:
                menu_div = normalize_case_menu_division(division_code)
                menu_typ = normalize_case_menu_type(type_code)
                if menu_div and menu_typ:
                    return and_(right_group_col == menu_div, matter_type_col == menu_typ)
                if menu_typ:
                    return matter_type_col == menu_typ
        except Exception:
            current_app.logger.debug("case menu list filter lookup failed", exc_info=True)

    if div == "ETC" and typ == "PCT":
        return or_(
            and_(right_group_col == "ETC", matter_type_col == "PCT"),
            and_(right_group_col == "OUT", matter_type_col == "PCT"),
        )
    if div == "ETC" and typ == "MADRID":
        return or_(
            and_(right_group_col == "ETC", matter_type_col == "MADRID"),
            and_(right_group_col == "OUT", matter_type_col == "TRADEMARK"),
        )
    if div == "ETC" and typ == "HAGUE":
        return or_(
            and_(right_group_col == "ETC", matter_type_col == "HAGUE"),
            and_(right_group_col == "OUT", matter_type_col == "DESIGN"),
        )
    if div == "ETC" and typ == "COPYRIGHT":
        return or_(
            and_(right_group_col == "ETC", matter_type_col == "COPYRIGHT"),
            matter_type_col == "MISC",
        )
    if div == "ETC" and typ == "LITIGATION":
        return or_(
            and_(right_group_col == "ETC", matter_type_col == "LITIGATION"),
            matter_type_col == "LITIGATION",
        )
    if div == "ETC" and typ == "MISC":
        return or_(
            and_(right_group_col == "ETC", matter_type_col == "MISC"),
            matter_type_col == "MISC",
        )
    if typ == "LITIGATION":
        return matter_type_col == "LITIGATION"
    if typ and div:
        return and_(right_group_col == div, matter_type_col == typ)
    if typ:
        return matter_type_col == typ
    if div:
        return right_group_col == div
    return None


def _case_list_xlsx_response(
    *,
    cases: list[VMatterOverview],
    case_extras: dict[str, dict[str, str]] | None,
    inventor_column_mode: str,
    inventor_column_label: str,
    filter_title: str,
    export_scope: str,
    total: int,
) -> object:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"

    headers = [
        "No",
        "Status",
        "Our Ref",
        "Your Ref",
        "Client",
        "Title",
        inventor_column_label or "Inventor",
        "Applicant",
        "App No",
        "App Date",
        "Matter ID",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    extras = case_extras or {}
    for idx, case in enumerate(cases or [], start=1):
        matter_id = str(getattr(case, "matter_id", "") or "")
        x = extras.get(matter_id, {}) if matter_id else {}
        red = str((x.get("display_red") or "")).strip()
        blue = str((x.get("display_blue") or "")).strip()
        if red and blue:
            status_text = f"{red} / {blue}"
        else:
            status_text = red or blue

        _profile_division, profile_type = resolve_profile_case_kind(
            getattr(case, "right_group", None),
            getattr(case, "matter_type", None),
        )
        is_tm = profile_type == "TRADEMARK"
        show_class = inventor_column_mode == "class" or (inventor_column_mode == "mixed" and is_tm)
        inventor_or_class = (
            str((x.get("trademark_classes") or "")).strip()
            if show_class
            else str((x.get("inventor_name") or "")).strip()
        )
        client_name = str((x.get("client_name") or getattr(case, "clients", "") or "")).strip()
        title = str((getattr(case, "right_name", "") or x.get("proposal_title", "") or "")).strip()
        applicant = str((x.get("applicant_name") or getattr(case, "applicants", "") or "")).strip()
        app_no = str((x.get("application_no") or "")).strip()
        app_date = str((x.get("application_date") or "")).strip()

        ws.append(
            [
                idx,
                status_text,
                str((getattr(case, "our_ref", "") or "")).strip(),
                str((getattr(case, "your_ref", "") or "")).strip(),
                client_name,
                title,
                inventor_or_class,
                applicant,
                app_no,
                app_date,
                matter_id,
            ]
        )

    ws.freeze_panes = "A2"

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            value_len = len(str(getattr(cell, "value", "") or ""))
            if value_len > max_length:
                max_length = value_len
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    meta_ws = wb.create_sheet("Meta")
    meta_ws.append(["Filter", filter_title or "Case List"])
    meta_ws.append(["Scope", "all" if export_scope == "all" else "page"])
    meta_ws.append(["Exported Rows", len(cases or [])])
    meta_ws.append(["Matched Total", int(total or 0)])
    meta_ws.append(["Export Date", date.today().isoformat()])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ts = date.today().isoformat()
    filename = f"case_list_{export_scope}_{ts}.xlsx"
    resp = current_app.response_class(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _render_by(
    division_code: str | None,
    type_code: str | None,
    title_label: str,
    *,
    query_modifier=None,
    create_matter_params: dict | None = None,
):
    query = VMatterOverview.query

    div = _normalize_case_division(division_code)
    typ = _normalize_case_type(type_code)

    base_filter = _case_kind_filter_expr(
        VMatterOverview.right_group,
        VMatterOverview.matter_type,
        division_code=division_code,
        type_code=type_code,
    )
    if base_filter is not None:
        query = query.filter(base_filter)

    if query_modifier:
        query = query_modifier(query)

    create_case_url = None
    try:
        if isinstance(create_matter_params, dict) and create_matter_params:
            create_case_url = url_for("case_work.create_matter", **create_matter_params)
        elif typ and div:
            create_case_url = url_for("case_work.create_matter", division=div, type=typ)
        elif typ:
            create_case_url = url_for("case_work.create_matter", type=typ)
    except Exception:
        create_case_url = None

    def _parse_int(name: str, default: int) -> int:
        try:
            return int(str(request.args.get(name) or "").strip() or default)
        except Exception:
            return default

    per_page_options = [20, 50, 100, 200, 500]
    per_page = _parse_int("per_page", 50)
    if per_page not in per_page_options:
        per_page = 50
    page = max(1, _parse_int("page", 1))
    export_format = (request.args.get("format") or "").strip().lower()
    export_scope_raw = request.args.get("export_scope")
    export_scope = (export_scope_raw or "page").strip().lower()
    if export_scope not in ("page", "all"):
        export_scope = "page"
    raw_export_flag = (request.args.get("export") or "").strip().lower()
    export_requested = export_format in ("xlsx", "csv") or raw_export_flag in (
        "1",
        "true",
        "yes",
        "on",
    )

    q = (request.args.get("q") or "").strip()
    attorney = (request.args.get("attorney") or "").strip()
    handler = (request.args.get("handler") or "").strip()
    manager = (request.args.get("manager") or "").strip()

    # Detailed search fields
    ds_client = (request.args.get("client") or "").strip()
    ds_applicant = (request.args.get("applicant") or "").strip()
    ds_our_ref = (request.args.get("our_ref") or "").strip()
    ds_status = (request.args.get("status") or "").strip()
    ds_right_name = (request.args.get("right_name") or "").strip()
    ds_inventor = (request.args.get("inventor") or "").strip()
    ds_app_no = (request.args.get("app_no") or "").strip()
    assigned_filter = (request.args.get("assigned") or "").strip().lower()
    if assigned_filter not in {"", "me", "mine"}:
        assigned_filter = ""
    due_filter = (request.args.get("due") or "").strip().lower()
    if due_filter not in {"", "due7", "this_week", "overdue"}:
        due_filter = ""
    search_expr = parse_search_expression(q, field_aliases=_CASE_QUERY_FIELD_ALIASES)

    filter_title = title_label
    special_parts = []
    if assigned_filter in {"me", "mine"}:
        special_parts.append(" Matter")
    if due_filter == "due7":
        special_parts.append(" Deadline")
    elif due_filter == "this_week":
        special_parts.append(" Deadline")
    elif due_filter == "overdue":
        special_parts.append(" Deadline")
    if special_parts:
        filter_title = f"{filter_title} · " + " · ".join(special_parts)
    staff_parts = []
    if attorney:
        staff_parts.append(f":{attorney}")
    if handler:
        staff_parts.append(f"Contact:{handler}")
    if manager:
        staff_parts.append(f":{manager}")
    if staff_parts:
        filter_title = f"{filter_title} · " + " · ".join(staff_parts)

    is_main_list = division_code is None and type_code is None
    _profile_division, profile_type = resolve_profile_case_kind(division_code, type_code)
    is_trademark_list = profile_type == "TRADEMARK"

    if is_main_list:
        inventor_column_mode = "mixed"
        inventor_column_label = "INVENTER/CLASS"
    elif is_trademark_list:
        inventor_column_mode = "class"
        inventor_column_label = ""
    else:
        inventor_column_mode = "inventor"
        inventor_column_label = "Inventor"

    needs_flat_index = bool(q or attorney or handler or manager or ds_inventor)

    def _apply_case_ordering(base_query):
        if is_main_list:
            # All  "Matter Create(created_at) " Defaultto .
            # entered/retained created_at  Legacy Data    .
            entered_at = func.nullif(VMatterOverview.entered_at, "")
            retained_at = func.nullif(VMatterOverview.retained_at, "")
            return base_query.order_by(
                VMatterOverview.created_at.desc().nulls_last(),
                entered_at.desc().nulls_last(),
                retained_at.desc().nulls_last(),
                VMatterOverview.matter_id.desc(),
            )
        return base_query.order_by(
            VMatterOverview.our_ref.desc(),
            VMatterOverview.matter_id.desc(),
        )

    def _apply_case_ordering_matter_id(base_query):
        if is_main_list:
            entered_at = func.nullif(Matter.entered_at, "")
            retained_at = func.nullif(Matter.retained_at, "")
            return base_query.order_by(
                Matter.created_at.desc().nulls_last(),
                entered_at.desc().nulls_last(),
                retained_at.desc().nulls_last(),
                Matter.matter_id.desc(),
            )
        return base_query.order_by(
            Matter.our_ref.desc(),
            Matter.matter_id.desc(),
        )

    def _party_name_match_exists(
        matter_id_expr,
        *,
        like: str,
        role_code: str | None = None,
    ):
        from app.models.party import Party

        q_party = (
            db.session.query(MatterPartyRole.mpr_id)
            .outerjoin(Party, Party.party_id == MatterPartyRole.party_id)
            .filter(MatterPartyRole.matter_id == matter_id_expr)
        )
        if role_code:
            q_party = q_party.filter(func.lower(MatterPartyRole.role_code) == role_code.lower())
        q_party = q_party.filter(
            or_(
                Party.name_display.ilike(like),
                MatterPartyRole.raw_text.ilike(like),
            )
        )
        return q_party.exists()

    def _fetch_case_rows_by_ids(matter_ids: list[str]):
        if not matter_ids:
            return []
        rows = Matter.query.filter(Matter.matter_id.in_(matter_ids)).all()
        by_mid = {str(getattr(r, "matter_id", "")): r for r in rows}
        return [by_mid[mid] for mid in matter_ids if mid in by_mid]

    def _split_csv_terms(raw_value: str) -> list[str]:
        raw_text = str(raw_value or "").strip()
        if not raw_text:
            return []
        parts = [t.strip() for t in raw_text.replace(";", ",").split(",") if t.strip()]
        if not parts:
            return [raw_text]
        seen: set[str] = set()
        out: list[str] = []
        for part in parts:
            if part in seen:
                continue
            seen.add(part)
            out.append(part)
        return out

    def _assigned_to_me_expr(matter_id_expr):
        if assigned_filter not in {"me", "mine"}:
            return None
        staff_pid = str(getattr(current_user, "staff_party_id", "") or "").strip()
        role_codes = get_self_assigned_role_codes()
        if not staff_pid or not role_codes:
            return false()
        role_expr = func.lower(func.trim(MatterStaffAssignment.staff_role_code))
        return (
            db.session.query(MatterStaffAssignment.msa_id)
            .filter(MatterStaffAssignment.matter_id == matter_id_expr)
            .filter(MatterStaffAssignment.staff_party_id == staff_pid)
            .filter(role_expr.in_(role_codes))
            .exists()
        )

    def _due_filter_expr(matter_id_expr):
        if not due_filter:
            return None

        today = date.today()
        if due_filter == "overdue":
            start_date = None
            end_date = today - timedelta(days=1)
        elif due_filter == "this_week":
            start_date = today
            end_date = today + timedelta(days=max(0, 6 - today.weekday()))
        else:
            start_date = today
            end_date = today + timedelta(days=7)

        dialect = getattr(db.engine.dialect, "name", "")
        due_text = effective_due_text_expr(DocketItem, dialect_name=dialect)
        if due_text is None:
            return false()

        q_due = (
            db.session.query(DocketItem.docket_id)
            .filter(DocketItem.matter_id == matter_id_expr)
            .filter(or_(DocketItem.done_date.is_(None), DocketItem.done_date == ""))
            .filter(or_(DocketItem.is_deleted.is_(False), DocketItem.is_deleted.is_(None)))
            .filter(due_text.isnot(None))
        )
        if start_date:
            q_due = q_due.filter(due_text >= start_date.isoformat())
        if end_date:
            q_due = q_due.filter(due_text <= end_date.isoformat())
        return q_due.exists()

    def _inventor_match_expr_for_matter(value: str):
        tokens = _split_csv_terms(value)
        if not tokens:
            return None

        from app.models.party import Party

        flat_like_exprs = [
            sqlalchemy_contains_query(CaseFlatIndex.inventor, token) for token in tokens
        ]
        inv_flat = or_(*flat_like_exprs) if len(flat_like_exprs) > 1 else flat_like_exprs[0]

        role_like_exprs = [
            or_(
                Party.name_display.ilike(f"%{token}%"),
                MatterPartyRole.raw_text.ilike(f"%{token}%"),
            )
            for token in tokens
        ]
        inv_party = (
            db.session.query(MatterPartyRole.mpr_id)
            .outerjoin(Party, Party.party_id == MatterPartyRole.party_id)
            .filter(MatterPartyRole.matter_id == Matter.matter_id)
            .filter(MatterPartyRole.role_code == "inventor")
            .filter(or_(*role_like_exprs))
            .exists()
        )

        raw_like_exprs = [RawImportField.value_text.ilike(f"%{token}%") for token in tokens]
        inv_raw = (
            db.session.query(RawImportField.raw_field_id)
            .filter(RawImportField.raw_id == Matter.raw_id)
            .filter(RawImportField.sheet_name == "Matter")
            .filter(RawImportField.source_column == "Inventor")
            .filter(or_(*raw_like_exprs))
            .exists()
        )

        return or_(inv_flat, inv_party, inv_raw)

    def _inventor_match_expr_for_overview(value: str):
        tokens = _split_csv_terms(value)
        if not tokens:
            return None

        from app.models.party import Party

        flat_like_exprs = [
            sqlalchemy_contains_query(CaseFlatIndex.inventor, token) for token in tokens
        ]
        inv_flat = or_(*flat_like_exprs) if len(flat_like_exprs) > 1 else flat_like_exprs[0]

        role_like_exprs = [
            or_(
                Party.name_display.ilike(f"%{token}%"),
                MatterPartyRole.raw_text.ilike(f"%{token}%"),
            )
            for token in tokens
        ]
        inv_party = (
            db.session.query(MatterPartyRole.mpr_id)
            .outerjoin(Party, Party.party_id == MatterPartyRole.party_id)
            .filter(MatterPartyRole.matter_id == VMatterOverview.matter_id)
            .filter(MatterPartyRole.role_code == "inventor")
            .filter(or_(*role_like_exprs))
            .exists()
        )

        raw_like_exprs = [RawImportField.value_text.ilike(f"%{token}%") for token in tokens]
        inv_raw = (
            db.session.query(RawImportField.raw_field_id)
            .join(Matter, Matter.raw_id == RawImportField.raw_id)
            .filter(Matter.matter_id == VMatterOverview.matter_id)
            .filter(RawImportField.sheet_name == "Matter")
            .filter(RawImportField.source_column == "Inventor")
            .filter(or_(*raw_like_exprs))
            .exists()
        )

        return or_(inv_flat, inv_party, inv_raw)

    def _app_no_expr_for_matter(value: str):
        clauses = [
            sqlalchemy_contains_query(CaseFlatIndex.application_no, value),
            (
                db.session.query(MatterIdentifier.mid_id)
                .filter(MatterIdentifier.matter_id == Matter.matter_id)
                .filter(sqlalchemy_contains_query(MatterIdentifier.id_value, value))
                .exists()
            ),
        ]
        return or_(*clauses)

    def _app_no_expr_for_overview(value: str):
        clauses = [
            sqlalchemy_contains_query(CaseFlatIndex.application_no, value),
            (
                db.session.query(MatterIdentifier.mid_id)
                .filter(MatterIdentifier.matter_id == VMatterOverview.matter_id)
                .filter(sqlalchemy_contains_query(MatterIdentifier.id_value, value))
                .exists()
            ),
        ]
        return or_(*clauses)

    def _case_text_expr_for_matter(term: str):
        value = str(term or "").strip()
        if not value:
            return None
        if _uses_compact_index_query(value):
            return CaseFlatIndex.search_compact.ilike(f"%{compact_search_text(value)}%")
        like = f"%{value}%"
        return or_(
            sqlalchemy_contains_query(CaseFlatIndex.search_text, value),
            sqlalchemy_contains_query(Matter.right_name, value),
            sqlalchemy_contains_query(Matter.our_ref, value),
            sqlalchemy_contains_query(Matter.old_our_ref, value),
            sqlalchemy_contains_query(Matter.your_ref, value),
            _status_display_match_exists(Matter.matter_id, value),
            _party_name_match_exists(Matter.matter_id, like=like, role_code=None),
            _staff_name_match_exists(
                Matter.matter_id,
                like=like,
                role_codes=(
                    "attorney",
                    "retainer",
                    "handler",
                    "staff",
                    "draftsman",
                    "manager",
                    "mgmt",
                ),
            ),
        )

    def _case_text_expr_for_overview(term: str):
        value = str(term or "").strip()
        if not value:
            return None
        if _uses_compact_index_query(value):
            return CaseFlatIndex.search_compact.ilike(f"%{compact_search_text(value)}%")
        like = f"%{value}%"
        return or_(
            sqlalchemy_contains_query(CaseFlatIndex.search_text, value),
            sqlalchemy_contains_query(VMatterOverview.clients, value),
            sqlalchemy_contains_query(VMatterOverview.applicants, value),
            sqlalchemy_contains_query(VMatterOverview.right_name, value),
            sqlalchemy_contains_query(VMatterOverview.our_ref, value),
            sqlalchemy_contains_query(VMatterOverview.old_our_ref, value),
            sqlalchemy_contains_query(VMatterOverview.your_ref, value),
            sqlalchemy_contains_query(VMatterOverview.attorneys, value),
            _status_display_match_exists(VMatterOverview.matter_id, value),
            _party_name_match_exists(VMatterOverview.matter_id, like=like, role_code=None),
            _staff_name_match_exists(
                VMatterOverview.matter_id,
                like=like,
                role_codes=(
                    "attorney",
                    "retainer",
                    "handler",
                    "staff",
                    "draftsman",
                    "manager",
                    "mgmt",
                ),
            ),
        )

    case_search_filter_for_matter = None
    case_search_filter_for_overview = None
    if q:
        case_search_filter_for_matter = build_sqlalchemy_search_filter(
            search_expr,
            general_term_builder=_case_text_expr_for_matter,
            field_builders={
                "client": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.client_name, value),
                    _party_name_match_exists(
                        Matter.matter_id, like=f"%{value}%", role_code="client"
                    ),
                ),
                "applicant": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.applicant, value),
                    _party_name_match_exists(
                        Matter.matter_id,
                        like=f"%{value}%",
                        role_code="applicant",
                    ),
                ),
                "party": lambda value: _party_name_match_exists(
                    Matter.matter_id,
                    like=f"%{value}%",
                    role_code=None,
                ),
                "attorney": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.attorney, value),
                    _staff_name_match_exists(
                        Matter.matter_id,
                        like=f"%{value}%",
                        role_codes=("attorney", "retainer"),
                    ),
                ),
                "handler": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.handler, value),
                    _staff_name_match_exists(
                        Matter.matter_id,
                        like=f"%{value}%",
                        role_codes=("handler", "staff", "draftsman"),
                    ),
                ),
                "manager": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.manager, value),
                    _staff_name_match_exists(
                        Matter.matter_id,
                        like=f"%{value}%",
                        role_codes=("manager", "mgmt"),
                    ),
                ),
                "our_ref": lambda value: sqlalchemy_contains_query(Matter.our_ref, value),
                "your_ref": lambda value: sqlalchemy_contains_query(Matter.your_ref, value),
                "matter_id": lambda value: sqlalchemy_contains_query(Matter.matter_id, value),
                "status": lambda value: or_(
                    sqlalchemy_contains_query(Matter.status_blue, value),
                    sqlalchemy_contains_query(Matter.status_red, value),
                    sqlalchemy_contains_query(Matter.inhouse_status, value),
                    _status_display_match_exists(Matter.matter_id, value),
                ),
                "right_name": lambda value: sqlalchemy_contains_query(Matter.right_name, value),
                "inventor": _inventor_match_expr_for_matter,
                "app_no": _app_no_expr_for_matter,
            },
        )
        case_search_filter_for_overview = build_sqlalchemy_search_filter(
            search_expr,
            general_term_builder=_case_text_expr_for_overview,
            field_builders={
                "client": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.client_name, value),
                    sqlalchemy_contains_query(VMatterOverview.clients, value),
                    _party_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_code="client",
                    ),
                ),
                "applicant": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.applicant, value),
                    sqlalchemy_contains_query(VMatterOverview.applicants, value),
                    _party_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_code="applicant",
                    ),
                ),
                "party": lambda value: or_(
                    sqlalchemy_contains_query(VMatterOverview.clients, value),
                    sqlalchemy_contains_query(VMatterOverview.applicants, value),
                    _party_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_code=None,
                    ),
                ),
                "attorney": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.attorney, value),
                    sqlalchemy_contains_query(VMatterOverview.attorneys, value),
                    _staff_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_codes=("attorney", "retainer"),
                    ),
                ),
                "handler": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.handler, value),
                    _staff_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_codes=("handler", "staff", "draftsman"),
                    ),
                ),
                "manager": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.manager, value),
                    _staff_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_codes=("manager", "mgmt"),
                    ),
                ),
                "our_ref": lambda value: sqlalchemy_contains_query(VMatterOverview.our_ref, value),
                "your_ref": lambda value: sqlalchemy_contains_query(
                    VMatterOverview.your_ref, value
                ),
                "matter_id": lambda value: sqlalchemy_contains_query(
                    VMatterOverview.matter_id,
                    value,
                ),
                "status": lambda value: or_(
                    sqlalchemy_contains_query(VMatterOverview.status_blue, value),
                    sqlalchemy_contains_query(VMatterOverview.status_red, value),
                    sqlalchemy_contains_query(VMatterOverview.inhouse_status, value),
                    _status_display_match_exists(VMatterOverview.matter_id, value),
                ),
                "right_name": lambda value: sqlalchemy_contains_query(
                    VMatterOverview.right_name,
                    value,
                ),
                "inventor": _inventor_match_expr_for_overview,
                "app_no": _app_no_expr_for_overview,
            },
        )

    use_split_query = bool(current_app.config.get("CASE_LIST_SPLIT_QUERY_ENABLED", True)) and (
        query_modifier is None
    )
    id_query = None

    if use_split_query:
        id_query = db.session.query(Matter.matter_id).filter(
            func.coalesce(Matter.is_deleted, False).is_(False)
        )

        id_base_filter = _case_kind_filter_expr(
            Matter.right_group,
            Matter.matter_type,
            division_code=division_code,
            type_code=type_code,
        )
        if id_base_filter is not None:
            id_query = id_query.filter(id_base_filter)

        if needs_flat_index:
            id_query = id_query.outerjoin(
                CaseFlatIndex,
                Matter.matter_id == CaseFlatIndex.matter_id,
            )

        if attorney:
            like = f"%{attorney}%"
            id_query = id_query.filter(
                or_(
                    CaseFlatIndex.attorney.ilike(like),
                    _staff_name_match_exists(
                        Matter.matter_id,
                        like=like,
                        role_codes=("attorney", "retainer"),
                    ),
                )
            )
        if handler:
            like = f"%{handler}%"
            id_query = id_query.filter(
                or_(
                    CaseFlatIndex.handler.ilike(like),
                    _staff_name_match_exists(
                        Matter.matter_id,
                        like=like,
                        role_codes=("handler", "staff", "draftsman"),
                    ),
                )
            )
        if manager:
            like = f"%{manager}%"
            id_query = id_query.filter(
                or_(
                    CaseFlatIndex.manager.ilike(like),
                    _staff_name_match_exists(
                        Matter.matter_id,
                        like=like,
                        role_codes=("manager", "mgmt"),
                    ),
                )
            )

        if ds_client:
            like = f"%{ds_client}%"
            id_query = id_query.filter(
                _party_name_match_exists(Matter.matter_id, like=like, role_code="client")
            )
        if ds_applicant:
            like = f"%{ds_applicant}%"
            id_query = id_query.filter(
                _party_name_match_exists(Matter.matter_id, like=like, role_code="applicant")
            )
        if ds_our_ref:
            id_query = id_query.filter(Matter.our_ref.ilike(f"%{ds_our_ref}%"))
        if ds_status:
            id_query = id_query.filter(
                or_(
                    Matter.status_blue.ilike(f"%{ds_status}%"),
                    Matter.status_red.ilike(f"%{ds_status}%"),
                    Matter.inhouse_status.ilike(f"%{ds_status}%"),
                    _status_display_match_exists(Matter.matter_id, ds_status),
                )
            )
        if ds_right_name:
            id_query = id_query.filter(Matter.right_name.ilike(f"%{ds_right_name}%"))
        if ds_inventor:
            raw = ds_inventor
            parts = [t.strip() for t in raw.replace(";", ",").split(",") if t.strip()]
            if not parts:
                parts = [raw]

            tokens = []
            seen = set()
            for t in parts:
                if t in seen:
                    continue
                seen.add(t)
                tokens.append(t)

            flat_like_exprs = [CaseFlatIndex.inventor.ilike(f"%{t}%") for t in tokens]
            inv_flat = or_(*flat_like_exprs) if len(flat_like_exprs) > 1 else flat_like_exprs[0]

            from app.models.party import Party

            role_like_exprs = [
                or_(
                    Party.name_display.ilike(f"%{t}%"),
                    MatterPartyRole.raw_text.ilike(f"%{t}%"),
                )
                for t in tokens
            ]
            inv_party = (
                db.session.query(MatterPartyRole.mpr_id)
                .outerjoin(Party, Party.party_id == MatterPartyRole.party_id)
                .filter(MatterPartyRole.matter_id == Matter.matter_id)
                .filter(MatterPartyRole.role_code == "inventor")
                .filter(or_(*role_like_exprs))
                .exists()
            )

            raw_like_exprs = [RawImportField.value_text.ilike(f"%{t}%") for t in tokens]
            inv_raw = (
                db.session.query(RawImportField.raw_field_id)
                .filter(RawImportField.raw_id == Matter.raw_id)
                .filter(RawImportField.sheet_name == "Matter")
                .filter(RawImportField.source_column == "Inventor")
                .filter(or_(*raw_like_exprs))
                .exists()
            )

            id_query = id_query.filter(or_(inv_flat, inv_party, inv_raw))

        if ds_app_no:
            id_match = (
                db.session.query(MatterIdentifier.mid_id)
                .filter(MatterIdentifier.matter_id == Matter.matter_id)
                .filter(MatterIdentifier.id_value.ilike(f"%{ds_app_no}%"))
                .exists()
            )
            id_query = id_query.filter(id_match)

        assigned_expr = _assigned_to_me_expr(Matter.matter_id)
        if assigned_expr is not None:
            id_query = id_query.filter(assigned_expr)

        due_expr = _due_filter_expr(Matter.matter_id)
        if due_expr is not None:
            id_query = id_query.filter(due_expr)

        if q and case_search_filter_for_matter is not None:
            id_query = id_query.filter(case_search_filter_for_matter)

        id_query = _apply_case_ordering_matter_id(id_query)
        total = id_query.order_by(None).count()
        pages = max(1, (total + per_page - 1) // per_page) if total else 1
        if page > pages:
            page = pages

        page_mid_rows = id_query.limit(per_page).offset((page - 1) * per_page).all()
        page_mids = [str(row[0]) for row in page_mid_rows if row and row[0]]
        cases = _fetch_case_rows_by_ids(page_mids)
    else:
        if needs_flat_index:
            query = query.outerjoin(
                CaseFlatIndex,
                VMatterOverview.matter_id == CaseFlatIndex.matter_id,
            )

        if attorney:
            like = f"%{attorney}%"
            query = query.filter(
                or_(
                    CaseFlatIndex.attorney.ilike(like),
                    VMatterOverview.attorneys.ilike(like),
                    _staff_name_match_exists(
                        VMatterOverview.matter_id,
                        like=like,
                        role_codes=("attorney", "retainer"),
                    ),
                )
            )
        if handler:
            like = f"%{handler}%"
            query = query.filter(
                or_(
                    CaseFlatIndex.handler.ilike(like),
                    _staff_name_match_exists(
                        VMatterOverview.matter_id,
                        like=like,
                        role_codes=("handler", "staff", "draftsman"),
                    ),
                )
            )
        if manager:
            like = f"%{manager}%"
            query = query.filter(
                or_(
                    CaseFlatIndex.manager.ilike(like),
                    _staff_name_match_exists(
                        VMatterOverview.matter_id,
                        like=like,
                        role_codes=("manager", "mgmt"),
                    ),
                )
            )

        if ds_client:
            query = query.filter(VMatterOverview.clients.ilike(f"%{ds_client}%"))
        if ds_applicant:
            query = query.filter(VMatterOverview.applicants.ilike(f"%{ds_applicant}%"))
        if ds_our_ref:
            query = query.filter(VMatterOverview.our_ref.ilike(f"%{ds_our_ref}%"))
        if ds_status:
            query = query.filter(
                or_(
                    VMatterOverview.status_blue.ilike(f"%{ds_status}%"),
                    VMatterOverview.status_red.ilike(f"%{ds_status}%"),
                    VMatterOverview.inhouse_status.ilike(f"%{ds_status}%"),
                    _status_display_match_exists(VMatterOverview.matter_id, ds_status),
                )
            )
        if ds_right_name:
            query = query.filter(VMatterOverview.right_name.ilike(f"%{ds_right_name}%"))
        if ds_inventor:
            raw = ds_inventor
            parts = [t.strip() for t in raw.replace(";", ",").split(",") if t.strip()]
            if not parts:
                parts = [raw]

            tokens = []
            seen = set()
            for t in parts:
                if t in seen:
                    continue
                seen.add(t)
                tokens.append(t)

            flat_like_exprs = [CaseFlatIndex.inventor.ilike(f"%{t}%") for t in tokens]
            inv_flat = or_(*flat_like_exprs) if len(flat_like_exprs) > 1 else flat_like_exprs[0]

            from app.models.party import Party

            role_like_exprs = [
                or_(
                    Party.name_display.ilike(f"%{t}%"),
                    MatterPartyRole.raw_text.ilike(f"%{t}%"),
                )
                for t in tokens
            ]
            inv_party = (
                db.session.query(MatterPartyRole.mpr_id)
                .outerjoin(Party, Party.party_id == MatterPartyRole.party_id)
                .filter(MatterPartyRole.matter_id == VMatterOverview.matter_id)
                .filter(MatterPartyRole.role_code == "inventor")
                .filter(or_(*role_like_exprs))
                .exists()
            )

            raw_like_exprs = [RawImportField.value_text.ilike(f"%{t}%") for t in tokens]
            inv_raw = (
                db.session.query(RawImportField.raw_field_id)
                .join(Matter, Matter.raw_id == RawImportField.raw_id)
                .filter(Matter.matter_id == VMatterOverview.matter_id)
                .filter(RawImportField.sheet_name == "Matter")
                .filter(RawImportField.source_column == "Inventor")
                .filter(or_(*raw_like_exprs))
                .exists()
            )

            query = query.filter(or_(inv_flat, inv_party, inv_raw))

        if ds_app_no:
            id_match = (
                db.session.query(MatterIdentifier.mid_id)
                .filter(MatterIdentifier.matter_id == VMatterOverview.matter_id)
                .filter(MatterIdentifier.id_value.ilike(f"%{ds_app_no}%"))
                .exists()
            )
            query = query.filter(id_match)

        assigned_expr = _assigned_to_me_expr(VMatterOverview.matter_id)
        if assigned_expr is not None:
            query = query.filter(assigned_expr)

        due_expr = _due_filter_expr(VMatterOverview.matter_id)
        if due_expr is not None:
            query = query.filter(due_expr)

        if q and case_search_filter_for_overview is not None:
            query = query.filter(case_search_filter_for_overview)

        query = _apply_case_ordering(query)
        total = query.order_by(None).count()
        pages = max(1, (total + per_page - 1) // per_page) if total else 1
        if page > pages:
            page = pages
        cases = query.limit(per_page).offset((page - 1) * per_page).all()

    base_args = {k: v for k, v in request.args.items() if v is not None}
    base_args["per_page"] = per_page
    prev_url = url_for(request.endpoint, **{**base_args, "page": page - 1}) if page > 1 else None
    next_url = (
        url_for(request.endpoint, **{**base_args, "page": page + 1}) if page < pages else None
    )

    case_ids = [str(getattr(c, "matter_id", "") or "") for c in (cases or []) if c]
    case_ids = [mid for mid in case_ids if mid]
    case_extras = _get_case_list_extras_cached(case_ids)
    if not isinstance(case_extras, dict):
        case_extras = _build_case_list_extras(cases)
        _set_case_list_extras_cached(case_ids, case_extras)

    if export_requested:
        export_cases = list(cases or [])
        export_case_extras = dict(case_extras or {})
        if export_scope == "all":
            max_rows = current_app.config.get("CASE_LIST_XLSX_MAX_ROWS", 10000)
            try:
                max_rows_int = int(max_rows or 10000)
            except Exception:
                max_rows_int = 10000
            max_rows_int = max(100, min(max_rows_int, 50000))

            if use_split_query and id_query is not None:
                all_mid_rows = id_query.limit(max_rows_int).all()
                all_mids = [str(row[0]) for row in all_mid_rows if row and row[0]]
                export_cases = _fetch_case_rows_by_ids(all_mids)
            else:
                export_cases = query.limit(max_rows_int).all()

            export_case_ids = [
                str(getattr(c, "matter_id", "") or "") for c in (export_cases or []) if c
            ]
            export_case_ids = [mid for mid in export_case_ids if mid]
            export_case_extras = _get_case_list_extras_cached(export_case_ids)
            if not isinstance(export_case_extras, dict):
                export_case_extras = _build_case_list_extras(export_cases)
                _set_case_list_extras_cached(export_case_ids, export_case_extras)

        return _case_list_xlsx_response(
            cases=export_cases,
            case_extras=export_case_extras,
            inventor_column_mode=inventor_column_mode,
            inventor_column_label=inventor_column_label,
            filter_title=filter_title,
            export_scope=export_scope,
            total=total,
        )

    export_args = {
        k: v
        for k, v in request.args.items()
        if v is not None and k not in {"format", "export_scope", "export"}
    }
    export_xlsx_page_url = url_for(
        request.endpoint,
        **{
            **export_args,
            "export": "1",
            "export_scope": "page",
        },
    )
    export_xlsx_all_url = url_for(
        request.endpoint,
        **{
            **export_args,
            "export": "1",
            "export_scope": "all",
        },
    )

    return render_template(
        "case/list.html",
        cases=cases,
        case_extras=case_extras,
        filter_title=filter_title,
        division=division_code,
        case_type=type_code,
        page=page,
        pages=pages,
        per_page=per_page,
        per_page_options=per_page_options,
        total=total,
        prev_url=prev_url,
        next_url=next_url,
        create_case_url=create_case_url,
        inventor_column_mode=inventor_column_mode,
        inventor_column_label=inventor_column_label,
        export_xlsx_page_url=export_xlsx_page_url,
        export_xlsx_all_url=export_xlsx_all_url,
        resolve_profile_case_kind=resolve_profile_case_kind,
    )


def _redirect_to_list_endpoint(endpoint: str):
    return redirect(url_for(endpoint, **request.args.to_dict(flat=True)), code=302)


@bp.route("/list")
@login_required
def case_list():
    return _render_by(None, None, "All Matters")


@bp.route("/kind/<string:division_code>/<path:type_code>")
@login_required
def list_custom_kind(division_code: str, type_code: str):
    try:
        from app.services.case.case_menu_config import find_case_menu_item

        menu_item = find_case_menu_item(division_code, type_code)
    except Exception:
        menu_item = None
    if not menu_item:
        abort(404)
    title = " · ".join(
        part
        for part in (
            str(menu_item.get("section_label") or "").strip(),
            str(menu_item.get("label") or "").strip(),
        )
        if part
    )
    return _render_by(
        str(menu_item.get("division") or ""),
        str(menu_item.get("type") or ""),
        title or "Matter List",
        create_matter_params={
            "division": str(menu_item.get("division") or ""),
            "type": str(menu_item.get("type") or ""),
        },
    )


@bp.route("/dom/patent")
@login_required
def list_dom_patent():
    return _render_by("DOM", "PATENT", "US · Patent")


@bp.route("/dom/utility")
@login_required
def list_dom_utility():
    return _render_by("DOM", "UTILITY", "US · Utility")


@bp.route("/dom/design")
@login_required
def list_dom_design():
    return _render_by("DOM", "DESIGN", "US · Design")


@bp.route("/dom/trademark")
@login_required
def list_dom_trademark():
    return _render_by("DOM", "TRADEMARK", "US · Trademark")


@bp.route("/dom/litigation")
@login_required
def list_dom_litigation():
    return _render_by("DOM", "LITIGATION", "US · Proceedings / Litigation")


@bp.route("/inc/patent")
@login_required
def list_inc_patent():
    return _render_by("INC", "PATENT", "Inbound US · Patent")


@bp.route("/inc/utility")
@login_required
def list_inc_utility():
    return _render_by("INC", "UTILITY", "Inbound US · Utility")


@bp.route("/inc/design")
@login_required
def list_inc_design():
    return _render_by("INC", "DESIGN", "Inbound US · Design")


@bp.route("/inc/trademark")
@login_required
def list_inc_trademark():
    return _render_by("INC", "TRADEMARK", "Inbound US · Trademark")


@bp.route("/inc/litigation")
@login_required
def list_inc_litigation():
    return _render_by("INC", "LITIGATION", "Inbound US · Proceedings / Litigation")


@bp.route("/out/patent")
@login_required
def list_out_patent():
    return _render_by("OUT", "PATENT", "Foreign · Patent")


@bp.route("/out/utility")
@login_required
def list_out_utility():
    return _render_by("OUT", "UTILITY", "Foreign · Utility")


@bp.route("/out/design")
@login_required
def list_out_design():
    return _render_by("OUT", "DESIGN", "Foreign · Design")


@bp.route("/out/trademark")
@login_required
def list_out_trademark():
    return _render_by("OUT", "TRADEMARK", "Foreign · Trademark")


@bp.route("/etc/pct")
@login_required
def list_pct():
    return _render_by(
        "ETC",
        "PCT",
        "Other Matter · PCT",
        create_matter_params={"division": "ETC", "type": "PCT"},
    )


@bp.route("/PCT")
@bp.route("/pct")
@login_required
def list_pct_legacy():
    return _redirect_to_list_endpoint("case_work.list_pct")


@bp.route("/out/litigation")
@login_required
def list_out_litigation():
    return _render_by(
        "OUT",
        "LITIGATION",
        "Foreign · Proceedings / Litigation",
        create_matter_params={"division": "ETC", "type": "LITIGATION"},
    )


@bp.route("/etc/litigation")
@login_required
def list_litigation():
    return _render_by(
        "ETC",
        "LITIGATION",
        "Other Matter · Proceedings / Litigation",
        create_matter_params={"division": "ETC", "type": "LITIGATION"},
    )


@bp.route("/litigation")
@login_required
def list_litigation_legacy():
    return _redirect_to_list_endpoint("case_work.list_litigation")


@bp.route("/etc/madrid")
@login_required
def list_madrid():
    return _render_by(
        "ETC",
        "MADRID",
        "Other Matter · Madrid",
        query_modifier=_apply_madrid_filter,
        create_matter_params={"division": "ETC", "type": "MADRID"},
    )


@bp.route("/madrid")
@login_required
def list_madrid_legacy():
    return _redirect_to_list_endpoint("case_work.list_madrid")


@bp.route("/etc/hague")
@login_required
def list_hague():
    return _render_by(
        "ETC",
        "HAGUE",
        "Other Matter · Hague",
        query_modifier=_apply_hague_filter,
        create_matter_params={"division": "ETC", "type": "HAGUE"},
    )


@bp.route("/hague")
@login_required
def list_hague_legacy():
    return _redirect_to_list_endpoint("case_work.list_hague")


@bp.route("/etc/copyright")
@login_required
def list_copyright():
    return _render_by(
        "ETC",
        "COPYRIGHT",
        "Other Matter · Copyright",
        query_modifier=_apply_copyright_filter,
        create_matter_params={"division": "ETC", "type": "COPYRIGHT"},
    )


@bp.route("/copyright")
@login_required
def list_copyright_legacy():
    return _redirect_to_list_endpoint("case_work.list_copyright")


@bp.route("/etc/misc")
@login_required
def list_misc():
    return _render_by(
        "ETC",
        "MISC",
        "Other Matter · Other",
        create_matter_params={"division": "ETC", "type": "MISC"},
    )


@bp.route("/misc")
@login_required
def list_misc_legacy():
    return _redirect_to_list_endpoint("case_work.list_misc")


@bp.route("/search")
@login_required
def search():
    q = (request.args.get("q") or "").strip()
    search_expr = parse_search_expression(q, field_aliases=_CASE_QUERY_FIELD_ALIASES)

    def _party_name_match_exists(
        matter_id_expr,
        *,
        like: str,
        role_code: str | None = None,
    ):
        from app.models.party import Party

        q_party = (
            db.session.query(MatterPartyRole.mpr_id)
            .outerjoin(Party, Party.party_id == MatterPartyRole.party_id)
            .filter(MatterPartyRole.matter_id == matter_id_expr)
        )
        if role_code:
            q_party = q_party.filter(func.lower(MatterPartyRole.role_code) == role_code.lower())
        q_party = q_party.filter(
            or_(
                Party.name_display.ilike(like),
                MatterPartyRole.raw_text.ilike(like),
            )
        )
        return q_party.exists()

    def _inventor_match_expr(value: str):
        raw_text = str(value or "").strip()
        if not raw_text:
            return None

        parts = [t.strip() for t in raw_text.replace(";", ",").split(",") if t.strip()]
        tokens = parts or [raw_text]
        seen: set[str] = set()
        tokens = [token for token in tokens if not (token in seen or seen.add(token))]
        if not tokens:
            return None

        from app.models.party import Party

        flat_like_exprs = [
            sqlalchemy_contains_query(CaseFlatIndex.inventor, token) for token in tokens
        ]
        inv_flat = or_(*flat_like_exprs) if len(flat_like_exprs) > 1 else flat_like_exprs[0]

        role_like_exprs = [
            or_(
                Party.name_display.ilike(f"%{token}%"),
                MatterPartyRole.raw_text.ilike(f"%{token}%"),
            )
            for token in tokens
        ]
        inv_party = (
            db.session.query(MatterPartyRole.mpr_id)
            .outerjoin(Party, Party.party_id == MatterPartyRole.party_id)
            .filter(MatterPartyRole.matter_id == VMatterOverview.matter_id)
            .filter(MatterPartyRole.role_code == "inventor")
            .filter(or_(*role_like_exprs))
            .exists()
        )

        raw_like_exprs = [RawImportField.value_text.ilike(f"%{token}%") for token in tokens]
        inv_raw = (
            db.session.query(RawImportField.raw_field_id)
            .join(Matter, Matter.raw_id == RawImportField.raw_id)
            .filter(Matter.matter_id == VMatterOverview.matter_id)
            .filter(RawImportField.sheet_name == "Matter")
            .filter(RawImportField.source_column == "Inventor")
            .filter(or_(*raw_like_exprs))
            .exists()
        )
        return or_(inv_flat, inv_party, inv_raw)

    def _app_no_expr(value: str):
        return or_(
            sqlalchemy_contains_query(CaseFlatIndex.application_no, value),
            (
                db.session.query(MatterIdentifier.mid_id)
                .filter(MatterIdentifier.matter_id == VMatterOverview.matter_id)
                .filter(sqlalchemy_contains_query(MatterIdentifier.id_value, value))
                .exists()
            ),
        )

    def _general_term_expr(term: str):
        value = str(term or "").strip()
        if not value:
            return None
        if _uses_compact_index_query(value):
            return CaseFlatIndex.search_compact.ilike(f"%{compact_search_text(value)}%")
        like = f"%{value}%"
        return or_(
            sqlalchemy_contains_query(CaseFlatIndex.search_text, value),
            sqlalchemy_contains_query(VMatterOverview.clients, value),
            sqlalchemy_contains_query(VMatterOverview.applicants, value),
            sqlalchemy_contains_query(VMatterOverview.right_name, value),
            sqlalchemy_contains_query(VMatterOverview.our_ref, value),
            sqlalchemy_contains_query(VMatterOverview.old_our_ref, value),
            sqlalchemy_contains_query(VMatterOverview.your_ref, value),
            sqlalchemy_contains_query(VMatterOverview.attorneys, value),
            _status_display_match_exists(VMatterOverview.matter_id, value),
            _party_name_match_exists(VMatterOverview.matter_id, like=like, role_code=None),
            _staff_name_match_exists(
                VMatterOverview.matter_id,
                like=like,
                role_codes=(
                    "attorney",
                    "retainer",
                    "handler",
                    "staff",
                    "draftsman",
                    "manager",
                    "mgmt",
                ),
            ),
        )

    # Hot-path cache (per-user, short TTL).
    try:
        user_id = getattr(current_user, "id", None)
    except Exception:
        user_id = None
    cache_key = None
    if user_id is not None:
        try:
            q_hash = hashlib.sha256(q.encode("utf-8", errors="ignore")).hexdigest()[:32]
        except Exception:
            q_hash = str(q)[:64]
        cache_key = f"case.search.v3:{int(user_id)}:{q_hash}"

        rc = _case_search_redis_client()
        if rc is not None:
            try:
                raw = rc.get(cache_key)
                if raw:
                    if isinstance(raw, (bytes, bytearray)):
                        raw = raw.decode("utf-8", errors="ignore")
                    cached = json.loads(raw)
                    if isinstance(cached, list):
                        return jsonify(cached)
            except Exception as exc:
                report_swallowed_exception(
                    exc,
                    context="case.routes.list.search.redis_get",
                    log_key="case.routes.list.search.redis_get",
                    log_window_seconds=60,
                )

        cached = _CASE_SEARCH_CACHE.get(cache_key)
        if isinstance(cached, list):
            return jsonify(cached)

    query = VMatterOverview.query

    if q:
        search_filter = build_sqlalchemy_search_filter(
            search_expr,
            general_term_builder=_general_term_expr,
            field_builders={
                "client": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.client_name, value),
                    sqlalchemy_contains_query(VMatterOverview.clients, value),
                    _party_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_code="client",
                    ),
                ),
                "applicant": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.applicant, value),
                    sqlalchemy_contains_query(VMatterOverview.applicants, value),
                    _party_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_code="applicant",
                    ),
                ),
                "party": lambda value: or_(
                    sqlalchemy_contains_query(VMatterOverview.clients, value),
                    sqlalchemy_contains_query(VMatterOverview.applicants, value),
                    _party_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_code=None,
                    ),
                ),
                "attorney": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.attorney, value),
                    sqlalchemy_contains_query(VMatterOverview.attorneys, value),
                    _staff_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_codes=("attorney", "retainer"),
                    ),
                ),
                "handler": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.handler, value),
                    _staff_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_codes=("handler", "staff", "draftsman"),
                    ),
                ),
                "manager": lambda value: or_(
                    sqlalchemy_contains_query(CaseFlatIndex.manager, value),
                    _staff_name_match_exists(
                        VMatterOverview.matter_id,
                        like=f"%{value}%",
                        role_codes=("manager", "mgmt"),
                    ),
                ),
                "our_ref": lambda value: sqlalchemy_contains_query(VMatterOverview.our_ref, value),
                "your_ref": lambda value: sqlalchemy_contains_query(
                    VMatterOverview.your_ref,
                    value,
                ),
                "matter_id": lambda value: sqlalchemy_contains_query(
                    VMatterOverview.matter_id,
                    value,
                ),
                "status": lambda value: or_(
                    sqlalchemy_contains_query(VMatterOverview.status_blue, value),
                    sqlalchemy_contains_query(VMatterOverview.status_red, value),
                    sqlalchemy_contains_query(VMatterOverview.inhouse_status, value),
                    _status_display_match_exists(VMatterOverview.matter_id, value),
                ),
                "right_name": lambda value: sqlalchemy_contains_query(
                    VMatterOverview.right_name,
                    value,
                ),
                "inventor": _inventor_match_expr,
                "app_no": _app_no_expr,
            },
        )

        query = query.outerjoin(CaseFlatIndex, VMatterOverview.matter_id == CaseFlatIndex.matter_id)
        if search_filter is not None:
            query = query.filter(search_filter)

        # For simple autocomplete, we just sort by new
        rows = (
            query.order_by(
                VMatterOverview.created_at.desc().nulls_last(),
                VMatterOverview.entered_at.desc(),
                VMatterOverview.matter_id.desc(),
            )
            .limit(20)
            .all()
        )
    else:
        rows = (
            query.order_by(
                VMatterOverview.created_at.desc().nulls_last(),
                VMatterOverview.entered_at.desc(),
                VMatterOverview.matter_id.desc(),
            )
            .limit(20)
            .all()
        )

    payload = [{"id": r.matter_id, "ref_no": r.our_ref, "title": r.right_name} for r in rows]
    if cache_key:
        rc = _case_search_redis_client()
        if rc is not None:
            try:
                rc.setex(cache_key, 5, json.dumps(payload, ensure_ascii=False))
            except Exception as exc:
                report_swallowed_exception(
                    exc,
                    context="case.routes.list.search.redis_set",
                    log_key="case.routes.list.search.redis_set",
                    log_window_seconds=60,
                )
        try:
            _CASE_SEARCH_CACHE.set(cache_key, payload, ttl_seconds=5.0)
        except Exception as exc:
            report_swallowed_exception(
                exc,
                context="case.routes.list.search.cache_set",
                log_key="case.routes.list.search.cache_set",
                log_window_seconds=60,
            )
    return jsonify(payload)


# --- All Letters & Notices ---


@bp.route("/all-letters")
@login_required
@require_permission("manage_case")
def all_letters():
    """List all letters (communications) across all matters."""
    page = _clamp_page(request.args.get("page", 1))
    per_page = _clamp_per_page(request.args.get("per_page", 50))
    search = request.args.get("q", "").strip()
    is_compact_q = search and _uses_compact_index_query(search)

    # Query communications with matter info - use LEFT JOIN for attach count instead of correlated subquery
    base_query = """
        SELECT
            c.comm_id, c.matter_id, c.note, c.comm_type,
            c.received_date, c.sent_date,
            m.our_ref,
            COALESCE(cfa.attach_count, 0) AS attach_count
        FROM communication c
        JOIN matter m ON m.matter_id = c.matter_id
        LEFT JOIN case_flat_index cfi ON cfi.matter_id = m.matter_id
        LEFT JOIN (
            SELECT comm_id, COUNT(*) AS attach_count
            FROM communication_file_asset
            GROUP BY comm_id
        ) cfa ON cfa.comm_id = c.comm_id
    """

    # Defensive classification: legacy data can have comm_type='R' even for emails.
    email_exists_sql = (
        "EXISTS ("
        " SELECT 1"
        " FROM communication_file_asset cf2"
        " JOIN file_asset fa2 ON fa2.file_asset_id = cf2.file_asset_id"
        " WHERE cf2.comm_id = c.comm_id"
        "   AND ("
        "     LOWER(COALESCE(fa2.original_name, '')) LIKE '%.eml'"
        "     OR LOWER(COALESCE(fa2.original_name, '')) LIKE '%.msg'"
        "     OR LOWER(COALESCE(fa2.mime_type, '')) IN ('message/rfc822', 'application/vnd.ms-outlook')"
        "     OR LOWER(COALESCE(fa2.file_path, '')) LIKE 'emails/%'"
        "   )"
        ")"
    )

    # Build WHERE clause
    where_clause = ""
    params = {}
    if search and not is_compact_q:
        where_clause = (
            " WHERE (c.note LIKE :search OR m.our_ref LIKE :search)"
            f" AND ({email_exists_sql} OR c.comm_type IS NULL OR c.comm_type != 'R')"
        )
        params["search"] = f"%{search}%"
    else:
        # Default filter: exclude 'R' responses, but keep email assets even if comm_type='R'.
        where_clause = f" WHERE ({email_exists_sql} OR c.comm_type IS NULL OR c.comm_type != 'R')"

    sort_expr = "COALESCE(NULLIF(c.received_date, ''), NULLIF(c.sent_date, ''))"
    order_clause = f" ORDER BY ({sort_expr} IS NULL) ASC, {sort_expr} DESC, c.comm_id DESC"

    if is_compact_q:
        rows, total, total_pages, page = _paginate_compact_search(
            base_query=base_query,
            where_clause=where_clause,
            params=params,
            order_clause=order_clause,
            search=search,
            page=page,
            per_page=per_page,
            build_search_text=None,
            compact_where_sql=(
                "COALESCE(c.search_compact, '') LIKE :like_compact "
                "OR COALESCE(cfi.search_compact, '') LIKE :like_compact"
            ),
        )
    else:
        # Count total - use simplified count query without attach_count subquery
        count_query = (
            "SELECT COUNT(*) FROM communication c JOIN matter m ON m.matter_id = c.matter_id"
            + where_clause
        )
        total = (
            db.session.execute(
                text(count_query).execution_options(policy_bypass=True), params
            ).scalar()
            or 0
        )
        total_pages = (total + per_page - 1) // per_page if total else 1
        if total_pages and page > total_pages:
            page = total_pages

        # Get paginated results
        data_query = base_query + where_clause + order_clause + " LIMIT :limit OFFSET :offset"
        params["limit"] = per_page
        params["offset"] = (page - 1) * per_page
        rows = (
            db.session.execute(text(data_query).execution_options(policy_bypass=True), params)
            .mappings()
            .all()
        )

    return render_template(
        "case/all_letters.html",
        rows=rows,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
        search=search,
    )


@bp.route("/all-responses")
@login_required
@require_permission("manage_case")
def all_responses():
    """List all response documents (comm_type='R') across all matters.

    Note: Emails (.eml/.msg) are always treated as letters, even if legacy data
    has comm_type='R'.
    """
    page = _clamp_page(request.args.get("page", 1))
    per_page = _clamp_per_page(request.args.get("per_page", 50))
    search = request.args.get("q", "").strip()
    is_compact_q = search and _uses_compact_index_query(search)

    # Query communications with matter info - use LEFT JOIN for attach count instead of correlated subquery
    base_query = """
        SELECT
            c.comm_id, c.matter_id, c.note, c.comm_type,
            c.received_date, c.sent_date,
            m.our_ref,
            COALESCE(cfa.attach_count, 0) AS attach_count
        FROM communication c
        JOIN matter m ON m.matter_id = c.matter_id
        LEFT JOIN case_flat_index cfi ON cfi.matter_id = m.matter_id
        LEFT JOIN (
            SELECT comm_id, COUNT(*) AS attach_count
            FROM communication_file_asset
            GROUP BY comm_id
        ) cfa ON cfa.comm_id = c.comm_id
    """

    # Build WHERE clause
    email_exists_sql = (
        "EXISTS ("
        " SELECT 1"
        " FROM communication_file_asset cf2"
        " JOIN file_asset fa2 ON fa2.file_asset_id = cf2.file_asset_id"
        " WHERE cf2.comm_id = c.comm_id"
        "   AND ("
        "     LOWER(COALESCE(fa2.original_name, '')) LIKE '%.eml'"
        "     OR LOWER(COALESCE(fa2.original_name, '')) LIKE '%.msg'"
        "     OR LOWER(COALESCE(fa2.mime_type, '')) IN ('message/rfc822', 'application/vnd.ms-outlook')"
        "     OR LOWER(COALESCE(fa2.file_path, '')) LIKE 'emails/%'"
        "   )"
        ")"
    )
    where_clause = f" WHERE c.comm_type = 'R' AND NOT ({email_exists_sql})"
    params = {}
    if search and not is_compact_q:
        where_clause += " AND (c.note LIKE :search OR m.our_ref LIKE :search)"
        params["search"] = f"%{search}%"

    sort_expr = "COALESCE(NULLIF(c.received_date, ''), NULLIF(c.sent_date, ''))"
    order_clause = f" ORDER BY ({sort_expr} IS NULL) ASC, {sort_expr} DESC, c.comm_id DESC"

    if is_compact_q:
        rows, total, total_pages, page = _paginate_compact_search(
            base_query=base_query,
            where_clause=where_clause,
            params=params,
            order_clause=order_clause,
            search=search,
            page=page,
            per_page=per_page,
            build_search_text=None,
            compact_where_sql=(
                "COALESCE(c.search_compact, '') LIKE :like_compact "
                "OR COALESCE(cfi.search_compact, '') LIKE :like_compact"
            ),
        )
    else:
        # Count total
        count_query = (
            "SELECT COUNT(*) FROM communication c JOIN matter m ON m.matter_id = c.matter_id"
            + where_clause
        )
        total = (
            db.session.execute(
                text(count_query).execution_options(policy_bypass=True), params
            ).scalar()
            or 0
        )
        total_pages = (total + per_page - 1) // per_page if total else 1
        if total_pages and page > total_pages:
            page = total_pages

        # Get paginated results
        data_query = base_query + where_clause + order_clause + " LIMIT :limit OFFSET :offset"
        params["limit"] = per_page
        params["offset"] = (page - 1) * per_page
        rows = (
            db.session.execute(text(data_query).execution_options(policy_bypass=True), params)
            .mappings()
            .all()
        )

    return render_template(
        "case/all_responses.html",
        rows=rows,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
        search=search,
    )


@bp.route("/all-notices")
@login_required
@require_permission("manage_case")
def all_notices():
    """List all office actions (notices) across all matters."""
    page = _clamp_page(request.args.get("page", 1))
    per_page = _clamp_per_page(request.args.get("per_page", 50))
    search = request.args.get("q", "").strip()
    is_compact_q = search and _uses_compact_index_query(search)

    # Query office_actions with matter info - use LEFT JOIN for attach count instead of correlated subquery
    base_query = """
        SELECT
            oa.oa_id, oa.matter_id, oa.doc_name,
            oa.received_date, oa.notified_date,
            oa.due_date, oa.extended_due_date, oa.done_date,
            m.our_ref,
            COALESCE(oafa.attach_count, 0) AS attach_count
        FROM office_action oa
        JOIN matter m ON m.matter_id = oa.matter_id
        LEFT JOIN case_flat_index cfi ON cfi.matter_id = m.matter_id
        LEFT JOIN (
            SELECT oa_id, COUNT(*) AS attach_count
            FROM office_action_file_asset
            GROUP BY oa_id
        ) oafa ON oafa.oa_id = oa.oa_id
    """

    # Build WHERE clause
    where_clause = ""
    params = {}
    if search and not is_compact_q:
        where_clause = (
            " WHERE (oa.doc_name LIKE :search OR m.our_ref LIKE :search)"
            " AND (oa.raw_id IS NULL OR oa.raw_id NOT LIKE 'MIGRATED_TO_COMM:%')"
        )
        params["search"] = f"%{search}%"
    else:
        where_clause = " WHERE (oa.raw_id IS NULL OR oa.raw_id NOT LIKE 'MIGRATED_TO_COMM:%')"

    sort_expr = (
        "COALESCE("
        "NULLIF(oa.received_date, ''), "
        "NULLIF(oa.notified_date, ''), "
        "NULLIF(oa.due_date, ''), "
        "NULLIF(oa.extended_due_date, ''), "
        "NULLIF(oa.done_date, '')"
        ")"
    )
    order_clause = f" ORDER BY ({sort_expr} IS NULL) ASC, {sort_expr} DESC, oa.oa_id DESC"

    if is_compact_q:
        rows, total, total_pages, page = _paginate_compact_search(
            base_query=base_query,
            where_clause=where_clause,
            params=params,
            order_clause=order_clause,
            search=search,
            page=page,
            per_page=per_page,
            build_search_text=None,
            compact_where_sql=(
                "COALESCE(oa.search_compact, '') LIKE :like_compact "
                "OR COALESCE(cfi.search_compact, '') LIKE :like_compact"
            ),
        )
    else:
        # Count total - use simplified count query without attach_count subquery
        count_query = (
            "SELECT COUNT(*) FROM office_action oa JOIN matter m ON m.matter_id = oa.matter_id"
            + where_clause
        )
        total = (
            db.session.execute(
                text(count_query).execution_options(policy_bypass=True), params
            ).scalar()
            or 0
        )
        total_pages = (total + per_page - 1) // per_page if total else 1
        if total_pages and page > total_pages:
            page = total_pages

        # Get paginated results
        data_query = base_query + where_clause + order_clause + " LIMIT :limit OFFSET :offset"
        params["limit"] = per_page
        params["offset"] = (page - 1) * per_page
        rows = (
            db.session.execute(text(data_query).execution_options(policy_bypass=True), params)
            .mappings()
            .all()
        )

    rows = [dict(r) for r in (rows or [])]

    return render_template(
        "case/all_notices.html",
        rows=rows,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
        search=search,
        today=date.today().isoformat(),
    )
