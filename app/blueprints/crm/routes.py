from __future__ import annotations

"""
CRM Routes for Client, Lead, Opportunity, Contact, and Activity management.
"""

import json
import mimetypes
import os
import re
import shutil
import tempfile
import uuid
import zipfile
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

from flask import (
    abort,
    current_app,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import String, cast, func, or_
from sqlalchemy.orm.attributes import flag_modified

from app.blueprints.crm import bp
from app.blueprints.crm.forms import (
    ActivityForm,
    ClientForm,
    ContactForm,
    LeadForm,
    OpportunityForm,
)
from app.extensions import db
from app.models.client import Client
from app.models.crm import CRMActivity, CRMContact, CRMLead, CRMOpportunity
from app.models.crm_client_merge_log import CRMClientMergeLog
from app.models.docket import DocketItem
from app.models.operation import Operation
from app.models.system_config import SystemConfig
from app.models.user import User
from app.services.audit.entity_audit import (
    diff_snapshots,
    record_entity_change_audit,
    snapshot_attrs,
)
from app.services.billing.db_core import _actual_table_name as _billing_table_name
from app.services.billing.db_core import get_db as get_billing_db
from app.services.billing.db_core import row_to_dict as billing_row_to_dict
from app.services.billing.db_core import unified_clients_enabled
from app.services.client.applicant_code_suggestion import (
    APPLICANT_CODE_LIMIT as _APPLICANT_CODE_LIMIT,
)
from app.services.client.applicant_code_suggestion import (
    build_applicant_code_suggestion as _build_applicant_code_suggestion,
)
from app.services.client.applicant_code_suggestion import (
    extract_client_applicant_codes as _extract_client_applicant_codes,
)
from app.services.client.applicant_code_suggestion import (
    normalize_applicant_code as _normalize_applicant_code,
)
from app.services.client.background_jobs import (
    create_customer_llm_parse_operation,
    enqueue_crm_client_post_save,
    enqueue_customer_llm_parse,
    set_crm_client_search_tags_fast,
)
from app.services.client.client_attachment_service import (
    ALLOWED_BIZREG_EXTS,
    client_attachment_dir,
    delete_client_attachment_for_crm_client,
    get_bizreg_attachment_for_crm_client,
    get_client_attachment_for_crm_client,
    list_client_attachments_for_crm_client,
    migrate_legacy_bizreg_for_crm_client,
    open_client_attachment_stream,
    resolve_attachment_client_id_for_crm_client,
    resolve_client_attachment_file_path,
    save_bizreg_attachment_for_crm_client,
    save_client_attachment_for_crm_client,
)
from app.services.client.client_merge_service import ClientMergeService
from app.services.client.client_tagging import build_client_search_tags_text
from app.services.core.llm_runtime import get_openai_api_key
from app.services.storage.file_asset_service import FileAssetService, UploadTooLargeError
from app.services.uploads.intake_security import (
    UploadSecurityError,
    scan_upload_path,
    validate_upload_path,
)
from app.utils.api_errors import json_error
from app.utils.error_logging import report_swallowed_exception
from app.utils.permissions import is_admin, is_invoice_manager, role_required
from app.utils.policy_sql import policy_text as text
from app.utils.search import sqlalchemy_contains_query
from app.utils.url_helpers import safe_referrer_path

# ============================================================================
# Helper: CRM Business Registration File Directory
# ============================================================================

_CRM_CLIENT_MATTER_NAMESPACES = (
    "domestic_patent",
    "domestic_design",
    "domestic_trademark",
    "incoming_patent",
    "incoming_design",
    "incoming_trademark",
    "outgoing_patent",
    "outgoing_design",
    "outgoing_trademark",
    "pct",
    "litigation",
    "misc",
    # Canonical helper namespace (best-effort)
    "basic",
)


def _truthy_debug_value(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on"}


def _count_debug_items(value: Any) -> int:
    if isinstance(value, dict):
        return len(value)
    if isinstance(value, (list, tuple, set)):
        return len(value)
    if value in (None, ""):
        return 0
    return 1


def _sanitize_applicant_code_debug_payload(payload: dict[str, Any] | None) -> dict[str, Any]:
    """Keep CRM applicant-code log signal without names, matter IDs, or customer codes."""
    data = payload if isinstance(payload, dict) else {}
    sanitized: dict[str, Any] = {
        "secured_debug": False,
        "client_id_present": data.get("client_id") is not None,
        "client_name_present": bool(str(data.get("client_name") or "").strip()),
    }

    keep_scalar_keys = {
        "slots",
        "matter_count",
        "candidates_primary_count",
        "candidates_name_match_count",
        "reason",
    }
    for key in keep_scalar_keys:
        if key in data:
            value = data.get(key)
            if isinstance(value, (int, float, bool)) or value is None:
                sanitized[key] = value
            else:
                sanitized[key] = str(value)[:80]

    sanitized["existing_codes_count"] = _count_debug_items(data.get("existing_codes"))
    sanitized["matter_ids_preview_count"] = _count_debug_items(data.get("matter_ids_preview"))
    sanitized["name_variants_count"] = _count_debug_items(data.get("name_variants"))

    suggestion = data.get("suggestion")
    if isinstance(suggestion, dict):
        sanitized["suggestion"] = {
            "codes_count": _count_debug_items(suggestion.get("codes")),
            "slots": suggestion.get("slots"),
            "extra_count": suggestion.get("extra_count"),
        }
    return sanitized


def _secured_applicant_code_debug_enabled() -> bool:
    if not _truthy_debug_value(request.args.get("debug_applicant_code")):
        return False
    if not _truthy_debug_value(current_app.config.get("CRM_APPLICANT_CODE_SECURED_DEBUG_ENABLED")):
        return False
    return bool(is_admin(current_user))


_CRM_CLIENT_PARTY_ROLE_CODES = ("client", "applicant")
_CLIENT_AUDIT_FIELDS = (
    "id",
    "party_id",
    "ipm_party_id",
    "ipm_client_id",
    "external_invoice_client_id",
    "name",
    "type",
    "registration_number",
    "contact_person",
    "manager",
    "email",
    "phone",
    "address",
    "notes",
    "biz_reg_number",
    "biz_company_name",
    "biz_representative_name",
    "biz_opening_date",
    "biz_corp_registration_number",
    "biz_business_location",
    "biz_head_office_location",
    "biz_business_type",
    "biz_tax_invoice_email",
    "extra",
    "is_deleted",
    "deleted_at",
    "deleted_by",
    "delete_reason",
)


def _client_audit_snapshot(client: Client) -> dict[str, Any]:
    return snapshot_attrs(client, _CLIENT_AUDIT_FIELDS)


def _record_client_audit(
    *,
    client: Client,
    action: str,
    before: dict[str, Any] | None = None,
    source: str,
    include_snapshots: bool = False,
) -> None:
    after = _client_audit_snapshot(client)
    changes = diff_snapshots(before or {}, after)
    if before is not None and not changes and not include_snapshots:
        return
    record_entity_change_audit(
        action=action,
        target_type="crm_client",
        target_id=client.id,
        actor_id=getattr(current_user, "id", None),
        changes=changes,
        before=before,
        after=after if include_snapshots else None,
        meta={"client_id": client.id, "client_name": client.name, "source": source},
        title=client.name,
        include_snapshots=include_snapshots,
    )


def _safe_positive_int(value) -> int | None:
    try:
        ivalue = int(value)
    except (TypeError, ValueError):
        return None
    return ivalue if ivalue > 0 else None


def _clean_text_value(value: Any) -> str:
    return str(value or "").strip()


def _crm_client_list_query(
    *,
    q: str,
    invoice_link: str,
    sort: str,
    direction: str,
):
    q = (q or "").strip()
    invoice_link = (invoice_link or "").strip().lower()
    sort = (sort or "id").strip()
    direction = (direction or "").strip().lower()

    query = Client.query.filter_by(is_deleted=False)

    if invoice_link == "missing":
        query = query.filter(Client.external_invoice_client_id.is_(None))
    elif invoice_link == "linked":
        query = query.filter(Client.external_invoice_client_id.isnot(None))
    else:
        invoice_link = ""

    if q:
        query = query.filter(
            db.or_(
                sqlalchemy_contains_query(Client.name, q),
                sqlalchemy_contains_query(Client.email, q),
                sqlalchemy_contains_query(Client.registration_number, q),
                sqlalchemy_contains_query(Client.phone, q),
                sqlalchemy_contains_query(Client.biz_reg_number, q),
                sqlalchemy_contains_query(Client.biz_company_name, q),
                sqlalchemy_contains_query(Client.biz_tax_invoice_email, q),
                sqlalchemy_contains_query(Client.search_tags, q),
            )
        )

    sort_map = {
        "name": Client.name,
        "id": Client.id,
        "registration_number": Client.registration_number,
        "email": Client.email,
        "phone": Client.phone,
    }
    if sort not in sort_map:
        sort = "id"
    if direction not in ("asc", "desc"):
        direction = "desc" if sort == "id" else "asc"

    sort_column = sort_map[sort]
    order_expr = sort_column.desc() if direction == "desc" else sort_column.asc()
    if sort == "id":
        query = query.order_by(order_expr)
    else:
        tie_breaker = Client.id.desc() if direction == "desc" else Client.id.asc()
        query = query.order_by(order_expr, tie_breaker)

    return query, sort, direction, invoice_link


def _active_matter_filter(Matter):
    return (Matter.is_deleted.is_(False)) | (Matter.is_deleted.is_(None))


def _client_party_ids(client: Client) -> list[str]:
    out: list[str] = []
    for pid in (getattr(client, "party_id", None), getattr(client, "ipm_party_id", None)):
        pid = _clean_text_value(pid)
        if pid and pid not in out:
            out.append(pid)
    return out


def _fetch_client_matter_map(
    clients: list[Client],
    *,
    limit_per_client: int | None = None,
) -> dict[int, tuple[list, int]]:
    """Resolve CRM clients to active Matter rows via custom-field and party links."""
    from app.models.ip_records import Matter, MatterCustomField, MatterPartyRole

    client_ids = [int(c.id) for c in (clients or []) if _safe_positive_int(getattr(c, "id", None))]
    if not client_ids:
        return {}

    client_id_set = set(client_ids)
    links_by_client: dict[int, set[str]] = {cid: set() for cid in client_ids}

    try:
        bind = db.session.get_bind()
        dialect = (getattr(bind.dialect, "name", "") or "").lower() if bind else ""
    except Exception:
        dialect = ""

    try:
        if dialect.startswith("postgres"):
            client_id_expr = MatterCustomField.data["client_id"].as_string()
        else:
            client_id_expr = func.json_extract(MatterCustomField.data, "$.client_id")
        client_id_expr = cast(client_id_expr, String)

        rows = (
            db.session.query(
                client_id_expr.label("client_id"),
                MatterCustomField.matter_id.label("matter_id"),
            )
            .join(Matter, Matter.matter_id == MatterCustomField.matter_id)
            .filter(MatterCustomField.namespace.in_(_CRM_CLIENT_MATTER_NAMESPACES))
            .filter(client_id_expr.in_([str(cid) for cid in client_ids]))
            .filter(_active_matter_filter(Matter))
            .distinct()
            .all()
        )
        for raw_cid, raw_mid in rows:
            cid = _safe_positive_int(raw_cid)
            mid = _clean_text_value(raw_mid)
            if cid in client_id_set and mid:
                links_by_client.setdefault(cid, set()).add(mid)
    except Exception:
        current_app.logger.exception(
            "Failed to query CRM client matter links from custom fields (dialect=%s)",
            dialect,
        )
        try:
            rows = MatterCustomField.query.filter(
                MatterCustomField.namespace.in_(_CRM_CLIENT_MATTER_NAMESPACES)
            ).all()
            for row in rows:
                data = row.data or {}
                if not isinstance(data, dict):
                    continue
                cid = _safe_positive_int(data.get("client_id"))
                mid = _clean_text_value(getattr(row, "matter_id", None))
                if cid in client_id_set and mid:
                    links_by_client.setdefault(cid, set()).add(mid)
        except Exception:
            current_app.logger.exception("Fallback CRM client matter custom-field lookup failed")

    party_to_clients: dict[str, list[int]] = {}
    for client in clients or []:
        cid = _safe_positive_int(getattr(client, "id", None))
        if cid not in client_id_set:
            continue
        for pid in _client_party_ids(client):
            party_to_clients.setdefault(pid, [])
            if cid not in party_to_clients[pid]:
                party_to_clients[pid].append(cid)

    if party_to_clients:
        try:
            rows = (
                db.session.query(MatterPartyRole.party_id, MatterPartyRole.matter_id)
                .join(Matter, Matter.matter_id == MatterPartyRole.matter_id)
                .filter(func.lower(MatterPartyRole.role_code).in_(_CRM_CLIENT_PARTY_ROLE_CODES))
                .filter(MatterPartyRole.party_id.in_(list(party_to_clients)))
                .filter(_active_matter_filter(Matter))
                .distinct()
                .all()
            )
            for raw_pid, raw_mid in rows:
                pid = _clean_text_value(raw_pid)
                mid = _clean_text_value(raw_mid)
                if not (pid and mid):
                    continue
                for cid in party_to_clients.get(pid, []):
                    links_by_client.setdefault(cid, set()).add(mid)
        except Exception:
            current_app.logger.exception("Fallback party-role lookup failed for CRM clients")

    all_matter_ids = sorted({mid for mids in links_by_client.values() for mid in mids if mid})
    if not all_matter_ids:
        return {cid: ([], 0) for cid in client_ids}

    matters = (
        Matter.query.filter(Matter.matter_id.in_(all_matter_ids))
        .filter(_active_matter_filter(Matter))
        .order_by(
            func.coalesce(Matter.our_ref, "").desc(),
            func.coalesce(Matter.entered_at, "").desc(),
            Matter.created_at.desc(),
            Matter.matter_id.desc(),
        )
        .all()
    )

    result: dict[int, tuple[list, int]] = {}
    for cid in client_ids:
        matter_ids = links_by_client.get(cid, set())
        selected: list = []
        for matter in matters:
            mid = _clean_text_value(getattr(matter, "matter_id", None))
            if mid not in matter_ids:
                continue
            selected.append(matter)
            if limit_per_client and len(selected) >= limit_per_client:
                break
        result[cid] = (selected, len(matter_ids))
    return result


def _fetch_client_matters_for_client(client: Client, *, limit: int = 30) -> tuple[list, int]:
    safe_limit = 30
    try:
        safe_limit = int(limit)
    except Exception:
        safe_limit = 30
    safe_limit = max(1, min(safe_limit, 200))
    return _fetch_client_matter_map([client], limit_per_client=safe_limit).get(
        int(client.id), ([], 0)
    )


def _safe_export_filename_part(value: Any, *, fallback: str) -> str:
    text_value = _clean_text_value(value) or fallback
    text_value = re.sub(r'[\\/:*?"<>|]+', "_", text_value)
    text_value = re.sub(r"\s+", " ", text_value).strip(" ._")
    if not text_value:
        text_value = fallback
    return text_value[:80]


def _autosize_worksheet(ws) -> None:
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            value_len = len(str(getattr(cell, "value", "") or ""))
            if value_len > max_length:
                max_length = value_len
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)


def _open_docket_summary_by_matter(matter_ids: list[str]) -> dict[str, dict[str, Any]]:
    out = {
        mid: {"open_count": 0, "next_due_date": "", "next_due_name": ""}
        for mid in matter_ids
        if mid
    }
    if not out:
        return out

    rows = (
        DocketItem.query.filter(DocketItem.matter_id.in_(list(out)))
        .filter((DocketItem.is_deleted.is_(False)) | (DocketItem.is_deleted.is_(None)))
        .filter((DocketItem.done_date.is_(None)) | (DocketItem.done_date == ""))
        .all()
    )
    for item in rows:
        mid = _clean_text_value(getattr(item, "matter_id", None))
        if mid not in out:
            continue
        info = out[mid]
        info["open_count"] = int(info.get("open_count") or 0) + 1
        due = _clean_text_value(getattr(item, "extended_due_date", None)) or _clean_text_value(
            getattr(item, "due_date", None)
        )
        if due and (not info.get("next_due_date") or due < str(info.get("next_due_date"))):
            info["next_due_date"] = due
            info["next_due_name"] = _clean_text_value(
                getattr(item, "name_ref", None)
            ) or _clean_text_value(getattr(item, "name_free", None))
    return out


def _case_export_custom_values_by_matter(matter_ids: list[str]) -> dict[str, dict[str, str]]:
    from app.models.ip_records import MatterCustomField

    target_ids = [mid for mid in matter_ids if mid]
    out = {
        mid: {"applicant_name": "", "application_no": "", "application_date": ""}
        for mid in target_ids
    }
    if not target_ids:
        return out

    applicant_keys = ("application_applicant_name", "applicant_name", "applicant_registrant")
    app_no_keys = (
        "application_no",
        "app_no",
        "pct_application_no",
        "madrid_application_no",
        "hague_application_no",
        "ep_application_no",
        "ctm_application_no",
    )
    app_date_keys = (
        "application_date",
        "app_date",
        "pct_application_date",
        "international_filing_date",
        "madrid_application_date",
        "hague_application_date",
    )

    rows = (
        MatterCustomField.query.filter(MatterCustomField.matter_id.in_(target_ids))
        .order_by((MatterCustomField.namespace == "basic").desc())
        .all()
    )
    for row in rows:
        mid = _clean_text_value(getattr(row, "matter_id", None))
        data = getattr(row, "data", None) or {}
        if mid not in out or not isinstance(data, dict):
            continue
        for target, keys in (
            ("applicant_name", applicant_keys),
            ("application_no", app_no_keys),
            ("application_date", app_date_keys),
        ):
            if out[mid].get(target):
                continue
            for key in keys:
                value = _clean_text_value(data.get(key))
                if value:
                    out[mid][target] = value
                    break
    return out


def _client_case_status_xlsx_bytes(
    *,
    client: Client,
    matters: list,
    total_matters: int,
    truncated: bool,
) -> bytes:
    from app.blueprints.case.helpers import _build_case_list_extras

    wb = Workbook()
    ws = wb.active
    ws.title = "Case Status"

    headers = [
        "No",
        "Our Ref",
        "Your Ref",
        "Matter",
        "Type",
        "Status(Red)",
        "Status(Blue)",
        "InternalStatus",
        "Applicant",
        "Application No.",
        "Filing date",
        "Next Deadline",
        "Next Deadline",
        "Done Deadline ",
        "Engagement date",
        "Entry date",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    extras = _build_case_list_extras(matters or [])
    matter_ids = [_clean_text_value(getattr(m, "matter_id", None)) for m in (matters or [])]
    docket_summary = _open_docket_summary_by_matter([mid for mid in matter_ids if mid])
    custom_fallbacks = _case_export_custom_values_by_matter([mid for mid in matter_ids if mid])

    for idx, matter in enumerate(matters or [], start=1):
        mid = _clean_text_value(getattr(matter, "matter_id", None))
        extra = extras.get(mid, {}) if mid else {}
        custom = custom_fallbacks.get(mid, {}) if mid else {}
        dockets = docket_summary.get(mid, {})
        division = _clean_text_value(getattr(matter, "right_group", None))
        matter_type = _clean_text_value(getattr(matter, "matter_type", None))
        kind = " / ".join([part for part in (division, matter_type) if part])

        ws.append(
            [
                idx,
                _clean_text_value(getattr(matter, "our_ref", None)),
                _clean_text_value(getattr(matter, "your_ref", None)),
                _clean_text_value(getattr(matter, "right_name", None)),
                kind,
                _clean_text_value(extra.get("display_red"))
                or _clean_text_value(getattr(matter, "status_red", None)),
                _clean_text_value(extra.get("display_blue"))
                or _clean_text_value(getattr(matter, "status_blue", None)),
                _clean_text_value(getattr(matter, "inhouse_status", None)),
                _clean_text_value(extra.get("applicant_name"))
                or _clean_text_value(custom.get("applicant_name")),
                _clean_text_value(extra.get("application_no"))
                or _clean_text_value(custom.get("application_no")),
                _clean_text_value(extra.get("application_date"))
                or _clean_text_value(custom.get("application_date")),
                _clean_text_value(dockets.get("next_due_date")),
                _clean_text_value(dockets.get("next_due_name")),
                int(dockets.get("open_count") or 0),
                _clean_text_value(getattr(matter, "retained_at", None)),
                _clean_text_value(getattr(matter, "entered_at", None)),
            ]
        )

    ws.freeze_panes = "A2"
    _autosize_worksheet(ws)

    meta = wb.create_sheet("Meta")
    meta.append(["Client ID", int(getattr(client, "id", 0) or 0)])
    meta.append(["Client name", _clean_text_value(getattr(client, "name", None))])
    meta.append(["All Link Matter ", int(total_matters or 0)])
    meta.append([" Matter ", len(matters or [])])
    meta.append(["Matter  to  ", "Y" if truncated else "N"])
    meta.append(["Export Date", datetime.utcnow().isoformat(timespec="seconds") + "Z"])
    _autosize_worksheet(meta)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _resolve_invoice_client_id_for_crm_client(client: Client) -> int | None:
    """Resolve the invoice-module client id without mutating CRM data."""
    try:
        if unified_clients_enabled():
            return _safe_positive_int(getattr(client, "id", None))
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_view.invoice_client_id.unified_mode",
            log_key="crm.client_view.invoice_client_id.unified_mode",
            log_window_seconds=300,
        )
        if current_app.config.get("INVOICEAPP_UNIFIED_CLIENTS"):
            return _safe_positive_int(getattr(client, "id", None))

    linked_id = _safe_positive_int(getattr(client, "external_invoice_client_id", None))
    if linked_id:
        return linked_id

    # Older/non-unified invoice rows can be linked only from billing clients.ipm_client_id.
    conn = None
    try:
        conn = get_billing_db()
        row = conn.execute(
            f"""
            SELECT id
            FROM {_billing_table_name('clients')}
            WHERE ipm_client_id = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (_safe_positive_int(getattr(client, "id", None)) or 0,),
        ).fetchone()
        if row:
            data = billing_row_to_dict(row)
            return _safe_positive_int(data.get("id"))
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_view.invoice_client_id.lookup",
            log_key="crm.client_view.invoice_client_id.lookup",
            log_window_seconds=300,
        )
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                _ = None

    return None


def _empty_client_invoice_snapshot(invoice_client_id: int | None = None) -> dict:
    return {
        "client_id": invoice_client_id,
        "items": [],
        "total_count": 0,
        "invoiced_by_currency": {},
        "outstanding_by_currency": {},
        "latest_limit": 20,
    }


def _fetch_client_invoice_snapshot(client: Client, *, limit: int = 20) -> dict:
    invoice_client_id = _resolve_invoice_client_id_for_crm_client(client)
    snapshot = _empty_client_invoice_snapshot(invoice_client_id)
    if not invoice_client_id:
        return snapshot

    safe_limit = max(1, min(int(limit or 20), 100))
    snapshot["latest_limit"] = safe_limit
    billing_expr = "COALESCE(i.billing_status, i.status)"
    payment_expr = (
        "COALESCE("
        "i.payment_status,"
        "CASE "
        " WHEN i.status='paid' THEN 'paid'"
        " WHEN i.status IN ('payment_pending','pre_overdue') THEN 'pending'"
        " WHEN i.status='void' THEN 'none'"
        " ELSE 'unpaid'"
        "END"
        ")"
    )
    active_outstanding_expr = (
        f"(({payment_expr} IN ('unpaid','pending')) OR ({billing_expr}='pre_overdue')) "
        f"AND NOT ({billing_expr}='void' OR {payment_expr}='none')"
    )

    conn = None
    try:
        conn = get_billing_db()
        invoices_tbl = _billing_table_name("invoices")
        profiles_tbl = _billing_table_name("business_profile")

        rows = conn.execute(
            f"""
            SELECT i.*, bp.name AS business_name
            FROM {invoices_tbl} i
            LEFT JOIN {profiles_tbl} bp ON bp.id = i.business_profile_id
            WHERE i.client_id = ?
            ORDER BY i.issue_date DESC, i.id DESC
            LIMIT ?
            """,
            (int(invoice_client_id), safe_limit),
        ).fetchall()
        snapshot["items"] = [billing_row_to_dict(r) for r in rows]

        stat_rows = conn.execute(
            f"""
            SELECT
                COALESCE(i.currency, 'USD') AS currency,
                COUNT(*) AS invoice_count,
                COALESCE(SUM(COALESCE(i.total, 0)), 0) AS invoiced_total,
                COALESCE(SUM(
                    CASE WHEN {active_outstanding_expr}
                         THEN COALESCE(i.total, 0)
                         ELSE 0
                    END
                ), 0) AS outstanding_total
            FROM {invoices_tbl} i
            WHERE i.client_id = ?
            GROUP BY COALESCE(i.currency, 'USD')
            ORDER BY COALESCE(i.currency, 'USD')
            """,
            (int(invoice_client_id),),
        ).fetchall()

        total_count = 0
        invoiced_by_currency = {}
        outstanding_by_currency = {}
        for row in stat_rows:
            data = billing_row_to_dict(row)
            currency = str(data.get("currency") or "USD").upper()
            count = int(data.get("invoice_count") or 0)
            total_count += count
            invoiced_by_currency[currency] = float(data.get("invoiced_total") or 0)
            outstanding = float(data.get("outstanding_total") or 0)
            if outstanding:
                outstanding_by_currency[currency] = outstanding

        snapshot["total_count"] = total_count
        snapshot["invoiced_by_currency"] = invoiced_by_currency
        snapshot["outstanding_by_currency"] = outstanding_by_currency
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_view.invoices",
            log_key="crm.client_view.invoices",
            log_window_seconds=300,
        )
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                _ = None

    return snapshot


def _crm_biz_reg_dir(client_id: int) -> str:
    """Return directory for storing CRM client's business registration files.
    Creates the directory if it doesn't exist.
    """
    base = current_app.config.get("CLIENT_ATTACHMENTS_DIR", "uploads/clients")
    path = os.path.join(base, f"crm_client_{client_id}")
    os.makedirs(path, exist_ok=True)
    return path


def _looks_like_bizreg_filename(name: str | None) -> bool:
    raw = (name or "").strip()
    if not raw:
        return False
    if "Business profileRegistration" in raw:
        return True
    lower = raw.lower()
    return (
        "biz_reg" in lower
        or "bizreg" in lower
        or "business registration" in lower
        or "business_registration" in lower
    )


def _save_biz_reg_file(client_id: int, file_obj, original_filename: str) -> dict:
    """Save a business registration file and return metadata.

    Returns:
        dict with 'original_name', 'stored_name', 'uploaded_at'
    """
    upload_dir = _crm_biz_reg_dir(client_id)
    root_dir = Path(current_app.config.get("CLIENT_ATTACHMENTS_DIR", "uploads/clients")).resolve()
    upload_dir_path = Path(upload_dir).resolve()
    try:
        rel_dir = upload_dir_path.relative_to(root_dir)
    except ValueError as exc:
        raise ValueError("invalid attachment path") from exc

    # Generate unique stored name
    ext = os.path.splitext(original_filename)[1].lower()
    stored_name = f"biz_reg_{uuid.uuid4().hex[:12]}{ext}"
    rel_path = rel_dir / stored_name
    service = FileAssetService(upload_root=root_dir)

    # If file_obj is a path string, copy; otherwise save directly
    if isinstance(file_obj, str):
        with open(file_obj, "rb") as handle:
            service.store_stream_to_path(handle, rel_path=rel_path, overwrite=True)
    else:
        service.store_upload_to_path(file_obj, rel_path=rel_path, overwrite=True)

    return {
        "original_name": original_filename,
        "stored_name": stored_name,
        "uploaded_at": datetime.utcnow().isoformat(),
    }


def _attach_biz_reg_file(client: Client, file_info: dict) -> dict:
    """Attach biz registration file metadata to client.extra safely."""
    extra = dict(client.extra or {})
    extra["biz_reg_file"] = file_info
    client.extra = extra
    flag_modified(client, "extra")
    return extra


def _reset_request_db_session() -> None:
    """Release any request-scoped DB connection before/after long external work."""
    try:
        db.session.rollback()
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.reset_request_db_session.rollback",
            log_key="crm.reset_request_db_session.rollback",
            log_window_seconds=300,
        )
    try:
        db.session.remove()
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.reset_request_db_session.remove",
            log_key="crm.reset_request_db_session.remove",
            log_window_seconds=300,
        )


def _set_client_search_tags(
    client: Client,
    extra: dict,
    *,
    api_key: str | None = None,
    use_llm: bool | None = None,
) -> None:
    if api_key is None:
        api_key = get_openai_api_key(allow_legacy=False) or None
    if use_llm is None:
        use_llm = bool(api_key)

    applicant_codes: list[Any] = []
    if isinstance(extra, dict):
        raw_codes = extra.get("applicant_codes") or []
        if isinstance(raw_codes, (list, tuple, set)):
            applicant_codes = list(raw_codes)
        else:
            applicant_codes = [raw_codes]

    values = [
        client.name,
        client.biz_company_name,
        (extra.get("name_en") if isinstance(extra, dict) else None),
        (extra.get("tax_company_name") if isinstance(extra, dict) else None),
        (extra.get("client_code") if isinstance(extra, dict) else None),
        *applicant_codes,
        getattr(client, "registration_number", None),
        getattr(client, "biz_reg_number", None),
        getattr(client, "phone", None),
    ]
    tags = build_client_search_tags_text(values, api_key=api_key, use_llm=bool(use_llm))
    client.search_tags = tags or None


_DUP_KIND_LABELS = {
    "email": "Email",
    "tax_email": "Billing tax email",
    "phone": "Phone",
    "registration": "Registration No.",
    "biz_reg": "Tax ID / EIN",
}
_DUP_KIND_WEIGHT = {
    "registration": 400,
    "biz_reg": 400,
    "email": 250,
    "tax_email": 250,
    "phone": 200,
}
_DUP_REJECTED_CONFIG_KEY = "CRM_DUPLICATE_REJECTED_GROUPS_JSON"
_DUP_REJECTED_MAX_ENTRIES = 500
_DUP_GROUP_SIGNATURE_RE = re.compile(r"^ids:\d+(?:,\d+)+$")


def _normalize_email(value: str | None) -> str:
    text = (value or "").strip().lower()
    if not text or "@" not in text:
        return ""
    local, _, domain = text.partition("@")
    local = local.strip()
    domain = domain.strip()
    if not local or not domain:
        return ""
    if "+" in local:
        local = local.split("+", 1)[0]
    if domain in ("gmail.com", "googlemail.com"):
        local = local.replace(".", "")
    return f"{local}@{domain}" if local else ""


def _normalize_digits(value: str | None, *, min_len: int = 6) -> str:
    digits = "".join(ch for ch in (value or "") if ch.isdigit())
    return digits if len(digits) >= min_len else ""


def _normalize_phone(value: str | None) -> str:
    digits = "".join(ch for ch in (value or "") if ch.isdigit())
    if not digits:
        return ""
    if digits.startswith("82") and len(digits) >= 11:
        tail = digits[2:]
        if tail.startswith("0"):
            tail = tail[1:]
        digits = f"0{tail}"
    return digits if len(digits) >= 7 else ""


def _build_duplicate_groups(rows: list) -> list[dict]:
    value_map: dict[tuple[str, str], set[int]] = {}
    client_names: dict[int, str] = {}

    for r in rows:
        try:
            cid = int(r.id)
        except (TypeError, ValueError):
            continue
        client_names[cid] = (r.name or "").strip()

        email = _normalize_email(getattr(r, "email", None))
        tax_email = _normalize_email(getattr(r, "biz_tax_invoice_email", None))
        phone = _normalize_phone(getattr(r, "phone", None))
        reg = _normalize_digits(getattr(r, "registration_number", None), min_len=6)
        biz_reg = _normalize_digits(getattr(r, "biz_reg_number", None), min_len=6)

        keys = [
            ("email", email),
            ("tax_email", tax_email),
            ("phone", phone),
            ("registration", reg),
            ("biz_reg", biz_reg),
        ]
        for kind, value in keys:
            if not value:
                continue
            value_map.setdefault((kind, value), set()).add(cid)

    dup_keys = {k: ids for k, ids in value_map.items() if len(ids) > 1}
    if not dup_keys:
        return []

    parent: dict[int, int] = {}

    def _find(x: int) -> int:
        parent.setdefault(x, x)
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def _union(a: int, b: int) -> None:
        ra = _find(a)
        rb = _find(b)
        if ra != rb:
            parent[rb] = ra

    for ids in dup_keys.values():
        ids_list = list(ids)
        if len(ids_list) < 2:
            continue
        base = ids_list[0]
        for other in ids_list[1:]:
            _union(base, other)

    groups: dict[int, dict] = {}
    for (kind, value), ids in dup_keys.items():
        root = _find(next(iter(ids)))
        group = groups.setdefault(root, {"ids": set(), "reasons": []})
        group["ids"].update(ids)
        group["reasons"].append(
            {
                "kind": kind,
                "kind_label": _DUP_KIND_LABELS.get(kind, kind),
                "value": value,
                "count": len(ids),
                "weight": _DUP_KIND_WEIGHT.get(kind, 0),
            }
        )

    results: list[dict] = []
    for group in groups.values():
        ids = sorted(group["ids"])
        reasons = sorted(
            group["reasons"],
            key=lambda r: (-r["weight"], -r["count"], r["kind_label"]),
        )
        if not reasons or len(ids) < 2:
            continue
        primary = reasons[0]
        display_reasons = [{k: v for k, v in r.items() if k != "weight"} for r in reasons]
        score = (
            len(ids) * 100
            + len(display_reasons) * 25
            + sum(_DUP_KIND_WEIGHT.get(r["kind"], 0) for r in display_reasons)
        )
        results.append(
            {
                "kind": primary["kind"],
                "kind_label": primary["kind_label"],
                "value": primary["value"],
                "client_ids": ids,
                "clients": [{"id": cid, "name": client_names.get(cid, "")} for cid in ids],
                "reasons": display_reasons,
                "score": score,
            }
        )

    results.sort(
        key=lambda g: (
            -int(g.get("score") or 0),
            -len(g.get("client_ids") or []),
            g.get("kind_label") or "",
        )
    )
    return results


def _duplicate_group_signature(client_ids: list[int] | tuple[int, ...] | set[int]) -> str:
    ids: list[int] = []
    for raw in client_ids or []:
        try:
            cid = int(raw)
        except (TypeError, ValueError):
            continue
        if cid > 0:
            ids.append(cid)
    ids = sorted(set(ids))
    if len(ids) < 2:
        return ""
    return "ids:" + ",".join(str(i) for i in ids)


def _normalize_duplicate_group_signature(value: str | None) -> str:
    raw = (value or "").strip()
    if not raw or not _DUP_GROUP_SIGNATURE_RE.fullmatch(raw):
        return ""
    prefix, _, body = raw.partition(":")
    if prefix != "ids" or not body:
        return ""
    tokens = []
    for part in body.split(","):
        if not part.isdigit():
            continue
        cid = int(part)
        if cid > 0:
            tokens.append(cid)
    if len(tokens) < 2:
        return ""
    return _duplicate_group_signature(tokens)


def _load_rejected_duplicate_group_signatures() -> list[str]:
    try:
        raw = SystemConfig.get_config(_DUP_REJECTED_CONFIG_KEY, "[]") or "[]"
        parsed = json.loads(raw)
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.duplicate_groups.load_rejected_signatures",
            log_key="crm.duplicate_groups.load_rejected_signatures",
            log_window_seconds=300,
        )
        return []

    candidates: list[str] = []
    if isinstance(parsed, list):
        candidates = [str(v) for v in parsed]
    elif isinstance(parsed, dict):
        candidates = [str(k) for k in parsed.keys()]

    cleaned: list[str] = []
    seen: set[str] = set()
    for cand in candidates:
        sig = _normalize_duplicate_group_signature(cand)
        if not sig or sig in seen:
            continue
        seen.add(sig)
        cleaned.append(sig)
    return cleaned


def _save_rejected_duplicate_group_signatures(signatures: list[str]) -> None:
    cleaned: list[str] = []
    seen: set[str] = set()
    for raw in signatures or []:
        sig = _normalize_duplicate_group_signature(raw)
        if not sig or sig in seen:
            continue
        seen.add(sig)
        cleaned.append(sig)
    if len(cleaned) > _DUP_REJECTED_MAX_ENTRIES:
        cleaned = cleaned[-_DUP_REJECTED_MAX_ENTRIES:]
    SystemConfig.set_config(_DUP_REJECTED_CONFIG_KEY, json.dumps(cleaned, ensure_ascii=False))


def _filter_duplicate_groups_by_rejections(groups: list[dict]) -> list[dict]:
    rejected = set(_load_rejected_duplicate_group_signatures())
    if not rejected:
        out: list[dict] = []
        for group in groups or []:
            sig = _duplicate_group_signature(group.get("client_ids") or [])
            if not sig:
                continue
            row = dict(group)
            row["signature"] = sig
            out.append(row)
        return out

    visible: list[dict] = []
    for group in groups or []:
        sig = _duplicate_group_signature(group.get("client_ids") or [])
        if not sig or sig in rejected:
            continue
        row = dict(group)
        row["signature"] = sig
        visible.append(row)
    return visible


# ============================================================================
# Shared Download: CRM <-> Invoice attachment consistency
# ============================================================================
@bp.route("/clients/<int:client_id>/biz-reg/shared-download")
@login_required
def biz_reg_download_shared(client_id: int):
    """
    shared download
    - Expenses client_attachments(Invoice Expenses)Expenses biz_reg Service Status
    - Payment Form legacy extra+crm_client_{id}New Admin Operations MenuGeneral Ledger New Expenses
    """
    debug = request.args.get("debug") == "1" and is_admin(current_user)
    debug_info = {
        "client_id": client_id,
        "attachment_client_id": None,
        "stored_name": None,
        "original_name": None,
        "bases": [],
        "candidates": [],
        "selected_path": None,
        "exists": None,
        "open_ok": None,
        "open_error": None,
    }

    def _list_files(dir_path: str):
        try:
            files = [f for f in os.listdir(dir_path) if os.path.isfile(os.path.join(dir_path, f))]
            return [
                f
                for f in files
                if (
                    f.rsplit(".", 1)[-1].lower() in ALLOWED_BIZREG_EXTS
                    and _looks_like_bizreg_filename(f)
                )
            ]
        except Exception:
            return []

    def _candidate_bases() -> list[str]:
        bases: list[str] = []
        configured = current_app.config.get("CLIENT_ATTACHMENTS_DIR", "uploads/clients")
        if configured:
            bases.append(os.path.abspath(configured))
        # Some older uploads (or FileAssetService local backend) may land under UPLOAD_FOLDER.
        upload_root = current_app.config.get("UPLOAD_FOLDER")
        if upload_root:
            upload_root_abs = os.path.abspath(upload_root)
            if upload_root_abs not in bases:
                bases.append(upload_root_abs)
            legacy_under_upload = os.path.abspath(
                os.path.join(upload_root_abs, "uploads", "clients")
            )
            if legacy_under_upload not in bases:
                bases.append(legacy_under_upload)
        # Legacy default relative to project root.
        try:
            project_root = Path(current_app.root_path).resolve().parent
            legacy = os.path.abspath(os.path.join(project_root, "uploads", "clients"))
            if legacy not in bases:
                bases.append(legacy)
        except Exception as exc:
            report_swallowed_exception(
                exc,
                context="crm.bizreg_attachment_download.candidate_bases.project_root",
                log_key="crm.bizreg_attachment_download.candidate_bases.project_root",
                log_window_seconds=300,
            )
        # Final fallback: relative path.
        rel = os.path.abspath("uploads/clients")
        if rel not in bases:
            bases.append(rel)
        return bases

    def _candidate_dirs(bases: list[str], att_cid: int, crm_cid: int) -> list[str]:
        dirs: list[str] = []
        for base in bases:
            dirs.extend(
                [
                    os.path.join(base, f"client_{att_cid}"),
                    os.path.join(base, f"client_{crm_cid}"),
                    os.path.join(base, f"crm_client_{crm_cid}"),
                ]
            )
        # Preserve order but remove duplicates.
        seen: set[str] = set()
        uniq: list[str] = []
        for d in dirs:
            if d in seen:
                continue
            seen.add(d)
            uniq.append(d)
        return uniq

    def _debug_response(ok: bool, status: int = 200):
        return jsonify({"ok": ok, "debug": debug_info}), status

    client = Client.query.get_or_404(client_id)
    info = get_bizreg_attachment_for_crm_client(client)

    if not info:
        bases = _candidate_bases()
        debug_info["bases"] = bases
        try:
            att_cid = resolve_attachment_client_id_for_crm_client(client)
        except Exception:
            att_cid = client_id
        candidates = _candidate_dirs(bases, att_cid, client_id)
        debug_info["attachment_client_id"] = att_cid
        debug_info["candidates"] = candidates

        latest_path = None
        latest_name = None
        latest_mtime = None
        for d in candidates:
            files = _list_files(d)
            for name in files:
                try:
                    path = os.path.join(d, name)
                    mtime = os.path.getmtime(path)
                except Exception:
                    continue
                if latest_mtime is None or mtime > latest_mtime:
                    latest_mtime = mtime
                    latest_path = path
                    latest_name = name

        if latest_path and latest_name:
            file_path = os.path.abspath(latest_path)
            debug_info["selected_path"] = file_path
            debug_info["stored_name"] = latest_name
            debug_info["original_name"] = latest_name
            debug_info["exists"] = os.path.exists(file_path)
            if debug:
                try:
                    with open(file_path, "rb"):
                        pass
                    debug_info["open_ok"] = True
                except Exception as exc:
                    debug_info["open_ok"] = False
                    debug_info["open_error"] = str(exc)
                return _debug_response(True)
            return send_file(
                file_path,
                as_attachment=True,
                download_name=latest_name,
                mimetype="application/octet-stream",
            )

        if debug:
            return _debug_response(False, 404)
        abort(404)

    if info.get("source") == "legacy_extra":
        try:
            migrated = migrate_legacy_bizreg_for_crm_client(
                client, uploaded_by=getattr(current_user, "id", None)
            )
            if migrated:
                info = migrated
            db.session.commit()
        except Exception as exc:
            report_swallowed_exception(
                exc,
                context="crm.routes.biz_reg_download_shared.migrate_legacy",
                log_key="crm.routes.biz_reg_download_shared.migrate_legacy",
                log_window_seconds=300,
            )
            try:
                db.session.rollback()
            except Exception as rollback_exc:
                report_swallowed_exception(
                    rollback_exc,
                    context="crm.routes.biz_reg_download_shared.rollback",
                    log_key="crm.routes.biz_reg_download_shared.rollback",
                    log_window_seconds=300,
                )

    bases = _candidate_bases()
    debug_info["bases"] = bases
    raw_stored = (info.get("stored_name") or "").strip()
    stored = os.path.basename(raw_stored)
    debug_info["stored_name"] = stored or None
    debug_info["original_name"] = info.get("original_name") or stored or None
    if not stored or stored != raw_stored:
        if debug:
            return _debug_response(False, 404)
        abort(404)

    try:
        att_cid = int(info.get("attachment_client_id") or client_id)
    except Exception:
        att_cid = client_id
    debug_info["attachment_client_id"] = att_cid

    candidates = _candidate_dirs(bases, att_cid, client_id)
    debug_info["candidates"] = candidates

    file_path = resolve_client_attachment_file_path(
        att_cid,
        stored,
        crm_client_id=client_id,
        include_legacy_crm=True,
        repair=True,
    )
    if file_path:
        debug_info["selected_path"] = file_path
        debug_info["exists"] = True
        if debug:
            try:
                with open(file_path, "rb"):
                    pass
                debug_info["open_ok"] = True
            except Exception as exc:
                debug_info["open_ok"] = False
                debug_info["open_error"] = str(exc)
            return _debug_response(True)
        return send_file(
            file_path,
            as_attachment=True,
            download_name=info.get("original_name") or stored,
            mimetype=info.get("content_type") or "application/octet-stream",
        )

    stream, key = open_client_attachment_stream(
        att_cid,
        stored,
        crm_client_id=client_id,
        include_legacy_crm=True,
    )
    if stream is not None:
        debug_info["selected_path"] = key
        debug_info["exists"] = True
        if debug:
            debug_info["open_ok"] = True
            return _debug_response(True)
        try:
            return send_file(
                stream,
                as_attachment=True,
                download_name=info.get("original_name") or stored,
                mimetype=info.get("content_type") or "application/octet-stream",
            )
        except Exception:
            try:
                data = stream.read()
                mime = (
                    info.get("content_type")
                    or mimetypes.guess_type(info.get("original_name") or stored)[0]
                    or "application/octet-stream"
                )
                return send_file(
                    BytesIO(data),
                    mimetype=mime,
                    as_attachment=True,
                    download_name=info.get("original_name") or stored,
                )
            except Exception:
                if debug:
                    debug_info["open_ok"] = False
                abort(404)

    # Fallback: if DB entry exists but the stored file is missing, pick the latest file.
    latest_path = None
    latest_name = None
    latest_mtime = None
    for d in candidates:
        files = _list_files(d)
        for name in files:
            try:
                path = os.path.join(d, name)
                mtime = os.path.getmtime(path)
            except Exception:
                continue
            if latest_mtime is None or mtime > latest_mtime:
                latest_mtime = mtime
                latest_path = path
                latest_name = name

    if latest_path and latest_name:
        repaired = resolve_client_attachment_file_path(
            att_cid,
            latest_name,
            crm_client_id=client_id,
            include_legacy_crm=True,
            repair=True,
        )
        file_path = os.path.abspath(repaired or latest_path)
        debug_info["selected_path"] = file_path
        debug_info["stored_name"] = latest_name
        debug_info["original_name"] = latest_name
        debug_info["exists"] = os.path.exists(file_path)
        if debug:
            try:
                with open(file_path, "rb"):
                    pass
                debug_info["open_ok"] = True
            except Exception as exc:
                debug_info["open_ok"] = False
                debug_info["open_error"] = str(exc)
            return _debug_response(True)
        return send_file(
            file_path,
            as_attachment=True,
            download_name=latest_name,
            mimetype=info.get("content_type") or "application/octet-stream",
        )

    debug_info["exists"] = False
    if debug:
        return _debug_response(False, 404)
    abort(404)


# ============================================================================
# Clients (existing functionality - preserved)
# ============================================================================


def _crm_clients_xlsx_response(
    *,
    clients: list[Client],
    q: str,
    sort: str,
    direction: str,
    scope: str,
    matched_total: int,
) -> object:
    wb = Workbook()
    ws = wb.active
    ws.title = "CRM Clients"

    headers = [
        "ID",
        "Name",
        "Type",
        "Registration No",
        "Email",
        "Phone",
        "Biz Reg No",
        "Invoice Client ID",
        "Address",
        "Manager",
        "Notes",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for c in clients or []:
        ws.append(
            [
                getattr(c, "id", None),
                str((getattr(c, "name", None) or "")).strip(),
                str((getattr(c, "type", None) or "")).strip(),
                str((getattr(c, "registration_number", None) or "")).strip(),
                str((getattr(c, "email", None) or "")).strip(),
                str((getattr(c, "phone", None) or "")).strip(),
                str((getattr(c, "biz_reg_number", None) or "")).strip(),
                getattr(c, "external_invoice_client_id", None),
                str((getattr(c, "address", None) or "")).strip(),
                str((getattr(c, "manager", None) or "")).strip(),
                str((getattr(c, "notes", None) or "")).strip(),
            ]
        )

    ws.freeze_panes = "A2"

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            value_len = len(str(getattr(cell, "value", "") or ""))
            if value_len > max_length:
                max_length = value_len
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    meta = wb.create_sheet("Meta")
    meta.append(["Scope", "all" if scope == "all" else "page"])
    meta.append(["Matched Total", int(matched_total or 0)])
    meta.append(["Exported Rows", len(clients or [])])
    meta.append(["Search", q or ""])
    meta.append(["Sort", sort or "id"])
    meta.append(["Direction", direction or "desc"])
    meta.append(["Export Date", datetime.utcnow().isoformat(timespec="seconds") + "Z"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"crm_clients_{scope}_{ts}.xlsx"
    resp = current_app.response_class(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@bp.route("/")
@login_required
def clients():
    """List all clients with search and pagination."""
    q = request.args.get("q", "").strip()
    page = request.args.get("page", 1, type=int)
    page = max(1, int(page or 1))
    per_page = request.args.get("per_page", 20, type=int)
    per_page = max(1, min(int(per_page or 20), 200))
    sort = request.args.get("sort", "id").strip()
    direction = request.args.get("direction", "").strip().lower()
    invoice_link = (request.args.get("invoice_link") or "").strip().lower()
    export_format = (request.args.get("format") or "").strip().lower()
    export_scope_raw = request.args.get("export_scope")
    export_scope = (export_scope_raw or "page").strip().lower()
    if export_scope not in ("page", "all"):
        export_scope = "page"
    raw_export_flag = (request.args.get("export") or "").strip().lower()
    export_requested = export_format in ("xlsx", "csv") or raw_export_flag in (
        "1",
        "true",
        "yes",
        "on",
    )

    query = Client.query.filter_by(is_deleted=False)

    if invoice_link == "missing":
        query = query.filter(Client.external_invoice_client_id.is_(None))
    elif invoice_link == "linked":
        query = query.filter(Client.external_invoice_client_id.isnot(None))
    else:
        invoice_link = ""

    if q:
        query = query.filter(
            db.or_(
                sqlalchemy_contains_query(Client.name, q),
                sqlalchemy_contains_query(Client.email, q),
                sqlalchemy_contains_query(Client.registration_number, q),
                sqlalchemy_contains_query(Client.phone, q),
                sqlalchemy_contains_query(Client.biz_reg_number, q),
                sqlalchemy_contains_query(Client.biz_company_name, q),
                sqlalchemy_contains_query(Client.biz_tax_invoice_email, q),
                sqlalchemy_contains_query(Client.search_tags, q),
            )
        )

    sort_map = {
        "name": Client.name,
        "id": Client.id,
        "registration_number": Client.registration_number,
        "email": Client.email,
        "phone": Client.phone,
    }
    if sort not in sort_map:
        sort = "id"
    if direction not in ("asc", "desc"):
        direction = "desc" if sort == "id" else "asc"

    sort_column = sort_map[sort]
    order_expr = sort_column.desc() if direction == "desc" else sort_column.asc()
    if sort == "id":
        query = query.order_by(order_expr)
    else:
        tie_breaker = Client.id.desc() if direction == "desc" else Client.id.asc()
        query = query.order_by(order_expr, tie_breaker)

    if export_requested:
        max_rows = current_app.config.get("CRM_CLIENT_XLSX_MAX_ROWS", 10000)
        try:
            max_rows_int = int(max_rows or 10000)
        except Exception:
            max_rows_int = 10000
        max_rows_int = max(100, min(max_rows_int, 50000))

        matched_total = query.order_by(None).count()
        page_count = max(1, (matched_total + per_page - 1) // per_page)
        if page > page_count:
            page = page_count
        if export_scope == "all":
            export_clients = query.limit(max_rows_int).all()
        else:
            export_clients = query.paginate(page=page, per_page=per_page, error_out=False).items

        return _crm_clients_xlsx_response(
            clients=export_clients,
            q=q,
            sort=sort,
            direction=direction,
            scope=export_scope,
            matched_total=matched_total,
        )

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    page_count = max(1, pagination.pages)
    if page > page_count:
        page = page_count
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        page_count = max(1, pagination.pages)

    bizreg_map = {}
    for c in pagination.items:
        try:
            info = get_bizreg_attachment_for_crm_client(c, verify_exists=True)
        except Exception:
            info = None
        if info:
            bizreg_map[int(c.id)] = info

    duplicate_groups = []
    try:
        rows = (
            Client.query.filter_by(is_deleted=False)
            .with_entities(
                Client.id,
                Client.name,
                Client.email,
                Client.phone,
                Client.registration_number,
                Client.biz_reg_number,
                Client.biz_tax_invoice_email,
            )
            .all()
        )
        duplicate_groups = _build_duplicate_groups(rows)
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.clients.duplicate_groups",
            log_key="crm.clients.duplicate_groups",
            log_window_seconds=300,
        )
    duplicate_groups = _filter_duplicate_groups_by_rejections(duplicate_groups)
    can_merge_admin = bool(is_admin(current_user))

    recent_merge_logs = []
    ttl_seconds = current_app.config.get("CRM_RECENT_MERGE_LOG_TTL_SECONDS", 0)
    try:
        ttl_seconds = int(ttl_seconds or 0)
    except (TypeError, ValueError):
        ttl_seconds = 0
    cutoff = datetime.utcnow() - timedelta(seconds=ttl_seconds) if ttl_seconds > 0 else None
    try:
        log_query = CRMClientMergeLog.query.filter(CRMClientMergeLog.undone_at.is_(None))
        if cutoff is not None:
            log_query = log_query.filter(CRMClientMergeLog.created_at >= cutoff)
        logs = log_query.order_by(CRMClientMergeLog.id.desc()).limit(5).all()
        for log in logs:
            src_ids = []
            try:
                src_ids = json.loads(log.source_client_ids_json) or []
            except Exception:
                src_ids = []
            recent_merge_logs.append(
                {
                    "id": log.id,
                    "target_id": log.target_client_id,
                    "source_ids": [int(i) for i in src_ids if str(i).isdigit()],
                    "source_count": len(src_ids),
                    "source_ids_display": ", ".join(str(i) for i in src_ids if str(i).isdigit()),
                    "created_at": log.created_at,
                }
            )
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.clients.recent_merge_logs",
            log_key="crm.clients.recent_merge_logs",
            log_window_seconds=300,
        )

    return render_template(
        "crm/list.html",
        clients=pagination.items,
        bizreg_map=bizreg_map,
        duplicate_groups=duplicate_groups[:10],
        can_merge_admin=can_merge_admin,
        recent_merge_logs=recent_merge_logs,
        page=page,
        pages=page_count,
        total=pagination.total,
        per_page=per_page,
        q=q,
        sort=sort,
        direction=direction,
        invoice_link=invoice_link,
    )


@bp.post("/clients/case-status-export")
@login_required
@role_required("admin")
def client_case_status_export():
    """Download one case-status workbook per selected CRM client as a zip."""
    export_scope = (request.form.get("export_scope") or "selected").strip().lower()
    if export_scope not in ("selected", "all"):
        export_scope = "selected"

    max_clients_raw = current_app.config.get("CRM_CLIENT_CASE_STATUS_XLSX_MAX_CLIENTS", 2000)
    try:
        max_clients = int(max_clients_raw or 2000)
    except Exception:
        max_clients = 2000
    max_clients = max(1, min(max_clients, 10000))

    max_cases_raw = current_app.config.get(
        "CRM_CLIENT_CASE_STATUS_XLSX_MAX_CASES_PER_CLIENT", 20000
    )
    try:
        max_cases_per_client = int(max_cases_raw or 20000)
    except Exception:
        max_cases_per_client = 20000
    max_cases_per_client = max(1, min(max_cases_per_client, 100000))

    matched_total = 0
    client_limit_truncated = False
    if export_scope == "all":
        q = (request.form.get("q") or "").strip()
        sort = (request.form.get("sort") or "id").strip()
        direction = (request.form.get("direction") or "").strip().lower()
        invoice_link = (request.form.get("invoice_link") or "").strip().lower()
        query, _sort, _direction, _invoice_link = _crm_client_list_query(
            q=q,
            invoice_link=invoice_link,
            sort=sort,
            direction=direction,
        )
        matched_total = int(query.order_by(None).count() or 0)
        clients_to_export = query.limit(max_clients).all()
        client_limit_truncated = matched_total > len(clients_to_export)
    else:
        raw_ids: list[str] = []
        raw_ids.extend(request.form.getlist("client_ids"))
        raw_ids.extend(request.form.getlist("selected_client_ids"))
        raw_ids.extend((request.form.get("client_ids_csv") or "").split(","))
        selected_ids: list[int] = []
        for raw in raw_ids:
            cid = _safe_positive_int(raw)
            if cid and cid not in selected_ids:
                selected_ids.append(cid)
        if not selected_ids:
            flash("Matter Download Client Select.", "warning")
            return redirect(safe_referrer_path() or url_for("customers.clients"))
        requested_count = len(selected_ids)
        selected_ids = selected_ids[:max_clients]
        matched_total = requested_count
        rows = Client.query.filter_by(is_deleted=False).filter(Client.id.in_(selected_ids)).all()
        by_id = {int(c.id): c for c in rows}
        clients_to_export = [by_id[cid] for cid in selected_ids if cid in by_id]
        client_limit_truncated = requested_count > len(selected_ids)

    if not clients_to_export:
        flash("Download Client not found.", "warning")
        return redirect(safe_referrer_path() or url_for("customers.clients"))

    matter_map = _fetch_client_matter_map(
        clients_to_export,
        limit_per_client=max_cases_per_client,
    )
    clients_with_matters = [
        client
        for client in clients_to_export
        if int((matter_map.get(int(client.id), ([], 0))[1]) or 0) > 0
    ]
    skipped_no_case_clients = len(clients_to_export) - len(clients_with_matters)
    if not clients_with_matters:
        flash("Matter Link Client  Matter Create .", "warning")
        return redirect(safe_referrer_path() or url_for("customers.clients"))

    output = BytesIO()
    used_names: set[str] = set()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for idx, client in enumerate(clients_with_matters, start=1):
            matters, total = matter_map.get(int(client.id), ([], 0))
            case_truncated = int(total or 0) > len(matters or [])
            client_name = _safe_export_filename_part(
                getattr(client, "name", None),
                fallback=f"client_{client.id}",
            )
            arcname = f"{idx:03d}_{client.id}_{client_name}_case_status.xlsx"
            while arcname in used_names:
                arcname = f"{idx:03d}_{client.id}_{client_name}_{uuid.uuid4().hex[:6]}.xlsx"
            used_names.add(arcname)

            zf.writestr(
                arcname,
                _client_case_status_xlsx_bytes(
                    client=client,
                    matters=matters,
                    total_matters=total,
                    truncated=case_truncated,
                ),
            )

    output.seek(0)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"crm_client_case_status_{export_scope}_{ts}.zip"
    resp = current_app.response_class(output.getvalue(), mimetype="application/zip")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.headers["X-CRM-Exported-Clients"] = str(len(clients_with_matters))
    resp.headers["X-CRM-Matched-Clients"] = str(matched_total)
    resp.headers["X-CRM-Skipped-No-Case-Clients"] = str(skipped_no_case_clients)
    resp.headers["X-CRM-Client-Limit-Truncated"] = "1" if client_limit_truncated else "0"
    return resp


@bp.route("/clients/create", methods=["GET", "POST"])
@login_required
def client_create():
    """Create a new client."""
    form = ClientForm()

    if request.method == "GET" and not form.idempotency_key.data:
        form.idempotency_key.data = uuid.uuid4().hex

    def _return_created(client: Client, *, duplicate: bool = False):
        if duplicate:
            flash(" Process .", "info")
        if request.args.get("popup") == "1":
            payload = {"client_id": client.id, "client_name": client.name}
            return render_template("crm/client_create_done.html", client_payload=payload)
        return redirect(url_for("customers.client_view", client_id=client.id))

    if form.validate_on_submit():
        # Build extra JSON from form fields
        extra = {
            "name_en": form.name_en.data,
            "nationality": form.nationality.data,
            "business_type": form.business_type.data,
            "client_code": form.client_code.data,
            "applicant_codes": [
                form.applicant_code1.data or "",
                form.applicant_code2.data or "",
                form.applicant_code3.data or "",
            ],
            "annuity_management_disabled": bool(form.annuity_management_disabled.data),
            "input_date": form.input_date.data.isoformat() if form.input_date.data else None,
            "category": form.category.data,
            "viewer_default": form.viewer_default.data,
            "main_phone": form.main_phone.data,
            "mobile_phone": form.mobile_phone.data,
            "main_fax": form.main_fax.data,
            "homepage": form.homepage.data,
            "other_contact": form.other_contact.data,
            "applicant_address": form.applicant_address.data,
            "applicant_email": form.applicant_email.data,
            "applicant_phone": form.applicant_phone.data,
            "applicant_fax": form.applicant_fax.data,
            "business_reg_no": form.business_reg_no.data,
            "tax_company_name": form.tax_company_name.data,
            "tax_ceo": form.tax_ceo.data,
            "tax_business_type": form.tax_business_type.data,
            "tax_business_item": form.tax_business_item.data,
            "tax_address": form.tax_address.data,
            "tax_manager": form.tax_manager.data,
            "tax_manager_email": form.tax_manager_email.data,
            "tax_manager_mobile": form.tax_manager_mobile.data,
            "mail_recv_address": form.mail_recv_address.data,
            "mail_receiver": form.mail_receiver.data,
            "personal_email": form.personal_email.data,
            "personal_phone": form.personal_phone.data,
            "personal_fax": form.personal_fax.data,
            "other_address": form.other_address.data,
            "other_email": form.other_email.data,
            "other_phone": form.other_phone.data,
            "other_fax": form.other_fax.data,
            "note": form.note.data,
            "special_note": form.special_note.data,
        }

        # Shared fields (Invoice module uses these columns directly)
        shared_address = (
            (form.address.data or "").strip()
            or (form.tax_address.data or "").strip()
            or (form.applicant_address.data or "").strip()
        )
        shared_manager = (
            (form.manager.data or "").strip()
            or (form.tax_manager.data or "").strip()
            or (form.tax_ceo.data or "").strip()
        )
        shared_notes = (form.notes.data or "").strip() or (form.note.data or "").strip()

        idem_key = (form.idempotency_key.data or "").strip()
        actor_id = getattr(current_user, "id", None)
        from app.services.ops.operation_log import namespace_idempotency_key

        op_request_id = namespace_idempotency_key(idem_key or None, actor_id)
        legacy_request_id = idem_key or None
        if op_request_id:
            existing = (
                Operation.query.filter(Operation.request_id == op_request_id)
                .filter(Operation.action == "client.create")
                .first()
            )
            if (
                not existing
                and legacy_request_id
                and actor_id is not None
                and legacy_request_id != op_request_id
            ):
                # Backward-compat: previously stored raw idempotency keys in Operation.request_id.
                # Only return a legacy op if it belongs to the same actor to avoid cross-user leaks.
                existing = (
                    Operation.query.filter(Operation.request_id == legacy_request_id)
                    .filter(Operation.action == "client.create")
                    .filter(Operation.actor_id == actor_id)
                    .first()
                )
            if existing and isinstance(existing.summary_json, dict):
                existing_id = existing.summary_json.get("client_id")
                if existing_id:
                    existing_client = Client.query.get(int(existing_id))
                    if existing_client:
                        return _return_created(existing_client, duplicate=True)

        try:
            client = Client(
                name=form.client_name.data,
                type=form.category.data,
                registration_number=form.registration_number.data,
                email=form.email.data,
                phone=form.main_phone.data,
                address=shared_address or None,
                manager=shared_manager or None,
                notes=shared_notes or None,
                extra=extra,
            )
            set_crm_client_search_tags_fast(client)

            db.session.add(client)
            db.session.flush()  # client.id 

            if op_request_id:
                op = Operation(
                    request_id=op_request_id,
                    actor_id=actor_id,
                    action="client.create",
                    risk_level="LOW",
                    status="applied",
                    summary_json={"client_id": client.id},
                    created_at=datetime.utcnow(),
                    applied_at=datetime.utcnow(),
                )
                db.session.add(op)

            # ✅ create from (  ) Business document File  Save + DB(extra) Link
            biz_reg_upload = request.files.get("biz_reg_file")
            if biz_reg_upload and getattr(biz_reg_upload, "filename", ""):
                allowed = {"pdf", "png", "jpg", "jpeg", "gif"}
                ext = (
                    biz_reg_upload.filename.rsplit(".", 1)[1].lower()
                    if "." in biz_reg_upload.filename
                    else ""
                )
                if ext not in allowed:
                    db.session.rollback()
                    flash("  Business document File .", "danger")
                    return render_template("crm/form.html", form=form)

                saved = save_bizreg_attachment_for_crm_client(
                    client,
                    biz_reg_upload,
                    uploaded_by=getattr(current_user, "id", None),
                )
                if saved:
                    _attach_biz_reg_file(
                        client,
                        {
                            "original_name": saved.get("original_name") or biz_reg_upload.filename,
                            "stored_name": saved.get("stored_name"),
                            "uploaded_at": datetime.utcnow().isoformat(),
                            "attachment_client_id": saved.get("attachment_client_id"),
                            "attachment_id": saved.get("id"),
                        },
                    )

            _record_client_audit(
                client=client,
                action="client.create",
                source="crm.client_create",
                include_snapshots=True,
            )
            db.session.commit()
            try:
                enqueue_crm_client_post_save(client.id)
            except Exception as exc:
                report_swallowed_exception(
                    exc,
                    context="crm.client_create.enqueue_post_save",
                    log_key="crm.client_create.enqueue_post_save",
                    log_window_seconds=300,
                )
        except UploadTooLargeError as exc:
            db.session.rollback()
            flash(str(exc) or "File  .", "danger")
            return render_template("crm/form.html", form=form)
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception("client_create failed while saving biz_reg_file")
            flash(f"Client Registration Failed: {e}", "danger")
            return render_template("crm/form.html", form=form)

        flash("Client Registration. Search   from Process.", "success")
        return _return_created(client)

    return render_template("crm/form.html", form=form)


@bp.route("/clients/<int:client_id>")
@login_required
def client_view(client_id):
    """View client details with contacts, opportunities, and activities."""
    client = Client.query.get_or_404(client_id)

    # Get related data
    contacts = client.contacts.order_by(CRMContact.is_primary.desc(), CRMContact.name).all()
    opportunities = client.opportunities.order_by(CRMOpportunity.created_at.desc()).all()
    activities = client.activities.order_by(CRMActivity.activity_date.desc()).limit(20).all()
    client_matters, client_matters_total = _fetch_client_matters_for_client(client, limit=50)
    client_invoices = _fetch_client_invoice_snapshot(client, limit=20)
    applicant_code_suggestion = None
    applicant_code_debug = {}
    try:
        matter_ids = [
            str(getattr(m, "matter_id", "") or "").strip()
            for m in (client_matters or [])
            if str(getattr(m, "matter_id", "") or "").strip()
        ]
        applicant_code_suggestion = _build_applicant_code_suggestion(
            client, matter_ids, debug=applicant_code_debug
        )
        debug_requested = _truthy_debug_value(request.args.get("debug_applicant_code"))
        secured_debug = _secured_applicant_code_debug_enabled()
        if debug_requested or not applicant_code_suggestion:
            debug_payload = (
                {**applicant_code_debug, "secured_debug": True}
                if secured_debug
                else _sanitize_applicant_code_debug_payload(applicant_code_debug)
            )
            current_app.logger.info("crm.applicant_code_suggestion debug=%s", debug_payload)
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_view.applicant_code_suggestion",
            log_key="crm.client_view.applicant_code_suggestion",
            log_window_seconds=300,
        )

    # Forms for modals
    contact_form = ContactForm()
    contact_form.client_id.data = client_id
    activity_form = ActivityForm()
    activity_form.client_id.data = client_id

    biz_reg_attachment = get_bizreg_attachment_for_crm_client(client, verify_exists=True)
    tax_suggestion = None
    biz_reg_suggestion = None
    attachment_client_id = None
    client_attachments = []
    try:
        extra = client.extra or {}
        if not isinstance(extra, dict):
            extra = {}
        raw_suggestion = extra.get("tax_invoice_suggestion")
        if isinstance(raw_suggestion, str):
            raw_suggestion = json.loads(raw_suggestion)
        if isinstance(raw_suggestion, dict):
            tax_suggestion = raw_suggestion
        raw_biz_suggestion = extra.get("biz_reg_suggestion")
        if isinstance(raw_biz_suggestion, str):
            raw_biz_suggestion = json.loads(raw_biz_suggestion)
        if isinstance(raw_biz_suggestion, dict):
            biz_reg_suggestion = raw_biz_suggestion
    except Exception:
        tax_suggestion = None
        biz_reg_suggestion = None
    try:
        attachment_client_id, client_attachments = list_client_attachments_for_crm_client(client)
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_view.attachments",
            log_key="crm.client_view.attachments",
            log_window_seconds=300,
        )
    return render_template(
        "crm/view.html",
        client=client,
        contacts=contacts,
        opportunities=opportunities,
        activities=activities,
        client_matters=client_matters,
        client_matters_total=client_matters_total,
        client_invoices=client_invoices,
        invoice_client_id=client_invoices.get("client_id"),
        contact_form=contact_form,
        activity_form=activity_form,
        biz_reg_attachment=biz_reg_attachment,
        tax_suggestion=tax_suggestion,
        biz_reg_suggestion=biz_reg_suggestion,
        applicant_code_suggestion=applicant_code_suggestion,
        attachment_client_id=attachment_client_id,
        client_attachments=client_attachments,
    )


@bp.post("/clients/<int:client_id>/attachments/upload")
@login_required
def client_attachment_upload(client_id: int):
    client = Client.query.get_or_404(client_id)

    files = request.files.getlist("files[]") or request.files.getlist("files")
    if not files and request.files.get("file"):
        files = [request.files.get("file")]
    if not files:
        return jsonify({"success": False, "error": "No upload file provided."}), 400

    saved = []
    skipped_invalid = 0
    last_invalid_msg = None
    try:
        for f in files:
            if not f or not getattr(f, "filename", ""):
                continue
            try:
                info = save_client_attachment_for_crm_client(
                    client, f, uploaded_by=getattr(current_user, "id", None)
                )
            except UploadSecurityError:
                raise
            except ValueError as exc:
                msg = str(exc)
                if " " in msg:
                    skipped_invalid += 1
                    last_invalid_msg = msg
                    continue
                raise
            if info:
                saved.append(
                    {
                        "id": info.get("id"),
                        "name": info.get("original_name"),
                        "size": info.get("size"),
                    }
                )

        if not saved:
            db.session.rollback()
            msg = last_invalid_msg or "No upload file provided."
            return jsonify({"success": False, "error": msg}), 400

        db.session.commit()
        record_entity_change_audit(
            action="client.attachment.upload",
            target_type="crm_client",
            target_id=client.id,
            actor_id=getattr(current_user, "id", None),
            meta={
                "client_id": client.id,
                "client_name": client.name,
                "attachments": saved,
                "source": "crm.client_attachment_upload",
            },
            title=client.name,
        )
        db.session.commit()
        return jsonify({"success": True, "attachments": saved})
    except UploadTooLargeError as exc:
        db.session.rollback()
        return jsonify({"success": False, "error": str(exc)}), 413
    except UploadSecurityError:
        db.session.rollback()
        return jsonify({"success": False, "error": "Upload Security  Failed."}), 400
    except Exception as exc:
        db.session.rollback()
        report_swallowed_exception(
            exc,
            context="crm.client_attachment_upload",
            log_key="crm.client_attachment_upload",
            log_window_seconds=300,
        )
        return jsonify({"success": False, "error": "Upload In Progress Error ."}), 500


@bp.get("/clients/<int:client_id>/attachments/<int:att_id>/download")
@login_required
def client_attachment_download(client_id: int, att_id: int):
    client = Client.query.get_or_404(client_id)
    info = get_client_attachment_for_crm_client(client, att_id)
    if not info:
        abort(404)

    stored = os.path.basename(str(info.get("stored_name") or ""))
    if not stored or stored != (info.get("stored_name") or ""):
        abort(404)

    att_cid = int(info.get("attachment_client_id") or client.id)
    path = resolve_client_attachment_file_path(
        att_cid,
        stored,
        crm_client_id=client.id,
        include_legacy_crm=False,
        repair=True,
    )
    if not path:
        stream, _ = open_client_attachment_stream(
            att_cid,
            stored,
            crm_client_id=client.id,
            include_legacy_crm=False,
        )
        if stream is None:
            abort(404)
        try:
            return send_file(
                stream,
                as_attachment=True,
                download_name=info.get("original_name") or stored,
                mimetype=info.get("content_type") or "application/octet-stream",
            )
        except Exception:
            try:
                data = stream.read()
                mime = (
                    info.get("content_type")
                    or mimetypes.guess_type(info.get("original_name") or stored)[0]
                    or "application/octet-stream"
                )
                return send_file(
                    BytesIO(data),
                    mimetype=mime,
                    as_attachment=True,
                    download_name=info.get("original_name") or stored,
                )
            except Exception:
                abort(404)

    return send_file(
        path,
        as_attachment=True,
        download_name=info.get("original_name") or stored,
        mimetype=info.get("content_type") or "application/octet-stream",
    )


@bp.post("/clients/<int:client_id>/attachments/<int:att_id>/delete")
@login_required
def client_attachment_delete(client_id: int, att_id: int):
    client = Client.query.get_or_404(client_id)
    try:
        ok = delete_client_attachment_for_crm_client(client, att_id)
        if not ok:
            abort(404)
        record_entity_change_audit(
            action="client.attachment.delete",
            target_type="crm_client",
            target_id=client.id,
            actor_id=getattr(current_user, "id", None),
            meta={
                "client_id": client.id,
                "client_name": client.name,
                "attachment_id": att_id,
                "source": "crm.client_attachment_delete",
            },
            title=client.name,
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        report_swallowed_exception(
            exc,
            context="crm.client_attachment_delete",
            log_key="crm.client_attachment_delete",
            log_window_seconds=300,
        )
        abort(500)

    if request.is_json:
        return jsonify({"success": True})
    return redirect(url_for("customers.client_view", client_id=client_id))


@bp.post("/clients/<int:client_id>/tax-suggestion/apply")
@login_required
def apply_tax_suggestion(client_id: int):
    client = Client.query.get_or_404(client_id)
    audit_before = _client_audit_snapshot(client)
    extra = dict(client.extra or {})
    suggestion = extra.get("tax_invoice_suggestion") or {}
    if isinstance(suggestion, str):
        try:
            suggestion = json.loads(suggestion)
        except Exception:
            suggestion = {}
    if not isinstance(suggestion, dict) or not suggestion.get("fields"):
        flash("No tax-document suggestion is available to apply.", "warning")
        return redirect(url_for("customers.client_view", client_id=client.id))

    fields = suggestion.get("fields") or {}

    def _val(key: str) -> str:
        value = fields.get(key)
        return (value or "").strip()

    corp_num = _val("corp_num")
    corp_name = _val("corp_name")
    ceo_name = _val("ceo_name")
    address = _val("address")
    biz_type = _val("biz_type")
    biz_class = _val("biz_class")
    contact_name = _val("contact_name")
    email = _val("email")
    tel = _val("tel")

    if corp_num:
        client.biz_reg_number = corp_num
    if corp_name:
        client.biz_company_name = corp_name
    if ceo_name:
        client.biz_representative_name = ceo_name
    if address:
        client.biz_business_location = address
    if biz_type:
        client.biz_business_type = biz_type
    if email:
        client.biz_tax_invoice_email = email

    if corp_name and not (client.name or "").strip():
        client.name = corp_name
    if corp_num and not (client.registration_number or "").strip():
        client.registration_number = corp_num
    if email and not (client.email or "").strip():
        client.email = email
    if tel and not (client.phone or "").strip():
        client.phone = tel
    if address and not (client.address or "").strip():
        client.address = address
    if contact_name and not (client.manager or "").strip():
        client.manager = contact_name

    if corp_num:
        extra["business_reg_no"] = corp_num
    if corp_name:
        extra["tax_company_name"] = corp_name
    if ceo_name:
        extra["tax_ceo"] = ceo_name
    if biz_type:
        extra["tax_business_type"] = biz_type
    if biz_class:
        extra["tax_business_item"] = biz_class
    if address:
        extra["tax_address"] = address
    if contact_name:
        extra["tax_manager"] = contact_name
    if email:
        extra["tax_manager_email"] = email
    if tel:
        extra["tax_manager_mobile"] = tel

    extra.pop("tax_invoice_suggestion", None)
    extra["tax_invoice_suggestion_applied_at"] = datetime.utcnow().isoformat()
    client.extra = extra
    flag_modified(client, "extra")
    _set_client_search_tags(client, extra)
    _record_client_audit(
        client=client,
        action="client.suggestion.apply",
        before=audit_before,
        source="crm.apply_tax_suggestion",
    )
    db.session.commit()

    flash("Applied tax-document suggestion to the client profile.", "success")
    return redirect(url_for("customers.client_view", client_id=client.id))


@bp.post("/clients/<int:client_id>/biz-reg-suggestion/apply")
@login_required
def apply_biz_reg_suggestion(client_id: int):
    client = Client.query.get_or_404(client_id)
    audit_before = _client_audit_snapshot(client)
    extra = dict(client.extra or {})
    suggestion = extra.get("biz_reg_suggestion") or {}
    if isinstance(suggestion, str):
        try:
            suggestion = json.loads(suggestion)
        except Exception:
            suggestion = {}
    if isinstance(suggestion, dict):
        fields = suggestion.get("fields") if isinstance(suggestion.get("fields"), dict) else {}
        if not fields and suggestion:
            fields = suggestion
    else:
        fields = {}

    if not fields:
        flash("Apply Business document   none.", "warning")
        return redirect(url_for("customers.client_view", client_id=client.id))

    selected_fields = [v for v in request.form.getlist("apply_fields") if v]
    selected_present = request.form.get("apply_fields_present")
    selected_set = set(selected_fields) if selected_fields else None
    if selected_present is not None and not selected_fields:
        flash("Select value none.", "warning")
        return redirect(url_for("customers.client_view", client_id=client.id))

    def _val(key: str) -> str:
        if selected_set is not None and key not in selected_set:
            return ""
        value = fields.get(key)
        return (value or "").strip()

    def _set_attr_if_blank(attr: str, value: str) -> None:
        if not value:
            return
        cur = getattr(client, attr, None)
        if isinstance(cur, str):
            if cur.strip():
                return
        elif cur not in (None, ""):
            return
        setattr(client, attr, value)

    def _set_extra_if_blank(key: str, value: str) -> None:
        if not value:
            return
        cur = extra.get(key)
        if isinstance(cur, str) and cur.strip():
            return
        if cur not in (None, ""):
            return
        extra[key] = value

    company_name = _val("company_name")
    reg_number = _val("reg_number")
    corp_reg_number = _val("corp_registration_number")
    rep_name = _val("representative_name")
    business_type = _val("business_type")
    business_location = _val("business_location")
    head_office_location = _val("head_office_location")
    opening_date = _val("opening_date")
    tax_email = _val("tax_invoice_email")
    address = business_location or head_office_location

    _set_attr_if_blank("name", company_name)
    _set_attr_if_blank("registration_number", corp_reg_number)
    _set_attr_if_blank("email", tax_email)
    _set_attr_if_blank("address", address)
    _set_attr_if_blank("manager", rep_name)
    _set_attr_if_blank("biz_reg_number", reg_number)
    _set_attr_if_blank("biz_company_name", company_name)
    _set_attr_if_blank("biz_representative_name", rep_name)
    _set_attr_if_blank("biz_business_type", business_type)
    _set_attr_if_blank("biz_business_location", business_location)
    _set_attr_if_blank("biz_head_office_location", head_office_location)
    _set_attr_if_blank("biz_opening_date", opening_date)
    _set_attr_if_blank("biz_corp_registration_number", corp_reg_number)
    _set_attr_if_blank("biz_tax_invoice_email", tax_email)

    _set_extra_if_blank("business_reg_no", reg_number)
    _set_extra_if_blank("tax_company_name", company_name)
    _set_extra_if_blank("tax_ceo", rep_name)
    _set_extra_if_blank("tax_business_type", business_type)
    _set_extra_if_blank("tax_address", address)
    _set_extra_if_blank("tax_manager_email", tax_email)
    _set_extra_if_blank("applicant_address", address)
    _set_extra_if_blank("applicant_email", tax_email)

    extra["biz_reg_parsed"] = fields
    extra.pop("biz_reg_suggestion", None)
    extra["biz_reg_suggestion_applied_at"] = datetime.utcnow().isoformat()
    client.extra = extra
    flag_modified(client, "extra")
    _set_client_search_tags(client, extra)
    _record_client_audit(
        client=client,
        action="client.suggestion.apply",
        before=audit_before,
        source="crm.apply_biz_reg_suggestion",
    )
    db.session.commit()

    flash("Business document   Client information Apply.", "success")
    return redirect(url_for("customers.client_view", client_id=client.id))


@bp.post("/clients/<int:client_id>/applicant-codes/apply")
@login_required
def apply_applicant_code_suggestion(client_id: int):
    client = Client.query.get_or_404(client_id)
    audit_before = _client_audit_snapshot(client)

    existing_codes = _extract_client_applicant_codes(client)
    slots = max(0, _APPLICANT_CODE_LIMIT - len(existing_codes))
    if slots <= 0:
        flash("PatentClient input field  .", "warning")
        return redirect(url_for("customers.client_view", client_id=client.id))

    selected = request.form.getlist("codes")
    if not selected:
        flash("Select value none.", "warning")
        return redirect(url_for("customers.client_view", client_id=client.id))

    selected_clean: list[str] = []
    seen: set[str] = set()
    for code in selected:
        norm = _normalize_applicant_code(code)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        selected_clean.append(norm)

    existing_norm = {_normalize_applicant_code(c) for c in existing_codes}
    filtered = [c for c in selected_clean if c not in existing_norm]
    if not filtered:
        flash("Apply PatentClient   none.", "warning")
        return redirect(url_for("customers.client_view", client_id=client.id))

    if len(filtered) > slots:
        filtered = filtered[:slots]
        flash("Save at least three patent-client fields before applying this item.", "warning")

    new_codes = (existing_codes + filtered)[:_APPLICANT_CODE_LIMIT]
    while len(new_codes) < _APPLICANT_CODE_LIMIT:
        new_codes.append("")

    extra = dict(client.extra or {})
    extra["applicant_codes"] = new_codes
    client.extra = extra
    flag_modified(client, "extra")
    _set_client_search_tags(client, extra)
    _record_client_audit(
        client=client,
        action="client.suggestion.apply",
        before=audit_before,
        source="crm.apply_applicant_code_suggestion",
    )
    db.session.commit()

    flash("PatentClient   Client information Apply.", "success")
    return redirect(url_for("customers.client_view", client_id=client.id))


@bp.route("/clients/<int:client_id>/edit", methods=["GET", "POST"])
@login_required
def client_edit(client_id):
    """Edit an existing client."""
    client = Client.query.get_or_404(client_id)
    form = ClientForm()

    if form.validate_on_submit():
        audit_before = _client_audit_snapshot(client)
        # Update extra JSON
        extra = dict(client.extra or {})
        extra.update(
            {
                "name_en": form.name_en.data,
                "nationality": form.nationality.data,
                "business_type": form.business_type.data,
                "client_code": form.client_code.data,
                "applicant_codes": [
                    form.applicant_code1.data or "",
                    form.applicant_code2.data or "",
                    form.applicant_code3.data or "",
                ],
                "annuity_management_disabled": bool(form.annuity_management_disabled.data),
                "input_date": form.input_date.data.isoformat() if form.input_date.data else None,
                "category": form.category.data,
                "viewer_default": form.viewer_default.data,
                "main_phone": form.main_phone.data,
                "mobile_phone": form.mobile_phone.data,
                "main_fax": form.main_fax.data,
                "homepage": form.homepage.data,
                "other_contact": form.other_contact.data,
                "applicant_address": form.applicant_address.data,
                "applicant_email": form.applicant_email.data,
                "applicant_phone": form.applicant_phone.data,
                "applicant_fax": form.applicant_fax.data,
                "business_reg_no": form.business_reg_no.data,
                "tax_company_name": form.tax_company_name.data,
                "tax_ceo": form.tax_ceo.data,
                "tax_business_type": form.tax_business_type.data,
                "tax_business_item": form.tax_business_item.data,
                "tax_address": form.tax_address.data,
                "tax_manager": form.tax_manager.data,
                "tax_manager_email": form.tax_manager_email.data,
                "tax_manager_mobile": form.tax_manager_mobile.data,
                "mail_recv_address": form.mail_recv_address.data,
                "mail_receiver": form.mail_receiver.data,
                "personal_email": form.personal_email.data,
                "personal_phone": form.personal_phone.data,
                "personal_fax": form.personal_fax.data,
                "other_address": form.other_address.data,
                "other_email": form.other_email.data,
                "other_phone": form.other_phone.data,
                "other_fax": form.other_fax.data,
                "note": form.note.data,
                "special_note": form.special_note.data,
            }
        )

        client.name = form.client_name.data
        client.type = form.category.data
        client.registration_number = form.registration_number.data
        client.email = form.email.data
        client.phone = form.main_phone.data
        # Shared columns used by Accounting/Invoice module
        client.address = form.address.data
        client.manager = form.manager.data
        client.notes = form.notes.data
        client.extra = extra
        flag_modified(client, "extra")
        set_crm_client_search_tags_fast(client)

        # ✅ edit from Business document upload Save(Existing create Save)
        biz_reg_upload = request.files.get("biz_reg_file")
        if biz_reg_upload and getattr(biz_reg_upload, "filename", ""):
            allowed = {"pdf", "png", "jpg", "jpeg", "gif"}
            ext = (
                biz_reg_upload.filename.rsplit(".", 1)[1].lower()
                if "." in biz_reg_upload.filename
                else ""
            )
            if ext not in allowed:
                flash("  Business document File .", "danger")
                biz_reg_attachment = get_bizreg_attachment_for_crm_client(
                    client, verify_exists=True
                )
                return render_template(
                    "crm/form.html",
                    form=form,
                    client=client,
                    edit_mode=True,
                    biz_reg_attachment=biz_reg_attachment,
                )
            try:
                saved = save_bizreg_attachment_for_crm_client(
                    client,
                    biz_reg_upload,
                    uploaded_by=getattr(current_user, "id", None),
                )
            except UploadTooLargeError as exc:
                flash(str(exc) or "File  .", "danger")
                biz_reg_attachment = get_bizreg_attachment_for_crm_client(
                    client, verify_exists=True
                )
                return render_template(
                    "crm/form.html",
                    form=form,
                    client=client,
                    edit_mode=True,
                    biz_reg_attachment=biz_reg_attachment,
                )
            if saved:
                _attach_biz_reg_file(
                    client,
                    {
                        "original_name": saved.get("original_name") or biz_reg_upload.filename,
                        "stored_name": saved.get("stored_name"),
                        "uploaded_at": datetime.utcnow().isoformat(),
                        "attachment_client_id": saved.get("attachment_client_id"),
                        "attachment_id": saved.get("id"),
                    },
                )

        _record_client_audit(
            client=client,
            action="client.update",
            before=audit_before,
            source="crm.client_edit",
        )
        db.session.commit()
        try:
            enqueue_crm_client_post_save(client.id)
        except Exception as exc:
            report_swallowed_exception(
                exc,
                context="crm.client_edit.enqueue_post_save",
                log_key="crm.client_edit.enqueue_post_save",
                log_window_seconds=300,
            )
        flash("Client information Edit. Search   from Process.", "success")
        return redirect(url_for("customers.client_view", client_id=client_id))

    # Populate form with existing data
    if request.method == "GET":
        extra = client.extra or {}
        form.client_name.data = client.name
        form.name_en.data = extra.get("name_en")
        form.category.data = client.type
        form.nationality.data = extra.get("nationality")
        form.business_type.data = extra.get("business_type")
        form.registration_number.data = client.registration_number
        form.email.data = client.email
        form.main_phone.data = client.phone or extra.get("main_phone")
        # Shared fields (Invoice )
        form.address.data = (
            client.address or extra.get("tax_address") or extra.get("applicant_address")
        )
        form.manager.data = client.manager or extra.get("tax_manager") or extra.get("tax_ceo")
        form.notes.data = client.notes or extra.get("note")
        form.mobile_phone.data = extra.get("mobile_phone")
        form.main_fax.data = extra.get("main_fax")
        form.homepage.data = extra.get("homepage")
        form.other_contact.data = extra.get("other_contact")
        form.client_code.data = extra.get("client_code")
        form.viewer_default.data = extra.get("viewer_default")
        form.annuity_management_disabled.data = bool(extra.get("annuity_management_disabled"))

        # Applicant codes
        codes = extra.get("applicant_codes", [])
        if len(codes) > 0:
            form.applicant_code1.data = codes[0]
        if len(codes) > 1:
            form.applicant_code2.data = codes[1]
        if len(codes) > 2:
            form.applicant_code3.data = codes[2]

        # Input date
        if extra.get("input_date"):
            try:
                from datetime import date

                form.input_date.data = date.fromisoformat(extra["input_date"])
            except (ValueError, TypeError) as exc:
                # Optional field: ignore invalid values, but keep a throttled record for debugging.
                report_swallowed_exception(
                    exc,
                    context="crm.client_form.parse_input_date",
                    log_key="crm.client_form.parse_input_date",
                    log_window_seconds=300,
                )

        # Other fields
        form.applicant_address.data = extra.get("applicant_address")
        form.applicant_email.data = extra.get("applicant_email")
        form.applicant_phone.data = extra.get("applicant_phone")
        form.applicant_fax.data = extra.get("applicant_fax")
        form.business_reg_no.data = extra.get("business_reg_no")
        form.tax_company_name.data = extra.get("tax_company_name")
        form.tax_ceo.data = extra.get("tax_ceo")
        form.tax_business_type.data = extra.get("tax_business_type")
        form.tax_business_item.data = extra.get("tax_business_item")
        form.tax_address.data = extra.get("tax_address")
        form.tax_manager.data = extra.get("tax_manager")
        form.tax_manager_email.data = extra.get("tax_manager_email")
        form.tax_manager_mobile.data = extra.get("tax_manager_mobile")
        form.mail_recv_address.data = extra.get("mail_recv_address")
        form.mail_receiver.data = extra.get("mail_receiver")
        form.personal_email.data = extra.get("personal_email")
        form.personal_phone.data = extra.get("personal_phone")
        form.personal_fax.data = extra.get("personal_fax")
        form.other_address.data = extra.get("other_address")
        form.other_email.data = extra.get("other_email")
        form.other_phone.data = extra.get("other_phone")
        form.other_fax.data = extra.get("other_fax")
        form.note.data = extra.get("note")
        form.special_note.data = extra.get("special_note")

    biz_reg_attachment = get_bizreg_attachment_for_crm_client(client, verify_exists=True)
    return render_template(
        "crm/form.html",
        form=form,
        client=client,
        edit_mode=True,
        biz_reg_attachment=biz_reg_attachment,
    )


@bp.route("/clients/<int:client_id>/delete", methods=["POST"])
@login_required
def client_delete(client_id):
    """Soft delete a client."""
    client = Client.query.get_or_404(client_id)
    audit_before = _client_audit_snapshot(client)
    client.is_deleted = True
    client.deleted_at = datetime.utcnow()
    client.deleted_by = getattr(current_user, "id", None)
    _record_client_audit(
        client=client,
        action="client.delete",
        before=audit_before,
        source="crm.client_delete",
        include_snapshots=True,
    )
    db.session.commit()
    flash("Client Delete.", "success")
    return redirect(url_for("customers.clients"))


@bp.route("/clients/<int:client_id>/sync-invoice", methods=["POST"])
@login_required
def client_sync_invoice(client_id):
    """Sync client with invoice system."""
    if not is_invoice_manager(current_user):
        abort(403, "You do not have permission to sync invoice clients.")
    client = Client.query.get_or_404(client_id)

    try:
        from app.services.billing.invoice_bridge import (
            InvoiceBridgeError,
            ensure_invoice_client_link,
        )

        invoice_client_id = ensure_invoice_client_link(client)
        flash(f"Invoice client link created. (Invoice Client #{invoice_client_id})", "success")
    except InvoiceBridgeError as exc:
        db.session.rollback()
        current_app.logger.exception("client_sync_invoice failed (client_id=%s)", client_id)
        flash(f"Invoice client link failed: {exc}", "danger")
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception(
            "client_sync_invoice unexpected error (client_id=%s)", client_id
        )
        flash(f"Invoice client link failed: {exc}", "danger")

    return redirect(safe_referrer_path() or url_for("customers.client_view", client_id=client_id))


@bp.route("/clients/sync_from_invoice_client", methods=["POST"])
@login_required
def sync_from_invoice_client():
    """Create/link CRM client from an invoice-module client_id (reverse sync)."""
    if not is_invoice_manager(current_user):
        abort(403, "You do not have permission to sync invoice clients.")
    raw = (
        request.form.get("invoice_client_id") or request.args.get("invoice_client_id") or ""
    ).strip()
    if not raw.isdigit():
        abort(400, "invoice_client_id is required.")
    invoice_client_id = int(raw)

    next_url = (request.form.get("next") or request.args.get("next") or "").strip()

    try:
        from app.services.billing.invoice_bridge import (
            InvoiceBridgeError,
            ensure_ipm_client_link_from_invoice_client,
        )

        ipm_client = ensure_ipm_client_link_from_invoice_client(invoice_client_id)
        flash(
            f"CRM Client Link. (CRM #{getattr(ipm_client, 'id', None)})",
            "success",
        )
        # Prefer explicit next (from invoice module UI), otherwise go to CRM client view.
        if next_url:
            try:
                next_url = next_url.replace("\r", "").replace("\n", "")
            except Exception:
                next_url = ""
            if next_url.startswith("/") and not next_url.startswith("//"):
                return redirect(next_url)
        return redirect(url_for("customers.client_view", client_id=int(getattr(ipm_client, "id"))))
    except InvoiceBridgeError as exc:
        db.session.rollback()
        current_app.logger.exception(
            "sync_from_invoice_client failed (invoice_client_id=%s)", invoice_client_id
        )
        flash(f"CRM Link Failed: {exc}", "danger")
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception(
            "sync_from_invoice_client unexpected error (invoice_client_id=%s)", invoice_client_id
        )
        flash(f"CRM Link Failed: {exc}", "danger")

    # Fall back to referrer or CRM list.
    return redirect(safe_referrer_path() or url_for("customers.clients"))


@bp.route("/clients/merge", methods=["GET", "POST"])
@login_required
@role_required("admin")
def client_merge():
    """Merge duplicate clients."""
    if request.method == "POST":
        source_ids = [int(x) for x in request.form.getlist("source_clients") if str(x).isdigit()]
        target_id = request.form.get("target_client", type=int)

        if not target_id or not source_ids:
            flash(" Client(1 )  Client Select.", "warning")
            return redirect(url_for("customers.client_merge"))

        source_ids = [i for i in source_ids if i != target_id]
        if not source_ids:
            flash(" Client Select  none.", "warning")
            return redirect(url_for("customers.client_merge"))

        try:
            result = ClientMergeService.merge_clients(
                target_client_id=target_id,
                source_client_ids=source_ids,
                merge_notes=True,
                merged_by=getattr(current_user, "id", None),
                reason="CRM UI merge",
                backup_required=True,
                backup_attachments=True,
            )
            attachment_issues = ClientMergeService.collect_attachment_move_issues(
                result.get("invoice", {})
            )
            flash(
                f"Client  Done: {len(result['source_client_ids'])} → #{result['target_client_id']}",
                "success",
            )
            if any(attachment_issues.values()):
                flash(
                    "File Go In Progress /Failed exists.   Confirm "
                    f"({attachment_issues['missing']} missing, "
                    f"{attachment_issues['copy_failures']} copy failure(s), "
                    f"{attachment_issues['delete_failures']} delete failure(s)).",
                    "warning",
                )
            return redirect(url_for("customers.client_view", client_id=target_id))
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception(
                "CRM merge failed (request_id=%s)",
                getattr(g, "request_id", None),
            )
            flash(f"Client  Failed: {e}", "danger")
            return redirect(url_for("customers.client_merge"))

    clients = Client.query.filter_by(is_deleted=False).order_by(Client.name).all()

    # Optional pre-selection via query string (useful from duplicate lists / deep links)
    preselected_target_id = request.args.get("target_client", type=int) or request.args.get(
        "target", type=int
    )
    raw_sources = []
    raw_sources.extend(request.args.getlist("source_clients"))
    raw_sources.extend(request.args.getlist("source"))
    preselected_source_ids: list[int] = []
    for v in raw_sources:
        if str(v).isdigit():
            preselected_source_ids.append(int(v))
    preselected_source_ids = sorted({i for i in preselected_source_ids if i})
    if preselected_target_id and preselected_target_id in preselected_source_ids:
        preselected_source_ids = [i for i in preselected_source_ids if i != preselected_target_id]

    # Relationship counts to help picking a "target" client.
    case_counts: dict[int, int] = {}
    contact_counts: dict[int, int] = {}
    opportunity_counts: dict[int, int] = {}
    activity_counts: dict[int, int] = {}
    try:
        # CRM "Matter" = IP matter. Count by:
        # 1) canonical MatterCustomField.data.client_id (new UI-created matters)
        # 2) migrated party-role mapping (matter_party_role.party_id == clients.party_id)
        from app.models.ip_records import Matter, MatterCustomField, MatterPartyRole

        try:
            bind = db.session.get_bind()
            dialect = (getattr(bind.dialect, "name", "") or "").lower() if bind else ""
        except Exception:
            dialect = ""

        if dialect.startswith("postgres"):
            client_id_expr = MatterCustomField.data["client_id"].as_string()
        else:
            client_id_expr = func.json_extract(MatterCustomField.data, "$.client_id")
        client_id_expr = cast(client_id_expr, String)

        is_matter_active = (Matter.is_deleted.is_(False)) | (Matter.is_deleted.is_(None))

        custom_links = (
            db.session.query(
                client_id_expr.label("client_id"),
                MatterCustomField.matter_id.label("matter_id"),
            )
            .join(Matter, Matter.matter_id == MatterCustomField.matter_id)
            .filter(MatterCustomField.namespace.in_(_CRM_CLIENT_MATTER_NAMESPACES))
            .filter(func.nullif(func.trim(client_id_expr), "").isnot(None))
            .filter(is_matter_active)
            .distinct()
        )

        client_party_id_expr = func.coalesce(
            func.nullif(Client.party_id, ""), func.nullif(Client.ipm_party_id, "")
        )
        party_links = (
            db.session.query(
                cast(Client.id, String).label("client_id"),
                MatterPartyRole.matter_id.label("matter_id"),
            )
            .join(MatterPartyRole, MatterPartyRole.party_id == client_party_id_expr)
            .join(Matter, Matter.matter_id == MatterPartyRole.matter_id)
            .filter(Client.is_deleted.is_(False))
            .filter(func.lower(MatterPartyRole.role_code).in_(_CRM_CLIENT_PARTY_ROLE_CODES))
            .filter(is_matter_active)
            .distinct()
        )

        links_subq = custom_links.union(party_links).subquery()
        rows = (
            db.session.query(
                links_subq.c.client_id.label("client_id"),
                func.count(func.distinct(links_subq.c.matter_id)).label("cnt"),
            )
            .group_by(links_subq.c.client_id)
            .all()
        )
        case_counts = {
            int(str(cid)): int(cnt or 0)
            for (cid, cnt) in rows
            if cid is not None and str(cid).isdigit()
        }
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_merge.case_counts",
            log_key="crm.client_merge.case_counts",
            log_window_seconds=300,
        )
    try:
        contact_counts = dict(
            db.session.query(CRMContact.client_id, func.count(CRMContact.id))
            .filter(CRMContact.client_id.isnot(None))
            .group_by(CRMContact.client_id)
            .all()
        )
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_merge.contact_counts",
            log_key="crm.client_merge.contact_counts",
            log_window_seconds=300,
        )
    try:
        opportunity_counts = dict(
            db.session.query(CRMOpportunity.client_id, func.count(CRMOpportunity.id))
            .filter(CRMOpportunity.client_id.isnot(None))
            .group_by(CRMOpportunity.client_id)
            .all()
        )
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_merge.opportunity_counts",
            log_key="crm.client_merge.opportunity_counts",
            log_window_seconds=300,
        )
    try:
        activity_counts = dict(
            db.session.query(CRMActivity.client_id, func.count(CRMActivity.id))
            .filter(CRMActivity.client_id.isnot(None))
            .group_by(CRMActivity.client_id)
            .all()
        )
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_merge.activity_counts",
            log_key="crm.client_merge.activity_counts",
            log_window_seconds=300,
        )

    # Duplicate hints (same logic as client list page; keep it small and safe).
    duplicate_groups = []
    try:
        rows = (
            Client.query.filter_by(is_deleted=False)
            .with_entities(
                Client.id,
                Client.name,
                Client.email,
                Client.phone,
                Client.registration_number,
                Client.biz_reg_number,
                Client.biz_tax_invoice_email,
            )
            .all()
        )
        duplicate_groups = _build_duplicate_groups(rows)
    except Exception as exc:
        report_swallowed_exception(
            exc,
            context="crm.client_merge.duplicate_groups",
            log_key="crm.client_merge.duplicate_groups",
            log_window_seconds=300,
        )
    duplicate_groups = _filter_duplicate_groups_by_rejections(duplicate_groups)

    return render_template(
        "crm/merge.html",
        clients=clients,
        case_counts=case_counts,
        contact_counts=contact_counts,
        opportunity_counts=opportunity_counts,
        activity_counts=activity_counts,
        duplicate_groups=duplicate_groups[:10],
        preselected_target_id=preselected_target_id,
        preselected_source_ids=preselected_source_ids,
    )


@bp.route("/clients/duplicates/reject", methods=["POST"])
@login_required
@role_required("admin")
def client_duplicate_reject():
    signature = _normalize_duplicate_group_signature(request.form.get("group_signature"))
    if not signature:
        flash("     .", "warning")
        return redirect(safe_referrer_path() or url_for("customers.clients"))

    try:
        signatures = _load_rejected_duplicate_group_signatures()
        if signature not in signatures:
            signatures.append(signature)
            _save_rejected_duplicate_group_signatures(signatures)
            db.session.commit()
            flash("  .", "success")
        else:
            flash("   .", "info")
    except Exception as exc:
        db.session.rollback()
        report_swallowed_exception(
            exc,
            context="crm.client_duplicate_reject",
            log_key="crm.client_duplicate_reject",
            log_window_seconds=300,
        )
        flash("   Process In Progress Error .", "danger")

    return redirect(safe_referrer_path() or url_for("customers.clients"))


@bp.route("/clients/parse-biz-reg", methods=["POST"])
@login_required
def parse_biz_reg():
    """Parse business registration document and extract fields.

    If client_id is provided, the file is saved permanently.
    """
    return json_error("not_available", "Business document parsing is not available.", status=404)

    try:
        f = request.files.get("biz_reg_file")
        if not f or not f.filename:
            return json_error("bad_request", "No file provided.", status=400)

        client_id = request.form.get("client_id", type=int)

        # Check extension
        allowed = {"pdf", "png", "jpg", "jpeg", "gif"}
        ext = f.filename.rsplit(".", 1)[1].lower() if "." in f.filename else ""
        if ext not in allowed:
            return json_error("invalid_file_type", "  File .", status=400)

        api_key = get_openai_api_key(allow_legacy=False)
        if not api_key:
            return json_error("config_missing", "OpenAI API  Settings .", status=500)

        tmp_path = None
        file_info = None
        try:
            suffix = os.path.splitext(f.filename)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp_path = tmp.name
            try:
                temp_root = Path(tmp_path).resolve().parent
                rel_tmp = Path(tmp_path).resolve().relative_to(temp_root)
                temp_service = FileAssetService(upload_root=temp_root)
                temp_service.store_upload_to_path(
                    f,
                    rel_path=rel_tmp,
                    overwrite=True,
                    track_pending=False,
                )
                validation = validate_upload_path(
                    tmp_path,
                    filename=f.filename,
                    allowed_exts={f".{item}" for item in allowed},
                )
                if not validation.ok:
                    return json_error(
                        "upload_security_failed",
                        "Upload Security  Failed.",
                        status=400,
                    )
                scan_upload_path(tmp_path, filename=f.filename)
            except UploadTooLargeError:
                return json_error("file_too_large", "File  .", status=413)
            except UploadSecurityError:
                return json_error(
                    "upload_security_failed",
                    "Upload Security  Failed.",
                    status=400,
                )

            # Release any checked-out request connection while calling the external analyzer.
            _reset_request_db_session()
            analysis_json = json.dumps({"biz_reg": {}}, ensure_ascii=False)

            if not analysis_json:
                return json_error(
                    "analysis_failed",
                    "analysis Failed (Data None).",
                    status=500,
                )

            data = json.loads(analysis_json)
            biz_reg = data.get("biz_reg", {})

            # Save file permanently if client_id is provided
            if client_id:
                try:
                    file_info = _save_biz_reg_file(client_id, tmp_path, f.filename)
                except UploadTooLargeError:
                    return json_error("file_too_large", "File  .", status=413)
                # Re-checkout a fresh DB connection after the long-running analysis call.
                _reset_request_db_session()
                client = db.session.get(Client, client_id)
                if client:
                    _attach_biz_reg_file(client, file_info)
                    db.session.commit()

            # Map to frontend fields
            fields = {
                "client_name": biz_reg.get("company_name") or "",
                "registration_number": biz_reg.get("corp_registration_number") or "",
                "business_reg_no": biz_reg.get("reg_number") or "",
                "main_phone": "",
                "tax_company_name": biz_reg.get("company_name") or "",
                "tax_ceo": biz_reg.get("representative_name") or "",
                "tax_business_type": biz_reg.get("business_type") or "",
                "tax_address": biz_reg.get("business_location")
                or biz_reg.get("head_office_location")
                or "",
                "tax_manager_email": biz_reg.get("tax_invoice_email") or "",
                "address": biz_reg.get("business_location")
                or biz_reg.get("head_office_location")
                or "",
                "manager": biz_reg.get("representative_name") or "",
                "applicant_address": biz_reg.get("business_location")
                or biz_reg.get("head_office_location")
                or "",
                "email": biz_reg.get("tax_invoice_email") or "",
            }

            response = {
                "success": True,
                "message": "Business document Analysis complete",
                "fields": fields,
                "raw": biz_reg,
            }
            if file_info:
                response["file_info"] = file_info

            return jsonify(response)

        except Exception as e:
            report_swallowed_exception(e, context="crm.parse_biz_reg")
            return json_error("server_error", f"Error : {str(e)}", status=500)
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception as exc:
                    report_swallowed_exception(exc, context="crm.parse_biz_reg.cleanup")
    except Exception as e:
        report_swallowed_exception(e, context="crm.parse_biz_reg.outer")
        return json_error("server_error", f" Error: {str(e)}", status=500)


@bp.route("/clients/<int:client_id>/biz-reg/upload", methods=["POST"])
@login_required
def biz_reg_upload(client_id):
    """Quick upload business registration file for a client."""
    client = Client.query.get_or_404(client_id)
    f = request.files.get("biz_reg_file")
    if not f or not f.filename:
        return json_error("bad_request", "No file provided.", status=400)

    ext = f.filename.rsplit(".", 1)[1].lower() if "." in f.filename else ""
    if ext not in ALLOWED_BIZREG_EXTS:
        return json_error(
            "invalid_file_type", "  Business document File .", status=400
        )

    try:
        saved = save_bizreg_attachment_for_crm_client(
            client,
            f,
            uploaded_by=getattr(current_user, "id", None),
        )
        if not saved:
            return json_error("upload_failed", "File Save Failed.", status=500)

        # save_bizreg_attachment_for_crm_client() may recreate the scoped session
        # when it has to recover from a stale DB connection. Re-attach the client
        # so the extra-field update participates in the same live session/tx.
        try:
            client = db.session.merge(client)
        except Exception:
            client = db.session.get(Client, client_id) or Client.query.get_or_404(client_id)

        extra = dict(client.extra or {})
        extra["biz_reg_file"] = {
            "original_name": saved.get("original_name") or f.filename,
            "stored_name": saved.get("stored_name"),
            "uploaded_at": datetime.utcnow().isoformat(),
            "attachment_client_id": saved.get("attachment_client_id"),
            "attachment_id": saved.get("id"),
        }
        biz_reg = {}
        analysis_meta = saved.get("analysis_meta")
        if analysis_meta:
            try:
                meta = (
                    json.loads(analysis_meta) if isinstance(analysis_meta, str) else analysis_meta
                )
                if isinstance(meta, dict):
                    biz_reg = meta.get("biz_reg") or {}
            except Exception:
                biz_reg = {}

        parsed = {}
        parsed_summary = ""
        if isinstance(biz_reg, dict) and biz_reg:
            company_name = biz_reg.get("company_name") or ""
            reg_number = biz_reg.get("reg_number") or ""
            corp_reg_number = biz_reg.get("corp_registration_number") or ""
            rep_name = biz_reg.get("representative_name") or ""
            business_type = biz_reg.get("business_type") or ""
            business_location = biz_reg.get("business_location") or ""
            head_office_location = biz_reg.get("head_office_location") or ""
            tax_email = biz_reg.get("tax_invoice_email") or ""
            address = business_location or head_office_location or ""

            parsed = {
                "company_name": company_name,
                "reg_number": reg_number,
                "corp_registration_number": corp_reg_number,
                "representative_name": rep_name,
                "business_type": business_type,
                "business_location": business_location or head_office_location,
                "tax_invoice_email": tax_email,
            }
            summary_parts = [p for p in [company_name, reg_number] if p]
            if not summary_parts and corp_reg_number:
                summary_parts.append(corp_reg_number)
            if summary_parts:
                parsed_summary = " / ".join(summary_parts[:2])
            extra["biz_reg_suggestion"] = {
                "source": "biz_reg_upload",
                "uploaded_at": datetime.utcnow().isoformat(),
                "attachment_id": saved.get("id"),
                "attachment_client_id": saved.get("attachment_client_id"),
                "file_name": saved.get("original_name") or f.filename,
                "fields": biz_reg,
                "summary": parsed_summary,
            }
            extra.pop("biz_reg_suggestion_applied_at", None)
        client.extra = extra
        flag_modified(client, "extra")
        db.session.commit()
    except UploadTooLargeError:
        db.session.rollback()
        return json_error("file_too_large", "File  .", status=413)
    except UploadSecurityError:
        db.session.rollback()
        return json_error(
            "upload_security_failed",
            "Upload Security  Failed.",
            status=400,
        )
    except ValueError as e:
        db.session.rollback()
        return json_error("bad_request", str(e), status=400)
    except Exception as e:
        db.session.rollback()
        report_swallowed_exception(e, context="crm.biz_reg_upload")
        return json_error("upload_failed", "Upload In Progress Error .", status=500)

    return jsonify(
        {
            "success": True,
            "message": "Upload complete (value Saved)",
            "file": {
                "original_name": saved.get("original_name") or f.filename,
            },
            "download_url": url_for("customers.biz_reg_download_shared", client_id=client.id),
            "parsed": parsed,
            "parsed_summary": parsed_summary,
            "suggestion": extra.get("biz_reg_suggestion"),
        }
    )


@bp.route("/clients/parse-customer-llm", methods=["POST"])
@login_required
def parse_customer_llm():
    """Queue email text parsing via LLM and return immediately."""
    try:
        if not request.is_json:
            return (
                jsonify({"success": False, "error": "Content-Type must be application/json"}),
                400,
            )

        data = request.get_json(silent=True) or {}
        email_text = (data.get("email_text") or "").strip()

        if not email_text:
            return jsonify({"success": False, "error": " Input."}), 400

        if not get_openai_api_key(allow_legacy=False):
            return (
                jsonify({"success": False, "error": "OpenAI API  Settings ."}),
                500,
            )

        op = create_customer_llm_parse_operation(
            actor_id=getattr(current_user, "id", None),
            email_text=email_text,
        )
        db.session.commit()
        job_id = enqueue_customer_llm_parse(op.id, email_text)
        return (
            jsonify(
                {
                    "success": True,
                    "queued": True,
                    "operation_id": op.id,
                    "job_id": job_id,
                    "status": "queued",
                    "status_url": url_for(
                        "customers.parse_customer_llm_status",
                        operation_id=op.id,
                    ),
                }
            ),
            202,
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("CRM LLM parsing failed")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/clients/parse-customer-llm/<int:operation_id>", methods=["GET"])
@login_required
def parse_customer_llm_status(operation_id: int):
    op = Operation.query.get_or_404(operation_id)
    actor_id = getattr(current_user, "id", None)
    if (
        op.actor_id
        and actor_id
        and int(op.actor_id) != int(actor_id)
        and not is_admin(current_user)
    ):
        abort(403)

    summary = op.summary_json if isinstance(op.summary_json, dict) else {}
    payload = {
        "success": True,
        "operation_id": op.id,
        "status": op.status,
        "queued": op.status in {"queued", "running"},
    }
    if op.status == "succeeded":
        payload["customer"] = summary.get("customer") or {}
    if op.status == "failed":
        payload["success"] = False
        payload["error"] = op.error_text or "LLM parsing failed"
    return jsonify(payload)


@bp.get("/clients/tax-suggest.json")
@login_required
def tax_suggest_clients_json():
    q = (request.args.get("q") or "").strip()
    limit = request.args.get("limit", type=int) or 6
    limit = max(1, min(limit, 20))
    if not q:
        return jsonify({"results": []})

    query = Client.query.filter_by(is_deleted=False).filter(
        db.or_(
            sqlalchemy_contains_query(Client.name, q),
            sqlalchemy_contains_query(Client.registration_number, q),
            sqlalchemy_contains_query(Client.biz_reg_number, q),
            sqlalchemy_contains_query(Client.biz_company_name, q),
            sqlalchemy_contains_query(Client.biz_tax_invoice_email, q),
            sqlalchemy_contains_query(Client.email, q),
            sqlalchemy_contains_query(Client.phone, q),
            sqlalchemy_contains_query(Client.search_tags, q),
        )
    )
    rows = query.order_by(Client.name).limit(limit).all()

    results = []
    for c in rows:
        extra = c.extra if isinstance(c.extra, dict) else {}

        def _extra_text(key: str) -> str:
            value = extra.get(key)
            if value is None:
                return ""
            if isinstance(value, str):
                return value.strip()
            return str(value).strip()

        results.append(
            {
                "id": c.id,
                "name": c.name or "",
                "registration_number": c.registration_number or "",
                "manager": c.manager or "",
                "email": c.email or "",
                "phone": c.phone or "",
                "address": c.address or "",
                "biz_reg_number": c.biz_reg_number or "",
                "biz_company_name": c.biz_company_name or "",
                "biz_representative_name": c.biz_representative_name or "",
                "biz_business_type": c.biz_business_type or "",
                "biz_business_location": c.biz_business_location or "",
                "biz_tax_invoice_email": c.biz_tax_invoice_email or "",
                # Dedicated tax-invoice fields stored in CRM extra payload.
                # Frontend should prioritize these fields when present.
                "tax_business_reg_no": _extra_text("business_reg_no"),
                "tax_company_name": _extra_text("tax_company_name"),
                "tax_ceo": _extra_text("tax_ceo"),
                "tax_business_type": _extra_text("tax_business_type"),
                "tax_business_item": _extra_text("tax_business_item"),
                "tax_address": _extra_text("tax_address"),
                "tax_manager": _extra_text("tax_manager"),
                "tax_manager_email": _extra_text("tax_manager_email"),
                "tax_manager_mobile": _extra_text("tax_manager_mobile"),
                "view_url": url_for("customers.client_view", client_id=c.id),
            }
        )
    return jsonify({"results": results})


@bp.route("/clients/<int:client_id>/biz-reg/download")
@login_required
def biz_reg_download(client_id):
    """Download the business registration file for a client."""
    return biz_reg_download_shared(client_id)


# ============================================================================
# Leads
# ============================================================================


@bp.route("/leads")
@login_required
def lead_list():
    """List all leads with search and filter."""
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "")
    page = request.args.get("page", 1, type=int)
    page = max(1, int(page or 1))
    per_page = request.args.get("per_page", 20, type=int)
    per_page = max(1, min(int(per_page or 20), 200))

    query = CRMLead.query

    if q:
        query = query.filter(
            db.or_(
                sqlalchemy_contains_query(CRMLead.name, q),
                sqlalchemy_contains_query(CRMLead.company_name, q),
                sqlalchemy_contains_query(CRMLead.email, q),
                sqlalchemy_contains_query(CRMLead.phone, q),
            )
        )

    if status:
        query = query.filter_by(status=status)

    query = query.order_by(CRMLead.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    page_count = max(1, pagination.pages)
    if page > page_count:
        page = page_count
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        page_count = max(1, pagination.pages)

    return render_template(
        "crm/lead_list.html",
        leads=pagination.items,
        page=page,
        pages=page_count,
        total=pagination.total,
        per_page=per_page,
        q=q,
        current_status=status,
        status_choices=CRMLead.STATUS_CHOICES,
    )


@bp.route("/leads/create", methods=["GET", "POST"])
@login_required
def lead_create():
    """Create a new lead."""
    form = LeadForm()

    # Populate assigned_to choices
    assigned_to_q = request.args.get("assigned_to_q", "").strip()
    users_query = User.query.filter_by(is_active=True)
    if assigned_to_q:
        users_query = users_query.filter(
            db.or_(
                sqlalchemy_contains_query(User.display_name, assigned_to_q),
                sqlalchemy_contains_query(User.username, assigned_to_q),
                sqlalchemy_contains_query(User.email, assigned_to_q),
            )
        ).order_by(User.display_name)
    else:
        users_query = users_query.order_by(User.display_name)
    users = users_query.limit(200).all()
    form.assigned_to.choices = [(0, "Select")] + [(u.id, u.display_name or u.username) for u in users]

    if form.validate_on_submit():
        lead = CRMLead(
            name=form.name.data,
            company_name=form.company_name.data,
            email=form.email.data,
            phone=form.phone.data,
            status=form.status.data,
            source=form.source.data if form.source.data else None,
            assigned_to=form.assigned_to.data if form.assigned_to.data else None,
            notes=form.notes.data,
        )
        db.session.add(lead)
        db.session.commit()

        flash(" Registration.", "success")
        return redirect(url_for("customers.lead_view", lead_id=lead.id))

    return render_template("crm/lead_form.html", form=form, assigned_to_q=assigned_to_q)


@bp.route("/leads/<int:lead_id>")
@login_required
def lead_view(lead_id):
    """View lead details with activities."""
    lead = CRMLead.query.get_or_404(lead_id)
    activities = lead.activities.order_by(CRMActivity.activity_date.desc()).all()

    activity_form = ActivityForm()
    activity_form.lead_id.data = lead_id

    return render_template(
        "crm/lead_view.html",
        lead=lead,
        activities=activities,
        activity_form=activity_form,
    )


@bp.route("/leads/<int:lead_id>/edit", methods=["GET", "POST"])
@login_required
def lead_edit(lead_id):
    """Edit an existing lead."""
    lead = CRMLead.query.get_or_404(lead_id)
    form = LeadForm(obj=lead)

    # Populate assigned_to choices
    assigned_to_q = request.args.get("assigned_to_q", "").strip()
    users_query = User.query.filter_by(is_active=True)
    if assigned_to_q:
        users_query = users_query.filter(
            db.or_(
                sqlalchemy_contains_query(User.display_name, assigned_to_q),
                sqlalchemy_contains_query(User.username, assigned_to_q),
                sqlalchemy_contains_query(User.email, assigned_to_q),
            )
        ).order_by(User.display_name)
    else:
        users_query = users_query.order_by(User.display_name)
    users = users_query.limit(200).all()
    if lead.assigned_to and all(u.id != lead.assigned_to for u in users):
        extra_user = User.query.get(lead.assigned_to)
        if extra_user:
            users.append(extra_user)
    users = sorted(users, key=lambda u: (u.display_name or u.username or ""))
    form.assigned_to.choices = [(0, "Select")] + [(u.id, u.display_name or u.username) for u in users]

    if form.validate_on_submit():
        lead.name = form.name.data
        lead.company_name = form.company_name.data
        lead.email = form.email.data
        lead.phone = form.phone.data
        lead.status = form.status.data
        lead.source = form.source.data if form.source.data else None
        lead.assigned_to = form.assigned_to.data if form.assigned_to.data else None
        lead.notes = form.notes.data

        db.session.commit()
        flash(" Edit.", "success")
        return redirect(url_for("customers.lead_view", lead_id=lead_id))

    return render_template(
        "crm/lead_form.html",
        form=form,
        lead=lead,
        edit_mode=True,
        assigned_to_q=assigned_to_q,
    )


@bp.route("/leads/<int:lead_id>/convert", methods=["POST"])
@login_required
def lead_convert(lead_id):
    """Convert a lead to a client."""
    lead = CRMLead.query.get_or_404(lead_id)

    if lead.status == "converted":
        flash("  .", "warning")
        return redirect(url_for("customers.lead_view", lead_id=lead_id))

    # Create a new client from lead
    client = Client(
        name=lead.company_name or lead.name,
        email=lead.email,
        phone=lead.phone,
        contact_person=lead.name if lead.company_name else None,
        notes=lead.notes,
        extra={
            "converted_from_lead": lead_id,
            "source": lead.source,
        },
    )
    db.session.add(client)
    db.session.flush()  # Get client.id

    # Update lead status
    lead.status = "converted"
    lead.converted_client_id = client.id

    # Move activities to client
    CRMActivity.query.filter_by(lead_id=lead_id).update({"client_id": client.id})

    db.session.commit()

    flash(f" Clientto : {client.name}", "success")
    return redirect(url_for("customers.client_view", client_id=client.id))


@bp.route("/leads/<int:lead_id>/delete", methods=["POST"])
@login_required
def lead_delete(lead_id):
    """Delete a lead."""
    lead = CRMLead.query.get_or_404(lead_id)

    # Delete associated activities
    CRMActivity.query.filter_by(lead_id=lead_id).delete()

    db.session.delete(lead)
    db.session.commit()

    flash(" Delete.", "success")
    return redirect(url_for("customers.lead_list"))


# ============================================================================
# Opportunities
# ============================================================================


@bp.route("/opportunities")
@login_required
def opportunity_list():
    """List all opportunities (pipeline view)."""
    stage = request.args.get("stage", "")
    client_id = request.args.get("client_id", type=int)
    page = request.args.get("page", 1, type=int)
    page = max(1, int(page or 1))
    per_page = request.args.get("per_page", 20, type=int)
    per_page = max(1, min(int(per_page or 20), 200))

    query = CRMOpportunity.query

    if stage:
        query = query.filter_by(stage=stage)

    if client_id:
        query = query.filter_by(client_id=client_id)

    query = query.order_by(CRMOpportunity.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    page_count = max(1, pagination.pages)
    if page > page_count:
        page = page_count
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        page_count = max(1, pagination.pages)

    # Get stage counts for summary
    stage_counts = {s: 0 for s, _label in CRMOpportunity.STAGE_CHOICES}
    for stage_value, count in (
        db.session.query(CRMOpportunity.stage, func.count()).group_by(CRMOpportunity.stage).all()
    ):
        if stage_value in stage_counts:
            stage_counts[stage_value] = count

    return render_template(
        "crm/opportunity_list.html",
        opportunities=pagination.items,
        page=page,
        pages=page_count,
        total=pagination.total,
        per_page=per_page,
        current_stage=stage,
        stage_choices=CRMOpportunity.STAGE_CHOICES,
        stage_counts=stage_counts,
    )


@bp.route("/opportunities/create", methods=["GET", "POST"])
@login_required
def opportunity_create():
    """Create a new opportunity."""
    form = OpportunityForm()

    # Populate client choices
    client_q = request.args.get("client_q", "").strip()
    clients_query = Client.query.filter_by(is_deleted=False)
    if client_q:
        clients_query = clients_query.filter(
            db.or_(
                sqlalchemy_contains_query(Client.name, client_q),
                sqlalchemy_contains_query(Client.registration_number, client_q),
                sqlalchemy_contains_query(Client.email, client_q),
            )
        ).order_by(Client.name)
    else:
        clients_query = clients_query.order_by(Client.id.desc())
    # Pre-select client if provided in query string
    preselect_client_id = request.args.get("client_id", type=int)
    clients = clients_query.limit(200).all()
    if preselect_client_id and all(c.id != preselect_client_id for c in clients):
        extra_client = Client.query.get(preselect_client_id)
        if extra_client:
            clients.append(extra_client)
    clients = sorted(clients, key=lambda c: (c.name or ""))
    form.client_id.choices = [(0, "Select")] + [(c.id, c.name) for c in clients]
    if preselect_client_id and request.method == "GET":
        form.client_id.data = preselect_client_id

    if form.validate_on_submit():
        opp = CRMOpportunity(
            client_id=form.client_id.data,
            name=form.name.data,
            stage=form.stage.data,
            amount=form.amount.data,
            probability=form.probability.data
            or CRMOpportunity.STAGE_PROBABILITY.get(form.stage.data, 10),
            expected_close_date=form.expected_close_date.data,
            notes=form.notes.data,
        )
        db.session.add(opp)
        db.session.commit()

        flash(" Registration.", "success")
        return redirect(url_for("customers.opportunity_view", opportunity_id=opp.id))

    return render_template(
        "crm/opportunity_form.html",
        form=form,
        client_q=client_q,
        preselect_client_id=preselect_client_id,
    )


@bp.route("/opportunities/<int:opportunity_id>")
@login_required
def opportunity_view(opportunity_id):
    """View opportunity details."""
    opportunity = CRMOpportunity.query.get_or_404(opportunity_id)
    activities = opportunity.activities.order_by(CRMActivity.activity_date.desc()).all()

    activity_form = ActivityForm()
    activity_form.opportunity_id.data = opportunity_id

    return render_template(
        "crm/opportunity_view.html",
        opportunity=opportunity,
        activities=activities,
        activity_form=activity_form,
    )


@bp.route("/opportunities/<int:opportunity_id>/edit", methods=["GET", "POST"])
@login_required
def opportunity_edit(opportunity_id):
    """Edit an existing opportunity."""
    opportunity = CRMOpportunity.query.get_or_404(opportunity_id)
    form = OpportunityForm(obj=opportunity)

    # Populate client choices
    client_q = request.args.get("client_q", "").strip()
    clients_query = Client.query.filter_by(is_deleted=False)
    if client_q:
        clients_query = clients_query.filter(
            db.or_(
                sqlalchemy_contains_query(Client.name, client_q),
                sqlalchemy_contains_query(Client.registration_number, client_q),
                sqlalchemy_contains_query(Client.email, client_q),
            )
        ).order_by(Client.name)
    else:
        clients_query = clients_query.order_by(Client.id.desc())
    clients = clients_query.limit(200).all()
    if opportunity.client_id and all(c.id != opportunity.client_id for c in clients):
        extra_client = Client.query.get(opportunity.client_id)
        if extra_client:
            clients.append(extra_client)
    clients = sorted(clients, key=lambda c: (c.name or ""))
    form.client_id.choices = [(0, "Select")] + [(c.id, c.name) for c in clients]

    if form.validate_on_submit():
        old_stage = opportunity.stage

        opportunity.client_id = form.client_id.data
        opportunity.name = form.name.data
        opportunity.stage = form.stage.data
        opportunity.amount = form.amount.data
        opportunity.probability = form.probability.data
        opportunity.expected_close_date = form.expected_close_date.data
        opportunity.notes = form.notes.data

        # Set closed_at when transitioning to closed state
        if opportunity.stage in ("closed_won", "closed_lost") and old_stage not in (
            "closed_won",
            "closed_lost",
        ):
            opportunity.closed_at = datetime.utcnow()
        elif opportunity.stage not in ("closed_won", "closed_lost"):
            opportunity.closed_at = None

        db.session.commit()
        flash(" Edit.", "success")
        return redirect(url_for("customers.opportunity_view", opportunity_id=opportunity_id))

    return render_template(
        "crm/opportunity_form.html",
        form=form,
        opportunity=opportunity,
        edit_mode=True,
        client_q=client_q,
        preselect_client_id=None,
    )


@bp.route("/opportunities/<int:opportunity_id>/update-stage", methods=["POST"])
@login_required
def opportunity_update_stage(opportunity_id):
    """Quick update opportunity stage (for Kanban drag-drop)."""
    opportunity = CRMOpportunity.query.get_or_404(opportunity_id)
    new_stage = request.form.get("stage") or request.json.get("stage")

    if new_stage not in dict(CRMOpportunity.STAGE_CHOICES):
        return jsonify({"success": False, "message": " ."}), 400

    old_stage = opportunity.stage
    opportunity.stage = new_stage
    opportunity.probability = CRMOpportunity.STAGE_PROBABILITY.get(
        new_stage, opportunity.probability
    )

    if new_stage in ("closed_won", "closed_lost") and old_stage not in (
        "closed_won",
        "closed_lost",
    ):
        opportunity.closed_at = datetime.utcnow()
    elif new_stage not in ("closed_won", "closed_lost"):
        opportunity.closed_at = None

    db.session.commit()

    return jsonify({"success": True, "message": " Change."})


@bp.route("/opportunities/<int:opportunity_id>/delete", methods=["POST"])
@login_required
def opportunity_delete(opportunity_id):
    """Delete an opportunity."""
    opportunity = CRMOpportunity.query.get_or_404(opportunity_id)
    client_id = opportunity.client_id

    # Delete associated activities
    CRMActivity.query.filter_by(opportunity_id=opportunity_id).delete()

    db.session.delete(opportunity)
    db.session.commit()

    flash(" Delete.", "success")
    return redirect(url_for("customers.client_view", client_id=client_id))


# ============================================================================
# Contacts
# ============================================================================


@bp.route("/contacts/create", methods=["POST"])
@login_required
def contact_create():
    """Create a new contact for a client."""
    form = ContactForm()

    if form.validate_on_submit():
        client_id = int(form.client_id.data)

        # If setting as primary, unset other primary contacts
        if form.is_primary.data:
            CRMContact.query.filter_by(client_id=client_id, is_primary=True).update(
                {"is_primary": False}
            )

        contact = CRMContact(
            client_id=client_id,
            name=form.name.data,
            title=form.title.data,
            email=form.email.data,
            phone=form.phone.data,
            mobile=form.mobile.data,
            is_primary=form.is_primary.data,
            notes=form.notes.data,
        )
        db.session.add(contact)
        db.session.commit()

        flash("Phone Add.", "success")
        return redirect(url_for("customers.client_view", client_id=client_id))

    flash("Phone Add Failed.", "danger")
    return redirect(safe_referrer_path() or url_for("customers.clients"))


@bp.route("/contacts/<int:contact_id>/edit", methods=["GET", "POST"])
@login_required
def contact_edit(contact_id):
    """Edit an existing contact."""
    contact = CRMContact.query.get_or_404(contact_id)
    form = ContactForm(obj=contact)
    form.client_id.data = contact.client_id

    if form.validate_on_submit():
        # If setting as primary, unset other primary contacts
        if form.is_primary.data and not contact.is_primary:
            CRMContact.query.filter_by(client_id=contact.client_id, is_primary=True).update(
                {"is_primary": False}
            )

        contact.name = form.name.data
        contact.title = form.title.data
        contact.email = form.email.data
        contact.phone = form.phone.data
        contact.mobile = form.mobile.data
        contact.is_primary = form.is_primary.data
        contact.notes = form.notes.data

        db.session.commit()
        flash("Phone Edit.", "success")
        return redirect(url_for("customers.client_view", client_id=contact.client_id))

    return render_template("crm/contact_form.html", form=form, contact=contact, edit_mode=True)


@bp.route("/contacts/<int:contact_id>/delete", methods=["POST"])
@login_required
def contact_delete(contact_id):
    """Delete a contact."""
    contact = CRMContact.query.get_or_404(contact_id)
    client_id = contact.client_id

    db.session.delete(contact)
    db.session.commit()

    flash("Phone Delete.", "success")
    return redirect(url_for("customers.client_view", client_id=client_id))


# ============================================================================
# Activities
# ============================================================================


@bp.route("/activities/create", methods=["POST"])
@login_required
def activity_create():
    """Log a new activity."""
    form = ActivityForm()

    if form.validate_on_submit():
        activity = CRMActivity(
            client_id=int(form.client_id.data) if form.client_id.data else None,
            lead_id=int(form.lead_id.data) if form.lead_id.data else None,
            opportunity_id=int(form.opportunity_id.data) if form.opportunity_id.data else None,
            type=form.type.data,
            summary=form.summary.data,
            description=form.description.data,
            activity_date=form.activity_date.data or datetime.utcnow(),
            user_id=current_user.id,
        )
        db.session.add(activity)
        db.session.commit()

        flash(" Log.", "success")

        # Redirect back to appropriate view
        if form.client_id.data:
            return redirect(url_for("customers.client_view", client_id=form.client_id.data))
        elif form.lead_id.data:
            return redirect(url_for("customers.lead_view", lead_id=form.lead_id.data))
        elif form.opportunity_id.data:
            return redirect(
                url_for("customers.opportunity_view", opportunity_id=form.opportunity_id.data)
            )

    flash(" Log Failed.", "danger")
    return redirect(safe_referrer_path() or url_for("customers.dashboard"))


@bp.route("/activities/<int:activity_id>/delete", methods=["POST"])
@login_required
def activity_delete(activity_id):
    """Delete an activity."""
    activity = CRMActivity.query.get_or_404(activity_id)

    # Remember where to redirect
    client_id = activity.client_id
    lead_id = activity.lead_id
    opportunity_id = activity.opportunity_id

    db.session.delete(activity)
    db.session.commit()

    flash(" Delete.", "success")

    if client_id:
        return redirect(url_for("customers.client_view", client_id=client_id))
    elif lead_id:
        return redirect(url_for("customers.lead_view", lead_id=lead_id))
    elif opportunity_id:
        return redirect(url_for("customers.opportunity_view", opportunity_id=opportunity_id))

    return redirect(url_for("customers.clients"))
