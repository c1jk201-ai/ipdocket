from __future__ import annotations

import hashlib
import io
import re
import threading
import time
import unicodedata
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from email.utils import parsedate_to_datetime
from typing import Any, Dict
from zoneinfo import ZoneInfo

import requests
from flask import (
  Blueprint,
  Response,
  abort,
  current_app,
  has_app_context,
  jsonify,
  render_template,
  request,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.billing.utils import (
  from_minor,
  is_compact_query,
  sql_ci_contains_any,
  to_compact,
  to_minor,
)
from app.utils.coercion import (
  coerce_decimal,
  coerce_float,
  coerce_int,
  load_json,
  parse_iso_datetime,
  zoneinfo_or_default,
)
from app.utils.error_logging import report_swallowed_exception
from app.utils.timezone import get_timezone_name

from ..db import (
  _actual_table_name,
  _ensure_column,
  _get_column_names,
  _is_postgres,
  get_db,
)


def _transaction_date_expr(conn) -> str:
  if _is_postgres(conn):
    return (
      "CASE "
      "WHEN trdate IS NOT NULL AND strpos(trdate,'-')=0 AND length(trdate)>=8 THEN substr(trdate,1,8) "
      "WHEN trdate IS NOT NULL AND strpos(trdate,'-')>0 AND length(trdate)>=10 THEN replace(substr(trdate,1,10),'-','') "
      "WHEN trdt IS NOT NULL AND strpos(trdt,'-')=0 AND length(trdt)>=8 THEN substr(trdt,1,8) "
      "WHEN trdt IS NOT NULL AND strpos(trdt,'-')>0 AND length(trdt)>=10 THEN replace(substr(trdt,1,10),'-','') "
      "ELSE NULL END"
    )
  return (
    "CASE "
    "WHEN trdate IS NOT NULL AND instr(trdate,'-')=0 AND length(trdate)>=8 THEN substr(trdate,1,8) "
    "WHEN trdate IS NOT NULL AND instr(trdate,'-')>0 AND length(trdate)>=10 THEN replace(substr(trdate,1,10),'-','') "
    "WHEN trdt IS NOT NULL AND instr(trdt,'-')=0 AND length(trdt)>=8 THEN substr(trdt,1,8) "
    "WHEN trdt IS NOT NULL AND instr(trdt,'-')>0 AND length(trdt)>=10 THEN replace(substr(trdt,1,10),'-','') "
    "ELSE NULL END"
  )


bp = Blueprint("bank_activity", __name__)
BANK_ACTIVITY_CURRENCY_OPTIONS = ("USD",)
ACCOUNT_INQUIRY_PROVIDER_OPTIONS = ("manual", "plaid")


# ---------- Helpers ----------


def _normalize_bank_activity_currency(value: str | None, default: str = "USD") -> str:
  raw = (value or default or "USD").strip().upper()
  if raw.startswith("C:"):
    raw = raw[2:].strip().upper()
  return raw if raw in BANK_ACTIVITY_CURRENCY_OPTIONS else default


def _bank_base_currency() -> str:
  try:
    configured = current_app.config.get("BANK_ACCOUNT_BASE_CURRENCY", "USD")
  except RuntimeError:
    configured = "USD"
  return _normalize_bank_activity_currency(str(configured or "USD"), default="USD")


def _normalize_account_provider(value: Any = None) -> str:
  try:
    configured = current_app.config.get("BANK_ACCOUNT_DATA_PROVIDER", "manual")
  except RuntimeError:
    configured = "manual"
  raw = str(value or configured or "manual").strip().lower()
  aliases = {
    "local": "manual",
    "manual_entry": "manual",
  }
  raw = aliases.get(raw, raw)
  return raw if raw in ACCOUNT_INQUIRY_PROVIDER_OPTIONS else "manual"


def _currency_template_context(selected: str | None = None) -> dict[str, Any]:
  currency = _normalize_bank_activity_currency(selected, default=_bank_base_currency())
  return {
    "selected_currency": currency,
    "currency_options": list(BANK_ACTIVITY_CURRENCY_OPTIONS),
    "bank_account_provider": _normalize_account_provider(),
  }


def _normalize_bank_activity_account_number(value: Any) -> str:
  return re.sub(r"\D+", "", str(value or ""))


def _normalize_bank_activity_bank_code(value: Any) -> str:
  raw = re.sub(r"\D+", "", str(value or ""))
  return raw.zfill(4) if raw else ""


def _infer_bank_activity_account_currency(
  *, bank_code: Any = None, account_number: Any = None, account_name: Any = None
) -> str:
  return "USD"

def _account_matches_currency(item: dict[str, Any], currency: str) -> bool:
  explicit = str(item.get("currency") or "").strip().upper()
  if explicit in BANK_ACTIVITY_CURRENCY_OPTIONS:
    return explicit == _normalize_bank_activity_currency(currency)
  inferred = _infer_bank_activity_account_currency(
    bank_code=item.get("bankCode") or item.get("bank_code"),
    account_number=item.get("accountNumber") or item.get("account_number"),
    account_name=item.get("accountName") or item.get("account_name"),
  )
  return inferred == _normalize_bank_activity_currency(currency)


def _now_in_tz():
  return datetime.now(_app_tz())


def _app_tz():
  return zoneinfo_or_default(get_timezone_name(), default="America/New_York")


def _datetime_format() -> str:
  if has_app_context():
    return current_app.config.get("DATETIME_FORMAT", "%m/%d/%Y %I:%M:%S %p")
  return "%Y-%m-%d %H:%M:%S"


def _format_kst_timestamp(value) -> str | None:
  if value in (None, ""):
    return None

  dt = None
  if isinstance(value, datetime):
    dt = value
  else:
    raw = str(value).strip()
    if not raw:
      return None
    dt = parse_iso_datetime(raw, None)
    if dt is None:
      try:
        dt = datetime.strptime(raw, "%Y-%m-%d %H:%M:%S")
      except ValueError:
        try:
          dt = parsedate_to_datetime(raw)
        except (TypeError, ValueError, IndexError):
          return raw

  if dt.tzinfo is None:
    dt = dt.replace(tzinfo=ZoneInfo("UTC"))
  fmt = _datetime_format()
  try:
    return dt.astimezone(_app_tz()).strftime(fmt)
  except ValueError:
    return dt.strftime(fmt)


def _add_months_clamped(base_date: date, months: int) -> date:
  """Add months while clamping to the last valid day of the target month."""
  month_index = (base_date.year * 12 + (base_date.month - 1)) + months
  year = month_index // 12
  month = (month_index % 12) + 1
  day = min(base_date.day, monthrange(year, month)[1])
  return date(year, month, day)



def _safe_int(value) -> int | None:
  return coerce_int(value)


def _safe_float(value, default: float = 0.0) -> float:
  parsed = coerce_float(value, default)
  return default if parsed is None else float(parsed)


def _clamp_page_for_total(page: int, total: int, per_page: int) -> tuple[int, int]:
  page_count = max(1, (int(total or 0) + int(per_page or 1) - 1) // int(per_page or 1))
  return min(max(int(page or 1), 1), page_count), page_count


def _parse_minor_amount(value: Any, default: int = 0) -> int:
  normalized = str(value or "").replace(",", "").replace(" ", "")
  parsed = coerce_int(normalized, default)
  return default if parsed is None else int(parsed)


_BANK_ACTIVITY_TRANSACTION_PROVIDER_COLUMNS = (
  ("account_name", "TEXT"),
  ("currency", "TEXT"),
  ("source_provider", "TEXT DEFAULT 'manual'"),
  ("external_id", "TEXT"),
)


def _ensure_bank_activity_transaction_provider_columns(conn) -> None:
  for column, column_type in _BANK_ACTIVITY_TRANSACTION_PROVIDER_COLUMNS:
    try:
      _ensure_column(conn, "bank_transactions", column, column_type)
    except Exception:
      continue


def _parse_bank_amount_to_storage(value: Any, *, currency: str, field_name: str) -> int:
  raw = str(value or "").strip().replace(",", "").replace("$", "")
  if raw == "":
    return 0
  try:
    amount = Decimal(raw)
  except (InvalidOperation, ValueError):
    abort(400, f"{field_name} must be a number")
  if amount < 0:
    abort(400, f"{field_name} must not be negative")
  try:
    return int(to_minor(amount, currency) or 0)
  except Exception:
    abort(400, f"{field_name} must be a valid {currency} amount")


def _storage_amount_to_display(value: Any, currency: str | None = None) -> int | float:
  cur = _normalize_bank_activity_currency(currency, default=_bank_base_currency())
  try:
    minor = int(value or 0)
  except (TypeError, ValueError):
    minor = 0
  amount = from_minor(minor, cur)
  if amount == amount.to_integral_value():
    return int(amount)
  return float(amount)


def _normalize_transaction_date(value: Any) -> str:
  raw = str(value or "").strip()
  if re.fullmatch(r"\d{8}", raw):
    try:
      datetime.strptime(raw, "%Y%m%d")
      return raw
    except ValueError:
      abort(400, "transactionDate must be a valid date")
  if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
    try:
      return datetime.strptime(raw, "%Y-%m-%d").strftime("%Y%m%d")
    except ValueError:
      abort(400, "transactionDate must be a valid date")
  abort(400, "transactionDate must be yyyy-mm-dd or yyyyMMdd")


def _manual_tid(record: dict[str, Any]) -> str:
  seed = "|".join(
    str(record.get(key) or "")
    for key in (
      "currency",
      "bank_code",
      "account_number",
      "trdt",
      "acc_in",
      "acc_out",
      "balance",
      "remark1",
      "remark2",
    )
  )
  digest = hashlib.sha1(seed.encode("utf-8", errors="ignore")).hexdigest()[:12]
  stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
  return f"manual:{stamp}:{digest}"


def _upsert_bank_activity_transaction_record(conn, record: dict[str, Any]) -> None:
  _ensure_bank_activity_transaction_provider_columns(conn)
  cols_available = _get_column_names(conn, "bank_transactions")
  preferred_cols = [
    "tid",
    "corp_num",
    "bank_code",
    "account_number",
    "account_name",
    "currency",
    "source_provider",
    "external_id",
    "trdate",
    "trdt",
    "trserial",
    "acc_in",
    "acc_out",
    "balance",
    "remark1",
    "remark2",
    "remark3",
    "memo",
    "reg_dt",
    "job_id",
  ]
  insert_cols = [col for col in preferred_cols if col in cols_available and col in record]
  if "tid" not in insert_cols:
    abort(500, "bank_transactions.tid column is missing")
  tbl_tx = _actual_table_name("bank_transactions")
  placeholders = ",".join(["?"] * len(insert_cols))
  columns_sql = ",".join(insert_cols)
  update_cols = [col for col in insert_cols if col != "tid"]
  set_sql = ", ".join(f"{col}=excluded.{col}" for col in update_cols)
  if "updated_at" in cols_available:
    set_sql = f"{set_sql}, updated_at=CURRENT_TIMESTAMP" if set_sql else "updated_at=CURRENT_TIMESTAMP"
  created_cols = ""
  created_vals = ""
  if "created_at" in cols_available:
    created_cols += ",created_at"
    created_vals += ",CURRENT_TIMESTAMP"
  if "updated_at" in cols_available:
    created_cols += ",updated_at"
    created_vals += ",CURRENT_TIMESTAMP"
  sql = (
    f"INSERT INTO {tbl_tx} ({columns_sql}{created_cols}) "
    f"VALUES ({placeholders}{created_vals}) "
    f"ON CONFLICT(tid) DO UPDATE SET {set_sql}"
  )
  conn.execute(sql, [record.get(col) for col in insert_cols])


def _local_account_items(conn, *, currency: str = "", where_sql: str = "", params=None) -> list[dict[str, Any]]:
  _ensure_bank_activity_transaction_provider_columns(conn)
  params = list(params or [])
  cols = _get_column_names(conn, "bank_transactions")
  currency_expr = "MAX(UPPER(COALESCE(currency,'')))" if "currency" in cols else "NULL"
  account_name_expr = "MAX(account_name)" if "account_name" in cols else "NULL"
  source_expr = "MAX(source_provider)" if "source_provider" in cols else "NULL"
  prefix = " WHERE " if not where_sql else f"{where_sql} AND "
  rows = conn.execute(
    f"""
    SELECT bank_code, account_number, {currency_expr}, {account_name_expr}, {source_expr}
    FROM bank_transactions
    {prefix}bank_code IS NOT NULL AND account_number IS NOT NULL
    GROUP BY bank_code, account_number
    ORDER BY bank_code, account_number
    """,
    params,
  ).fetchall()
  out: list[dict[str, Any]] = []
  for row in rows:
    item = {
      "bankCode": row[0] or "",
      "accountNumber": row[1] or "",
      "currency": (
        str(row[2] or "").strip().upper()
        or _infer_bank_activity_account_currency(bank_code=row[0], account_number=row[1])
      ),
      "accountName": row[3] or "",
      "sourceProvider": row[4] or "manual",
      "state": 1,
    }
    if not (item["bankCode"] and item["accountNumber"]):
      continue
    if currency and not _account_matches_currency(item, currency):
      continue
    out.append(item)
  return out


def _plaid_base_url() -> str:
  env = str(current_app.config.get("PLAID_ENV") or "sandbox").strip().lower()
  if env == "production":
    return "https://production.plaid.com"
  if env == "development":
    return "https://development.plaid.com"
  return "https://sandbox.plaid.com"


def _plaid_access_token() -> str:
  return str(current_app.config.get("PLAID_ACCESS_TOKEN") or "").strip()


def _plaid_post(path: str, payload: dict[str, Any]) -> dict[str, Any]:
  client_id = str(current_app.config.get("PLAID_CLIENT_ID") or "").strip()
  secret = str(current_app.config.get("PLAID_SECRET") or "").strip()
  access_token = _plaid_access_token()
  if not (client_id and secret and access_token):
    abort(400, "Plaid client_id, secret, and access token are required")
  body = dict(payload)
  body.update({"client_id": client_id, "secret": secret, "access_token": access_token})
  try:
    response = requests.post(f"{_plaid_base_url()}{path}", json=body, timeout=30)
  except requests.RequestException as exc:
    return {"error": True, "message": f"Plaid connection error: {type(exc).__name__}"}
  try:
    data = response.json()
  except ValueError:
    data = {"error": True, "message": response.text}
  if response.status_code >= 400:
    data.setdefault("error", True)
    data.setdefault("message", data.get("error_message") or response.reason)
  return data


def _plaid_account_items(currency: str = "") -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
  data = _plaid_post("/accounts/get", {})
  if data.get("error"):
    return [], data
  configured_ids = {
    item.strip()
    for item in str(current_app.config.get("PLAID_ACCOUNT_IDS") or "").split(",")
    if item.strip()
  }
  items: list[dict[str, Any]] = []
  for account in data.get("accounts") or []:
    account_id = str(account.get("account_id") or "").strip()
    if configured_ids and account_id not in configured_ids:
      continue
    balances = account.get("balances") or {}
    cur = str(balances.get("iso_currency_code") or _bank_base_currency()).upper()
    item = {
      "accountNumber": account_id,
      "bankCode": "PLAID",
      "accountName": account.get("name") or account.get("official_name") or account_id,
      "accountType": account.get("type"),
      "state": 1,
      "currency": _normalize_bank_activity_currency(cur, default=_bank_base_currency()),
      "sourceProvider": "plaid",
      "baseDate": _now_in_tz().strftime("%Y%m%d"),
    }
    if currency and not _account_matches_currency(item, currency):
      continue
    items.append(item)
  return items, None


def _sync_plaid_transactions(start_date: str, end_date: str, currency: str) -> dict[str, Any]:
  start_iso = datetime.strptime(start_date, "%Y%m%d").strftime("%Y-%m-%d")
  end_iso = datetime.strptime(end_date, "%Y%m%d").strftime("%Y-%m-%d")
  data = _plaid_post(
    "/transactions/get",
    {
      "start_date": start_iso,
      "end_date": end_iso,
      "options": {"count": 500, "offset": 0},
    },
  )
  if data.get("error"):
    return {"ok": False, "error": data.get("message") or data.get("error_message") or "Plaid error"}

  accounts_by_id = {a.get("account_id"): a for a in data.get("accounts") or []}
  conn = get_db()
  saved = 0
  skipped = 0
  try:
    for tx in data.get("transactions") or []:
      tx_currency = str(tx.get("iso_currency_code") or currency or _bank_base_currency()).upper()
      tx_currency = _normalize_bank_activity_currency(tx_currency, default=_bank_base_currency())
      if currency and tx_currency != _normalize_bank_activity_currency(currency, default=_bank_base_currency()):
        skipped += 1
        continue
      account_id = str(tx.get("account_id") or "").strip()
      account = accounts_by_id.get(account_id) or {}
      amount = Decimal(str(tx.get("amount") or "0"))
      units = int(to_minor(amount.copy_abs(), tx_currency) or 0)
      record = {
        "tid": f"plaid:{tx.get('transaction_id')}",
        "bank_code": "PLAID",
        "account_number": account_id,
        "account_name": account.get("name") or account.get("official_name") or "",
        "currency": tx_currency,
        "source_provider": "plaid",
        "external_id": tx.get("transaction_id"),
        "trdate": str(tx.get("date") or "").replace("-", ""),
        "trdt": str(tx.get("datetime") or tx.get("date") or "").replace("-", "").replace(":", "").replace("T", "").replace("Z", "")[:14],
        "trserial": tx.get("transaction_id"),
        "acc_in": units if amount < 0 else 0,
        "acc_out": units if amount > 0 else 0,
        "balance": 0,
        "remark1": tx.get("merchant_name") or tx.get("name"),
        "remark2": tx.get("payment_channel"),
        "remark3": f"plaid_amount={amount} {tx_currency}",
        "memo": "",
        "reg_dt": _now_in_tz().isoformat(),
        "job_id": "plaid",
      }
      if not record["trdate"]:
        skipped += 1
        continue
      _upsert_bank_activity_transaction_record(conn, record)
      saved += 1
    conn.commit()
  finally:
    conn.close()
  return {"ok": True, "provider": "plaid", "rowsPersisted": saved, "skipped": skipped}


_COMPANY_TOKENS = (
  "Company",
  "Company",
  "Company",
  "peopleCompany",
  "",
  "",
  "Company",
  "",
  "",
  "()",
  "㈜",
  "()",
  "()",
  "()",
  "corp",
  "corporation",
  "inc",
  "co",
  "ltd",
  "limited",
  "llc",
)
_REC_SCORE_THRESHOLD = 48
_REC_MIN_AMOUNT_SCORE = 12
_CLIENT_PAYER_HISTORY_CACHE_TTL_SECONDS = 300
_CLIENT_PAYER_HISTORY_CACHE_MAX_CLIENTS = 512
_CLIENT_PAYER_HISTORY_CACHE: dict[tuple[str, int], tuple[float, list[dict[str, Any]]]] = {}
_CLIENT_PAYER_HISTORY_CACHE_LOCK = threading.Lock()


def _safe_normalize_name(value: str | None) -> str:
  raw = str(value or "")
  raw = unicodedata.normalize("NFKC", raw)
  return re.sub(r"\s+", " ", re.sub(r"[|/\\,.;:_-]+", " ", raw.lower())).strip()


def _remove_company_tokens(value: str | None) -> str:
  out = _safe_normalize_name(value)
  for token in _COMPANY_TOKENS:
    cleaned_token = _safe_normalize_name(token)
    if cleaned_token:
      out = out.replace(cleaned_token, " ")
  out = re.sub(r"\b(?:co|corp|inc|ltd|llc)\b", " ", out)
  return re.sub(r"\s+", " ", out).strip()


def _compact_entity_name(value: str | None) -> str:
  return re.sub(r"[^0-9a-z-]+", "", _remove_company_tokens(value))


def _build_name_variants(value: str | None) -> list[str]:
  raw = _safe_normalize_name(value)
  if not raw:
    return []
  before_paren = raw.split("(", 1)[0].strip()
  no_paren_chars = re.sub(r"[()\[\]{}]", " ", raw)
  candidates = {
    raw,
    before_paren,
    re.sub(r"\s+", " ", no_paren_chars).strip(),
    _remove_company_tokens(raw),
    _remove_company_tokens(before_paren),
    _compact_entity_name(raw),
    _compact_entity_name(before_paren),
  }
  return [item for item in candidates if item]


def _build_name_token_set(value: str | None) -> set[str]:
  out: set[str] = set()
  for variant in _build_name_variants(value):
    cleaned = _remove_company_tokens(variant)
    for part in cleaned.split():
      token = re.sub(r"[^0-9a-z-]+", "", part).strip()
      if len(token) >= 2:
        out.add(token)
    compact = _compact_entity_name(variant)
    if len(compact) >= 2:
      out.add(compact)
  return out


def _bigram_set(value: str | None) -> set[str]:
  compact = _compact_entity_name(value)
  if not compact:
    return set()
  if len(compact) < 2:
    return {compact}
  return {compact[i : i + 2] for i in range(len(compact) - 1)}


def _bigram_similarity(a: str | None, b: str | None) -> float:
  sa = _bigram_set(a)
  sb = _bigram_set(b)
  if not sa or not sb:
    return 0.0
  return len(sa & sb) / max(len(sa), len(sb))


def _token_overlap_ratio(a: str | None, b: str | None) -> float:
  sa = _build_name_token_set(a)
  sb = _build_name_token_set(b)
  if not sa or not sb:
    return 0.0
  return len(sa & sb) / min(len(sa), len(sb))


def _matching_name_score(client_name: str | None, payer_name: str | None) -> tuple[int, str]:
  client_core = _compact_entity_name(client_name)
  payer_core = _compact_entity_name(payer_name)
  if not client_core or not payer_core:
    return 0, ""
  if client_core == payer_core:
    return 35, "Client name match"
  shorter = min(len(client_core), len(payer_core))
  if shorter >= 3 and (client_core in payer_core or payer_core in client_core):
    return 28, "Client name match"
  overlap = _token_overlap_ratio(client_name, payer_name)
  if overlap >= 1:
    return 24, "Client name token match"
  if overlap >= 0.5:
    return 18, "Client name partial token match"
  client_cho = re.sub(r"\s+", "", to_compact(client_core))
  payer_cho = re.sub(r"\s+", "", to_compact(payer_core))
  if client_cho and payer_cho and (client_cho in payer_cho or payer_cho in client_cho):
    return 16, "Client name initial match"
  similarity = _bigram_similarity(client_core, payer_core)
  if similarity >= 0.72:
    return 16, "Client name fuzzy match"
  if similarity >= 0.56:
    return 10, "Client name weak fuzzy match"
  return 0, ""


def _matching_number(value: Any) -> float:
  try:
    raw = str(value or "0").replace(",", "").replace(" ", "")
    return float(raw or 0)
  except (TypeError, ValueError):
    return 0.0


def _matching_amount_score(
  target_amount: int | float | None,
  deposit_amount: int | float | None,
  currency: str | None = "USD",
) -> tuple[int, str]:
  target = _matching_number(target_amount)
  deposit = _matching_number(deposit_amount)
  if not target or not deposit:
    return 0, ""
  diff = abs(target - deposit)
  pct = (diff / target) if target > 0 else 999.0
  cur = _normalize_bank_activity_currency(currency)
  if cur == "USD":
    near_abs, similar_abs, tolerated_abs = 10000, 50000, 100000
  else:
    near_abs, similar_abs, tolerated_abs = 1, 10, 50
  if diff == 0:
    return 45, "Amount match"
  if pct <= 0.01 or (diff <= near_abs and pct <= 0.05):
    return 40, "Amount near match"
  if pct <= 0.05 or (diff <= similar_abs and pct <= 0.1):
    return 32, "Amount similar match"
  if pct <= 0.1 or (diff <= tolerated_abs and pct <= 0.2):
    return 22, "Amount tolerance match"
  ratio = (min(target, deposit) / target) if target > 0 else 0.0
  if deposit < target and ratio >= 0.4:
    return 16, "Partial deposit amount"
  if deposit > target and pct <= 0.2:
    return 12, "Deposit amount slightly exceeds invoice"
  return 0, ""


def _amount_for_matching(total_minor: int | float | None, currency: str | None) -> int | float:
  cur = _normalize_bank_activity_currency(currency)
  try:
    minor_int = int(total_minor or 0)
  except (TypeError, ValueError):
    minor_int = 0
  if cur == "USD":
    return minor_int
  amount = from_minor(minor_int, cur)
  if amount == amount.to_integral_value():
    return int(amount)
  return float(amount)


def _payment_meta_amount_value(value: Any, currency: str | None) -> int | float:
  cur = _normalize_bank_activity_currency(currency)
  if cur == "USD":
    return _parse_minor_amount(value, 0)
  amount = _matching_number(value)
  return int(amount) if float(amount).is_integer() else amount


def _parse_matching_issue_date(value: str | None) -> datetime | None:
  raw = (value or "").strip()
  if not raw:
    return None
  if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
    try:
      return datetime.strptime(raw, "%Y-%m-%d")
    except ValueError:
      return None
  parsed = parse_iso_datetime(raw, None)
  return parsed.replace(tzinfo=None) if parsed else None


def _parse_matching_trdt(value: str | None) -> datetime | None:
  raw = (value or "").strip()
  if not raw:
    return None
  if re.fullmatch(r"\d{14}", raw):
    try:
      return datetime.strptime(raw, "%Y%m%d%H%M%S")
    except ValueError:
      return None
  if re.fullmatch(r"\d{8}", raw):
    try:
      return datetime.strptime(raw, "%Y%m%d")
    except ValueError:
      return None
  parsed = parse_iso_datetime(raw, None)
  return parsed.replace(tzinfo=None) if parsed else None


def _matching_date_score(issue_date: str | None, deposit_date: str | None) -> tuple[int, str]:
  inv_dt = _parse_matching_issue_date(issue_date)
  dep_dt = _parse_matching_trdt(deposit_date)
  if inv_dt is None or dep_dt is None:
    return 0, ""
  diff = (dep_dt.date() - inv_dt.date()).days
  if 0 <= diff <= 3:
    return 20, "Issued within 3 days"
  if 4 <= diff <= 7:
    return 16, "Issued within 1 week"
  if 8 <= diff <= 15:
    return 12, "Issued within 2 weeks"
  if 16 <= diff <= 30:
    return 8, "Issued within 30 days"
  if 31 <= diff <= 60:
    return 4, "Issued within 60 days"
  if -7 <= diff < 0:
    return 10, "Deposit before invoice issue date"
  if -30 <= diff < -7:
    return 4, "Deposit within 30 days before invoice"
  return 0, ""


def _safe_json_value(value: Any) -> Any:
  if value is None:
    return None
  if isinstance(value, (dict, list)):
    return value
  raw = str(value).strip()
  if not raw:
    return None
  return load_json(raw, None)


def _payment_meta_deposit_records(value: Any) -> list[dict[str, Any]]:
  parsed = _safe_json_value(value)
  if not isinstance(parsed, dict):
    return []
  deposits = parsed.get("deposits")
  if isinstance(deposits, list):
    return [item for item in deposits if isinstance(item, dict)]
  if any(parsed.get(key) not in (None, "") for key in ("tid", "summary", "deposit")):
    return [parsed]
  return []


def _payment_meta_amount_minor(value: Any, currency: str | None) -> int:
  cur = _normalize_bank_activity_currency(currency)
  parsed = coerce_decimal(str(value or "").replace(",", ""), None)
  if parsed is None:
    return 0
  return int(to_minor(parsed, cur) or 0)


def _not_deleted_sql(column: str = "is_deleted") -> str:
  return (
    f"COALESCE(LOWER(CAST({column} AS TEXT)), 'false') "
    "NOT IN ('1', 'true', 't', 'yes', 'y')"
  )


def _payment_tid_allocated_minor(conn, tid: str | None, currency: str | None) -> int:
  key = str(tid or "").strip()
  if not key:
    return 0
  like_a = f'%"tid": "{key}"%'
  like_b = f'%"tid":"{key}"%'
  try:
    rows = conn.execute(
      f"""
      SELECT currency, payment_meta
      FROM invoices
      WHERE (payment_meta LIKE ? OR payment_meta LIKE ?)
       AND {_not_deleted_sql()}
       AND COALESCE(billing_status, '') != 'void'
      """,
      (like_a, like_b),
    ).fetchall()
  except Exception:
    return 0

  total_minor = 0
  for row in rows or []:
    try:
      row_currency = row[0] or currency
      raw_meta = row[1]
    except Exception:
      row_currency = currency
      raw_meta = None
    for record in _payment_meta_deposit_records(raw_meta):
      if str(record.get("tid") or "").strip() != key:
        continue
      total_minor += _payment_meta_amount_minor(
        record.get("deposit"),
        record.get("currency") or row_currency or currency,
      )
  return int(total_minor)


def _annotate_deposit_allocations(
  conn,
  items: list[dict[str, Any]],
  currency: str | None,
) -> list[dict[str, Any]]:
  fallback_currency = _normalize_bank_activity_currency(currency) if currency else ""
  for item in items or []:
    cur = fallback_currency or _infer_bank_activity_account_currency(
      bank_code=item.get("bankCode") or item.get("bank_code"),
      account_number=item.get("accountNumber") or item.get("account_number"),
    )
    raw_tx_minor = item.get("_acc_in_minor")
    if raw_tx_minor is None:
      tx_minor = _payment_meta_amount_minor(item.get("accIn"), cur)
    else:
      try:
        tx_minor = int(raw_tx_minor or 0)
      except (TypeError, ValueError):
        tx_minor = 0
    allocated_minor = _payment_tid_allocated_minor(conn, item.get("tid"), cur)
    remaining_minor = max(0, int(tx_minor) - int(allocated_minor))
    item["matched_amount"] = _storage_amount_to_display(allocated_minor, cur)
    item["remainingAccIn"] = _storage_amount_to_display(remaining_minor, cur)
    if tx_minor > 0 and allocated_minor >= tx_minor:
      item["match_state"] = "full"
    elif allocated_minor > 0:
      item["match_state"] = "partial"
    else:
      item["match_state"] = "unmatched"
    item.pop("_acc_in_minor", None)
  return items


def _summary_primary_payer_name(value: str | None) -> str:
  raw = str(value or "").strip()
  if not raw:
    return ""
  first_chunk = raw.split("+", 1)[0].strip()
  parts = [part.strip() for part in first_chunk.split("|") if part.strip()]
  return parts[0] if parts else first_chunk


def _history_bucket_key(value: str | None) -> str:
  compact = _compact_entity_name(value)
  if compact:
    return compact
  simplified = _remove_company_tokens(value)
  if simplified:
    return simplified
  return _safe_normalize_name(value)


def _history_payer_name_from_record(
  record: dict[str, Any], tx_lookup: dict[str, dict[str, Any]]
) -> str:
  tid = str(record.get("tid") or "").strip()
  if tid:
    tx = tx_lookup.get(tid) or {}
    payer_name = str(tx.get("payer_name") or "").strip()
    if payer_name:
      return payer_name
  return _summary_primary_payer_name(record.get("summary"))


def invalidate_client_payer_history_cache(client_ids: list[int | str | None] | None = None) -> None:
  with _CLIENT_PAYER_HISTORY_CACHE_LOCK:
    if client_ids is None:
      _CLIENT_PAYER_HISTORY_CACHE.clear()
      return
    drop_keys: list[tuple[str, int]] = []
    for raw in client_ids:
      client_id = coerce_int(raw, 0) or 0
      if client_id > 0:
        drop_keys.extend(
          key for key in _CLIENT_PAYER_HISTORY_CACHE.keys() if key[1] == client_id
        )
    for key in drop_keys:
      _CLIENT_PAYER_HISTORY_CACHE.pop(key, None)


def _get_cached_client_payer_history(
  client_ids: list[int],
  currency: str = "USD",
) -> tuple[dict[int, list[dict[str, Any]]], list[int]]:
  now = time.monotonic()
  cur = _normalize_bank_activity_currency(currency)
  cached: dict[int, list[dict[str, Any]]] = {}
  missing: list[int] = []
  expired_keys: list[tuple[str, int]] = []
  with _CLIENT_PAYER_HISTORY_CACHE_LOCK:
    for client_id in client_ids:
      cache_key = (cur, client_id)
      entry = _CLIENT_PAYER_HISTORY_CACHE.get(cache_key)
      if not entry:
        missing.append(client_id)
        continue
      loaded_at, items = entry
      if now - loaded_at > _CLIENT_PAYER_HISTORY_CACHE_TTL_SECONDS:
        expired_keys.append(cache_key)
        missing.append(client_id)
        continue
      cached[client_id] = items
    for cache_key in expired_keys:
      _CLIENT_PAYER_HISTORY_CACHE.pop(cache_key, None)
  return cached, missing


def _set_cached_client_payer_history(
  history_by_client: dict[int, list[dict[str, Any]]],
  currency: str = "USD",
) -> None:
  if not history_by_client:
    return
  now = time.monotonic()
  cur = _normalize_bank_activity_currency(currency)
  with _CLIENT_PAYER_HISTORY_CACHE_LOCK:
    for client_id, items in history_by_client.items():
      _CLIENT_PAYER_HISTORY_CACHE[(cur, int(client_id))] = (now, items)
    overflow = len(_CLIENT_PAYER_HISTORY_CACHE) - _CLIENT_PAYER_HISTORY_CACHE_MAX_CLIENTS
    if overflow > 0:
      oldest = sorted(
        _CLIENT_PAYER_HISTORY_CACHE.items(),
        key=lambda item: item[1][0],
      )[:overflow]
      for client_id, _entry in oldest:
        _CLIENT_PAYER_HISTORY_CACHE.pop(client_id, None)


def _load_client_payer_history_index(
  conn,
  client_ids: list[int | str | None],
  currency: str = "USD",
) -> dict[int, list[dict[str, Any]]]:
  cur = _normalize_bank_activity_currency(currency)
  normalized_ids: list[int] = []
  for raw in client_ids:
    client_id = coerce_int(raw, 0) or 0
    if client_id > 0 and client_id not in normalized_ids:
      normalized_ids.append(client_id)
  if not normalized_ids:
    return {}

  cached, missing_ids = _get_cached_client_payer_history(normalized_ids, cur)
  if not missing_ids:
    return cached

  qmarks = ",".join(["?"] * len(missing_ids))
  rows = conn.execute(
    f"""
    SELECT client_id, payment_meta
    FROM invoices
    WHERE UPPER(COALESCE(currency, 'USD')) = ?
     AND payment_meta IS NOT NULL
     AND {_not_deleted_sql()}
     AND client_id IN ({qmarks})
    """,
    [cur] + missing_ids,
  ).fetchall()

  history_records: list[tuple[int, dict[str, Any]]] = []
  tids: list[str] = []
  for row in rows:
    client_id = coerce_int(row[0], 0) or 0
    if client_id <= 0:
      continue
    for record in _payment_meta_deposit_records(row[1]):
      history_records.append((client_id, record))
      tid = str(record.get("tid") or "").strip()
      if tid and tid not in tids:
        tids.append(tid)

  tx_lookup: dict[str, dict[str, Any]] = {}
  if tids:
    qmarks = ",".join(["?"] * len(tids))
    tx_rows = conn.execute(
      f"""
      SELECT tid, remark1, remark2, remark3
      FROM bank_transactions
      WHERE tid IN ({qmarks})
      """,
      tids,
    ).fetchall()
    for row in tx_rows:
      payer_name = row[1] or " | ".join([str(x) for x in (row[1], row[2], row[3]) if x]) or ""
      tx_lookup[str(row[0])] = {"payer_name": payer_name}

  history_index: dict[int, dict[str, dict[str, Any]]] = {
    client_id: {} for client_id in missing_ids
  }
  for client_id, record in history_records:
    payer_name = _history_payer_name_from_record(record, tx_lookup)
    key = _history_bucket_key(payer_name)
    if not key:
      continue
    by_name = history_index.setdefault(client_id, {})
    item = by_name.get(key)
    if item is None:
      by_name[key] = {"payer_name": payer_name, "count": 1}
      continue
    item["count"] = int(item.get("count") or 0) + 1
    current_name = str(item.get("payer_name") or "")
    if len(_compact_entity_name(payer_name)) > len(_compact_entity_name(current_name)):
      item["payer_name"] = payer_name

  fetched = {
    client_id: sorted(
      values.values(),
      key=lambda item: (int(item.get("count") or 0), str(item.get("payer_name") or "")),
      reverse=True,
    )
    for client_id, values in history_index.items()
  }
  _set_cached_client_payer_history(fetched, cur)
  merged = dict(cached)
  merged.update(fetched)
  return merged


def _matching_history_score(
  client_id: int | str | None,
  payer_name: str | None,
  history_by_client: dict[int, list[dict[str, Any]]] | None,
) -> tuple[int, str]:
  if not history_by_client:
    return 0, ""
  client_id_int = coerce_int(client_id, 0) or 0
  if client_id_int <= 0:
    return 0, ""
  current_payer_name = str(payer_name or "").strip()
  if not current_payer_name:
    return 0, ""

  best_bonus = 0
  best_count = 0
  for item in history_by_client.get(client_id_int) or []:
    name_score, _ = _matching_name_score(item.get("payer_name"), current_payer_name)
    if name_score >= 28:
      bonus = 18
    elif name_score >= 18:
      bonus = 12
    elif name_score >= 10:
      bonus = 8
    else:
      continue
    count = coerce_int(item.get("count"), 0) or 0
    if count > 1:
      bonus += min(6, (count - 1) * 2)
    if bonus > best_bonus or (bonus == best_bonus and count > best_count):
      best_bonus = bonus
      best_count = count
  if best_bonus <= 0:
    return 0, ""
  if best_count >= 2:
    return best_bonus, f"Historical payer match ({best_count} times)"
  return best_bonus, "Historical payer match"


def _deposit_amount_for_recommendation(deposit: dict[str, Any]) -> float:
  """Use unallocated deposit balance when a bank transaction was split across invoices."""
  state = str(deposit.get("match_state") or deposit.get("matchState") or "").strip().lower()
  matched = _matching_number(deposit.get("matched_amount") or deposit.get("matchedAmount"))
  if "remainingAccIn" in deposit and (state in ("unmatched", "partial", "full") or matched > 0):
    return _matching_number(deposit.get("remainingAccIn"))
  return _matching_number(deposit.get("accIn"))


def _score_invoice_deposit_recommendation(
  *,
  invoice: dict[str, Any],
  deposit: dict[str, Any],
  history_by_client: dict[int, list[dict[str, Any]]] | None = None,
) -> dict[str, Any]:
  target_amount = _matching_number(invoice.get("remaining")) or _matching_number(
    invoice.get("total")
  )
  amount_score, amount_reason = _matching_amount_score(
    target_amount,
    _deposit_amount_for_recommendation(deposit),
    invoice.get("currency"),
  )
  date_score, date_reason = _matching_date_score(
    invoice.get("issue_date"),
    deposit.get("trdt") or deposit.get("trdate"),
  )
  name_score, name_reason = _matching_name_score(
    invoice.get("client_name"),
    deposit.get("payer_name"),
  )
  history_score, history_reason = _matching_history_score(
    invoice.get("client_id"),
    deposit.get("payer_name"),
    history_by_client,
  )
  total = amount_score + date_score + name_score + history_score
  reasons = [
    reason for reason in (amount_reason, date_reason, name_reason, history_reason) if reason
  ]
  return {
    "score": total,
    "reasons": reasons,
    "recommended": total >= _REC_SCORE_THRESHOLD and amount_score >= _REC_MIN_AMOUNT_SCORE,
  }


def _invoice_item_from_matching_row(row) -> dict[str, Any]:
  saved_minor = int(row[3] or 0)
  currency = row[4] or "USD"
  vat_rate = _safe_float(row[9], 0.0)
  service_total = _safe_float(row[10], 0.0)
  admin_total = _safe_float(row[11], 0.0)
  foreign_total = _safe_float(row[12], 0.0)
  foreign_taxable_total = _safe_float(row[13], 0.0)
  vat_multiplier = (vat_rate / 100.0) if vat_rate > 1 else vat_rate
  tax_dynamic = vat_multiplier * (service_total + foreign_taxable_total)
  total_display = service_total + admin_total + foreign_total + tax_dynamic
  decimal_total = coerce_decimal(total_display, None)
  computed_minor = (
    to_minor(decimal_total, currency)
    if decimal_total is not None
    else int(round(total_display))
  )
  total_minor_for_ui = computed_minor if computed_minor > 0 else saved_minor
  total_amount = _storage_amount_to_display(total_minor_for_ui, currency)

  paid_sum: int | float = 0
  deposit_count = 0
  payment_meta_raw = row[14] if len(row) > 14 else None
  payment_meta = load_json(payment_meta_raw, None)
  if isinstance(payment_meta, dict):
    deposits = payment_meta.get("deposits")
    if isinstance(deposits, list):
      paid_sum = sum(
        _payment_meta_amount_value((d or {}).get("deposit"), currency)
        for d in deposits
        if isinstance(d, dict)
      )
      deposit_count = len(deposits)
    elif payment_meta.get("deposit") is not None:
      paid_sum = _payment_meta_amount_value(payment_meta.get("deposit"), currency)
      deposit_count = 1
  remaining = _matching_number(total_amount) - _matching_number(paid_sum)
  if abs(remaining) < 0.000001:
    remaining = 0
  if _matching_number(paid_sum) <= 0:
    match_state = "unmatched"
  elif remaining == 0:
    match_state = "full"
  elif remaining > 0:
    match_state = "partial"
  else:
    match_state = "over"

  return {
    "id": row[0],
    "number": row[1],
    "issue_date": row[2],
    "total": total_amount,
    "currency": currency,
    "status": row[5],
    "client_name": row[6],
    "biz_name": row[7],
    "biz_tax_id": row[8],
    "paid_sum": paid_sum,
    "remaining": remaining,
    "deposit_count": deposit_count,
    "match_state": match_state,
    "client_id": row[15],
    "ipm_case_id": row[16],
    "ipm_case_ref": row[17],
    "billing_status": row[18],
    "payment_status": row[19],
  }


def _deposit_item_from_matching_row(row, currency: str | None = None) -> dict[str, Any]:
  payer_name = row[7] or " | ".join([str(x) for x in (row[7], row[8], row[9]) if x]) or ""
  try:
    acc_in_minor = int(row[3] or 0)
  except (TypeError, ValueError):
    acc_in_minor = 0
  return {
    "tid": row[0],
    "trdt": row[1],
    "trdate": row[2],
    "accIn": _storage_amount_to_display(acc_in_minor, currency),
    "balance": _storage_amount_to_display(row[4], currency),
    "bankCode": row[5],
    "accountNumber": row[6],
    "remark1": row[7],
    "remark2": row[8],
    "remark3": row[9],
    "memo": row[10],
    "payer_name": payer_name,
    "_acc_in_minor": acc_in_minor,
  }


def _load_recommend_deposit_context(
  conn, tid: str | None, currency: str | None = None
) -> dict[str, Any] | None:
  key = str(tid or "").strip()
  if not key:
    return None
  row = conn.execute(
    """
    SELECT tid, trdt, trdate, acc_in, balance, bank_code, account_number, remark1, remark2, remark3, memo
    FROM bank_transactions
    WHERE tid = ?
    LIMIT 1
    """,
    [key],
  ).fetchone()
  return _deposit_item_from_matching_row(row, currency) if row else None


def _load_recommend_invoice_context(conn, invoice_id: int | None) -> dict[str, Any] | None:
  try:
    invoice_id_int = int(invoice_id or 0)
  except Exception:
    return None
  if invoice_id_int <= 0:
    return None
  row = conn.execute(
    """
    SELECT invoices.id, invoices.number, invoices.issue_date, invoices.total_minor, invoices.currency, invoices.status,
        clients.name as client_name, bp.name as biz_name, bp.tax_id as biz_tax_id,
        invoices.vat_rate,
        (SELECT COALESCE(SUM(li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0)), 0)
         FROM line_items li
         WHERE li.invoice_id = invoices.id AND li.item_type = 'service'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)) as service_total,
        (SELECT COALESCE(SUM(li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0)), 0)
         FROM line_items li
         WHERE li.invoice_id = invoices.id AND li.item_type = 'admin'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)) as admin_total,
        (SELECT COALESCE(SUM(
             CASE WHEN COALESCE(li.fx_rate_used, 0) > 0 THEN
               (COALESCE(li.fx_fee,0) + COALESCE(li.fx_gov,0))
               * COALESCE(li.fx_rate_used, 0)
               * (1 + COALESCE(li.fx_markup,0)/100.0)
             ELSE
               (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
             END
           ), 0)
         FROM line_items li
         WHERE li.invoice_id = invoices.id AND li.item_type = 'foreign'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)) as foreign_total,
        (SELECT COALESCE(SUM(
             CASE WHEN COALESCE(li.fx_rate_used, 0) > 0 THEN
               (COALESCE(li.fx_fee,0) + COALESCE(li.fx_gov,0))
               * COALESCE(li.fx_rate_used, 0)
               * (1 + COALESCE(li.fx_markup,0)/100.0)
             ELSE
               (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
             END
           ), 0)
         FROM line_items li
         WHERE li.invoice_id = invoices.id AND li.item_type = 'foreign'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)
          AND COALESCE(li.is_taxable,0)=1) as foreign_taxable_total,
        invoices.payment_meta,
        invoices.client_id,
        invoices.ipm_case_id,
        invoices.ipm_case_ref,
        invoices.billing_status,
        invoices.payment_status
    FROM invoices
    JOIN clients ON clients.id = invoices.client_id
    LEFT JOIN business_profile bp ON bp.id = invoices.business_profile_id
    WHERE invoices.id = ?
    LIMIT 1
    """,
    [invoice_id_int],
  ).fetchone()
  return _invoice_item_from_matching_row(row) if row else None



def _extract_invoice_number_from_memo(memo: str | None) -> str | None:
  try:
    s = (memo or "").strip()
  except Exception:
    s = ""
  if not s:
    return None
  matches = re.findall(r"INV:([^\s|]+)", s)
  if not matches:
    return None
  inv_no = (matches[-1] or "").strip()
  return inv_no or None


def _fetch_unique_invoice_row_by_number(
  conn, invoice_number: str | None
) -> tuple[tuple[int, str, str] | None, bool]:
  """Return (row, ambiguous) where row=(id, number, status).

  invoices.number is *not* globally unique (unique constraint is (business_profile_id, number)),
  so lookups by number alone must fail closed if multiple rows exist.
  """
  no = (invoice_number or "").strip()
  if not no:
    return None, False
  rows = conn.execute(
    "SELECT id, number, status FROM invoices WHERE number=? ORDER BY id DESC LIMIT 2",
    (no,),
  ).fetchall()
  if not rows:
    return None, False
  if len(rows) > 1:
    return None, True
  r = rows[0]
  try:
    return (int(r[0]), str(r[1] or ""), str(r[2] or "")), False
  except (TypeError, ValueError):
    return None, False


def _fetch_unique_invoice_row_by_payment_tid(
  conn, tid: str | None
) -> tuple[tuple[int, str, str] | None, bool]:
  """Return (row, ambiguous) where row=(id, number, status) by searching payment_meta."""
  t = (tid or "").strip()
  if not t:
    return None, False
  like_a = f'%"tid": "{t}"%'
  like_b = f'%"tid":"{t}"%'
  rows = conn.execute(
    """
    SELECT id, number, status
    FROM invoices
    WHERE (payment_meta LIKE ? OR payment_meta LIKE ?)
    ORDER BY issue_date DESC, id DESC
    LIMIT 2
    """,
    (like_a, like_b),
  ).fetchall()
  if not rows:
    return None, False
  if len(rows) > 1:
    return None, True
  r = rows[0]
  try:
    return (int(r[0]), str(r[1] or ""), str(r[2] or "")), False
  except (TypeError, ValueError):
    return None, False


def _resolve_unique_invoice_row_for_transaction(
  conn, *, tid: str | None, memo_invoice_number: str | None
) -> tuple[tuple[int, str, str] | None, bool]:
  """Resolve which invoice a transaction memo intends to reference (fail closed if ambiguous).

  Preference:
   1) If payment_meta contains this tid and it matches memo_invoice_number, use it.
   2) Otherwise, fall back to memo_invoice_number if it resolves to a single invoice row.
  """
  memo_no = (memo_invoice_number or "").strip() or None

  # Try tid-based linkage first, but only if it matches memo number (prevents mismatched updates).
  tid_row, tid_amb = _fetch_unique_invoice_row_by_payment_tid(conn, tid)
  if tid_amb:
    return None, True
  if tid_row and memo_no and (tid_row[1] or "").strip() == memo_no:
    return tid_row, False

  # Fall back to memo invoice number.
  no_row, no_amb = _fetch_unique_invoice_row_by_number(conn, memo_no)
  if no_amb:
    return None, True
  return no_row, False


def _invoice_status_map_for_unique_numbers(conn, invoice_numbers: set[str]) -> dict[str, str]:
  """Build a number->status map only for numbers that resolve to a single invoice row."""
  if not invoice_numbers:
    return {}
  nums = [n for n in {str(x or "").strip() for x in invoice_numbers} if n]
  if not nums:
    return {}
  qmarks = ",".join(["?"] * len(nums))
  rows = conn.execute(
    f"""
    SELECT number, COUNT(*) AS cnt, MAX(status) AS status
    FROM invoices
    WHERE number IN ({qmarks})
    GROUP BY number
    """,
    nums,
  ).fetchall()
  out: dict[str, str] = {}
  for r in rows or []:
    try:
      no = str(r[0] or "").strip()
      cnt = int(r[1] or 0)
      st = str(r[2] or "")
    except (TypeError, ValueError):
      continue
    if no and cnt == 1:
      out[no] = st
  return out


def _compute_effective_tax_invoice(
  override_val,
  stored_issued,
  stored_at,
  memo,
  invoice_status_map: dict[str, str] | None,
) -> tuple[int, str | None, str | None]:
  """Return (effectiveIssued, effectiveAt, invoiceNumber).

  Priority:
   1) override: 1 => issued, 0 => not issued
   2) invoice status: if memo has INV:<no> and invoice.status == tax_issued => issued
   3) stored flag tax_invoice_issued
  """
  try:
    if override_val is not None:
      ov = int(override_val)
      if ov == 1:
        return 1, stored_at, _extract_invoice_number_from_memo(memo)
      if ov == 0:
        return 0, None, _extract_invoice_number_from_memo(memo)
  except (TypeError, ValueError):
    pass

  inv_no = _extract_invoice_number_from_memo(memo)
  if inv_no and invoice_status_map and inv_no in invoice_status_map:
    st = (invoice_status_map.get(inv_no) or "").strip().lower()
    if st == "tax_issued":
      return 1, stored_at, inv_no

  try:
    return (1 if int(stored_issued or 0) == 1 else 0), stored_at, inv_no
  except (TypeError, ValueError):
    return 0, stored_at, inv_no




# ---------- Routes ----------


@bp.get("")
def index():
  # Compute default recent 1-month range in ET (Eastern time).
  end_date = _now_in_tz().date()
  start_date = _add_months_clamped(end_date, -1)
  currency_raw = (request.args.get("currency") or "").strip()
  currency = (
    _normalize_bank_activity_currency(currency_raw, default=_bank_base_currency())
    if currency_raw
    else None
  )
  return render_template(
    "bank_activity.html",
    default_start=start_date.isoformat(),
    default_end=end_date.isoformat(),
    **_currency_template_context(currency),
  )


@bp.get("/page")
def page():
  # Alias to support url_for('billing_invoices.bank_activity.page') used by templates/nav
  end_date = _now_in_tz().date()
  start_date = _add_months_clamped(end_date, -1)
  currency = _normalize_bank_activity_currency(
    request.args.get("currency"), default=_bank_base_currency()
  )
  return render_template(
    "bank_activity.html",
    default_start=start_date.isoformat(),
    default_end=end_date.isoformat(),
    **_currency_template_context(currency),
  )


@bp.get("/matching")
def matching_page():
  end_date = _now_in_tz().date()
  start_date = _add_months_clamped(end_date, -3)
  currency = _normalize_bank_activity_currency(
    request.args.get("currency"), default=_bank_base_currency()
  )
  return render_template(
    "bank_activity_matching.html",
    default_start=start_date.isoformat(),
    default_end=end_date.isoformat(),
    **_currency_template_context(currency),
  )


@bp.get("/matching/invoice/<int:invoice_id>")
def matching_invoice_detail(invoice_id: int):
  """Render deep matching page for a specific invoice.
  Shows invoice summary and a powerful deposit search UI.
  """
  conn = get_db()
  row = conn.execute(
    """
    SELECT i.id, i.number, i.issue_date, i.total_minor, i.currency, i.status,
        c.name as client_name, bp.id as bp_id, bp.name as biz_name, bp.tax_id as biz_tax_id,
        i.vat_rate,
        i.payment_meta,
        -- Aggregated totals (FX-aware for foreign)
        (SELECT COALESCE(SUM(li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0)), 0)
         FROM line_items li
         WHERE li.invoice_id = i.id AND li.item_type = 'service'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)) as service_total,
        (SELECT COALESCE(SUM(li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0)), 0)
         FROM line_items li
         WHERE li.invoice_id = i.id AND li.item_type = 'admin'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)) as admin_total,
        (SELECT COALESCE(SUM(
             CASE WHEN COALESCE(li.fx_rate_used, 0) > 0 THEN
               (COALESCE(li.fx_fee,0) + COALESCE(li.fx_gov,0))
               * COALESCE(li.fx_rate_used, 0)
               * (1 + COALESCE(li.fx_markup,0)/100.0)
             ELSE
               (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
             END
           ), 0)
         FROM line_items li
         WHERE li.invoice_id = i.id AND li.item_type = 'foreign'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)) as foreign_total,
        (SELECT COALESCE(SUM(
             CASE WHEN COALESCE(li.fx_rate_used, 0) > 0 THEN
               (COALESCE(li.fx_fee,0) + COALESCE(li.fx_gov,0))
               * COALESCE(li.fx_rate_used, 0)
               * (1 + COALESCE(li.fx_markup,0)/100.0)
             ELSE
               (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
             END
           ), 0)
         FROM line_items li
         WHERE li.invoice_id = i.id AND li.item_type = 'foreign'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)
          AND COALESCE(li.is_taxable,0)=1) as foreign_taxable_total,
        i.client_id,
        i.ipm_case_id,
        i.ipm_case_ref,
        i.billing_status,
        i.payment_status
    FROM invoices i
    JOIN clients c ON c.id = i.client_id
    LEFT JOIN business_profile bp ON bp.id = i.business_profile_id
    WHERE i.id = ?
    """,
    (invoice_id,),
  ).fetchone()
  # conn remains open for now; we don't need it further
  conn.close()
  if not row:
    abort(404)

  # Compute default window from issue_date to +30 days. Fallback: last 30 days.
  from datetime import datetime as _dt

  def _safe_iso(date_str: str):
    try:
      return _dt.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
      return None

  issue = _safe_iso(row[2] or "")
  if issue is None:
    now = _now_in_tz().date()
    default_start = _add_months_clamped(now, -1).isoformat()
    default_end = now.isoformat()
  else:
    default_start = issue.isoformat()
    default_end = _add_months_clamped(issue, 1).isoformat()

  # Compute dynamic FX-aware total (minor units) to avoid 0 saved totals
  saved_minor = int(row[3] or 0)
  currency = row[4] or "USD"
  vt = _safe_float(row[10], 0.0)
  svc = _safe_float(row[12], 0.0)
  adm = _safe_float(row[13], 0.0)
  frn = _safe_float(row[14], 0.0)
  frn_taxable = _safe_float(row[15], 0.0)
  vm = (vt / 100.0) if vt > 1 else vt
  tax_dynamic = vm * (svc + frn_taxable)
  total_display = svc + adm + frn + tax_dynamic
  decimal_total = coerce_decimal(total_display, None)
  computed_minor = (
    to_minor(decimal_total, currency)
    if decimal_total is not None
    else int(round(total_display))
  )
  total_minor_for_ui = computed_minor if computed_minor > 0 else saved_minor
  total_amount = _storage_amount_to_display(total_minor_for_ui, currency)

  inv = {
    "id": row[0],
    "number": row[1],
    "issue_date": row[2],
    "total": total_amount,
    "currency": currency,
    "status": row[5],
    "client_name": row[6],
    "client_id": row[16],
    "bp_id": row[7],
    "biz_name": row[8],
    "biz_tax_id": row[9],
    "ipm_case_id": row[17],
    "ipm_case_ref": row[18],
    "billing_status": row[19],
    "payment_status": row[20],
  }

  matched_deposits = []
  paid_sum: int | float = 0
  raw_pm = row[11] if len(row) > 11 else None
  pm = load_json(raw_pm, None)
  if isinstance(pm, dict):
    deps = pm.get("deposits")
    if isinstance(deps, list):
      for d in deps:
        if not isinstance(d, dict):
          continue
        amt = _payment_meta_amount_value(d.get("deposit"), currency)
        paid_sum = _matching_number(paid_sum) + _matching_number(amt)
        matched_deposits.append(
          {
            "tid": (d.get("tid") or ""),
            "date": (d.get("date") or ""),
            "deposit": amt,
            "account_alias": (d.get("account_alias") or ""),
            "summary": (d.get("summary") or ""),
          }
        )
    elif pm.get("deposit") is not None:
      amt = _payment_meta_amount_value(pm.get("deposit"), currency)
      paid_sum = amt
      matched_deposits.append(
        {
          "tid": "",
          "date": (pm.get("date") or ""),
          "deposit": amt,
          "account_alias": (pm.get("account_alias") or ""),
          "summary": (pm.get("summary") or ""),
        }
      )

  remaining = _matching_number(total_amount) - _matching_number(paid_sum)
  if abs(remaining) < 0.000001:
    remaining = 0
  inv["paid_sum"] = int(paid_sum) if float(paid_sum or 0).is_integer() else paid_sum
  inv["remaining"] = int(remaining) if float(remaining or 0).is_integer() else remaining
  inv["deposit_count"] = len(matched_deposits)
  return render_template(
    "bank_activity_matching_detail.html",
    inv=inv,
    matched_deposits=matched_deposits,
    default_start=default_start,
    default_end=default_end,
    **_currency_template_context(currency),
  )


@bp.get("/matching/deposit/<tid>")
def matching_deposit_detail(tid: str):
  """Render a small detail page for a specific deposit transaction (bank_transactions.tid).

  This is mainly used by the matching UIs to provide deep-links from a TID to:
  - the raw Bank activity transaction fields
  - linked invoice(s) if memo includes INV:<number> or payment_meta contains this tid
  """
  tid = (tid or "").strip()
  if not tid:
    abort(404)

  conn = get_db()
  row = conn.execute(
    """
    SELECT
     tid, trdt, trdate, acc_in, acc_out, balance,
     bank_code, account_number,
     remark1, remark2, remark3, memo,
     tax_invoice_issued, tax_invoice_issued_at, tax_invoice_override,
     reg_dt, job_id,
     created_at, updated_at
    FROM bank_transactions
    WHERE tid = ?
    """,
    (tid,),
  ).fetchone()
  if not row:
    conn.close()
    abort(404)

  tx = {
    "tid": row[0],
    "trdt": row[1],
    "trdate": row[2],
    "acc_in": _storage_amount_to_display(row[3], "USD"),
    "acc_out": _storage_amount_to_display(row[4], "USD"),
    "balance": _storage_amount_to_display(row[5], "USD"),
    "bank_code": row[6],
    "account_number": row[7],
    "remark1": row[8],
    "remark2": row[9],
    "remark3": row[10],
    "memo": row[11],
    "tax_invoice_issued": row[12],
    "tax_invoice_issued_at": row[13],
    "tax_invoice_override": row[14],
    "reg_dt": row[15],
    "job_id": row[16],
    "created_at": row[17],
    "updated_at": row[18],
  }

  inv_no = _extract_invoice_number_from_memo(tx.get("memo"))

  invoices: list[dict[str, Any]] = []
  seen_inv_ids: set[int] = set()

  def _add_invoice_row(r) -> None:
    if not r:
      return
    try:
      inv_id = int(r[0])
    except Exception:
      return
    if inv_id in seen_inv_ids:
      return
    seen_inv_ids.add(inv_id)
    invoices.append(
      {
        "id": inv_id,
        "number": r[1],
        "client_id": r[2],
        "client_name": r[3],
        "ipm_case_id": r[4],
        "ipm_case_ref": r[5],
        "billing_status": r[6],
        "payment_status": r[7],
      }
    )

  # 1) Prefer memo INV:<number> linkage.
  if inv_no:
    try:
      rows = conn.execute(
        """
        SELECT i.id, i.number, i.client_id, c.name as client_name,
            i.ipm_case_id, i.ipm_case_ref, i.billing_status, i.payment_status
        FROM invoices i
        LEFT JOIN clients c ON c.id = i.client_id
        WHERE i.number = ?
        ORDER BY i.issue_date DESC, i.id DESC
        LIMIT 20
        """,
        (inv_no,),
      ).fetchall()
      for r in rows or []:
        _add_invoice_row(r)
    except Exception as exc:
      report_swallowed_exception(
        exc,
        context="bank_activity.matching_deposit_detail.lookup_invoice_by_number",
      )

  # 2) Fallback: reverse lookup by tid inside invoices.payment_meta (best-effort).
  try:
    like_a = f'%"tid": "{tid}"%'
    like_b = f'%"tid":"{tid}"%'
    rows = conn.execute(
      """
      SELECT i.id, i.number, i.client_id, c.name as client_name,
          i.ipm_case_id, i.ipm_case_ref, i.billing_status, i.payment_status
      FROM invoices i
      LEFT JOIN clients c ON c.id = i.client_id
      WHERE (i.payment_meta LIKE ? OR i.payment_meta LIKE ?)
      ORDER BY i.issue_date DESC, i.id DESC
      LIMIT 20
      """,
      (like_a, like_b),
    ).fetchall()
    for r in rows or []:
      _add_invoice_row(r)
  except Exception as exc:
    report_swallowed_exception(
      exc,
      context="bank_activity.matching_deposit_detail.lookup_invoice_by_tid",
    )

  conn.close()
  return render_template(
    "bank_activity_deposit_detail.html",
    tx=tx,
    inv_no=inv_no,
    invoices=invoices,
  )


@bp.get("/db_accounts")
def db_accounts():
  """Return distinct bank accounts present in bank_transactions.
  Optional params:
   sdate: yyyyMMdd
   edate: yyyyMMdd
   depositsOnly: 1 to only include accounts with deposits (acc_in>0) in range
  """
  currency_raw = (request.args.get("currency") or "").strip()
  currency = (
    _normalize_bank_activity_currency(currency_raw, default=_bank_base_currency())
    if currency_raw
    else ""
  )
  sdate = (request.args.get("sdate") or "").strip()
  edate = (request.args.get("edate") or "").strip()
  deposits_only = (request.args.get("depositsOnly") or "").strip().lower() in (
    "1",
    "true",
    "yes",
    "y",
  )

  conn = get_db()
  date_expr = _transaction_date_expr(conn)
  wh = []
  params: list[Any] = [] # type: ignore[name-defined]
  if sdate and len(sdate) == 8:
    wh.append(f"{date_expr} >= ?")
    params.append(sdate)
  if edate and len(edate) == 8:
    wh.append(f"{date_expr} <= ?")
    params.append(edate)
  if deposits_only:
    wh.append("acc_in > 0")
  where_sql = (" WHERE " + " AND ".join(wh)) if wh else ""
  items = _local_account_items(conn, currency=currency, where_sql=where_sql, params=params)
  conn.close()
  return jsonify(
    {
      "items": [
        {
          "bankCode": item.get("bankCode") or "",
          "accountNumber": item.get("accountNumber") or "",
          "currency": item.get("currency") or "",
        }
        for item in items
      ]
    }
  )


@bp.get("/matching/biz_profiles")
def matching_biz_profiles():
  """Return business profiles for the selected matching currency (JSON)."""
  currency = _normalize_bank_activity_currency(request.args.get("currency"), default=_bank_base_currency())
  conn = get_db()
  rows = conn.execute(
    """
    SELECT id, name, tax_id
    FROM business_profile
    WHERE UPPER(COALESCE(currency,'USD')) = ?
    ORDER BY name
    """,
    (currency,),
  ).fetchall()
  conn.close()
  items = [{"id": r[0], "name": r[1], "tax_id": r[2]} for r in rows]
  return jsonify({"items": items, "currency": currency})


@bp.get("/accounts")
def list_accounts():
    currency_raw = (request.args.get("currency") or "").strip()
    currency = (
        _normalize_bank_activity_currency(currency_raw, default=_bank_base_currency())
        if currency_raw
        else ""
    )
    provider = _normalize_account_provider(request.args.get("provider"))
    if provider == "manual":
        conn = get_db()
        try:
            items = _local_account_items(conn, currency=currency)
        finally:
            conn.close()
        return jsonify(
            {
                "all": items,
                "active": items,
                "closed": [],
                "currency": currency or None,
                "provider": "manual",
            }
        )
    if provider == "plaid":
        items, error = _plaid_account_items(currency)
        if error:
            return jsonify(error), 400
        return jsonify(
            {
                "all": items,
                "active": items,
                "closed": [],
                "currency": currency or None,
                "provider": "plaid",
            }
        )
    abort(400, "Unsupported bank account data provider")




@bp.post("/manual_transaction")
def manual_transaction():
  data = request.get_json(silent=True) or {}
  currency = _normalize_bank_activity_currency(data.get("currency"), default=_bank_base_currency())
  account_number = str(data.get("accountNumber") or data.get("account_number") or "").strip()
  if not account_number:
    abort(400, "accountNumber is required")
  bank_code = str(data.get("bankCode") or data.get("bank_code") or "MANUAL").strip() or "MANUAL"
  account_name = str(data.get("accountName") or data.get("account_name") or "").strip()
  trdate = _normalize_transaction_date(
    data.get("transactionDate") or data.get("trdate") or data.get("date")
  )
  time_raw = re.sub(r"\D+", "", str(data.get("transactionTime") or data.get("time") or ""))
  time_part = (time_raw + "000000")[:6]
  trdt = f"{trdate}{time_part}"
  acc_in = _parse_bank_amount_to_storage(
    data.get("accIn") or data.get("deposit") or data.get("inflow"),
    currency=currency,
    field_name="accIn",
  )
  acc_out = _parse_bank_amount_to_storage(
    data.get("accOut") or data.get("withdraw") or data.get("outflow"),
    currency=currency,
    field_name="accOut",
  )
  if acc_in and acc_out:
    abort(400, "Only one of accIn or accOut can be entered")
  if not (acc_in or acc_out):
    abort(400, "accIn or accOut is required")
  balance = _parse_bank_amount_to_storage(data.get("balance"), currency=currency, field_name="balance")
  record = {
    "tid": str(data.get("tid") or "").strip(),
    "bank_code": bank_code,
    "account_number": account_number,
    "account_name": account_name,
    "currency": currency,
    "source_provider": "manual",
    "external_id": str(data.get("externalId") or data.get("external_id") or "").strip(),
    "trdate": trdate,
    "trdt": trdt,
    "trserial": str(data.get("serial") or data.get("trserial") or "").strip(),
    "acc_in": acc_in,
    "acc_out": acc_out,
    "balance": balance,
    "remark1": str(data.get("payerName") or data.get("remark1") or "").strip(),
    "remark2": str(data.get("reference") or data.get("remark2") or "").strip(),
    "remark3": str(data.get("description") or data.get("remark3") or "").strip(),
    "memo": str(data.get("memo") or "").strip(),
    "reg_dt": _now_in_tz().isoformat(),
    "job_id": "manual",
  }
  if not record["tid"]:
    record["tid"] = _manual_tid(record)

  conn = get_db()
  try:
    _upsert_bank_activity_transaction_record(conn, record)
    conn.commit()
  finally:
    conn.close()
  return jsonify(
    {
      "ok": True,
      "tid": record["tid"],
      "currency": currency,
      "provider": "manual",
      "bankCode": bank_code,
      "accountNumber": account_number,
    }
  )


@bp.post("/sync_provider")
def sync_provider():
  data = request.get_json(silent=True) or {}
  provider = _normalize_account_provider(data.get("provider"))
  if provider != "plaid":
    abort(400, "Only provider='plaid' supports provider sync")
  currency = _normalize_bank_activity_currency(data.get("currency"), default=_bank_base_currency())
  sdate = _normalize_transaction_date(data.get("startDate") or data.get("sdate"))
  edate = _normalize_transaction_date(data.get("endDate") or data.get("edate"))
  if sdate > edate:
    abort(400, "startDate must be before endDate")
  result = _sync_plaid_transactions(sdate, edate, currency)
  status = 200 if result.get("ok") else 400
  return jsonify(result), status




@bp.get("/export")
def export_csv():
    trade_type_raw = request.args.get("tradeType", "").strip().upper()
    local_trade_type = trade_type_raw if trade_type_raw in ("I", "O") else ""
    search_string = request.args.get("searchString") or None
    currency = _normalize_bank_activity_currency(request.args.get("currency"), default=_bank_base_currency())
    order = (request.args.get("order") or "D").upper()
    if order not in ("D", "A"):
        order = "D"

    sdate = (request.args.get("sdate") or "").strip()
    edate = (request.args.get("edate") or "").strip()
    accounts_param = (request.args.get("accounts") or "").strip()
    if not sdate or not edate or len(sdate) != 8 or len(edate) != 8:
        abort(400, "sdate and edate must be yyyyMMdd")

    acct_pairs: list[tuple[str, str]] = []
    for tok in accounts_param.split(","):
        token = tok.strip()
        if not token or "|" not in token:
            continue
        bank_code, account_number = token.split("|", 1)
        acct_pairs.append((bank_code.strip(), account_number.strip()))
    if not acct_pairs:
        abort(400, "accounts is required")

    conn = get_db()

    def _ts_from(trdt: str | None, trdate: str | None) -> float:
        if trdt:
            s = str(trdt)
            try:
                if len(s) == 14 and s.isdigit():
                    dt = datetime(
                        int(s[0:4]),
                        int(s[4:6]),
                        int(s[6:8]),
                        int(s[8:10]),
                        int(s[10:12]),
                        int(s[12:14]),
                    )
                    return dt.timestamp()
                s2 = s.replace("Z", "+00:00").replace(" ", "T")
                dt = datetime.fromisoformat(s2)
                return dt.timestamp()
            except (TypeError, ValueError):
                pass
        if trdate and len(str(trdate)) == 8 and str(trdate).isdigit():
            try:
                s = str(trdate)
                dt = datetime(int(s[0:4]), int(s[4:6]), int(s[6:8]))
                return dt.timestamp()
            except (TypeError, ValueError):
                return 0.0
        return 0.0

    is_compact_q = bool(search_string and is_compact_query(search_string))
    date_expr = _transaction_date_expr(conn)
    wh = [f"{date_expr} >= ?", f"{date_expr} <= ?"]
    params: list[Any] = [sdate, edate]
    if local_trade_type == "I":
        wh.append("acc_in > 0")
    elif local_trade_type == "O":
        wh.append("acc_out > 0")
    pair_ors = []
    for bank_code, account_number in acct_pairs:
        pair_ors.append("(bank_code=? AND account_number=?)")
        params += [bank_code, account_number]
    wh.append("(" + " OR ".join(pair_ors) + ")")
    if search_string and not is_compact_q:
        search_clause, search_params = sql_ci_contains_any(
            ["remark1", "remark2", "remark3", "memo"],
            search_string,
        )
        if search_clause:
            wh.append(search_clause)
            params += search_params
    where_sql = " WHERE " + " AND ".join(wh)
    order_sql = "ASC" if order == "A" else "DESC"
    raw_rows = conn.execute(
        f"""
        SELECT tid, trdt, trdate, acc_in, acc_out, balance, remark1, remark2, remark3, memo,
               bank_code, account_number
        FROM bank_transactions
        {where_sql}
        ORDER BY COALESCE(trdt, trdate) {order_sql}, tid {order_sql}
        """,
        params,
    ).fetchall()

    rows = []
    q_compact = to_compact(search_string or "") if is_compact_q else ""
    for raw in raw_rows:
        if is_compact_q:
            text_for_search = " ".join(
                [str(raw[6] or ""), str(raw[7] or ""), str(raw[8] or ""), str(raw[9] or "")]
            )
            if q_compact not in to_compact(text_for_search):
                continue
        trdt = raw[1]
        trdate = raw[2]
        bank_code = raw[10] or ""
        account_no = raw[11] or ""
        rows.append(
            {
                "ts": _ts_from(trdt, trdate),
        "tid": raw[0],
        "trdt": trdt,
        "trdate": trdate,
        "accIn": _storage_amount_to_display(raw[3], currency),
        "accOut": _storage_amount_to_display(raw[4], currency),
        "balance": _storage_amount_to_display(raw[5], currency),
                "remark1": raw[6],
                "remark2": raw[7],
                "remark3": raw[8],
                "memo": raw[9],
                "bank_code": bank_code,
                "account_no": account_no,
                "alias": f"{bank_code} {account_no}".strip(),
            }
        )
    conn.close()

    rows.sort(key=lambda x: x["ts"], reverse=(order == "D"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Activity"
    headers = [
        "No",
        "Alias",
        "Date",
        "Time",
        "In",
        "Out",
        "Balance",
        "Currency",
        "Remark1",
        "Remark2",
        "Remark3",
        "Memo",
        "BankCode",
        "AccountNumber",
        "TID",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    def _format_dt(trdt: str | None, trdate: str | None) -> tuple[str, str]:
        try:
            if trdt and len(str(trdt)) == 14 and str(trdt).isdigit():
                s = str(trdt)
                dt = datetime(
                    int(s[0:4]),
                    int(s[4:6]),
                    int(s[6:8]),
                    int(s[8:10]),
                    int(s[10:12]),
                    int(s[12:14]),
                )
                return (dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S"))
            if trdt:
                s2 = str(trdt).replace("Z", "+00:00").replace(" ", "T")
                dt = datetime.fromisoformat(s2)
                return (dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S"))
            if trdate and len(str(trdate)) == 8 and str(trdate).isdigit():
                s = str(trdate)
                dt = datetime(int(s[0:4]), int(s[4:6]), int(s[6:8]))
                return (dt.strftime("%Y-%m-%d"), "00:00:00")
        except (TypeError, ValueError):
            return (str(trdate or ""), "")
        return (str(trdate or ""), "")

    for i, row in enumerate(rows, start=1):
        dt_str, tm_str = _format_dt(row.get("trdt"), row.get("trdate"))
        ws.append(
            [
                i,
                row.get("alias") or "",
                dt_str,
                tm_str,
                row.get("accIn") or 0,
                row.get("accOut") or 0,
                row.get("balance") or 0,
                currency,
                row.get("remark1") or "",
                row.get("remark2") or "",
                row.get("remark3") or "",
                row.get("memo") or "",
                row.get("bank_code") or "",
                row.get("account_no") or "",
                row.get("tid") or "",
            ]
        )

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                value_text = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value_text))
            except Exception as exc:
                report_swallowed_exception(
                    exc,
                    context="billing_invoices.bank_activity.export.auto_width",
                    log_key="billing_invoices.bank_activity.export.auto_width",
                    log_window_seconds=300,
                )
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"bank_activity_{_now_in_tz().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )




# ---------- Matching data endpoints ----------


@bp.get("/matching/invoices")
def matching_invoices():
  """Return invoices in issued state for the selected matching currency.
  Optional filters: date_from, date_to (issue_date), q (search)
  Paging: page (default 1), perPage (default 15, max 200)
  """
  currency = _normalize_bank_activity_currency(request.args.get("currency"), default=_bank_base_currency())
  date_from = (request.args.get("date_from") or "").strip()
  date_to = (request.args.get("date_to") or "").strip()
  q = (request.args.get("q") or "").strip()
  recommend_tid = (request.args.get("recommend_tid") or "").strip()
  is_compact_q = q and is_compact_query(q)
  try:
    page = max(int(request.args.get("page", 1) or 1), 1)
  except Exception:
    page = 1
  try:
    per_page = int(request.args.get("perPage", 15) or 15)
  except Exception:
    per_page = 15
  per_page = max(1, min(per_page, 200))

  conn = get_db()
  where, params = [
    "UPPER(COALESCE(invoices.currency, 'USD')) = ?",
    "(COALESCE(invoices.billing_status, '') NOT IN ('draft','void'))",
    "(invoices.payment_status IS NULL OR invoices.payment_status != 'paid')",
  ], [currency]

  # Optional business profile filter (comma-separated IDs)
  bp_ids_param = (request.args.get("bpIds") or "").strip()
  bp_ids = []
  if bp_ids_param:
    for tok in bp_ids_param.split(","):
      tok = tok.strip()
      if not tok:
        continue
      try:
        bp_ids.append(int(tok))
      except ValueError:
        pass
  if date_from:
    where.append("invoices.issue_date >= ?")
    params.append(date_from)
  if date_to:
    where.append("invoices.issue_date <= ?")
    params.append(date_to)
  if q and not is_compact_q:
    search_clause, search_params = sql_ci_contains_any(
      ["invoices.number", "clients.name", "invoices.internal_reference"],
      q,
    )
    if search_clause:
      where.append(search_clause)
      params += search_params
  if bp_ids:
    qmarks = ",".join(["?"] * len(bp_ids))
    where.append(f"invoices.business_profile_id IN ({qmarks})")
    params += bp_ids
  where_sql = (" WHERE " + " AND ".join(where)) if where else ""

  base_sql = f"""
    SELECT invoices.id, invoices.number, invoices.issue_date, invoices.total_minor, invoices.currency, invoices.status,
        clients.name as client_name, bp.name as biz_name, bp.tax_id as biz_tax_id,
        invoices.vat_rate,
        -- Aggregated totals (FX-aware for foreign)
        (SELECT COALESCE(SUM(li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0)), 0)
         FROM line_items li
         WHERE li.invoice_id = invoices.id AND li.item_type = 'service'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)) as service_total,
        (SELECT COALESCE(SUM(li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0)), 0)
         FROM line_items li
         WHERE li.invoice_id = invoices.id AND li.item_type = 'admin'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)) as admin_total,
        (SELECT COALESCE(SUM(
             CASE WHEN COALESCE(li.fx_rate_used, 0) > 0 THEN
               (COALESCE(li.fx_fee,0) + COALESCE(li.fx_gov,0))
               * COALESCE(li.fx_rate_used, 0)
               * (1 + COALESCE(li.fx_markup,0)/100.0)
             ELSE
               (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
             END
           ), 0)
         FROM line_items li
         WHERE li.invoice_id = invoices.id AND li.item_type = 'foreign'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)) as foreign_total,
        (SELECT COALESCE(SUM(
             CASE WHEN COALESCE(li.fx_rate_used, 0) > 0 THEN
               (COALESCE(li.fx_fee,0) + COALESCE(li.fx_gov,0))
               * COALESCE(li.fx_rate_used, 0)
               * (1 + COALESCE(li.fx_markup,0)/100.0)
             ELSE
               (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
             END
           ), 0)
         FROM line_items li
         WHERE li.invoice_id = invoices.id AND li.item_type = 'foreign'
          AND (li.is_estimated IS NULL OR li.is_estimated = 0)
          AND COALESCE(li.is_taxable,0)=1) as foreign_taxable_total
        , invoices.payment_meta
        , invoices.client_id
        , invoices.ipm_case_id
        , invoices.ipm_case_ref
        , invoices.billing_status
        , invoices.payment_status
    FROM invoices JOIN clients ON clients.id = invoices.client_id
    LEFT JOIN business_profile bp ON bp.id = invoices.business_profile_id
    {where_sql}
  """

  recommend_deposit = _load_recommend_deposit_context(conn, recommend_tid, currency)
  if recommend_tid and recommend_deposit is None:
    conn.close()
    return jsonify({"items": [], "total": 0, "pageNum": 1, "pageCount": 1})
  if recommend_deposit is not None:
    _annotate_deposit_allocations(conn, [recommend_deposit], currency)

  if is_compact_q or recommend_deposit is not None:
    # -only Search: SQLfrom q Filters  times from  Filters
    rows_all = conn.execute(
      base_sql + " ORDER BY invoices.issue_date DESC, invoices.id DESC", params
    ).fetchall()
    history_by_client = (
      _load_client_payer_history_index(conn, [row[15] for row in rows_all], currency)
      if recommend_deposit is not None
      else {}
    )
    q_compact = to_compact(q)
    filtered_items = []
    for r in rows_all:
      # r[1]=number, r[6]=client_name, r[7]=biz_name
      if is_compact_q:
        text = " ".join(
          [
            str(r[1] or ""),
            str(r[6] or ""),
            str(r[7] or ""),
          ]
        )
        if q_compact not in to_compact(text):
          continue
      item = _invoice_item_from_matching_row(r)
      if recommend_deposit is not None:
        recommend = _score_invoice_deposit_recommendation(
          invoice=item,
          deposit=recommend_deposit,
          history_by_client=history_by_client,
        )
        if not recommend["recommended"]:
          continue
        item["recommended"] = True
        item["recommend_score"] = recommend["score"]
        item["recommend_reasons"] = recommend["reasons"]
      filtered_items.append(item)
    if recommend_deposit is not None:
      filtered_items.sort(
        key=lambda item: (
          int(item.get("recommend_score") or 0),
          str(item.get("issue_date") or ""),
          int(item.get("id") or 0),
        ),
        reverse=True,
      )
    total = len(filtered_items)
    page, page_count = _clamp_page_for_total(page, total, per_page)
    offset = (page - 1) * per_page
    out = filtered_items[offset : offset + per_page]
  else:
    total = conn.execute(
      f"SELECT COUNT(*) FROM invoices JOIN clients ON clients.id = invoices.client_id LEFT JOIN business_profile bp ON bp.id = invoices.business_profile_id {where_sql}",
      params,
    ).fetchone()[0]
    page, page_count = _clamp_page_for_total(page, total, per_page)
    offset = (page - 1) * per_page
    rows = conn.execute(
      base_sql + " ORDER BY invoices.issue_date DESC, invoices.id DESC LIMIT ? OFFSET ?",
      params + [per_page, offset],
    ).fetchall()
    out = [_invoice_item_from_matching_row(r) for r in rows]
  conn.close()
  return jsonify(
    {
      "items": out,
      "total": total,
      "pageNum": page,
      "pageCount": page_count,
      "currency": currency,
    }
  )


@bp.get("/matching/deposits")
def matching_deposits():
  """Return deposit transactions (acc_in>0) from local DB within date range and selected accounts.
  Params:
   sdate: yyyyMMdd
   edate: yyyyMMdd
   accounts: comma-separated list of bankCode|accountNumber
   limit: max rows (default 500)
  """
  sdate = (request.args.get("sdate") or "").strip()
  edate = (request.args.get("edate") or "").strip()
  accounts_param = (request.args.get("accounts") or "").strip()
  recommend_invoice_id_raw = (request.args.get("recommend_invoice_id") or "").strip()
  currency = _normalize_bank_activity_currency(request.args.get("currency"), default=_bank_base_currency())
  memo_mode = (request.args.get("memoMode") or "empty").strip().lower()
  if memo_mode not in ("unmatched", "empty", "all"):
    memo_mode = "empty"
  page = max(coerce_int(request.args.get("page"), 1) or 1, 1)
  per_page = coerce_int(request.args.get("perPage"), 15) or 15
  per_page = max(1, min(per_page, 200))

  if not sdate or not edate or len(sdate) != 8 or len(edate) != 8:
    abort(400, "sdate and edate must be yyyyMMdd")

  acct_pairs = []
  if accounts_param:
    for tok in accounts_param.split(","):
      tok = tok.strip()
      if not tok:
        continue
      if "|" in tok:
        b, a = tok.split("|", 1)
        acct_pairs.append((b.strip(), a.strip()))
  if not acct_pairs:
    abort(400, "accounts is required")

  conn = get_db()
  # Derive a date key from trdate or the date part of trdt to ensure inclusive filtering for end date
  date_expr = _transaction_date_expr(conn)
  wh = [
    "acc_in > 0",
    f"{date_expr} >= ?",
    f"{date_expr} <= ?",
  ]
  # Treat tax-invoice-issued transactions as processed; hide them unless memoMode='all'
  if memo_mode != "all":
    wh.append("(COALESCE(tax_invoice_override,0) != 1 AND COALESCE(tax_invoice_issued,0) = 0)")
  if memo_mode == "unmatched":
    wh.append("(memo IS NULL OR memo NOT LIKE '%INV:%')")
  elif memo_mode == "empty":
    wh.append("((memo IS NULL OR trim(memo) = '') OR memo LIKE '%INV:%')")
  params = [sdate, edate]
  # Auto-hide keyword matches in remarks
  auto_hide_keywords = ["Cancel", "", "CancelDeposit"]
  hide_terms = []
  for kw in auto_hide_keywords:
    if not str(kw or "").strip():
      continue
    like = f"%{kw}%"
    hide_terms.append("COALESCE(remark1,'') LIKE ?")
    hide_terms.append("COALESCE(remark2,'') LIKE ?")
    hide_terms.append("COALESCE(remark3,'') LIKE ?")
    params += [like, like, like]
  if hide_terms:
    wh.append("NOT (" + " OR ".join(hide_terms) + ")")
  # Build dynamic OR clauses for (bank_code, account_number)
  pair_ors = []
  for b, a in acct_pairs:
    pair_ors.append("(bank_code=? AND account_number=?)")
    params += [b, a]
  wh.append("(" + " OR ".join(pair_ors) + ")")
  where_sql = " WHERE " + " AND ".join(wh)
  partition_cols = "bank_code, account_number, trdt, trserial, acc_in, acc_out, balance, remark1, remark2, remark3"
  order_expr = (
    "CASE WHEN memo IS NOT NULL AND trim(memo) <> '' THEN 0 ELSE 1 END, "
    "COALESCE(updated_at, created_at) DESC, tid DESC"
  )
  recommend_invoice_id = coerce_int(recommend_invoice_id_raw, 0) or 0
  recommend_invoice = _load_recommend_invoice_context(conn, recommend_invoice_id)
  if recommend_invoice_id_raw and recommend_invoice is None:
    conn.close()
    return jsonify({"items": [], "total": 0, "pageNum": 1, "pageCount": 1})

  if recommend_invoice is not None:
    recommend_currency = _normalize_bank_activity_currency(recommend_invoice.get("currency"))
    history_by_client = _load_client_payer_history_index(
      conn,
      [recommend_invoice.get("client_id")],
      recommend_currency,
    )
    rows_all = conn.execute(
      f"""
      WITH ranked AS (
       SELECT tid, trdt, trdate, acc_in, balance, bank_code, account_number,
           remark1, remark2, remark3, memo,
           ROW_NUMBER() OVER (PARTITION BY {partition_cols} ORDER BY {order_expr}) AS rn
       FROM bank_transactions
       {where_sql}
      )
      SELECT tid, trdt, trdate, acc_in, balance, bank_code, account_number, remark1, remark2, remark3, memo
      FROM ranked
      WHERE rn = 1
      ORDER BY trdt DESC
      """,
      params,
    ).fetchall()
    candidate_items = [_deposit_item_from_matching_row(row, recommend_currency) for row in rows_all]
    _annotate_deposit_allocations(conn, candidate_items, recommend_currency)
    filtered_items = []
    for item in candidate_items:
      recommend = _score_invoice_deposit_recommendation(
        invoice=recommend_invoice,
        deposit=item,
        history_by_client=history_by_client,
      )
      if not recommend["recommended"]:
        continue
      item["recommended"] = True
      item["recommend_score"] = recommend["score"]
      item["recommend_reasons"] = recommend["reasons"]
      filtered_items.append(item)
    if memo_mode == "empty":
      filtered_items = [
        item
        for item in filtered_items
        if not str(item.get("memo") or "").strip() or item.get("match_state") == "partial"
      ]
    filtered_items.sort(
      key=lambda item: (
        int(item.get("recommend_score") or 0),
        str(item.get("trdt") or item.get("trdate") or ""),
        str(item.get("tid") or ""),
      ),
      reverse=True,
    )
    total = len(filtered_items)
    page, page_count = _clamp_page_for_total(page, total, per_page)
    offset = (page - 1) * per_page
    items = filtered_items[offset : offset + per_page]
  elif memo_mode == "empty":
    rows_all = conn.execute(
      f"""
      WITH ranked AS (
       SELECT tid, trdt, trdate, acc_in, balance, bank_code, account_number,
           remark1, remark2, remark3, memo,
           ROW_NUMBER() OVER (PARTITION BY {partition_cols} ORDER BY {order_expr}) AS rn
       FROM bank_transactions
       {where_sql}
      )
      SELECT tid, trdt, trdate, acc_in, balance, bank_code, account_number, remark1, remark2, remark3, memo
      FROM ranked
      WHERE rn = 1
      ORDER BY trdt DESC
      """,
      params,
    ).fetchall()
    filtered_items = [_deposit_item_from_matching_row(r, currency) for r in rows_all]
    _annotate_deposit_allocations(conn, filtered_items, currency)
    filtered_items = [
      item
      for item in filtered_items
      if not str(item.get("memo") or "").strip() or item.get("match_state") == "partial"
    ]
    total = len(filtered_items)
    page, page_count = _clamp_page_for_total(page, total, per_page)
    offset = (page - 1) * per_page
    items = filtered_items[offset : offset + per_page]
  else:
    total = conn.execute(
      f"""
      WITH ranked AS (
       SELECT tid, trdt, trdate, acc_in, balance, bank_code, account_number,
           remark1, remark2, remark3, memo,
           ROW_NUMBER() OVER (PARTITION BY {partition_cols} ORDER BY {order_expr}) AS rn
       FROM bank_transactions
       {where_sql}
      )
      SELECT COUNT(*) FROM ranked WHERE rn = 1
      """,
      params,
    ).fetchone()[0]
    page, page_count = _clamp_page_for_total(page, total, per_page)
    offset = (page - 1) * per_page
    rows = conn.execute(
      f"""
      WITH ranked AS (
       SELECT tid, trdt, trdate, acc_in, balance, bank_code, account_number,
           remark1, remark2, remark3, memo,
           ROW_NUMBER() OVER (PARTITION BY {partition_cols} ORDER BY {order_expr}) AS rn
       FROM bank_transactions
       {where_sql}
      )
      SELECT tid, trdt, trdate, acc_in, balance, bank_code, account_number, remark1, remark2, remark3, memo
      FROM ranked
      WHERE rn = 1
      ORDER BY trdt DESC
      LIMIT ? OFFSET ?
      """,
      params + [per_page, offset],
    ).fetchall()
    items = [_deposit_item_from_matching_row(r, currency) for r in rows]
    _annotate_deposit_allocations(conn, items, currency)
  conn.close()
  return jsonify({"items": items, "total": total, "pageNum": page, "pageCount": page_count})


@bp.get("/local_search")
def local_search():
  """Search locally persisted transactions from local transaction data.
  Query params:
   sdate: yyyyMMdd
   edate: yyyyMMdd
   accounts: comma-separated bankCode|accountNumber
   tradeType: '' (all), 'I' (deposit only), 'O' (withdraw only)
   searchString: text to search in remark1/2/3/memo
   page: 1-based
   perPage: items per page (max 200)
   order: 'A' asc, 'D' desc by trdt (fallback trdate)
  """
  sdate = (request.args.get("sdate") or "").strip()
  edate = (request.args.get("edate") or "").strip()
  accounts_param = (request.args.get("accounts") or "").strip()
  trade_type = (request.args.get("tradeType") or "").strip().upper()
  currency = _normalize_bank_activity_currency(request.args.get("currency"), default=_bank_base_currency())
  # Optional server-side filters to keep pagination consistent with results
  exclude_matched_raw = (request.args.get("excludeMatched") or "").strip().lower()
  exclude_matched = exclude_matched_raw in ("1", "true", "yes", "y")
  amount_min_str = (request.args.get("amountMin") or "").strip()
  amount_max_str = (request.args.get("amountMax") or "").strip()
  amount_min = (
    _parse_bank_amount_to_storage(amount_min_str, currency=currency, field_name="amountMin")
    if amount_min_str
    else None
  )
  amount_max = (
    _parse_bank_amount_to_storage(amount_max_str, currency=currency, field_name="amountMax")
    if amount_max_str
    else None
  )
  search_string = (request.args.get("searchString") or "").strip()
  is_compact_q = bool(search_string and is_compact_query(search_string))
  page = max(coerce_int(request.args.get("page"), 1) or 1, 1)
  per_page = coerce_int(request.args.get("perPage"), 100) or 100
  per_page = max(1, min(per_page, 200))
  order = (request.args.get("order") or "D").upper()
  if order not in ("A", "D"):
    order = "D"

  if not sdate or not edate or len(sdate) != 8 or len(edate) != 8:
    abort(400, "sdate and edate must be yyyyMMdd")

  acct_pairs = []
  if accounts_param:
    for tok in accounts_param.split(","):
      tok = tok.strip()
      if not tok:
        continue
      if "|" in tok:
        b, a = tok.split("|", 1)
        acct_pairs.append((b.strip(), a.strip()))
  if not acct_pairs:
    abort(400, "accounts is required")

  conn = get_db()
  # Filter by derived date key (trdate or date part of trdt)
  date_expr = _transaction_date_expr(conn)
  wh = [f"{date_expr} >= ?", f"{date_expr} <= ?"]
  params = [sdate, edate]
  # trade type filter
  if trade_type == "I":
    wh.append("acc_in > 0")
  elif trade_type == "O":
    wh.append("acc_out > 0")
  # exclude already matched (memo contains 'INV:')
  if exclude_matched:
    wh.append("(memo IS NULL OR memo NOT LIKE '%INV:%')")
  # amount range filter (primarily for deposits)
  amount_field = "acc_in" if trade_type != "O" else "acc_out"
  if amount_min is not None:
    wh.append(f"{amount_field} >= ?")
    params.append(amount_min)
  if amount_max is not None:
    wh.append(f"{amount_field} <= ?")
    params.append(amount_max)
  # accounts OR list
  pair_ors = []
  for b, a in acct_pairs:
    pair_ors.append("(bank_code=? AND account_number=?)")
    params += [b, a]
  wh.append("(" + " OR ".join(pair_ors) + ")")
  # search string
  if search_string and not is_compact_q:
    search_clause, search_params = sql_ci_contains_any(
      ["remark1", "remark2", "remark3", "memo"],
      search_string,
    )
    if search_clause:
      wh.append(search_clause)
      params += search_params
  where_sql = " WHERE " + " AND ".join(wh)

  order_sql = "ASC" if order == "A" else "DESC"
  if is_compact_q:
    rows_all = conn.execute(
      f"""
      SELECT tid, trdt, trdate, trserial, acc_in, acc_out, balance, remark1, remark2, remark3, memo, tax_invoice_issued, tax_invoice_issued_at, tax_invoice_override, bank_code, account_number
      FROM bank_transactions
      {where_sql}
      ORDER BY COALESCE(trdt, trdate) {order_sql}, tid {order_sql}
      """,
      params,
    ).fetchall()
    q_compact = to_compact(search_string)
    filtered_rows = []
    for r in rows_all:
      text = " ".join(
        [
          str(r[7] or ""),
          str(r[8] or ""),
          str(r[9] or ""),
          str(r[10] or ""),
        ]
      )
      if q_compact in to_compact(text):
        filtered_rows.append(r)
    total = len(filtered_rows)
    page, page_count = _clamp_page_for_total(page, total, per_page)
    offset = (page - 1) * per_page
    rows = filtered_rows[offset : offset + per_page]
  else:
    total = conn.execute(
      f"SELECT COUNT(*) FROM bank_transactions {where_sql}", params
    ).fetchone()[0]
    page, page_count = _clamp_page_for_total(page, total, per_page)
    offset = (page - 1) * per_page
    rows = conn.execute(
      f"""
      SELECT tid, trdt, trdate, trserial, acc_in, acc_out, balance, remark1, remark2, remark3, memo, tax_invoice_issued, tax_invoice_issued_at, tax_invoice_override, bank_code, account_number
      FROM bank_transactions
      {where_sql}
      ORDER BY COALESCE(trdt, trdate) {order_sql}, tid {order_sql}
      LIMIT ? OFFSET ?
      """,
      params + [per_page, offset],
    ).fetchall()

  # Preload invoice statuses for INV:... memos
  invoice_numbers = set()
  for r in rows:
    inv_no = _extract_invoice_number_from_memo(r[10])
    if inv_no:
      invoice_numbers.add(inv_no)
  try:
    invoice_status_map = _invoice_status_map_for_unique_numbers(conn, invoice_numbers)
  except Exception:
    invoice_status_map = {}

  items = []
  for r in rows:
    eff, eff_at, _inv_no = _compute_effective_tax_invoice(
      r[13], r[11], r[12], r[10], invoice_status_map
    )
    items.append(
      {
        "tid": r[0],
        "trdt": r[1],
        "trdate": r[2],
        "trserial": r[3],
        "accIn": _storage_amount_to_display(r[4], currency),
        "accOut": _storage_amount_to_display(r[5], currency),
        "balance": _storage_amount_to_display(r[6], currency),
        "remark1": r[7],
        "remark2": r[8],
        "remark3": r[9],
        "memo": r[10],
        "taxInvoiceIssued": int(eff),
        "taxInvoiceIssuedAt": eff_at,
        "bankCode": r[14],
        "accountNumber": r[15],
      }
    )

  _annotate_deposit_allocations(conn, items, currency)
  conn.close()

  return jsonify(
    {
      "total": total,
      "pageNum": page,
      "perPage": per_page,
      "pageCount": page_count,
      "list": items,
    }
  )


@bp.get("/local_summary")
def local_summary():
  """Account-level aggregates from local DB from local transaction data.
  Query params are same as /local_search (subset used):
   sdate, edate (yyyyMMdd)
   accounts (comma-separated bankCode|accountNumber)
   tradeType: '' (all), 'I' (deposit only), 'O' (withdraw only)
   searchString: text to search in remark1/2/3/memo
   excludeMatched: 1/true to exclude rows with memo like 'INV:%'
   amountMin/amountMax: optional amount filter applied to acc_in (or acc_out when tradeType='O')
  Returns per account: count, totalIn, totalOut, lastBalance.
  """
  sdate = (request.args.get("sdate") or "").strip()
  edate = (request.args.get("edate") or "").strip()
  accounts_param = (request.args.get("accounts") or "").strip()
  trade_type = (request.args.get("tradeType") or "").strip().upper()
  currency = _normalize_bank_activity_currency(request.args.get("currency"), default=_bank_base_currency())
  exclude_matched_raw = (request.args.get("excludeMatched") or "").strip().lower()
  exclude_matched = exclude_matched_raw in ("1", "true", "yes", "y")
  amount_min_str = (request.args.get("amountMin") or "").strip()
  amount_max_str = (request.args.get("amountMax") or "").strip()
  amount_min = (
    _parse_bank_amount_to_storage(amount_min_str, currency=currency, field_name="amountMin")
    if amount_min_str
    else None
  )
  amount_max = (
    _parse_bank_amount_to_storage(amount_max_str, currency=currency, field_name="amountMax")
    if amount_max_str
    else None
  )
  search_string = (request.args.get("searchString") or "").strip()
  is_compact_q = bool(search_string and is_compact_query(search_string))

  if not sdate or not edate or len(sdate) != 8 or len(edate) != 8:
    abort(400, "sdate and edate must be yyyyMMdd")

  acct_pairs: list[tuple[str, str]] = []
  if accounts_param:
    for tok in accounts_param.split(","):
      tok = tok.strip()
      if not tok:
        continue
      if "|" in tok:
        b, a = tok.split("|", 1)
        acct_pairs.append((b.strip(), a.strip()))
  if not acct_pairs:
    abort(400, "accounts is required")

  conn = get_db()
  # Build filters (reuse logic from local_search)
  date_expr = _transaction_date_expr(conn)
  wh = [f"{date_expr} >= ?", f"{date_expr} <= ?"]
  params: list[Any] = [sdate, edate]
  if trade_type == "I":
    wh.append("acc_in > 0")
  elif trade_type == "O":
    wh.append("acc_out > 0")
  if exclude_matched:
    wh.append("(memo IS NULL OR memo NOT LIKE '%INV:%')")
  # amount range applies to deposits unless withdraw-only
  amount_field = "acc_out" if trade_type == "O" else "acc_in"
  if amount_min is not None:
    wh.append(f"{amount_field} >= ?")
    params.append(amount_min)
  if amount_max is not None:
    wh.append(f"{amount_field} <= ?")
    params.append(amount_max)
  # accounts OR list
  pair_ors = []
  for b, a in acct_pairs:
    pair_ors.append("(bank_code=? AND account_number=?)")
    params += [b, a]
  wh.append("(" + " OR ".join(pair_ors) + ")")
  if search_string and not is_compact_q:
    search_clause, search_params = sql_ci_contains_any(
      ["remark1", "remark2", "remark3", "memo"],
      search_string,
    )
    if search_clause:
      wh.append(search_clause)
      params += search_params
  where_sql = " WHERE " + " AND ".join(wh)

  if is_compact_q:
    rows_all = conn.execute(
      f"""
      SELECT tid, trdt, trdate, balance, acc_in, acc_out, remark1, remark2, remark3, memo, bank_code, account_number
      FROM bank_transactions
      {where_sql}
      ORDER BY COALESCE(trdt, trdate) DESC, tid DESC
      """,
      params,
    ).fetchall()
    q_compact = to_compact(search_string)
    agg: dict[tuple[str, str], dict[str, Any]] = {}
    for r in rows_all:
      text = " ".join(
        [
          str(r[6] or ""),
          str(r[7] or ""),
          str(r[8] or ""),
          str(r[9] or ""),
        ]
      )
      if q_compact not in to_compact(text):
        continue
      bcode = str(r[10] or "")
      accno = str(r[11] or "")
      key = (bcode, accno)
      if key not in agg:
        agg[key] = {
          "bankCode": bcode,
          "accountNumber": accno,
          "count": 0,
          "totalIn": 0,
          "totalOut": 0,
          "lastBalance": _storage_amount_to_display(r[3], currency),
        }
      agg[key]["count"] += 1
      agg[key]["totalIn"] += _storage_amount_to_display(r[4], currency)
      agg[key]["totalOut"] += _storage_amount_to_display(r[5], currency)
    out = sorted(
      agg.values(),
      key=lambda x: (x.get("bankCode") or "", x.get("accountNumber") or ""),
    )
    conn.close()
    return jsonify({"items": out})

  # Group aggregates
  rows = conn.execute(
    f"""
    SELECT bank_code, account_number,
        COUNT(*) as cnt,
        SUM(acc_in) as sum_in,
        SUM(acc_out) as sum_out
    FROM bank_transactions t
    {where_sql}
    GROUP BY bank_code, account_number
    ORDER BY bank_code, account_number
    """,
    params,
  ).fetchall()

  # For each account, fetch last balance within the same filter window
  out: list[dict[str, Any]] = []
  for r in rows:
    bcode, accno, cnt, sum_in, sum_out = (
      r[0],
      r[1],
      int(r[2] or 0),
      int(r[3] or 0),
      int(r[4] or 0),
    )
    wh_last = list(wh)
    params_last = list(params)
    wh_last.append("(bank_code=? AND account_number=?)")
    params_last += [bcode, accno]
    where_sql_last = " WHERE " + " AND ".join(wh_last)
    bal_row = conn.execute(
      f"""
      SELECT balance
      FROM bank_transactions
      {where_sql_last}
      ORDER BY COALESCE(trdt, trdate) DESC, tid DESC
      LIMIT 1
      """,
      params_last,
    ).fetchone()
    last_balance = (
      _storage_amount_to_display(bal_row[0], currency)
      if bal_row and bal_row[0] is not None
      else 0
    )
    out.append(
      {
        "bankCode": bcode,
        "accountNumber": accno,
        "count": cnt,
        "totalIn": _storage_amount_to_display(sum_in, currency),
        "totalOut": _storage_amount_to_display(sum_out, currency),
        "lastBalance": last_balance,
      }
    )

  conn.close()
  return jsonify({"items": out})


from app.blueprints.billing_invoices.routes import bank_activity_status_routes # noqa: E402,F401
