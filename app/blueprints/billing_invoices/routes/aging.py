from __future__ import annotations

import io
import json
from datetime import date, datetime

from flask import Blueprint, current_app, flash, redirect, render_template, request, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.billing.utils import (
  invoice_case_link_filter_sql,
  is_compact_query,
  normalize_case_linked_filter,
  sql_ci_contains_any,
  to_compact,
)
from app.utils.error_logging import report_swallowed_exception

from ..auth import log_audit
from ..db import get_all_business_profiles, get_business_profile, get_db, row_to_dict
from ..services.aging_invoices_service import (
  build_aging_invoices_export,
  build_aging_invoices_result,
  parse_aging_date,
  parse_aging_deposit_info,
)

bp = Blueprint("aging", __name__)


@bp.route("/aging", methods=["GET"])
def aging_report():
  """Aging report for outstanding invoices grouped by client and currency.
  Buckets: current (<=0), 1-30, 31-60, 61-90, >90 days past due.
  Supports HTML plus Excel/JSON export via ?format=xlsx|json.
  """
  # Filters
  bp_raw = (request.args.get("business_profile_id") or "").strip()
  as_of_str = (request.args.get("as_of") or "").strip()
  q = (request.args.get("q") or "").strip()
  is_compact_q = q and is_compact_query(q)
  sort_by = (request.args.get("sort") or "name").strip().lower()
  case_linked = normalize_case_linked_filter(request.args.get("case_linked"))
  try:
    as_of_date = date.fromisoformat(as_of_str) if as_of_str else date.today()
  except Exception:
    as_of_date = date.today()

  # Overdue-only filter: default OFF; enable only when 'overdue_only' truthy param is present
  overdue_only_values = [(v or "").strip().lower() for v in request.args.getlist("overdue_only")]
  if overdue_only_values:
    # Treat any truthy value as enabling overdue_only
    overdue_only = any(v in ("1", "true", "yes", "on") for v in overdue_only_values)
  else:
    overdue_only = False

  # Fetch relevant invoices (outstanding only)
  conn = get_db()
  params = []
  # Outstanding when NOT fully paid AND (either payment pending/unpaid OR billing shows non-finalized group)
  where = [
    "(invoices.payment_status IN ('unpaid','pending') OR invoices.billing_status = 'pre_overdue')"
  ]
  # business_profile filter: when empty or 'all/combined/*' -> combined view (no filter)
  bp_ids = []
  low = (bp_raw or "").lower()
  if low and low not in ("all", "combined", "*"):
    for part in bp_raw.split(","):
      part = part.strip()
      if part.isdigit():
        bp_ids.append(int(part))
  if bp_ids:
    placeholders = ",".join(["?"] * len(bp_ids))
    where.append(f"invoices.business_profile_id IN ({placeholders})")
    params.extend(bp_ids)
  if q and not is_compact_q:
    search_clause, search_params = sql_ci_contains_any(["clients.name", "invoices.number"], q)
    if search_clause:
      where.append(search_clause)
      params.extend(search_params)
  case_link_clause = invoice_case_link_filter_sql(case_linked, "invoices")
  if case_link_clause:
    where.append(case_link_clause)
  where_sql = " WHERE " + " AND ".join(where)

  rows_sql = f"""
   SELECT invoices.id, invoices.client_id, clients.name AS client_name,
       invoices.issue_date, invoices.due_date, invoices.total, invoices.currency, invoices.status,
       invoices.billing_status, invoices.payment_status, invoices.payment_meta
   FROM invoices
   JOIN clients ON clients.id = invoices.client_id
   {where_sql}
  """
  db_rows = conn.execute(rows_sql, params).fetchall()
  all_profiles = get_all_business_profiles()
  # Only set bp_row when a single profile is selected; otherwise None (combined)
  if len(bp_ids) == 1:
    bp_row = get_business_profile(bp_ids[0])
  else:
    bp_row = None

  # Aggregate per (client, currency)
  rows_map = {}
  totals_by_currency = {}

  def _ensure_totals(cur: str):
    if cur not in totals_by_currency:
      totals_by_currency[cur] = {
        "current": 0.0,
        "d1_30": 0.0,
        "d31_60": 0.0,
        "d61_90": 0.0,
        "d91_plus": 0.0,
        "total": 0.0,
      }

  # Build buckets per (client,currency)
  for row in db_rows:
    r = row_to_dict(row)
    cur = (r.get("currency") or "USD").upper()
    try:
      total_amt = float(r.get("total") or 0.0)
    except Exception:
      total_amt = 0.0

    dep = 0.0
    try:
      dep, _ = parse_aging_deposit_info(
        r.get("payment_meta"),
        cur,
        invoice_id=r.get("id"),
        log=current_app.logger,
      )
    except Exception as exc:
      report_swallowed_exception(
        exc,
        context="billing_invoices.aging.aging_report.parse_payment_meta",
        log_key="billing_invoices.aging.aging_report.parse_payment_meta",
        log_window_seconds=300,
      )
      dep = 0.0

    amt = max(0.0, total_amt - float(dep or 0.0))
    if amt <= 0:
      continue

    # Base date for overdue
    default_due = parse_aging_date(r.get("due_date")) or parse_aging_date(r.get("issue_date"))
    if default_due is None:
      default_due = as_of_date
    base_date = default_due
    days_over = (as_of_date - base_date).days
    # pre_overdue -> at least 1 day overdue
    try:
      bs = r.get("billing_status") if "billing_status" in r else r.get("status")
      if (str(bs or "").lower() == "pre_overdue") and days_over <= 0:
        days_over = 1
    except Exception as exc:
      report_swallowed_exception(
        exc,
        context="billing_invoices.aging.aging_report.pre_overdue_adjust",
        log_key="billing_invoices.aging.aging_report.pre_overdue_adjust",
        log_window_seconds=300,
      )

    bucket = (
      "current"
      if days_over <= 0
      else (
        "d1_30"
        if days_over <= 30
        else (
          "d31_60" if days_over <= 60 else ("d61_90" if days_over <= 90 else "d91_plus")
        )
      )
    )

    key = (r["client_id"], cur)
    if key not in rows_map:
      rows_map[key] = {
        "client_id": r["client_id"],
        "client_name": r["client_name"],
        "currency": cur,
        "current": 0.0,
        "d1_30": 0.0,
        "d31_60": 0.0,
        "d61_90": 0.0,
        "d91_plus": 0.0,
        "total": 0.0,
        "count": 0,
      }

    if overdue_only and bucket == "current":
      continue

    rows_map[key][bucket] += amt
    rows_map[key]["total"] += amt
    rows_map[key]["count"] += 1

    _ensure_totals(cur)
    totals_by_currency[cur][bucket] += amt
    totals_by_currency[cur]["total"] += amt

  rows = list(rows_map.values())
  if overdue_only:
    rows = [r for r in rows if (r["d1_30"] + r["d31_60"] + r["d61_90"] + r["d91_plus"]) > 0]

  # -only Search:  Name to Filters  
  if q and is_compact_q:
    q_compact = to_compact(q)
    rows = [r for r in rows if q_compact in to_compact(str(r.get("client_name") or ""))]
    totals_by_currency = {}
    for r in rows:
      cur = r["currency"]
      _ensure_totals(cur)
      totals_by_currency[cur]["current"] += float(r.get("current") or 0.0)
      totals_by_currency[cur]["d1_30"] += float(r.get("d1_30") or 0.0)
      totals_by_currency[cur]["d31_60"] += float(r.get("d31_60") or 0.0)
      totals_by_currency[cur]["d61_90"] += float(r.get("d61_90") or 0.0)
      totals_by_currency[cur]["d91_plus"] += float(r.get("d91_plus") or 0.0)
      totals_by_currency[cur]["total"] += float(r.get("total") or 0.0)

  # Sorting
  if sort_by == "overdue_desc":
    rows.sort(
      key=lambda r: (
        r["d1_30"] + r["d31_60"] + r["d61_90"] + r["d91_plus"],
        (r["client_name"] or "").lower(),
        r["currency"],
      ),
      reverse=True,
    )
  else:
    rows.sort(key=lambda r: ((r["client_name"] or "").lower(), r["currency"]))

  # Close DB connection now that processing is complete
  try:
    conn.close()
  except Exception as exc:
    report_swallowed_exception(
      exc,
      context="billing_invoices.aging.aging_report.close_conn",
      log_key="billing_invoices.aging.aging_report.close_conn",
      log_window_seconds=300,
    )

  fmt = (request.args.get("format") or "").lower()
  ts = datetime.now().strftime("%Y%m%d%H%M%S")

  # Exports
  if fmt in ("json", "xlsx"):
    try:
      log_audit(
        "invoice.aging_export",
        "invoice",
        None,
        f'{{"rows": {len(rows)}, "format": "{fmt}"}}',
      )
    except Exception as exc:
      report_swallowed_exception(
        exc,
        context="billing_invoices.aging.aging_report.log_audit_export",
        log_key="billing_invoices.aging.aging_report.log_audit_export",
        log_window_seconds=300,
      )
    if fmt == "json":
      payload = json.dumps(
        {
          "as_of": as_of_date.isoformat(),
          "overdue_only": overdue_only,
          "case_linked": case_linked,
          "q": q,
          "sort": sort_by,
          "rows": rows,
          "totals_by_currency": totals_by_currency,
        },
        ensure_ascii=False,
      )
      resp = current_app.response_class(payload, mimetype="application/json; charset=utf-8")
      resp.headers["Content-Disposition"] = f'attachment; filename="aging_{ts}.json"'
      return resp
    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Aging Report"

    headers = [
      "Client ID",
      "Client name",
      "Currency",
      "Current",
      "1-30",
      "31-60",
      "61-90",
      ">90",
      "Total",
      "Count",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
      cell.fill = header_fill
      cell.font = header_font

    for v in rows:
      ws.append(
        [
          v["client_id"],
          v["client_name"],
          v["currency"],
          v["current"],
          v["d1_30"],
          v["d31_60"],
          v["d61_90"],
          v["d91_plus"],
          v["total"],
          v["count"],
        ]
      )

    for column in ws.columns:
      max_length = 0
      column_letter = column[0].column_letter
      for cell in column:
        value_len = len(str(getattr(cell, "value", "") or ""))
        if value_len > max_length:
          max_length = value_len
      adjusted_width = min(max_length + 2, 50)
      ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = current_app.response_class(
      output.getvalue(),
      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="aging_{ts}.xlsx"'
    return resp

  return render_template(
    "invoices_aging.html",
    bp=bp_row,
    all_profiles=all_profiles,
    as_of=as_of_date.isoformat(),
    rows=rows,
    totals_by_currency=totals_by_currency,
    overdue_only=overdue_only,
    case_linked=case_linked,
    sort_by=sort_by,
    q=q,
  )


@bp.route("/aging/details", methods=["GET"])
def aging_details():
  """Per-invoice aging details for a client (and optional currency)."""
  client_id = request.args.get("client_id", type=int)
  if not client_id:
    return "client_id is required", 400
  currency = (request.args.get("currency") or "").strip().upper()
  bp_id = request.args.get("business_profile_id", type=int)
  as_of = (request.args.get("as_of") or "").strip()
  overdue_only_values = request.args.getlist("overdue_only")
  overdue_only = any(str(v).lower() in ("1", "true", "yes", "on") for v in overdue_only_values)
  case_linked = normalize_case_linked_filter(request.args.get("case_linked"))
  try:
    as_of_date = date.fromisoformat(as_of) if as_of else date.today()
  except (TypeError, ValueError):
    as_of_date = date.today()

  conn = get_db()
  where = [
    "invoices.client_id = ?",
    "(invoices.payment_status IN ('unpaid','pending') OR invoices.billing_status = 'pre_overdue')",
  ]
  params = [client_id]
  if bp_id:
    where.append("invoices.business_profile_id = ?")
    params.append(bp_id)
  if currency:
    where.append("UPPER(invoices.currency) = ?")
    params.append(currency)
  case_link_clause = invoice_case_link_filter_sql(case_linked, "invoices")
  if case_link_clause:
    where.append(case_link_clause)
  where_sql = " WHERE " + " AND ".join(where)

  sql = f"""
   SELECT invoices.id, invoices.number, invoices.issue_date, invoices.due_date,
       invoices.total, invoices.tax, invoices.currency, invoices.status,
       invoices.payment_meta, invoices.tax_issued_at,
       clients.name AS client_name,
       (
        SELECT COALESCE(SUM(qty * unit_price * (1 - discount/100.0)), 0)
        FROM line_items
        WHERE line_items.invoice_id = invoices.id
         AND line_items.item_type = 'service'
         AND (line_items.is_estimated IS NULL OR line_items.is_estimated = 0)
       ) AS service_total,
       (
        SELECT COALESCE(SUM(qty * unit_price * (1 - discount/100.0)), 0)
        FROM line_items
        WHERE line_items.invoice_id = invoices.id
         AND line_items.item_type = 'admin'
         AND (line_items.is_estimated IS NULL OR line_items.is_estimated = 0)
       ) AS admin_total,
       (
        SELECT COALESCE(SUM(qty * unit_price * (1 - discount/100.0)), 0)
        FROM line_items
        WHERE line_items.invoice_id = invoices.id
         AND line_items.item_type = 'foreign'
         AND (line_items.is_estimated IS NULL OR line_items.is_estimated = 0)
       ) AS foreign_total
   FROM invoices
   JOIN clients ON clients.id = invoices.client_id
   {where_sql}
   ORDER BY invoices.issue_date DESC, invoices.id DESC
  """

  rows = []
  cur = conn.cursor()
  for r in cur.execute(sql, params).fetchall():
    d = row_to_dict(r)
    cur_code = (d.get("currency") or "USD").upper()
    # Aggregates
    addl = float(d.get("admin_total") or 0.0) + float(d.get("foreign_total") or 0.0)
    d["additional_charges"] = addl
    # Deposit info (all currencies)
    dep = 0.0
    dep_date = ""
    try:
      dep, dep_date = parse_aging_deposit_info(
        d.get("payment_meta"),
        cur_code,
        invoice_id=d.get("id"),
        log=current_app.logger,
      )
    except Exception as exc:
      report_swallowed_exception(
        exc,
        context="billing_invoices.aging.aging_details.parse_payment_meta",
        log_key="billing_invoices.aging.aging_details.parse_payment_meta",
        log_window_seconds=300,
      )
    d["deposit_amount"] = dep
    d["deposit_date"] = dep_date
    try:
      d["outstanding"] = max(0.0, float(d.get("total") or 0.0) - float(dep or 0.0))
    except Exception:
      d["outstanding"] = float(d.get("total") or 0.0)
    if d["outstanding"] <= 0:
      continue
    if overdue_only:
      default_due = parse_aging_date(d.get("due_date")) or parse_aging_date(
        d.get("issue_date")
      )
      base_date = default_due or as_of_date
      days_over = (as_of_date - base_date).days
      if (
        str(d.get("billing_status") or d.get("status") or "").lower() == "pre_overdue"
        and days_over <= 0
      ):
        days_over = 1
      if days_over <= 0:
        continue
    rows.append(d)

  # Client/business profile rows for header context
  bp_row = get_business_profile(bp_id) if bp_id else None
  client_name = None
  try:
    client_name = conn.execute("SELECT name FROM clients WHERE id=?", (client_id,)).fetchone()[
      0
    ]
  except Exception:
    client_name = None
  conn.close()

  try:
    log_audit(
      "invoice.aging_details",
      "invoice",
      None,
      json.dumps(
        {"client_id": client_id, "currency": currency, "count": len(rows)},
        ensure_ascii=False,
      ),
    )
  except Exception as exc:
    report_swallowed_exception(
      exc,
      context="billing_invoices.aging.aging_details.log_audit",
      log_key="billing_invoices.aging.aging_details.log_audit",
      log_window_seconds=300,
    )

  return render_template(
    "invoices_aging_details.html",
    rows=rows,
    client_name=client_name,
    currency=currency,
    bp=bp_row,
    as_of=as_of or "",
    overdue_only=overdue_only,
    case_linked=case_linked,
  )


@bp.route("/aging/mark_pre_overdue/<int:invoice_id>", methods=["POST"])
def aging_mark_pre_overdue(invoice_id: int):
  """Mark a single invoice as pre_overdue from aging details view."""
  conn = get_db()
  updated = False
  try:
    row = conn.execute(
      "SELECT id, status, billing_status, payment_status FROM invoices WHERE id=?",
      (invoice_id,),
    ).fetchone()
    invoice = row_to_dict(row)
    if not invoice:
      flash("Invoice not found.", "warning")
    else:
      status = str(invoice.get("status") or "").lower()
      billing_status = str(invoice.get("billing_status") or "").lower()
      payment_status = str(invoice.get("payment_status") or "").lower()
      if status in {"paid", "void"} or payment_status == "paid":
        flash("Done/Cancel Invoice Advanced costto change target not available.", "warning")
      elif status == "pre_overdue" and billing_status == "pre_overdue":
        flash(" Advanced costto Display Invoice.", "info")
      else:
        conn.execute(
          "UPDATE invoices SET status=?, billing_status=? WHERE id=?",
          ("pre_overdue", "pre_overdue", invoice_id),
        )
        conn.commit()
        updated = True
        flash("Advanced costto Display.", "success")
  finally:
    try:
      conn.close()
    except Exception as exc:
      report_swallowed_exception(
        exc,
        context="billing_invoices.aging.aging_mark_pre_overdue.close_conn",
        log_key="billing_invoices.aging.aging_mark_pre_overdue.close_conn",
        log_window_seconds=300,
      )

  if updated:
    try:
      log_audit(
        "invoice.status_change",
        "invoice",
        invoice_id,
        '{"new_status":"pre_overdue","source":"aging"}',
      )
    except Exception as exc:
      report_swallowed_exception(
        exc,
        context="billing_invoices.aging.aging_mark_pre_overdue.log_audit",
        log_key="billing_invoices.aging.aging_mark_pre_overdue.log_audit",
        log_window_seconds=300,
      )

  # Redirect back to details
  client_id = request.form.get("client_id", type=int)
  currency = (request.form.get("currency") or "").strip()
  bp_id = request.form.get("business_profile_id", type=int)
  as_of = (request.form.get("as_of") or "").strip()
  overdue_only = any(
    str(v).lower() in ("1", "true", "yes", "on") for v in request.form.getlist("overdue_only")
  )
  case_linked = normalize_case_linked_filter(request.form.get("case_linked"))
  if not client_id:
    return redirect(
      url_for(
        "billing_invoices.aging.aging_report",
        business_profile_id=bp_id,
        as_of=as_of,
        overdue_only=1 if overdue_only else 0,
        case_linked=case_linked,
      )
    )
  return redirect(
    url_for(
      "billing_invoices.aging.aging_details",
      client_id=client_id,
      currency=currency,
      business_profile_id=bp_id,
      as_of=as_of,
      overdue_only=1 if overdue_only else 0,
      case_linked=case_linked,
    )
  )


@bp.route("/aging/invoices", methods=["GET"])
def aging_invoices():
  """Invoice-level list of outstanding (current + overdue) with option to show overdue-only.
  Includes deposits (USD) and tax-document record date.
  """
  bp_raw = (request.args.get("business_profile_id") or "").strip()
  as_of_str = (request.args.get("as_of") or "").strip()
  if not as_of_str:
    as_of_str = date.today().isoformat()
  q = (request.args.get("q") or "").strip()
  is_compact_q = q and is_compact_query(q)
  sort_by = (request.args.get("sort") or "issue_date").strip().lower()
  case_linked = normalize_case_linked_filter(request.args.get("case_linked"))
  overdue_only_values = request.args.getlist("overdue_only")
  overdue_only = any(str(v).lower() in ("1", "true", "yes", "on") for v in overdue_only_values)
  try:
    as_of_date = date.fromisoformat(as_of_str)
  except (ValueError, TypeError):
    as_of_date = date.today()
    as_of_str = as_of_date.isoformat()

  bp_ids = []
  low = (bp_raw or "").lower()
  if low and low not in ("all", "combined", "*"):
    for part in bp_raw.split(","):
      part = part.strip()
      if part.isdigit():
        bp_ids.append(int(part))

  result = build_aging_invoices_result(
    bp_ids=bp_ids,
    q=q,
    is_compact_q=is_compact_q,
    as_of_date=as_of_date,
    overdue_only=overdue_only,
    case_linked=case_linked,
    sort_by=sort_by,
    log=current_app.logger,
  )

  fmt = (request.args.get("format") or "").lower()
  export = build_aging_invoices_export(
    rows=result.rows,
    fmt=fmt,
    as_of_date=as_of_date,
    overdue_only=overdue_only,
    case_linked=case_linked,
    q=q,
    sort_by=sort_by,
    log=current_app.logger,
  )
  if export is not None:
    resp = current_app.response_class(export.payload, mimetype=export.mimetype)
    resp.headers["Content-Disposition"] = f'attachment; filename="{export.filename}"'
    return resp

  return render_template(
    "invoices_aging_invoices.html",
    rows=result.rows,
    bp=result.bp_row,
    all_profiles=result.all_profiles,
    as_of=as_of_date.isoformat(),
    overdue_only=overdue_only,
    case_linked=case_linked,
    sort_by=sort_by,
    q=q,
  )
