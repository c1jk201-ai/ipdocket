from __future__ import annotations

import io
import json
from datetime import datetime
from decimal import Decimal
from zoneinfo import ZoneInfo

from flask import current_app, flash, redirect, render_template, request, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.billing.utils import is_compact_query, sql_ci_contains_any, to_compact
from app.utils.error_logging import report_swallowed_exception
from app.utils.url_helpers import safe_referrer_path

from ..auth import log_audit
from ..db import (
  _get_column_names,
  get_all_business_profiles,
  get_db,
  row_get,
  row_to_dict,
  safe_json_parse,
)
from ..settlement import is_default_settlement_split
from ..services.tax_issue_readiness_service import evaluate_tax_issue_readiness
from .invoices import (
  _BILLING_TRANSITIONS,
  _ensure_external_invoice_case_map,
  _get_settlement_share_ratio,
  _outgoing_invoice_filter_sql,
  _parse_int_amount_usd,
  _resolve_billing_status,
  _resolve_payment_status,
  _safe_float,
  _safe_int,
  _sync_legacy_status,
  _transition_allowed,
  bp,
)


def _digits_only(value: object) -> str:
  return "".join(ch for ch in str(value or "") if ch.isdigit())


def _looks_like_business_registration_number(value: object) -> bool:
  return len(_digits_only(value)) == 10


def _first_business_registration_number(*values: object) -> str:
  for value in values:
    text = str(value or "").strip()
    if text and _looks_like_business_registration_number(text):
      return text
  return ""


def _client_is_individual(value: object) -> bool:
  normalized = str(value or "").strip().lower()
  return normalized in {"individual", "person", "personal", "items"}


def _json_export_default(value: object) -> object:
  if isinstance(value, Decimal):
    return float(value)
  if isinstance(value, datetime):
    return value.isoformat()
  return str(value)


def _first_personal_registration_number(client_type: object, *values: object) -> str:
  if not _client_is_individual(client_type):
    return ""
  for value in values:
    text = str(value or "").strip()
    if text and not _looks_like_business_registration_number(text):
      return text
  return ""


@bp.route("/tax_issue", methods=["GET"])
def tax_issue():
  """Tax documentation queue page, including payment-complete filters."""
  all_profiles = get_all_business_profiles()
  try:
    currencies = sorted({((p["currency"] or "USD").upper()) for p in all_profiles})
  except Exception:
    currencies = ["USD"]

  bp_id = request.args.get("business_profile_id", "").strip()
  if not bp_id and "USD" in currencies:
    bp_id = "C:USD"
  q = request.args.get("q", "").strip()
  is_compact_q = q and is_compact_query(q)
  sort = request.args.get("sort", "issue_date").strip()
  date_from = request.args.get("date_from", "").strip()
  date_to = request.args.get("date_to", "").strip()
  min_amount = request.args.get("min_amount", "").strip()
  max_amount = request.args.get("max_amount", "").strip()

  if sort not in {"issue_date", "due_date", "total"}:
    sort = "issue_date"

  conn = get_db()
  _ensure_external_invoice_case_map(conn)
  client_cols = _get_column_names(conn, "clients")
  client_registration_expr = (
    "COALESCE(c.registration_number, '')" if "registration_number" in client_cols else "''"
  )
  client_biz_reg_expr = (
    "COALESCE(c.biz_reg_number, '')" if "biz_reg_number" in client_cols else "''"
  )
  client_type_expr = "COALESCE(c.type, '')" if "type" in client_cols else "''"
  client_extra_expr = "c.extra" if "extra" in client_cols else "NULL"
  client_address_expr = "COALESCE(c.address, '')" if "address" in client_cols else "''"
  client_biz_business_location_expr = (
    "COALESCE(c.biz_business_location, '')" if "biz_business_location" in client_cols else "''"
  )
  client_biz_head_office_location_expr = (
    "COALESCE(c.biz_head_office_location, '')"
    if "biz_head_office_location" in client_cols
    else "''"
  )
  params, where = [], []
  where.append(
    "(i.payment_status = 'paid' OR i.status = 'paid' OR COALESCE(i.payment_verified,0)=1)"
  )
  where.append(
    "(i.billing_status IS NULL OR i.billing_status NOT IN ('tax_issued', 'cash_issued', 'processed'))"
  )
  where.append("(i.status IS NULL OR i.status NOT IN ('tax_issued', 'cash_issued', 'processed'))")
  where.append(
    "EXISTS (SELECT 1 FROM line_items li WHERE li.invoice_id = i.id "
    "AND (li.item_type = 'service' OR (li.item_type = 'foreign' AND COALESCE(li.is_taxable,0)=1)))"
  )

  if q and not is_compact_q:
    search_clause, search_params = sql_ci_contains_any(
      ["i.number", "i.internal_reference", "c.name", "i.notes"],
      q,
    )
    if search_clause:
      where.append(search_clause)
      params += search_params

  if bp_id:
    if bp_id.upper().startswith("C:"):
      where.append("i.currency = ?")
      params.append(bp_id.split(":", 1)[1].upper())
    else:
      where.append("i.business_profile_id = ?")
      bp_id_int = _safe_int(bp_id, None)
      if bp_id_int is not None:
        params.append(int(bp_id_int))

  if date_from:
    where.append("i.issue_date >= ?")
    params.append(date_from)
  if date_to:
    where.append("i.issue_date <= ?")
    params.append(date_to)

  min_amount_v = _safe_float(min_amount, None)
  max_amount_v = _safe_float(max_amount, None)

  where_sql = (" WHERE " + " AND ".join(where)) if where else ""
  if sort == "issue_date":
    order_clause = "i.issue_date DESC, i.id DESC"
  elif sort == "due_date":
    order_clause = "i.due_date DESC, i.id DESC"
  elif sort == "total":
    order_clause = "i.total DESC, i.id DESC"
  else:
    order_clause = "i.id DESC"

  sql = f"""
    SELECT i.id, i.number, i.internal_reference, i.issue_date, i.currency, i.payment_meta, i.vat_rate,
        i.status, i.billing_status, i.payment_status, COALESCE(i.payment_verified,0) AS payment_verified,
        COALESCE(bp.name, '') AS business_name,
        COALESCE(c.name, '') AS client_name,
        c.id AS client_id,
        {client_type_expr} AS client_type,
        {client_biz_reg_expr} AS client_biz_reg_number,
        {client_registration_expr} AS client_registration_number,
        {client_extra_expr} AS client_extra,
        {client_address_expr} AS client_address,
        {client_biz_business_location_expr} AS client_biz_business_location,
        {client_biz_head_office_location_expr} AS client_biz_head_office_location,
        COALESCE(c.biz_tax_invoice_email, '') AS tax_email,
        COALESCE(c.email, '') AS client_email,
        COALESCE(it.service_total, 0) AS service_total,
        COALESCE(it.admin_total, 0) AS admin_total,
        COALESCE(it.foreign_total, 0) AS foreign_total,
        COALESCE(it.foreign_taxable_total, 0) AS foreign_taxable_total
    FROM invoices i
    LEFT JOIN business_profile bp ON bp.id = i.business_profile_id
    LEFT JOIN clients c ON c.id = i.client_id
    LEFT JOIN (
      SELECT li.invoice_id AS invoice_id,
          COALESCE(SUM(CASE
            WHEN li.item_type = 'service' AND COALESCE(li.is_estimated,0) = 0
            THEN (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
            ELSE 0
          END), 0) AS service_total,
          COALESCE(SUM(CASE
            WHEN li.item_type = 'admin' AND COALESCE(li.is_estimated,0) = 0
            THEN (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
            ELSE 0
          END), 0) AS admin_total,
          COALESCE(SUM(CASE
            WHEN li.item_type = 'foreign' AND COALESCE(li.is_estimated,0) = 0 THEN
              CASE WHEN COALESCE(li.fx_rate_used, 0) > 0 THEN
                (COALESCE(li.fx_fee,0) + COALESCE(li.fx_gov,0))
                * COALESCE(li.fx_rate_used, 0)
                * (1 + COALESCE(li.fx_markup,0)/100.0)
              ELSE
                (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
              END
            ELSE 0
          END), 0) AS foreign_total,
          COALESCE(SUM(CASE
            WHEN li.item_type = 'foreign'
            AND COALESCE(li.is_estimated,0) = 0
            AND COALESCE(li.is_taxable,0) = 1 THEN
              CASE WHEN COALESCE(li.fx_rate_used, 0) > 0 THEN
                (COALESCE(li.fx_fee,0) + COALESCE(li.fx_gov,0))
                * COALESCE(li.fx_rate_used, 0)
                * (1 + COALESCE(li.fx_markup,0)/100.0)
              ELSE
                (li.qty * li.unit_price * (1 - COALESCE(li.discount,0)/100.0))
              END
            ELSE 0
          END), 0) AS foreign_taxable_total
       FROM line_items li
       GROUP BY li.invoice_id
    ) AS it ON it.invoice_id = i.id
    {where_sql}
    ORDER BY {order_clause}"""
  rows = conn.execute(sql, params).fetchall()

  invoices = []
  for r in rows:
    inv = row_to_dict(r)
    client_extra = safe_json_parse(inv.get("client_extra"), {})
    if not isinstance(client_extra, dict):
      client_extra = {}
    inv["client_biz_reg_number"] = _first_business_registration_number(
      inv.get("client_biz_reg_number"),
      client_extra.get("business_reg_no"),
      client_extra.get("tax_business_reg_no"),
      client_extra.get("reg_number"),
      client_extra.get("biz_reg_number"),
      inv.get("client_registration_number"),
    )
    inv["client_personal_registration_number"] = _first_personal_registration_number(
      inv.get("client_type"),
      inv.get("client_registration_number"),
      client_extra.get("personal_registration_number"),
      client_extra.get("resident_registration_number"),
      client_extra.get("resident_reg_no"),
      client_extra.get("rrn"),
    )
    inv["client_tax_address"] = (
      str(client_extra.get("tax_address") or "").strip()
      or str(inv.get("client_biz_business_location") or "").strip()
      or str(inv.get("client_biz_head_office_location") or "").strip()
      or str(inv.get("client_address") or "").strip()
    )
    readiness = evaluate_tax_issue_readiness(conn, inv)
    inv["tax_issue_ready"] = readiness.ready
    inv["tax_issue_blockers"] = readiness.reasons
    inv["tax_issue_blocker_text"] = " / ".join(readiness.reasons)
    inv["tax_issue_case_matched"] = readiness.case_matched
    inv["tax_issue_payment_verified"] = readiness.payment_verified
    deposit = 0
    deposit_date = ""
    try:
      cur_code = (inv.get("currency") or "USD").upper()
      if inv.get("payment_meta") and cur_code == "USD":
        meta = json.loads(inv.get("payment_meta"))
        try:
          deposit = _parse_int_amount_usd(meta.get("deposit"))
        except Exception:
          deposit = 0
        deposit_date = str(meta.get("date") or "")
    except Exception as exc:
      report_swallowed_exception(
        exc,
        context="billing_invoices.invoices_export.tax_issue.parse_payment_meta",
        log_key="billing_invoices.invoices_export.tax_issue.parse_payment_meta",
        log_window_seconds=300,
      )
    inv["deposit_amount"] = deposit
    inv["deposit_date"] = deposit_date

    items_sql = """
      SELECT description, qty, unit_price, COALESCE(discount,0) AS discount,
          (qty * unit_price * (1 - COALESCE(discount,0)/100.0)) AS amount
      FROM line_items
      WHERE invoice_id = ? AND (item_type = 'service' OR (item_type = 'foreign' AND COALESCE(is_taxable,0)=1))
      ORDER BY id
      """
    inv_items = [
      {
        "description": ir[0],
        "qty": ir[1],
        "unit_price": ir[2],
        "discount": ir[3],
        "amount": ir[4],
      }
      for ir in conn.execute(items_sql, (inv["id"],)).fetchall()
    ]
    inv["items"] = inv_items

    try:
      inv["taxable_subtotal"] = sum((it.get("amount") or 0.0) for it in inv_items)
      vr = float(inv.get("vat_rate") or 0.0)
      vm = (vr / 100.0) if vr > 1 else vr
      taxable_vat = vm * float(inv["taxable_subtotal"] or 0.0)
      inv["taxable_vat"] = taxable_vat
      inv["taxable_total_incl_vat"] = float(inv["taxable_subtotal"]) + taxable_vat
      inv["taxable_total_incl_vat_items_sum"] = sum(
        (float(it.get("amount") or 0.0) * (1.0 + vm)) for it in inv_items
      )
      svc = float(inv.get("service_total") or 0.0)
      adm = float(inv.get("admin_total") or 0.0)
      frn = float(inv.get("foreign_total") or 0.0)
      frn_taxable = float(inv.get("foreign_taxable_total") or 0.0)
      tax_dynamic = vm * (svc + frn_taxable)
      inv["total"] = svc + adm + frn + tax_dynamic
    except Exception:
      inv["taxable_subtotal"] = 0.0
      inv["taxable_vat"] = 0.0
      inv["taxable_total_incl_vat"] = 0.0
      inv["taxable_total_incl_vat_items_sum"] = 0.0
      inv["total"] = 0.0

    invoices.append(inv)

  if is_compact_q:
    q_compact = to_compact(q)
    filtered = []
    for inv in invoices:
      text = " ".join(
        [
          str(inv.get("number") or ""),
          str(inv.get("internal_reference") or ""),
          str(inv.get("client_name") or ""),
          str(inv.get("notes") or ""),
        ]
      )
      if q_compact in to_compact(text):
        filtered.append(inv)
    invoices = filtered

  if min_amount_v is not None:
    invoices = [
      inv for inv in invoices if float(inv.get("total") or 0.0) >= float(min_amount_v)
    ]
  if max_amount_v is not None:
    invoices = [
      inv for inv in invoices if float(inv.get("total") or 0.0) <= float(max_amount_v)
    ]

  if sort == "total":
    invoices.sort(key=lambda x: float(x.get("total") or 0.0), reverse=True)

  conn.close()
  ready_count = sum(1 for inv in invoices if inv.get("tax_issue_ready"))
  blocked_count = len(invoices) - ready_count
  return render_template(
    "invoices_tax_issue.html",
    invoices=invoices,
    all_profiles=all_profiles,
    currencies=currencies,
    selected_business_profile_id=bp_id,
    ready_count=ready_count,
    blocked_count=blocked_count,
  )


def _current_tax_issue_confirmed_at() -> str:
  try:
    return datetime.now(ZoneInfo(current_app.config.get("TIMEZONE", "America/New_York"))).isoformat(
      timespec="seconds"
    )
  except Exception:
    return datetime.now().isoformat(timespec="seconds")


@bp.route("/tax_issue/confirm", methods=["POST"])
def confirm_tax_issue():
  """Mark selected tax-document targets as completed after manual review."""
  raw_ids = request.form.getlist("invoice_ids[]")
  invoice_ids: list[int] = []
  for raw_id in raw_ids:
    try:
      invoice_id = int(str(raw_id or "").strip())
    except (TypeError, ValueError):
      continue
    if invoice_id > 0 and invoice_id not in invoice_ids:
      invoice_ids.append(invoice_id)

  if not invoice_ids:
    flash("Confirm Process Select an invoice.", "warning")
    return redirect(safe_referrer_path() or url_for("billing_invoices.invoices.tax_issue"))

  conn = get_db()
  _ensure_external_invoice_case_map(conn)
  placeholders = ",".join(["?"] * len(invoice_ids))
  rows = conn.execute(
    f"""
    SELECT id, number, status, billing_status, payment_status, payment_verified, currency
    FROM invoices
    WHERE id IN ({placeholders})
    """,
    invoice_ids,
  ).fetchall()
  invoice_map = {int(row["id"]): row_to_dict(row) for row in rows}

  updated_ids: list[int] = []
  updated_numbers: list[str] = []
  already_numbers: list[str] = []
  skipped_details: list[str] = []

  for invoice_id in invoice_ids:
    row = invoice_map.get(invoice_id)
    if not row:
      skipped_details.append(f"ID {invoice_id}: Invoice ")
      continue

    number = str(row.get("number") or invoice_id)
    readiness = evaluate_tax_issue_readiness(conn, row)
    current_billing = _resolve_billing_status(row)
    if current_billing == "tax_issued":
      already_numbers.append(number)
      continue
    if not readiness.ready:
      reason = " / ".join(readiness.reasons) or " items"
      skipped_details.append(f"{number}: {reason}")
      continue
    if not _transition_allowed(current_billing, "tax_issued", _BILLING_TRANSITIONS, "draft"):
      skipped_details.append(f"{number}: status change ")
      continue
    now = _current_tax_issue_confirmed_at()
    conn.execute(
      """
      UPDATE invoices
      SET tax_issued_at=?,
        billing_status='tax_issued',
        tax_issue_type='tax_invoice',
        tax_issue_source='tax_issue_page',
        tax_issue_note=NULL
      WHERE id=?
      """,
      (now, invoice_id),
    )
    _sync_legacy_status(
      conn,
      invoice_id,
      billing_status="tax_issued",
      payment_status=_resolve_payment_status(row),
    )
    updated_ids.append(invoice_id)
    updated_numbers.append(number)

  conn.commit()
  conn.close()

  log_audit(
    "invoice.tax_issue_page_confirm",
    "invoice",
    None,
    json.dumps(
      {
        "count": len(updated_ids),
        "invoice_ids": updated_ids,
        "invoice_numbers": updated_numbers,
        "already_numbers": already_numbers[:100],
        "skipped_preview": skipped_details[:20],
      },
      ensure_ascii=False,
    ),
  )

  if updated_ids:
    try:
      from app.services.billing.invoice_manager_followup_service import (
        maybe_notify_manager_followup_for_invoice,
      )

      for updated_id in updated_ids:
        try:
          maybe_notify_manager_followup_for_invoice(
            action="invoice.tax_issued",
            invoice_id=int(updated_id),
          )
        except Exception:
          continue
    except Exception as exc:
      report_swallowed_exception(
        exc,
        context="billing_invoices.invoices_export.confirm_tax_issue.followup_notice",
        log_key="billing_invoices.invoices_export.confirm_tax_issue.followup_notice",
        log_window_seconds=300,
      )

  if updated_numbers:
    flash(f"Recorded tax documentation for {len(updated_numbers)} invoice(s).", "success")
  else:
    flash("New Mark tax recorded Invoice not available.", "warning")
  if already_numbers:
    flash(
      f" tax-recorded status Invoice {len(already_numbers)}items .",
      "warning",
    )
  if skipped_details:
    preview = " / ".join(skipped_details[:5])
    suffix = " / ..." if len(skipped_details) > 5 else ""
    flash(f"items Item {len(skipped_details)}items: {preview}{suffix}", "warning")

  return redirect(safe_referrer_path() or url_for("billing_invoices.invoices.tax_issue"))


@bp.route("/export", methods=["GET"])
def export_invoices():
  """Export currently filtered invoices as Excel or JSON."""
  bp_id = request.args.get("business_profile_id", "").strip()
  basis = (request.args.get("basis") or "issued").strip().lower()
  if basis not in ("issued", "settlement"):
    basis = "issued"
  q = request.args.get("q", "").strip()
  is_compact_q = q and is_compact_query(q)
  status = request.args.get("status", "").strip()
  outgoing_filter = request.args.get("outgoing", "").strip()
  sort = request.args.get("sort", "issue_date").strip()
  date_from = request.args.get("date_from", "").strip()
  date_to = request.args.get("date_to", "").strip()
  min_amount = request.args.get("min_amount", "").strip()
  max_amount = request.args.get("max_amount", "").strip()
  min_amount_v = _safe_float(min_amount, None) if min_amount else None
  max_amount_v = _safe_float(max_amount, None) if max_amount else None

  if sort not in {"issue_date", "due_date", "total"}:
    sort = "issue_date"

  conn = get_db()
  if outgoing_filter == "1":
    _ensure_external_invoice_case_map(conn)
  params, where = [], []
  use_settlement_filter = False
  target_bp_id = None
  if q and not is_compact_q:
    search_clause, search_params = sql_ci_contains_any(
      ["invoices.number", "invoices.internal_reference", "clients.name", "invoices.notes"],
      q,
    )
    if search_clause:
      where.append(search_clause)
      params += search_params
  if status:
    if status in ("draft", "sent", "void", "tax_issued", "processed", "pre_overdue"):
      where.append("invoices.billing_status = ?")
      params.append(status)
    elif status == "sent_unpaid":
      where.append(
        "(invoices.billing_status = 'sent' AND invoices.payment_status = 'unpaid')"
      )
    elif status == "payment_pending":
      where.append(
        "(invoices.billing_status = 'sent' AND invoices.payment_status = 'pending')"
      )
    elif status == "sent_unpaid_or_pending":
      where.append(
        "(invoices.billing_status = 'sent' AND invoices.payment_status IN ('unpaid','pending'))"
      )
    elif status == "paid":
      where.append("invoices.payment_status = 'paid'")
    elif status == "paid_no_tax":
      where.append(
        "(invoices.payment_status = 'paid' AND (invoices.billing_status IS NULL OR invoices.billing_status NOT IN ('tax_issued','cash_issued','processed')))"
      )
    else:
      where.append("invoices.status = ?")
      params.append(status)
  if outgoing_filter == "1":
    where.append(_outgoing_invoice_filter_sql("invoices"))
  if bp_id:
    if bp_id.upper().startswith("C:"):
      where.append("invoices.currency = ?")
      params.append(bp_id.split(":", 1)[1].upper())
    else:
      bp_id_int = _safe_int(bp_id, None)
      if bp_id_int is not None:
        if basis == "settlement":
          use_settlement_filter = True
          target_bp_id = int(bp_id_int)
        else:
          where.append("invoices.business_profile_id = ?")
          params.append(int(bp_id_int))
  if date_from:
    where.append("invoices.issue_date >= ?")
    params.append(date_from)
  if date_to:
    where.append("invoices.issue_date <= ?")
    params.append(date_to)
  if not (use_settlement_filter and target_bp_id is not None):
    if min_amount_v is not None:
      where.append("invoices.total >= ?")
      params.append(float(min_amount_v))
    if max_amount_v is not None:
      where.append("invoices.total <= ?")
      params.append(float(max_amount_v))

  where_sql = (" WHERE " + " AND ".join(where)) if where else ""
  if sort == "issue_date":
    order_clause = "invoices.issue_date DESC, invoices.id DESC"
  elif sort == "due_date":
    order_clause = "invoices.due_date DESC, invoices.id DESC"
  elif sort == "total":
    order_clause = "invoices.total DESC, invoices.id DESC"
  else:
    order_clause = "invoices.id DESC"

  sql = f"""
   SELECT invoices.*, clients.name as client_name, business_profile.name as business_name
   FROM invoices
   JOIN clients ON clients.id=invoices.client_id
   LEFT JOIN business_profile ON business_profile.id=invoices.business_profile_id
   {where_sql}
   ORDER BY {order_clause}"""
  rows = conn.execute(sql, params).fetchall()
  rows = [row_to_dict(r) for r in rows]

  if is_compact_q:
    q_compact = to_compact(q)
    filtered_rows = []
    for r in rows:
      text = " ".join(
        [
          str(r["number"] or ""),
          str(r["internal_reference"] or ""),
          str(r["client_name"] or ""),
          str(r["notes"] or ""),
        ]
      )
      if q_compact in to_compact(text):
        filtered_rows.append(r)
    rows = filtered_rows

  if use_settlement_filter and target_bp_id is not None:
    adjusted_rows = []
    for r in rows:
      ratio = _get_settlement_share_ratio(r, target_bp_id)
      if ratio <= 0:
        continue
      for amount_key in ("subtotal", "tax", "total"):
        try:
          r[amount_key] = float(r.get(amount_key) or 0.0) * ratio
        except (TypeError, ValueError):
          r[amount_key] = 0.0
      if min_amount_v is not None and float(r.get("total") or 0.0) < float(min_amount_v):
        continue
      if max_amount_v is not None and float(r.get("total") or 0.0) > float(max_amount_v):
        continue
      adjusted_rows.append(r)
    rows = adjusted_rows
    if sort == "total":
      rows.sort(
        key=lambda row: (float(row.get("total") or 0.0), int(row["id"])), reverse=True
      )

  try:
    all_profiles = get_all_business_profiles()
    bp_name_map = {p["id"]: p["name"] for p in all_profiles}
  except Exception:
    bp_name_map = {}

  def _settlement_summary_for_row(r):
    settlement_summary = None
    try:
      meta_s = r["settlement_meta"]
    except Exception:
      meta_s = None
    if meta_s:
      try:
        parsed = json.loads(meta_s)
      except Exception:
        parsed = None
      if isinstance(parsed, list) and not is_default_settlement_split(
        parsed,
        row_get(r, "business_profile_id", default=None),
      ):
        parts = []
        for rec in parsed:
          try:
            bpv = int(rec.get("business_profile_id"))
            pctv = float(rec.get("percent"))
          except Exception:
            continue
          if pctv <= 0:
            continue
          name = bp_name_map.get(bpv) or r["business_name"] or "-"
          try:
            if float(pctv).is_integer():
              pct_str = f"{int(pctv)}%"
            else:
              pct_str = f"{pctv:.1f}%"
          except Exception:
            pct_str = f"{pctv}%"
          parts.append(f"{name} {pct_str}")
        if parts:
          settlement_summary = ", ".join(parts)
    if not settlement_summary:
      try:
        settlement_summary = r["business_name"] or "-"
      except Exception:
        settlement_summary = "-"
    return settlement_summary

  fmt = (request.args.get("format", "xlsx") or "xlsx").lower()
  ts = datetime.now().strftime("%Y%m%d%H%M%S")

  try:
    from ..auth import log_audit

    log_audit(
      "invoice.export",
      "invoice",
      None,
      f'{{"count": {len(rows)}, "format": "{fmt}"}}',
    )
  except Exception as exc:
    report_swallowed_exception(
      exc,
      context="billing_invoices.invoices_export.export.log_audit",
      log_key="billing_invoices.invoices_export.export.log_audit",
      log_window_seconds=300,
    )

  if fmt == "json":
    data = []
    for r in rows:
      data.append(
        {
          "id": r["id"],
          "number": r["number"],
          "internal_reference": r["internal_reference"],
          "issue_date": r["issue_date"],
          "due_date": r["due_date"],
          "status": r["status"],
          "client_name": r["client_name"],
          "business_name": r["business_name"],
          "subtotal": r["subtotal"],
          "tax": r["tax"],
          "total": r["total"],
          "currency": r["currency"],
          "vat_rate": r["vat_rate"],
          "language": row_get(r, "language", default=None),
          "settlement_summary": _settlement_summary_for_row(r),
        }
      )
    payload = json.dumps(data, ensure_ascii=False, indent=2, default=_json_export_default)
    resp = current_app.response_class(payload, mimetype="application/json; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="invoices_export_{ts}.json"'
    return resp

  wb = Workbook()
  ws = wb.active
  ws.title = "Invoices"

  headers = [
    "ID",
    "Number",
    "Internal Reference",
    "Issue Date",
    "Due Date",
    "Status",
    "Client Name",
    "Business Name",
    "Settlement Summary",
    "Subtotal",
    "Tax",
    "Total",
    "Currency",
    "Sales tax Rate",
  ]
  ws.append(headers)

  header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
  header_font = Font(bold=True, color="FFFFFF")
  for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

  for r in rows:
    ws.append(
      [
        r["id"],
        r["number"],
        r["internal_reference"],
        r["issue_date"],
        r["due_date"],
        r["status"],
        r["client_name"],
        r["business_name"],
        _settlement_summary_for_row(r),
        r["subtotal"],
        r["tax"],
        r["total"],
        r["currency"],
        r["vat_rate"],
      ]
    )

  for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
      value_len = len(str(getattr(cell, "value", "") or ""))
      if value_len > max_length:
        max_length = value_len
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

  output = io.BytesIO()
  wb.save(output)
  output.seek(0)

  resp = current_app.response_class(
    output.getvalue(),
    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
  )
  resp.headers["Content-Disposition"] = f'attachment; filename="invoices_export_{ts}.xlsx"'
  return resp
