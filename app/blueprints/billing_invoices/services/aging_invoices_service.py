from __future__ import annotations

import io
import json
import logging
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.billing.utils import to_compact

from ..auth import log_audit
from ..db import get_all_business_profiles, get_business_profile
from ..repos.aging_invoices_repo import fetch_aging_invoices_rows

logger = logging.getLogger(__name__)
EMPTY_AGING_EXPORT_HEADERS = [
  "client_name",
  "invoice_number",
  "currency",
  "service_total",
  "tax",
  "additional_charges",
  "total",
  "deposit_amount",
  "deposit_date",
  "outstanding",
  "tax_issued_at",
  "status",
]


@dataclass(frozen=True)
class AgingInvoicesResult:
  rows: list[dict]
  all_profiles: list[dict]
  bp_row: dict | None


@dataclass(frozen=True)
class AgingInvoicesExport:
  payload: str | bytes
  mimetype: str
  filename: str


def parse_aging_date(value) -> date | None:
  if not value:
    return None
  if isinstance(value, date) and not isinstance(value, datetime):
    return value
  if isinstance(value, datetime):
    return value.date()

  raw = str(value).strip()
  if not raw:
    return None

  # Accept common DB datetime strings (e.g. "YYYY-MM-DD HH:MM:SS").
  if len(raw) >= 10:
    try:
      return date.fromisoformat(raw[:10])
    except ValueError:
      logger.debug("Failed ISO date slice parse for value=%r", value, exc_info=True)

  try:
    normalized = raw.replace("Z", "+00:00")
    return datetime.fromisoformat(normalized).date()
  except ValueError:
    return None


def _to_float(value, default: float = 0.0) -> float:
  try:
    return float(value or 0.0)
  except (TypeError, ValueError):
    return float(default)


def _parse_deposit_amount(raw_amount, cur_code: str) -> float:
  normalized = str(raw_amount or "").replace(",", "").strip()
  if not normalized:
    return 0.0

  try:
    dep_decimal = Decimal(normalized)
  except (InvalidOperation, ValueError, TypeError):
    return 0.0

  if cur_code in ("USD", "JPY"):
    return float(int(dep_decimal))
  return float(dep_decimal)


def parse_aging_deposit_info(
  payment_meta,
  cur_code: str,
  invoice_id: int | None = None,
  log: logging.Logger | None = None,
) -> tuple[float, str]:
  log = log or logger
  if not payment_meta:
    return 0.0, ""

  meta = None
  if isinstance(payment_meta, dict):
    meta = payment_meta
  elif isinstance(payment_meta, str):
    try:
      meta = json.loads(payment_meta)
    except (TypeError, ValueError):
      log.debug(
        "Failed to parse payment_meta for invoice %s",
        invoice_id,
        exc_info=True,
      )
      return 0.0, ""
  else:
    log.debug(
      "Unexpected payment_meta type for invoice %s: %s",
      invoice_id,
      type(payment_meta).__name__,
    )
    return 0.0, ""

  if not isinstance(meta, dict):
    return 0.0, ""

  dep = _parse_deposit_amount(meta.get("deposit"), cur_code)
  if dep == 0.0 and str((meta.get("deposit") if meta else "") or "").strip():
    log.debug(
      "Failed to parse deposit amount for invoice %s",
      invoice_id,
    )
  if dep < 0:
    dep = 0.0

  dep_date = str(meta.get("date") or meta.get("time") or "")
  return dep, dep_date


def _calculate_days_over(row: dict, as_of_date: date) -> int:
  default_due = parse_aging_date(row.get("due_date")) or parse_aging_date(row.get("issue_date"))
  base_date = default_due or as_of_date
  days_over = (as_of_date - base_date).days
  if (row.get("billing_status") or "").lower() == "pre_overdue" and days_over <= 0:
    return 1
  return days_over


def _matches_compact_query(row: dict, q_compact: str) -> bool:
  search_text = " ".join([str(row.get("client_name") or ""), str(row.get("number") or "")])
  return q_compact in to_compact(search_text)


def _sort_aging_rows(rows: list[dict], sort_by: str) -> None:
  if sort_by == "outstanding":
    rows.sort(
      key=lambda r: (
        float(r.get("outstanding") or 0.0),
        int(r.get("days_over") or 0),
        int(r.get("id") or 0),
      ),
      reverse=True,
    )
    return

  if sort_by == "client_name":
    rows.sort(
      key=lambda r: (
        (r.get("client_name") or "").lower(),
        parse_aging_date(r.get("issue_date")) or date.min,
        int(r.get("id") or 0),
      ),
      reverse=False,
    )
    return

  if sort_by == "days_over":
    rows.sort(
      key=lambda r: (
        int(r.get("days_over") or 0),
        float(r.get("outstanding") or 0.0),
        int(r.get("id") or 0),
      ),
      reverse=True,
    )
    return

  rows.sort(
    key=lambda r: (
      parse_aging_date(r.get("issue_date")) or date.min,
      int(r.get("id") or 0),
    ),
    reverse=True,
  )


def _append_excel_header(ws, headers: list[str]) -> None:
  ws.append(headers)
  header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
  header_font = Font(bold=True, color="FFFFFF")
  for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font


def _autosize_excel_columns(ws, *, max_width: int = 50, padding: int = 2) -> None:
  for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
      value_len = len(str(getattr(cell, "value", "") or ""))
      if value_len > max_length:
        max_length = value_len
    ws.column_dimensions[column_letter].width = min(max_length + padding, max_width)


def build_aging_invoices_result(
  *,
  bp_ids: list[int],
  q: str,
  is_compact_q: bool,
  as_of_date: date,
  overdue_only: bool,
  case_linked: str,
  sort_by: str,
  log: logging.Logger | None = None,
) -> AgingInvoicesResult:
  log = log or logger
  raw_rows = fetch_aging_invoices_rows(
    bp_ids=bp_ids,
    q=q,
    is_compact_q=is_compact_q,
    case_linked=case_linked,
  )

  rows: list[dict] = []
  for r in raw_rows:
    d = dict(r)
    cur_code = (d.get("currency") or "USD").upper()
    addl = _to_float(d.get("admin_total")) + _to_float(d.get("foreign_total"))
    d["additional_charges"] = addl

    dep, dep_date = parse_aging_deposit_info(
      d.get("payment_meta"),
      cur_code,
      invoice_id=d.get("id"),
      log=log,
    )
    d["deposit_amount"] = dep
    d["deposit_date"] = dep_date
    d["outstanding"] = max(0.0, _to_float(d.get("total")) - float(dep or 0.0))
    if d["outstanding"] <= 0:
      continue

    d["days_over"] = _calculate_days_over(d, as_of_date)

    rows.append(d)

  if overdue_only:
    rows = [r for r in rows if r.get("days_over", 0) > 0 and r.get("outstanding", 0) > 0]

  if is_compact_q:
    q_compact = to_compact(q)
    rows = [r for r in rows if _matches_compact_query(r, q_compact)]

  _sort_aging_rows(rows, sort_by)

  all_profiles = get_all_business_profiles()
  bp_row = get_business_profile(bp_ids[0]) if len(bp_ids) == 1 else None
  return AgingInvoicesResult(rows=rows, all_profiles=all_profiles, bp_row=bp_row)


def build_aging_invoices_export(
  *,
  rows: list[dict],
  fmt: str,
  as_of_date: date,
  overdue_only: bool,
  case_linked: str,
  q: str,
  sort_by: str,
  log: logging.Logger | None = None,
) -> AgingInvoicesExport | None:
  if fmt not in ("json", "xlsx"):
    return None

  log = log or logger
  ts = datetime.now().strftime("%Y%m%d%H%M%S")

  try:
    log_audit(
      "invoice.aging_invoices_export",
      "invoice",
      None,
      f'{{"rows": {len(rows)}, "format": "{fmt}"}}',
    )
  except Exception:
    log.warning("Failed to log aging_invoices export audit.", exc_info=True)

  if fmt == "json":
    payload = json.dumps(
      {
        "as_of": as_of_date.isoformat(),
        "overdue_only": overdue_only,
        "case_linked": case_linked,
        "q": q,
        "sort": sort_by,
        "rows": rows,
      },
      ensure_ascii=False,
    )
    return AgingInvoicesExport(
      payload=payload,
      mimetype="application/json; charset=utf-8",
      filename=f"aging_invoices_{ts}.json",
    )

  wb = Workbook()
  ws = wb.active
  ws.title = "Aging Invoices"

  if rows:
    header = list(rows[0].keys())
    _append_excel_header(ws, header)
    for r in rows:
      ws.append([r.get(key, "") for key in header])
  else:
    _append_excel_header(ws, EMPTY_AGING_EXPORT_HEADERS)

  _autosize_excel_columns(ws)

  output = io.BytesIO()
  wb.save(output)
  output.seek(0)

  return AgingInvoicesExport(
    payload=output.getvalue(),
    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    filename=f"aging_invoices_{ts}.xlsx",
  )
